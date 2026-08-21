#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CHECKBOOK — Lector inverso: del Excel lleno a filas normalizadas para FinPlan.

Devuelve el detalle completo (cuenta x detalle x mes) y el rollup por cuenta,
para que FinPlan pueda cargar hoy a nivel cuenta y manana a nivel detalle sin
volver a tocar el Excel.

Uso:
    python read_checkbook.py CHECKBOOK_MADRESAL_2027.xlsx --out-dir ./salida
    python read_checkbook.py archivo.xlsx --json      # imprime JSON a stdout
"""

import argparse
import csv
import json
import os
import re
import sys

from openpyxl import load_workbook

COL_MES_INI, COL_MES_FIN, COL_TOTAL = 7, 18, 19
MESES = list(range(1, 13))


def _num(v):
    return v if isinstance(v, (int, float)) else None


def leer(path):
    wb = load_workbook(path, data_only=True)
    hoja = next((n for n in wb.sheetnames if n.upper().startswith('BUDGET')), None)
    if hoja is None:
        raise ValueError(f"No encontre el tab de detalle (BUDGET ... Detail) en {path}")
    ws = wb[hoja]
    m = re.search(r'(\d{4})', hoja)
    anio = int(m.group(1)) if m else None

    depto = ws.cell(3, 3).value
    cod = None

    detalle, referencias, cuentas = [], [], {}
    cuenta_actual = None

    for r in range(15, ws.max_row + 1):
        b, c, d, e = (ws.cell(r, col).value for col in (2, 3, 4, 5))

        # Fila de detalle: # cuenta, departamento y detalle numericos
        if isinstance(b, (int, float)) and isinstance(e, (int, float)):
            cuenta_actual = int(b)
            cod = cod if cod is not None else d
            cuentas.setdefault(int(b), c)
            montos = [_num(ws.cell(r, col).value)
                      for col in range(COL_MES_INI, COL_MES_FIN + 1)]
            desc_det = ws.cell(r, 6).value
            tiene_desc = isinstance(desc_det, str) and desc_det.strip() != ''
            if any(v is not None for v in montos) or tiene_desc:
                detalle.append({
                    'fila': r,
                    'cuenta': int(b),
                    'descripcion_cuenta': c,
                    'departamento': d,
                    'detalle': int(e),
                    'descripcion_detalle': desc_det,
                    'montos': [v or 0 for v in montos],
                    'total': round(sum(v or 0 for v in montos), 2),
                })
            continue

        # Fila TOTAL {anio}: referencia de un anio anterior
        if isinstance(c, str) and c.upper().startswith('TOTAL ') and cuenta_actual:
            m2 = re.search(r'(\d{4})', c)
            if not m2:
                continue
            a = int(m2.group(1))
            if a == anio:
                continue                       # el TOTAL del anio de version es formula
            montos = [_num(ws.cell(r, col).value)
                      for col in range(COL_MES_INI, COL_MES_FIN + 1)]
            if any(v is not None for v in montos):
                referencias.append({
                    'cuenta': cuenta_actual,
                    'anio': a,
                    'montos': [v or 0 for v in montos],
                    'total': round(sum(v or 0 for v in montos), 2),
                })

    # Rollup por cuenta
    rollup = {}
    for row in detalle:
        acc = rollup.setdefault(row['cuenta'], {
            'cuenta': row['cuenta'],
            'descripcion_cuenta': row['descripcion_cuenta'],
            'montos': [0.0] * 12, 'total': 0.0, 'lineas': 0})
        for i in range(12):
            acc['montos'][i] += row['montos'][i]
        acc['total'] = round(acc['total'] + row['total'], 2)
        acc['lineas'] += 1

    estadisticas = {
        'noches_disponibles': [_num(ws.cell(5, col).value) or 0
                               for col in range(COL_MES_INI, COL_MES_FIN + 1)],
        'noches_ocupadas': [_num(ws.cell(6, col).value) or 0
                            for col in range(COL_MES_INI, COL_MES_FIN + 1)],
    }
    estadisticas['ocupacion'] = [
        round(o / dsp, 4) if dsp else 0
        for o, dsp in zip(estadisticas['noches_ocupadas'], estadisticas['noches_disponibles'])]

    gran_total = round(sum(a['total'] for a in rollup.values()), 2)
    cache_gt = _num(ws.cell(9, COL_TOTAL).value)
    cuadra = cache_gt is None or abs(cache_gt - gran_total) < 0.01

    return {
        'archivo': os.path.basename(path),
        'tab': hoja,
        'departamento': depto,
        'codigo_departamento': cod,
        'anio_version': anio,
        'estadisticas': estadisticas,
        'detalle': detalle,
        'por_cuenta': sorted(rollup.values(), key=lambda x: x['cuenta']),
        'referencias': referencias,
        'gran_total': gran_total,
        'gran_total_en_hoja': cache_gt,
        'cuadra': cuadra,
    }


def _csv(path, filas, campos):
    with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
        w = csv.DictWriter(fh, fieldnames=campos)
        w.writeheader()
        w.writerows(filas)


def exportar(data, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    base = f"{data['departamento'] or 'DEPTO'}_{data['anio_version']}".replace(' ', '_')

    largo = []
    for row in data['detalle']:
        for i, mes in enumerate(MESES):
            if row['montos'][i]:
                largo.append({
                    'departamento': data['departamento'],
                    'codigo_departamento': data['codigo_departamento'],
                    'anio': data['anio_version'], 'mes': mes,
                    'cuenta': row['cuenta'],
                    'descripcion_cuenta': row['descripcion_cuenta'],
                    'detalle': row['detalle'],
                    'descripcion_detalle': row['descripcion_detalle'],
                    'monto': round(row['montos'][i], 2),
                })
    p1 = os.path.join(out_dir, f'{base}_detalle.csv')
    _csv(p1, largo, ['departamento', 'codigo_departamento', 'anio', 'mes', 'cuenta',
                     'descripcion_cuenta', 'detalle', 'descripcion_detalle', 'monto'])

    cta = []
    for row in data['por_cuenta']:
        for i, mes in enumerate(MESES):
            if row['montos'][i]:
                cta.append({
                    'departamento': data['departamento'],
                    'codigo_departamento': data['codigo_departamento'],
                    'anio': data['anio_version'], 'mes': mes,
                    'cuenta': row['cuenta'],
                    'descripcion_cuenta': row['descripcion_cuenta'],
                    'monto': round(row['montos'][i], 2),
                })
    p2 = os.path.join(out_dir, f'{base}_por_cuenta.csv')
    _csv(p2, cta, ['departamento', 'codigo_departamento', 'anio', 'mes', 'cuenta',
                   'descripcion_cuenta', 'monto'])

    p3 = os.path.join(out_dir, f'{base}.json')
    with open(p3, 'w', encoding='utf-8') as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    return p1, p2, p3


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('archivo')
    ap.add_argument('--out-dir', default='.')
    ap.add_argument('--json', action='store_true')
    a = ap.parse_args()

    data = leer(a.archivo)
    if a.json:
        print(json.dumps(data, ensure_ascii=False, indent=2))
        sys.exit(0)

    paths = exportar(data, a.out_dir)
    print(f"Departamento : {data['departamento']} ({data['codigo_departamento']})")
    print(f"Anio version : {data['anio_version']}")
    print(f"Lineas detalle capturadas : {len(data['detalle'])}")
    print(f"Cuentas con monto         : {len(data['por_cuenta'])}")
    print(f"Referencias historicas    : {len(data['referencias'])} filas")
    print(f"GRAN TOTAL calculado      : {data['gran_total']:,.2f}")
    print(f"GRAN TOTAL en la hoja     : {data['gran_total_en_hoja']}")
    print(f"CUADRA                    : {'SI' if data['cuadra'] else 'NO — revisar'}")
    for p in paths:
        print('->', p)
