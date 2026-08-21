#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CHECKBOOK — Generador de formato estandar de gastos por departamento.
Proyecto: FINPLAN / The Costa Rica Collection

Geometria derivada de CHECKBOOK_MADRESAL__2026.xlsx, parametrizada:
  - N cuentas -> N bloques (paso fijo de 18 filas con 11 detalles).
  - Cada cuenta trae su TOTAL del anio de version + 3 anios de referencia
    (anio-1, anio-2, anio-3) precargados desde FinPlan.
  - Estadisticas de ocupacion en el encabezado.
  - Hoja protegida: solo quedan editables las celdas de captura.

Uso:
    python build_checkbook.py config.json CHECKBOOK_<DEPTO>_<ANIO>.xlsx [--force]
    python build_checkbook.py                # usa CONFIG_DEMO
"""

import json
import os
import sys
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# FORMATOS
# ---------------------------------------------------------------------------
FMT_MONEY = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
FMT_MONEY_SUM = '[$$-409]#,##0.00_);[Red]\\([$$-409]#,##0.00\\)'
FMT_MES = 'mmm-yy'
FMT_NOCHES = '#,##0'
FMT_PCT = '0.0%'

# Paleta
NAVY = 'FF1F3864'
FILL_TITULO = PatternFill('solid', fgColor=NAVY)
FILL_HEADER = PatternFill('solid', fgColor=NAVY)
FILL_MESES = PatternFill('solid', fgColor='FFFAE2D5')
FILL_GRAN = PatternFill('solid', fgColor='FFBDD7EE')
# El TOTAL de cada cuenta en el anio de version: pastel, no gris. «Dale color
# al total por cuenta del 2027, un color pastel» (owner, 18-ago-2026). Era el
# mismo gris que las referencias, asi que la fila que MAS importa de cada bloque
# —la unica que suma— se veia igual que las que no suman.
FILL_TOTAL = PatternFill('solid', fgColor='FFDDEBF7')
FILL_REF = PatternFill('solid', fgColor='FFFAFAFA')
FILL_STAT = PatternFill('solid', fgColor='FFE2EFDA')

F10 = Font(name='Arial', size=10)
F10B = Font(name='Arial', size=10, bold=True)
F14B = Font(name='Arial', size=14, bold=True)
F18B = Font(name='Arial', size=18, bold=True)
F10_IN = Font(name='Arial', size=10, color='FF0000FF')          # captura manual
F10_REF = Font(name='Arial', size=10, color='FF0000FF', italic=True)  # precarga FinPlan
F9I = Font(name='Arial', size=9, italic=True)
F12W = Font(name='Arial', size=12, bold=True, color='FFFFFFFF')
F10W = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
UNLOCKED = Protection(locked=False)

_S_THIN = Side('thin', color='FFD9D9D9')
_S_MED = Side('medium', color=NAVY)
_S_GRUESO = Side('thick', color=NAVY)
B_CELDA = Border(left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_THIN)
B_TOTAL = Border(left=_S_THIN, right=_S_THIN, top=_S_MED, bottom=_S_THIN)

# ⚠️ El TOTAL del anio de version CIERRA las lineas de detalle de arriba; los
# anios de abajo son referencia y no entran en esa suma. Sin nada que lo separe
# el bloque se lee ambiguo —«no se si la cuenta suma arriba o abajo» (owner,
# 18-ago-2026)—, asi que el TOTAL lleva linea gruesa arriba Y abajo: arriba
# cierra lo que suma, abajo corta con lo que no.
B_TOTAL_CIERRE = Border(left=_S_THIN, right=_S_THIN, top=_S_MED, bottom=_S_GRUESO)
# Y la ultima referencia cierra el bloque entero, para que se vea donde termina
# una cuenta y empieza la siguiente.
B_FIN_BLOQUE = Border(left=_S_THIN, right=_S_THIN, top=_S_THIN, bottom=_S_MED)


def _celda_activa(ws, celda):
    """Deja la seleccion DENTRO del panel desplazable.

    ⚠️ Al congelar, openpyxl escribe `<selection pane="bottomRight"
    activeCell="A1" sqref="A1"/>`: el panel activo empieza en la celda
    congelada —G16— pero su celda activa apunta a A1, que esta FUERA de ese
    panel. Excel no puede resolver la contradiccion y devuelve la vista al
    origen cada vez que uno se mueve: «llego al final y se vuelve al inicio»
    (owner, 18-ago-2026).

    Se ve normal al abrir el archivo. Solo aparece al intentar recorrerlo.
    """
    for sel in ws.sheet_view.selection:
        if sel.pane == 'bottomRight':
            sel.activeCell = celda
            sel.sqref = celda


def _bordes(ws, fila, c_ini, c_fin, borde=B_CELDA):
    for col in range(c_ini, c_fin + 1):
        ws.cell(fila, col).border = borde

# ---------------------------------------------------------------------------
# GEOMETRIA FIJA
# ---------------------------------------------------------------------------
FILA_DEPTO = 3           # C3 = nombre del departamento
FILA_MESES_TOP = 4       # C4 = "ESTADISTICAS", G4:R4 = meses
FILA_DISP = 5            # noches disponibles (captura)
FILA_OCUP = 6            # noches ocupadas   (captura)
FILA_PCT = 7             # % ocupacion       (formula)
FILA_GT_VER = 9          # GRAN TOTAL anio de version (formula)
FILA_GT_REF = 10         # 10, 11 = GRAN TOTAL anio-1, -2 (formula)
FILA_ETIQ_ANIO = 13
FILA_HDR1 = 15
# Anios de referencia por cuenta. Bajo de 3 a 2 el 18-ago-2026: «no creo que
# ocupemos mas de 2 anios; para 2027 solo ocupamos Forecast 2026 y Actual 2025»
# (owner).
#
# ⚠️ FILAS_BLANCO sube de 2 a 3 para compensar: el PASO del bloque tiene que
# seguir siendo 18 filas. Es lo que mantiene los headers en 15/33/51 y los TOTAL
# en 27/45/63, que es la geometria del archivo original del owner y de la que
# cuelgan las referencias directas del SUMMARY.
N_REFS = 2               # anios de referencia por cuenta
FILAS_BLANCO = 3         # blancos al final de cada bloque (ver N_REFS)

COL_MES_INI, COL_MES_FIN, COL_TOTAL = 7, 18, 19   # G, R, S


def _meses(anio):
    return [datetime(anio, m, 1) for m in range(1, 13)]


def _set(ws, row, col, value, font=F10, fmt=None, fill=None,
         align=None, unlock=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if unlock:
        c.protection = UNLOCKED
    return c


def _fila_meses(ws, fila, desde=2):
    for col in range(desde, COL_MES_INI):
        _set(ws, fila, col, None, F10W, None, FILL_HEADER)
    for i, col in enumerate(range(COL_MES_INI, COL_MES_FIN + 1)):
        _set(ws, fila, col, ws._meses_cache[i], F10W, FMT_MES, FILL_HEADER, CENTER)
    _set(ws, fila, COL_TOTAL, 'TOTAL', F10W, None, FILL_HEADER, CENTER)
    _bordes(ws, fila, desde, COL_TOTAL)
    ws.row_dimensions[fila].height = 18


# ---------------------------------------------------------------------------
# TAB 1 — BUDGET {anio} Detail
# ---------------------------------------------------------------------------
def _build_detalle(ws, cfg):
    anio = cfg['anio_version']
    cod = cfg['codigo_departamento']
    cuentas = cfg['cuentas']
    ndet = cfg.get('detalles_por_cuenta', 11)
    det_ini = cfg.get('detalle_inicial', 800)
    refs = cfg.get('referencias', {})          # {"2026": {"7030": [12 montos]}}
    # Lo que el departamento YA tiene cargado en el ano de version:
    #     {"7030": {"800": {"descripcion": "...", "montos": [12]}}}
    #
    # ⚠️ AGREGADO el 18-ago-2026, y es lo que convierte el archivo en una ida y
    # vuelta de verdad. Sin esto, bajar el checkbook de un departamento que ya
    # tiene presupuesto daba una hoja EN BLANCO: habia que volver a teclear lo
    # que FinPlan ya sabia, y las descripciones de detalle que el owner tenia
    # escritas en su archivo original (columna F) se perdian en cada regeneracion.
    #
    # Es OPCIONAL: si el config no lo trae, el archivo sale identico al de
    # antes, que es como quedo validado contra el original del owner.
    lineas = cfg.get('lineas', {}) or {}
    stats = cfg.get('estadisticas', {})        # {"noches_disponibles": [...], "noches_ocupadas": [...]}
    anios_ref = [anio - k for k in range(1, N_REFS + 1)]
    ws._meses_cache = _meses(anio)

    # Banda de titulo
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=COL_TOTAL)
    _set(ws, 1, 2, f"{cfg['departamento'].upper()}   |   PRESUPUESTO {anio}   |   "
                   f"CHECKBOOK DE GASTOS", F12W, None, FILL_TITULO, CENTER)
    _bordes(ws, 1, 2, COL_TOTAL)
    ws.row_dimensions[1].height = 26

    if cfg.get('incluir_leyenda', True):
        _set(ws, 2, 2, 'Capturar solo las celdas AZULES: columna F (descripcion del detalle) y '
                       'G:R (montos por mes). Azul cursiva = referencia precargada desde FinPlan. '
                       'Negro = formula bloqueada.', F9I)

    # --- Encabezado y estadisticas ------------------------------------------
    _set(ws, FILA_DEPTO, 3, cfg['departamento'], F14B)
    _set(ws, FILA_DEPTO, 6, f'{len(cuentas)} cuentas  ·  {ndet} detalles por cuenta', F9I)
    _set(ws, FILA_MESES_TOP, 3, 'ESTADISTICAS', F10W, None, FILL_HEADER)
    _fila_meses(ws, FILA_MESES_TOP, desde=3)

    for fila, etiqueta, clave in (
        (FILA_DISP, 'Noches Disponibles', 'noches_disponibles'),
        (FILA_OCUP, 'Noches Ocupadas', 'noches_ocupadas'),
    ):
        _set(ws, fila, 3, etiqueta, F10B, None, FILL_STAT)
        vals = stats.get(clave) or [None] * 12
        for i, col in enumerate(range(COL_MES_INI, COL_MES_FIN + 1)):
            _set(ws, fila, col, vals[i], F10_IN, FMT_NOCHES, FILL_STAT, unlock=True)
        _set(ws, fila, COL_TOTAL, f'=SUM(G{fila}:R{fila})', F10B, FMT_NOCHES, FILL_STAT)
        _bordes(ws, fila, 3, COL_TOTAL)

    _set(ws, FILA_PCT, 3, '% Ocupacion', F10B, None, FILL_STAT)
    for col in range(COL_MES_INI, COL_MES_FIN + 1):
        L = get_column_letter(col)
        _set(ws, FILA_PCT, col,
             f'=IFERROR({L}{FILA_OCUP}/{L}{FILA_DISP},0)', F10B, FMT_PCT, FILL_STAT)
    _set(ws, FILA_PCT, COL_TOTAL,
         f'=IFERROR(S{FILA_OCUP}/S{FILA_DISP},0)', F10B, FMT_PCT, FILL_STAT)
    _bordes(ws, FILA_PCT, 3, COL_TOTAL)

    _set(ws, FILA_ETIQ_ANIO, 3, anio, F10B)

    # --- Bloques por cuenta --------------------------------------------------
    paso = ndet + 2 + N_REFS + FILAS_BLANCO       # header + detalles + TOTAL + refs + blancos
    filas_total, filas_ref = [], {a: [] for a in anios_ref}

    for idx, cta in enumerate(cuentas):
        hdr = FILA_HDR1 + idx * paso
        ini, fin = hdr + 1, hdr + ndet
        tot = hdr + ndet + 1
        filas_total.append(tot)

        if idx == 0:
            for col, txt in ((2, '# Cuenta'), (3, 'Descripcion de Cuenta'),
                             (4, 'Departamento'), (5, 'Detalle'),
                             (6, 'Detalle Descripcion')):
                _set(ws, hdr, col, txt, F10W, None, FILL_HEADER)
        _fila_meses(ws, hdr)

        # Lo que la cuenta YA tiene cargado, por codigo de detalle. Ver `lineas`
        # en la nota de arriba: si el config no lo trae, esto queda vacio y el
        # archivo sale igual que antes.
        ya = (lineas.get(str(cta['cuenta'])) or {})

        for j in range(ndet):
            r = ini + j
            det = det_ini + j
            previo = ya.get(str(det)) or {}
            _set(ws, r, 2, cta['cuenta'], F10)
            _set(ws, r, 3, cta['descripcion'], F10)
            _set(ws, r, 4, cod, F10)
            _set(ws, r, 5, det, F10)
            _set(ws, r, 6, previo.get('descripcion') or None, F10_IN, unlock=True)
            montos = previo.get('montos') or [None] * 12
            for i, col in enumerate(range(COL_MES_INI, COL_MES_FIN + 1)):
                _set(ws, r, col, montos[i] or None, F10_IN, FMT_MONEY, unlock=True)
            _set(ws, r, COL_TOTAL, f'=SUM(G{r}:R{r})', F10, FMT_MONEY)
            _bordes(ws, r, 2, COL_TOTAL)

        # TOTAL del anio de version
        for col in (2, 4, 5):
            _set(ws, tot, col, None, F10B, None, FILL_TOTAL)
        _set(ws, tot, 3, f'TOTAL {anio}', F10B, None, FILL_TOTAL)
        _set(ws, tot, 6, f'suma las {ndet} lineas de arriba', F9I, None, FILL_TOTAL)
        for col in range(COL_MES_INI, COL_MES_FIN + 1):
            L = get_column_letter(col)
            _set(ws, tot, col, f'=SUM({L}{ini}:{L}{fin})', F10B, FMT_MONEY, FILL_TOTAL)
        _set(ws, tot, COL_TOTAL, f'=SUM(G{tot}:R{tot})', F10B, FMT_MONEY, FILL_TOTAL)
        _bordes(ws, tot, 2, COL_TOTAL, B_TOTAL_CIERRE)

        # Referencias: anio-1, anio-2, anio-3 (precarga FinPlan, editables)
        for k, aref in enumerate(anios_ref):
            r = tot + 1 + k
            filas_ref[aref].append(r)
            _set(ws, r, 3, f'TOTAL {aref}', F10, None, FILL_REF)
            for col in (2, 4, 5):
                _set(ws, r, col, None, F10, None, FILL_REF)
            # Un rotulo que quite la duda: la linea gruesa dice DONDE corta, esto
            # dice POR QUE. «No se si la cuenta suma arriba o abajo» (owner).
            _set(ws, r, 6, f'referencia {aref} — no suma al total {anio}',
                 F9I, None, FILL_REF)
            # ⚠️ CERO, no vacio. «Pon al menos 0.00 para que vean que ahi va un
            # dato que no hay en ese momento» (owner, 18-ago-2026): una celda en
            # blanco en la fila de un anio se lee como «no aplica», y lo que
            # dice es «ese anio no trajo esta cuenta». Un 0.00 deja ver que el
            # renglon existe y esta esperando dato.
            vals = (refs.get(str(aref), {}) or {}).get(str(cta['cuenta'])) or [0] * 12
            for i, col in enumerate(range(COL_MES_INI, COL_MES_FIN + 1)):
                _set(ws, r, col, vals[i] or 0, F10_REF, FMT_MONEY, FILL_REF, unlock=True)
            _set(ws, r, COL_TOTAL, f'=SUM(G{r}:R{r})', F10, FMT_MONEY, FILL_REF)
            _bordes(ws, r, 2, COL_TOTAL,
                    B_FIN_BLOQUE if k == len(anios_ref) - 1 else B_CELDA)

    # --- GRAN TOTALES --------------------------------------------------------
    def _gran(fila, etiqueta, filas, fill):
        _set(ws, fila, 3, etiqueta, F10B, None, fill)
        for col in range(COL_MES_INI, COL_MES_FIN + 1):
            L = get_column_letter(col)
            _set(ws, fila, col, '=' + '+'.join(f'{L}{r}' for r in filas),
                 F10B, FMT_MONEY, fill)
        _set(ws, fila, COL_TOTAL, f'=SUM(G{fila}:R{fila})', F10B, FMT_MONEY, fill)
        _bordes(ws, fila, 3, COL_TOTAL, B_TOTAL if fill is FILL_GRAN else B_CELDA)

    _gran(FILA_GT_VER, f'GRAN TOTAL {anio}', filas_total, FILL_GRAN)
    for k, aref in enumerate(anios_ref):
        _gran(FILA_GT_REF + k, f'GRAN TOTAL {aref}', filas_ref[aref], FILL_REF)

    # --- Presentacion y proteccion ------------------------------------------
    # ⚠️ La F baja de 64.29 a 42. Congelada, A-F sumaban ~122 caracteres: mas
    # ancho que la ventana, y Excel deja de poder desplazarse — «no logro ver
    # los 12 meses, se devuelve al principio el cursor» (owner, 18-ago-2026).
    # 42 deja el texto legible y ~60 caracteres de aire para los meses.
    for col, w in (('A', 2.0), ('B', 9.0), ('C', 27.57), ('D', 14.0),
                   ('E', 7.43), ('F', 42.0), ('S', 16.71)):
        ws.column_dimensions[col].width = w
    for col in range(COL_MES_INI, COL_MES_FIN + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13.0
    # ⚠️ G16, no G8. Congela las columnas A-F **y** las filas hasta la 15, que
    # es donde estan los meses. Con G8 las columnas quedaban fijas pero al bajar
    # se perdian los nombres de mes: «se mueve todo y no se ve nada» (owner).
    ws.freeze_panes = 'G16'
    _celda_activa(ws, 'G16')
    ws.sheet_view.showGridLines = False
    if cfg.get('proteger', True):
        ws.protection.sheet = True
        ws.protection.password = cfg.get('password_proteccion', 'FINPLAN')
        ws.protection.formatCells = False
        # ⚠️ FALSE, y es lo contrario de lo que parece. En Excel
        # `selectLockedCells=True` significa «el usuario NO puede seleccionar
        # celdas bloqueadas»: las flechas y el Tab saltan solo entre celdas de
        # captura y, al llegar a la ultima, VUELVEN A LA PRIMERA — «llego al
        # final y se vuelve al inicio» (owner, 18-ago-2026). No se podia
        # recorrer la hoja ni leer los totales.
        #
        # La proteccion tiene que impedir EDITAR, no NAVEGAR. Con False se puede
        # seleccionar y copiar cualquier celda; escribir en las bloqueadas sigue
        # siendo imposible, que es lo unico que se queria.
        ws.protection.selectLockedCells = False

    return {'filas_total': filas_total, 'filas_ref': filas_ref, 'paso': paso}


# ---------------------------------------------------------------------------
# TAB 2 — SUMMARY
# ---------------------------------------------------------------------------
def _build_summary(ws, cfg, nombre_detalle, geo):
    anio = cfg['anio_version']
    cuentas = cfg['cuentas']
    anios_ref = [anio - k for k in range(1, N_REFS + 1)]
    ws._meses_cache = _meses(anio)
    ref = f"'{nombre_detalle}'"
    # {anio: "Forecast 2026"} — como se llama el escenario del que salio cada
    # referencia. Si no viene, se cae a «TOTAL {anio}».
    etiquetas = cfg.get('etiquetas_ref') or {}
    C_REF0 = 21                      # U
    C_VAR = C_REF0 + N_REFS          # X

    # ── Geometria del SUMMARY, calculada y no quemada ────────────────────────
    #
    # Filas 5..    Rooms Occupied del anio de version y de cada referencia
    # +1 por anio  Cost per Room Occupied de cada uno
    # +2           encabezado de la tabla de cuentas
    #
    # Estaba quemado en la fila 9. Al agregar el costo de los anios de
    # referencia el encabezado paso a la 12, y una fila fija habria escrito la
    # tabla encima de las estadisticas sin que nada fallara.
    FILA_OCUP0 = 5
    FILA_CPO0 = FILA_OCUP0 + 1 + N_REFS
    FILA_HEAD = FILA_CPO0 + N_REFS + 2
    FILL_BANDA = PatternFill('solid', fgColor='FFF7F9FC')

    # Banda de titulo
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=C_VAR)
    _set(ws, 2, 4, f"{cfg['departamento'].upper()}   |   RESUMEN PRESUPUESTO {anio}",
         F12W, None, FILL_TITULO, CENTER)
    _bordes(ws, 2, 4, C_VAR)
    ws.row_dimensions[2].height = 26

    _set(ws, 4, 4, 'Cuentas', F9I)
    _set(ws, 4, 5, f'=COUNT(D{FILA_HEAD + 1}:D{FILA_HEAD + len(cuentas)})', F10B, '#,##0')
    _set(ws, 4, 7, 'Gran total del ano de version', F9I)

    # ── Estadisticas por anio + costo por habitacion ocupada ─────────────────
    #
    # «En summary debes empezar con las estadisticas por anio, Rooms Occupied
    # para los ultimos 2 anios mas el corriente, para sacar un cost per room
    # occupied» (owner, 18-ago-2026).
    #
    # Va en las filas 5-8, que estaban libres entre «Cuentas» y el encabezado de
    # la fila 9: asi la tabla de cuentas no se mueve ni una fila y el Detail
    # sigue apuntando donde apuntaba.
    est = cfg.get('estadisticas') or {}
    est_ref = cfg.get('estadisticas_ref') or {}       # {anio: [12 noches ocupadas]}
    filas_ocup = {}

    def _ocupadas(fila, etiqueta, serie, fuente_fill):
        _set(ws, fila, 4, etiqueta, F10B, None, fuente_fill, LEFT)
        _set(ws, fila, 5, 'Habitaciones ocupadas', F9I, None, fuente_fill, LEFT)
        for i, col in enumerate(range(COL_MES_INI, COL_MES_FIN + 1)):
            v = (serie or [0] * 12)[i] if i < len(serie or []) else 0
            _set(ws, fila, col, v or 0, F10, '#,##0', fuente_fill, CENTER)
        L0 = get_column_letter(COL_MES_INI)
        L1 = get_column_letter(COL_MES_FIN)
        _set(ws, fila, COL_TOTAL, f'=SUM({L0}{fila}:{L1}{fila})', F10B, '#,##0', fuente_fill, CENTER)
        _bordes(ws, fila, 4, COL_TOTAL)

    _ocupadas(FILA_OCUP0, f'Rooms Occupied {anio}', est.get('noches_ocupadas'), FILL_STAT)
    filas_ocup[anio] = FILA_OCUP0
    for k, aref in enumerate(anios_ref):
        _ocupadas(FILA_OCUP0 + 1 + k, f'Rooms Occupied {aref}', (est_ref.get(str(aref)) or {}).get('noches_ocupadas'), FILL_REF)
        filas_ocup[aref] = FILA_OCUP0 + 1 + k

    # Costo por habitacion ocupada, de CADA anio. «Mete tambien para 2026-2025
    # el costo, tal como 2027» (owner, 18-ago-2026).
    #
    # Cada uno divide el gran total de SU anio por las noches ocupadas de SU
    # anio: comparar el costo de 2027 contra el de 2026 solo significa algo si
    # los dos estan por habitacion ocupada — un ano con menos ocupacion gasta
    # menos en total sin ser mas eficiente.
    #
    # Sale del Detail (filas de GRAN TOTAL) para no depender de que la tabla de
    # cuentas de abajo este completa.
    f_cpo0 = FILA_CPO0
    filas_gt = [FILA_GT_VER] + [FILA_GT_REF + k for k in range(len(anios_ref))]
    for k, (a_, f_gt) in enumerate(zip([anio] + anios_ref, filas_gt)):
        f = f_cpo0 + k
        relleno = FILL_GRAN if k == 0 else FILL_REF
        fuente = F10B if k == 0 else F10
        _set(ws, f, 4, f'Cost per Room Occupied {a_}', fuente, None, relleno, LEFT)
        _set(ws, f, 5, f'Gran total {a_} / habitaciones ocupadas {a_}', F9I, None, relleno, LEFT)
        f_noches = filas_ocup[a_]
        for col in range(COL_MES_INI, COL_MES_FIN + 1):
            L = get_column_letter(col)
            _set(ws, f, col, f'=IFERROR({ref}!{L}{f_gt}/{L}{f_noches},0)',
                 fuente, FMT_MONEY, relleno, CENTER)
        LT = get_column_letter(COL_TOTAL)
        _set(ws, f, COL_TOTAL, f'=IFERROR({ref}!{LT}{f_gt}/{LT}{f_noches},0)',
             fuente, FMT_MONEY, relleno, CENTER)
        _bordes(ws, f, 4, COL_TOTAL, B_TOTAL if k == 0 else B_CELDA)


    # Encabezado de columnas
    _set(ws, FILA_HEAD, 4, '# Cuenta', F10W, None, FILL_HEADER, CENTER)
    _set(ws, FILA_HEAD, 5, 'Descripcion de Cuenta', F10W, None, FILL_HEADER, LEFT)
    _set(ws, FILA_HEAD, 6, None, F10W, None, FILL_HEADER)
    for i, col in enumerate(range(COL_MES_INI, COL_MES_FIN + 1)):
        _set(ws, FILA_HEAD, col, ws._meses_cache[i], F10W, FMT_MES, FILL_HEADER, CENTER)
    _set(ws, FILA_HEAD, COL_TOTAL, f'TOTAL {anio}', F10W, None, FILL_HEADER, CENTER)
    for k, aref in enumerate(anios_ref):
        # El nombre del escenario, no «TOTAL 2026»: un mismo anio tiene
        # Working, Draft y Final, y la columna tiene que decir cual es.
        _set(ws, FILA_HEAD, C_REF0 + k, etiquetas.get(str(aref), f'TOTAL {aref}'),
             F10W, None, FILL_HEADER, CENTER)
    _set(ws, FILA_HEAD, C_VAR, f'Var % vs {anios_ref[0]}', F10W, None, FILL_HEADER, CENTER)
    _bordes(ws, FILA_HEAD, 4, C_VAR)
    ws.row_dimensions[FILA_HEAD].height = 18

    # Una fila por cuenta
    r0 = FILA_HEAD + 1
    for i, cta in enumerate(cuentas):
        r = r0 + i
        banda = FILL_BANDA if i % 2 else None
        # Sin separador de miles: «7,030» no es una cuenta, es un numero. El
        # SUMIF del Detail busca por el valor, asi que el formato es solo lo
        # que se ve — pero lo que se ve tiene que parecerse a una cuenta.
        _set(ws, r, 4, cta['cuenta'], F10, '0', banda, CENTER)
        _set(ws, r, 5, cta['descripcion'], F10, None, banda)
        _set(ws, r, 6, None, F10, None, banda)
        for col in range(COL_MES_INI, COL_MES_FIN + 1):
            L = get_column_letter(col)
            _set(ws, r, col, f'=SUMIF({ref}!$B:$B,$D{r},{ref}!{L}:{L})',
                 F10, FMT_MONEY_SUM, banda)
        _set(ws, r, COL_TOTAL, f'=SUM(G{r}:R{r})', F10B, FMT_MONEY_SUM, banda)
        for k, aref in enumerate(anios_ref):
            _set(ws, r, C_REF0 + k, f'={ref}!S{geo["filas_ref"][aref][i]}',
                 F10, FMT_MONEY_SUM, banda)
        LR = get_column_letter(C_REF0)
        _set(ws, r, C_VAR, f'=IFERROR(S{r}/{LR}{r}-1,0)', F10, FMT_PCT, banda)
        _bordes(ws, r, 4, C_VAR)

    # Fila TOTAL: suma por columna (cada mes) y total general
    tot = r0 + len(cuentas) + 1
    _set(ws, tot, 4, None, F10B, None, FILL_TOTAL)
    _set(ws, tot, 5, 'TOTAL DEPARTAMENTO', F10B, None, FILL_TOTAL)
    _set(ws, tot, 6, None, F10B, None, FILL_TOTAL)
    for col in list(range(COL_MES_INI, COL_TOTAL + 1)) + [C_REF0 + k for k in range(N_REFS)]:
        L = get_column_letter(col)
        _set(ws, tot, col, f'=SUM({L}{r0}:{L}{tot - 1})', F10B, FMT_MONEY_SUM, FILL_TOTAL)
    LR = get_column_letter(C_REF0)
    _set(ws, tot, C_VAR, f'=IFERROR(S{tot}/{LR}{tot}-1,0)', F10B, FMT_PCT, FILL_TOTAL)
    _bordes(ws, tot, 4, C_VAR, B_TOTAL)
    ws.row_dimensions[tot].height = 18

    # Chequeo de cuadre contra el tab de detalle
    chk = tot + 2
    _set(ws, chk, 5, 'Cuadre vs. detalle', F9I)
    _set(ws, chk, COL_TOTAL, f'=S{tot}-{ref}!S{FILA_GT_VER}', F9I, FMT_MONEY_SUM)
    _set(ws, chk, COL_TOTAL + 1, 'debe ser 0.00', F9I)

    _set(ws, 4, 8, f'=S{tot}', Font(name='Arial', size=12, bold=True, color=NAVY), FMT_MONEY_SUM)

    # La D llevaba 12 y ahora tiene que caber «Cost per Room Occupied 2027»;
    # la E, la explicacion de cada fila. Con los anchos viejos el rotulo salia
    # cortado («Rooms Occu|», «Cost per Roo|») y el conteo daba ####.
    for col, w in (('A', 2.0), ('B', 3.0), ('C', 3.0), ('D', 30.0),
                   ('E', 40.0), ('F', 3.0), ('S', 16.0), ('T', 3.0)):
        ws.column_dimensions[col].width = w
    for col in list(range(COL_MES_INI, COL_MES_FIN + 1)) + list(range(C_REF0, C_VAR + 1)):
        ws.column_dimensions[get_column_letter(col)].width = 13.0
    # Congelado en Enero y despues del encabezado: las etiquetas de la
    # izquierda y los nombres de mes quedan fijos al desplazarse.
    ws.freeze_panes = f'G{FILA_HEAD + 1}'
    _celda_activa(ws, f'G{FILA_HEAD + 1}')
    ws.sheet_view.showGridLines = False
    if cfg.get('proteger', True):
        ws.protection.sheet = True
        ws.protection.password = cfg.get('password_proteccion', 'FINPLAN')
        # ⚠️ FALSE, y es lo contrario de lo que parece. En Excel
        # `selectLockedCells=True` significa «el usuario NO puede seleccionar
        # celdas bloqueadas»: las flechas y el Tab saltan solo entre celdas de
        # captura y, al llegar a la ultima, VUELVEN A LA PRIMERA — «llego al
        # final y se vuelve al inicio» (owner, 18-ago-2026). No se podia
        # recorrer la hoja ni leer los totales.
        #
        # La proteccion tiene que impedir EDITAR, no NAVEGAR. Con False se puede
        # seleccionar y copiar cualquier celda; escribir en las bloqueadas sigue
        # siendo imposible, que es lo unico que se queria.
        ws.protection.selectLockedCells = False


# ---------------------------------------------------------------------------
def validar(cfg):
    faltantes = [k for k in ('departamento', 'codigo_departamento', 'anio_version', 'cuentas')
                 if not cfg.get(k)]
    if faltantes:
        raise ValueError(f"Config incompleto, faltan: {', '.join(faltantes)}")
    nums = [c['cuenta'] for c in cfg['cuentas']]
    dups = [n for n, k in Counter(nums).items() if k > 1]
    if dups:
        raise ValueError(f"Cuentas duplicadas en el config: {dups}. El SUMIF del SUMMARY "
                         f"las sumaria juntas y el TOTAL quedaria doblado.")
    sin_desc = [c['cuenta'] for c in cfg['cuentas'] if not str(c.get('descripcion', '')).strip()]
    if sin_desc:
        raise ValueError(f"Cuentas sin descripcion: {sin_desc}")
    for clave, serie in (cfg.get('estadisticas') or {}).items():
        if serie is not None and len(serie) != 12:
            raise ValueError(f"estadisticas.{clave} debe traer 12 valores (Ene-Dic), trae {len(serie)}")
    for anio, mapa in (cfg.get('referencias') or {}).items():
        for cta, serie in (mapa or {}).items():
            if serie is not None and len(serie) != 12:
                raise ValueError(f"referencias.{anio}.{cta} debe traer 12 valores, trae {len(serie)}")


def build(cfg, salida, force=False):
    validar(cfg)
    if os.path.exists(salida) and not force:
        raise FileExistsError(f"{salida} ya existe. Si ya tiene montos capturados, "
                              f"regenerarlo los borra. Usa --force o cambia el nombre.")
    nombre_detalle = f"BUDGET {cfg['anio_version']} Detail"
    wb = Workbook()
    ws_det = wb.active
    ws_det.title = nombre_detalle
    geo = _build_detalle(ws_det, cfg)
    _build_summary(wb.create_sheet('SUMMARY'), cfg, nombre_detalle, geo)
    wb.save(salida)
    return salida


CONFIG_DEMO = {
    "departamento": "Gastos Operativos Club Madresal",
    "codigo_departamento": 600,
    "anio_version": 2027,
    "detalles_por_cuenta": 11,
    "detalle_inicial": 800,
    "incluir_leyenda": True,
    "proteger": True,
    "estadisticas": {},
    "referencias": {},
    "cuentas": [
        {"cuenta": 7030, "descripcion": "Building"},
        {"cuenta": 7050, "descripcion": "Centralized Accounting Charges"},
        {"cuenta": 7065, "descripcion": "Cleaning Supplies"},
        {"cuenta": 7105, "descripcion": "Contract Services"},
        {"cuenta": 7155, "descripcion": "Electrical and Mechanical Equipment"},
        {"cuenta": 7160, "descripcion": "Electricity"},
        {"cuenta": 7170, "descripcion": "Engineering Supplies"},
        {"cuenta": 7175, "descripcion": "Entertainment-In-House"},
        {"cuenta": 7180, "descripcion": "Equipment"},
        {"cuenta": 7240, "descripcion": "Grounds Maintenance and Landscaping"},
        {"cuenta": 7310, "descripcion": "Laundry and Dry Cleaning"},
        {"cuenta": 7335, "descripcion": "Licenses and Permits"},
        {"cuenta": 7340, "descripcion": "Life/Safety"},
        {"cuenta": 7380, "descripcion": "Miscellaneous"},
        {"cuenta": 7400, "descripcion": "Operating Supplies"},
        {"cuenta": 7420, "descripcion": "Other Fuels"},
        {"cuenta": 7490, "descripcion": "Printing and Stationery"},
        {"cuenta": 7500, "descripcion": "Promotion"},
        {"cuenta": 7555, "descripcion": "Swimming Pool"},
        {"cuenta": 7605, "descripcion": "System Expenses: Information Systems"},
        {"cuenta": 7635, "descripcion": "System Expenses: Telecom and Information Systems"},
        {"cuenta": 7680, "descripcion": "Uniform Costs"},
        {"cuenta": 7710, "descripcion": "Water/Sewer"},
        {"cuenta": 8005, "descripcion": "Owners Fees/Management Fee"},
        {"cuenta": 8015, "descripcion": "Property Insurance"},
    ],
}

if __name__ == '__main__':
    args = [a for a in sys.argv[1:] if a != '--force']
    force = '--force' in sys.argv[1:]
    if args and args[0].endswith('.json'):
        with open(args[0], encoding='utf-8') as fh:
            cfg = json.load(fh)
        out = args[1] if len(args) > 1 else 'CHECKBOOK.xlsx'
    else:
        cfg = CONFIG_DEMO
        out = args[0] if args else 'CHECKBOOK.xlsx'
    print(build(cfg, out, force=force))
