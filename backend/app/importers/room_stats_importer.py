"""Parser de la hoja 'Room Stats' del paquete ejecutivo (datos de Opera/PMS).

La hoja tiene un bloque YTD arriba y luego un bloque por mes (Enero..Mayo...).
Cada bloque: encabezado con 'Room Category' y filas por tipo de habitación con
Units (C), Total Noches (D), Total Noches Ocupadas (E), Total Revenue (G),
Total Pax (I), y el nombre del mes en la col J de las filas de datos. La fila
'Other Rooms Revenue' trae solo revenue.

Devuelve una lista de filas planas {month, room_type_name, units,
nights_available, nights_occupied, revenue, pax} — solo de los bloques MENSUALES
(el bloque YTD se ignora, se reconstruye sumando los meses).
"""
from __future__ import annotations
from io import BytesIO

_MONMAP = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def _month_num(v) -> int | None:
    if not v:
        return None
    s = str(v).lower()
    if "ytd" in s:
        return None
    for k, n in _MONMAP.items():
        if k in s:
            return n
    return None


def _num(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_room_stats(file_bytes: bytes, sheet_name: str = "Room Stats") -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        # fuzzy: alguna hoja que contenga 'room' y 'stat'
        cand = [s for s in wb.sheetnames if "room" in s.lower() and "stat" in s.lower()]
        if not cand:
            raise ValueError(f"No se encontró la hoja '{sheet_name}'. Hojas: {wb.sheetnames}")
        sheet_name = cand[0]
    ws = wb[sheet_name]

    out: list[dict] = []
    r = 1
    maxr = ws.max_row
    while r <= maxr:
        b = ws.cell(r, 2).value
        if b and "Room Category" in str(b):
            # leer filas de datos del bloque
            rr = r + 1
            block_month = None
            rows = []
            while rr <= maxr:
                bb = ws.cell(rr, 2).value
                if bb and ("TOTAL" in str(bb) or "Room Category" in str(bb)):
                    break
                if bb:
                    mn = _month_num(ws.cell(rr, 10).value)
                    if mn:
                        block_month = mn
                    rows.append({
                        "name": str(bb).strip(),
                        "units": int(_num(ws.cell(rr, 3).value)),
                        "nights_available": _num(ws.cell(rr, 4).value),
                        "nights_occupied": _num(ws.cell(rr, 5).value),
                        "revenue": _num(ws.cell(rr, 7).value),
                        "pax": _num(ws.cell(rr, 9).value),
                    })
                rr += 1
            # solo bloques mensuales (con mes detectado y NO ytd)
            if block_month:
                for row in rows:
                    out.append({"month": block_month, "room_type_name": row["name"],
                                "units": row["units"], "nights_available": row["nights_available"],
                                "nights_occupied": row["nights_occupied"], "revenue": row["revenue"],
                                "pax": row["pax"]})
            r = rr
        else:
            r += 1
    return out
