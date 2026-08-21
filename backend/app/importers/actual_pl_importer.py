"""
Actual P&L importer — reads the CWL working workbook ("Budget 2025W" tab) and
extracts the per-account, per-department, per-month detail for every version.

Layout discovered in FCST May(500) 2026.xlsx, sheet "Budget 2025W":
  - Detail section starts at the "Gl Account" header row (~247).
  - Per detail row: col C = account_code, col E = description, col J = department.
  - Month data is in 12-column blocks; the version of each block is labelled in
    row 15. Decoded blocks (1-indexed columns):
        Actual 2024        K..V   (11..22)
        Actual 2025        W..AH  (23..34)
        Actual 2026        AI..AM (35..39)   Jan-May only (rest is forecast)
        Forecast 2026 FY   AI..AT (35..46)   Jan-May actual + Jun-Dec forecast
        Forecast Apr 2026  AV..BG (48..59)
        Budget 2026        BI..BT (61..72)

Output: {block_name: [ {dept_code, account_code, account_name, jan..dec}, ... ]}
aggregated by (dept_code, account_code). Feed these rows to the actuals API
(account-level) or to pl_engine.build_actual_inputs for validation.
"""
from __future__ import annotations

from decimal import Decimal
import openpyxl

ZERO = Decimal("0")
MONTH_KEYS = ["jan", "feb", "mar", "apr", "may", "jun",
              "jul", "aug", "sep", "oct", "nov", "dec"]

DEFAULT_SHEET = "Budget 2025W"

# Version blocks: name -> (scenario_type, year, start_col_1indexed, n_months)
VERSION_BLOCKS: dict[str, tuple[str, int, int, int]] = {
    "ACTUAL_2024":       ("ACTUAL",   2024, 11, 12),
    "ACTUAL_2025":       ("ACTUAL",   2025, 23, 12),
    "ACTUAL_2026":       ("ACTUAL",   2026, 35, 5),    # Jan-May actual only
    "FORECAST_2026":     ("FORECAST", 2026, 35, 12),   # Jan-May actual + Jun-Dec fcst
    "FORECAST_APR_2026": ("FORECAST", 2026, 48, 12),   # April reforecast
    "BUDGET_2026":       ("BUDGET",   2026, 61, 12),
}

# Department name (col J) -> dept_code. Matched by keyword (robust to accents /
# encoding). The 5 ambiguous ones (claro huerta / beneficios / property / misc)
# only affect the per-dept split, never the GOP/EBITDA/Net totals.
_DEPT_KEYWORDS: list[tuple[str, str]] = [
    ("habitacion", "0110"),
    ("a&b", "0120"),
    # 0140, el PADRE del Spa. Este importador mandaba «spa» al 0130 y el del
    # detalle del GL al 0140: el mismo Spa entraba por dos departamentos según
    # por dónde se cargara. Ahora los dos van al padre, que es donde viven las
    # reglas (owner, 2026-08-13).
    # ⚠️ Los DOS importadores tienen que mandar el mismo nombre al mismo
    # departamento. Cuando no coinciden, el gasto cae en una linea u otra
    # segun por donde se cargue el archivo — y las dos versiones se ven
    # bien por separado (auditoria, 2026-08-14).
    ("spa", "0140"),
    ("tours", "0150"),
    ("gift", "0165"),
    ("transport", "0152"),
    ("lavander", "0161"),
    ("innoceana", "0155"),
    ("crowther", "0156"),
    ("administ", "0180"),
    ("ventas", "0190"),
    ("mercadeo", "0190"),
    ("mantenim", "0200"),
    ("claro huerta", "0205"),
    # ⚠️ Utilities es el 0210, NO el 0205. El 0205 es Claro del Bosque, que
    # tiene sus propias lineas (OH_/COH_/REV_CLARO_HUERTA). Con esto los
    # actuales cargados por el resumen dejaban `OH_UTILITIES` en CERO e
    # inflaban Claro del Bosque. El importador del GL siempre mando a 0210.
    ("utility", "0210"),
    ("utilit", "0210"),
    ("cafeteria", "0220"),
    # Employee Benefits: 0181. El importador del GL lo mandaba al 0220
    # (Cafeteria), que ademas se descarta entero en el camino legacy — el mismo
    # gasto caia en tres sitios distintos segun por donde entrara.
    ("beneficios", "0181"),
    ("property expenses", "0250"),
    ("management fees", "0250"),
    ("miscelane", "0240"),
]


def dept_code_for_name(name: str | None) -> str:
    if not name:
        return ""
    low = str(name).strip().lower()
    # "Departamento de TI" — match the IT dept explicitly (avoid 'ti' substrings)
    if low.endswith(" ti") or low.endswith("de ti"):
        return "0230"
    for kw, code in _DEPT_KEYWORDS:
        if kw in low:
            return code
    return ""  # unknown -> pl_engine routes to OTHER_OVERHEAD


def _norm_account(v) -> str | None:
    """Normalise an account-code cell to a string like '4000' (or None)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s or not s[0].isdigit():
        return None
    # USALI codes are 4xxx..9xxx; ignore obvious non-codes
    head = s.split(".")[0]
    if not head.isdigit() or not (4 <= int(head[0]) <= 9):
        return None
    return head


def _find_detail_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=400, min_col=3, max_col=3):
        v = row[0].value
        if isinstance(v, str) and v.strip().lower() == "gl account":
            return row[0].row
    raise ValueError("No se encontró la fila de encabezado 'Gl Account'")


def import_actual_pl(path: str, sheet: str = DEFAULT_SHEET) -> dict[str, list[dict]]:
    """
    Parse the workbook and return per-version aggregated account rows.
    {block_name: [{dept_code, account_code, account_name, jan..dec}]}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    header_row = _find_detail_header_row(ws)

    # max column we ever read
    max_col = max(start + n - 1 for _, _, start, n in VERSION_BLOCKS.values())

    # accumulators: block -> {(dept,acct): {"name":..., "m":[12 Decimals]}}
    acc: dict[str, dict[tuple[str, str], dict]] = {b: {} for b in VERSION_BLOCKS}

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=max_col):
        account = _norm_account(row[2].value)        # col C
        if account is None:
            continue
        desc = row[4].value if len(row) > 4 else ""   # col E
        dept_name = row[9].value if len(row) > 9 else ""  # col J
        dept_code = dept_code_for_name(dept_name)
        key = (dept_code, account)

        for block, (_t, _y, start, nmonths) in VERSION_BLOCKS.items():
            bucket = acc[block].setdefault(
                key, {"name": str(desc or ""), "m": [ZERO] * 12}
            )
            for i in range(nmonths):
                cell = row[start - 1 + i]
                val = cell.value
                if val is None or val == "":
                    continue
                try:
                    bucket["m"][i] += Decimal(str(val))
                except (ValueError, ArithmeticError):
                    pass

    out: dict[str, list[dict]] = {}
    for block, rows in acc.items():
        lst = []
        for (dept_code, account), data in rows.items():
            rec = {"dept_code": dept_code, "account_code": account,
                   "account_name": data["name"]}
            for i, mk in enumerate(MONTH_KEYS):
                rec[mk] = data["m"][i]
            lst.append(rec)
        out[block] = lst
    return out
