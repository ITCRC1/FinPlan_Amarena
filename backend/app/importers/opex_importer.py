"""
OPEX checkbook importer — Phase 6.

Reads all OPEXC_2026__[DEPT]__BUDGET.xlsx files and returns OpexEntry objects.

File structure (common across all files):
  - Some header/summary rows at the top
  - A row where col B == '# Cuenta' → identifies the label header
  - The row immediately after contains month datetime objects in cols G-R (indices 6-17)
    Exception: in some files the dates are in the same row as '# Cuenta'
  - Data rows follow: col B = account_code (int), C = account_name,
    D = dept_code, E = detail_code (int), F = detail_desc, G-R = months 1-12

Call: import_opex_for_scenario(scenario_id, hotel_id, base_dir) → list[OpexEntry]
"""
import uuid
from decimal import Decimal
from pathlib import Path
from typing import Optional
import datetime

import openpyxl

from app.models.opex_entry import OpexEntry

# Map short key → (filename_glob_pattern, expected_dept_override_or_None)
# dept_override forces the dept_code if the file's col D is unreliable or missing
OPEX_FILE_MAP: dict[str, tuple[str, Optional[str]]] = {
    "ROOMS":   ("OPEXCR 2026  (ROOMS) BUDGET.xlsx",     None),
    "F_B":     ("OPEXC 2026 (F&B) BUDGET.xlsx",         None),
    "SPA":     ("OPEXC 2026 (SPA) BUDGET.xlsx",         None),
    "ADMIN":   ("OPEXC  2026 (ADMIN) BUDGET.xlsx",      None),
    "IT":      ("OPEXC  2026 (IT) BUDGET.xlsx",         None),
    "OWN":     ("OPEXC  2026 (OWN) BUDGET.xlsx",        None),
    "SALES":   ("OPEXC  2026 (SALES) BUDGET.xlsx",      None),
    "UTILITY": ("OPEXC  2026 (UTILITY) BUDGET.xlsx",    None),
    "LAUND":   ("OPEXC 2025 (LAUND) BUDGET.xlsx",       None),
    "ACT":     ("OPEXC 2026 (ACT) BUDGET.xlsx",         None),
    "C_BOSQ":  ("OPEXC 2026 (C BOSQ) BUDGET.xlsx",      None),
    "CAF":     ("OPEXC 2026 (CAF) BUD.xlsx",            None),
    "CROW":    ("OPEXC 2026 (CROW) BUDGET.xlsx",        None),
    "INNO":    ("OPEXC 2026 (INNO) BUDGET.xlsx",        None),
    "MAINT":   ("OPEXC 2026 (MAINT) BUDGET FINAL.xlsx", None),
    "RETAIL":  ("OPEXC 2026 (RETAIL) BUDGET.xlsx",      None),
    "TRANSP":  ("OPEXC 2026 (TRANSP) BUDGET.xlsx",      None),
}


def _clean_dept(raw: object) -> str:
    """Strip apostrophes and whitespace that sometimes appear in dept codes."""
    if raw is None:
        return ""
    s = str(raw).strip().lstrip("'").strip()
    return s


def _to_decimal(val: object) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(float(val))).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def _find_header_and_data_start(ws) -> tuple[int, int]:
    """
    Returns (header_row_idx, data_start_row_idx) — 1-based row numbers.
    header_row_idx: the row where col B == '# Cuenta'
    data_start_row_idx: first row of actual data (after month labels)
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        if row[1] == "# Cuenta":
            header_row = i
            # Check if THIS row has datetime objects in cols 6-17
            has_dates = any(
                isinstance(row[c], datetime.datetime) for c in range(6, 18)
            )
            if has_dates:
                return header_row, header_row + 1
            else:
                return header_row, header_row + 2   # dates in next row, data starts +2
    raise ValueError("Header row ('# Cuenta') not found in first 40 rows")


def _import_single_file(
    filepath: Path,
    scenario_id: str,
    hotel_id: str,
    dept_override: Optional[str],
) -> list[OpexEntry]:
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    try:
        _, data_start = _find_header_and_data_start(ws)
    except ValueError:
        return []

    entries: list[OpexEntry] = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        # col B = account_code (must be numeric)
        raw_acct = row[1]
        if raw_acct is None:
            continue
        try:
            account_code = str(int(float(str(raw_acct))))
        except (ValueError, TypeError):
            continue

        # col C = account_name, D = dept_code, E = detail_code, F = detail_desc
        account_name = str(row[2] or "").strip()
        dept_code = dept_override if dept_override else _clean_dept(row[3])
        raw_detail = row[4]
        detail_code = str(int(float(str(raw_detail)))) if raw_detail is not None else ""
        detail_desc = str(row[5] or "").strip()

        # cols G-R (indices 6-17) = months 1-12
        month_vals = [_to_decimal(row[c]) for c in range(6, 18)]

        # Skip entirely-zero rows
        if all(v == Decimal("0") for v in month_vals):
            continue

        entry = OpexEntry(
            id=str(uuid.uuid4()),
            scenario_id=scenario_id,
            hotel_id=hotel_id,
            dept_code=dept_code,
            account_code=account_code,
            account_name=account_name,
            detail_code=detail_code,
            detail_desc=detail_desc,
            jan=month_vals[0],  feb=month_vals[1],  mar=month_vals[2],
            apr=month_vals[3],  may=month_vals[4],  jun=month_vals[5],
            jul=month_vals[6],  aug=month_vals[7],  sep=month_vals[8],
            oct=month_vals[9],  nov=month_vals[10], dec=month_vals[11],
        )
        entries.append(entry)

    return entries


def import_opex_for_scenario(
    scenario_id: str,
    hotel_id: str,
    base_dir: Path,
) -> list[OpexEntry]:
    """
    Import all OPEX checkbook files found in base_dir.
    Returns flat list of OpexEntry objects (not yet persisted).
    Files not found are skipped with a warning printed to stdout.
    """
    all_entries: list[OpexEntry] = []
    for key, (filename, dept_override) in OPEX_FILE_MAP.items():
        filepath = base_dir / filename
        if not filepath.exists():
            print(f"[opex_importer] WARNING: {filename} not found — skipping {key}")
            continue
        entries = _import_single_file(filepath, scenario_id, hotel_id, dept_override)
        all_entries.extend(entries)
        print(f"[opex_importer] {key}: {len(entries)} non-zero lines imported")

    return all_entries


def import_single_opex_file(
    filepath: Path,
    scenario_id: str,
    hotel_id: str,
    dept_override: Optional[str] = None,
) -> list[OpexEntry]:
    """Import a single OPEX file. Used for targeted re-import."""
    return _import_single_file(filepath, scenario_id, hotel_id, dept_override)
