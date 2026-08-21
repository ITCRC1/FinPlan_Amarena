"""
Load report_line_config and account_mapping from the mapping Excel file.

Usage:
    cd backend
    python -m app.importers.mapping_loader --file "C:/FinPlan_CWL/data/formato_mapping_reporte_app.xlsx"

Or as a library:
    from app.importers.mapping_loader import load_mapping_workbook
    await load_mapping_workbook(session, path)
"""
from __future__ import annotations

import argparse
import asyncio
import uuid
from pathlib import Path

import openpyxl
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.mapping import AccountMapping, ReportLineConfig

REPORT_CONFIG_SHEET = "02_Report_Config"
MAPPING_SHEET = "03_Mapping_Upload"
REPORT_ID = "P&L_DETAIL_OWNERS"


def _dept_code_from_name(name: str | None) -> str | None:
    """Nombre de depto del mapeo ('Departamento de A&B') → código ('0120').

    El motor del P&L rutea por `dept_code`; el archivo trae el nombre. Sin esta
    traducción toda cuenta que mapee a distinta línea según depto (las 17 de
    planilla, el opex compartido) colapsa en la línea del primer depto.
    """
    if not name or not str(name).strip():
        return None
    from app.importers.gl_detail_importer import dept_code_from_name
    return dept_code_from_name(str(name).strip()) or None


def _cell(row, idx: int):
    """Return stripped string or None for a 0-based column index."""
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_report_config(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[REPORT_CONFIG_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[1:]:   # skip header
        if all(v is None for v in row):
            continue
        report_id = _cell(row, 0)
        if not report_id:
            continue
        records.append({
            "report_id":          report_id,
            "display_order":      int(row[1]) if row[1] is not None else 0,
            "line_code":          _cell(row, 2) or "",
            "section":            _cell(row, 3) or "",
            "line_name":          _cell(row, 4) or "",
            "line_type":          _cell(row, 5) or "MAPPED",
            "parent_line_code":   _cell(row, 6),
            "calculation_logic":  _cell(row, 7),
            "active_status":      _cell(row, 8) or "YES",
            "format_hint":        _cell(row, 9),
            "notes":              _cell(row, 10),
        })
    return records


def parse_mapping_upload(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[MAPPING_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        mapping_id = _cell(row, 0)
        if not mapping_id:
            continue
        records.append({
            "mapping_id":          mapping_id,
            "active_status":       _cell(row, 1) or "YES",
            "report_id":           _cell(row, 2) or REPORT_ID,
            "report_line_code":    _cell(row, 3) or "",
            "report_line_name":    _cell(row, 4),
            "report_section":      _cell(row, 5),
            "display_order":       int(row[6]) if row[6] is not None else None,
            "source_origin":       _cell(row, 7),
            "source_department":   _cell(row, 8),
            "account_code":        str(_cell(row, 9) or "").strip(),
            "account_name_example":_cell(row, 10),
            "financial_nature":    _cell(row, 11) or "Expense",
            "rollup_operator":     _cell(row, 12) or "SUM",
            "sign_rule":           _cell(row, 13),
            "notes":               _cell(row, 15),
        })
    return records


async def load_mapping_workbook(session, path: str) -> dict:
    """
    (Re)load report_line_config and account_mapping from the Excel file.
    Replaces all existing rows for REPORT_ID before inserting.
    Returns a summary dict.
    """
    config_rows = parse_report_config(path)
    mapping_rows = parse_mapping_upload(path)

    # --- report_line_config ---
    await session.execute(
        delete(ReportLineConfig).where(ReportLineConfig.report_id == REPORT_ID)
    )
    for r in config_rows:
        active = str(r.get("active_status", "YES")).upper() == "YES"
        obj = ReportLineConfig(
            id=str(uuid.uuid4()),
            report_id=r["report_id"],
            display_order=r["display_order"],
            line_code=r["line_code"],
            section=r["section"],
            line_name=r["line_name"],
            line_type=r["line_type"],
            parent_line_code=r["parent_line_code"],
            calculation_logic=r["calculation_logic"],
            format_hint=r["format_hint"],
            active=active,
        )
        session.add(obj)

    # --- account_mapping ---
    await session.execute(
        delete(AccountMapping).where(AccountMapping.report_id == REPORT_ID)
    )
    loaded = 0
    skipped = 0
    for r in mapping_rows:
        if not r["account_code"] or not r["report_line_code"]:
            skipped += 1
            continue
        obj = AccountMapping(
            id=str(uuid.uuid4()),
            active_status=r["active_status"],
            report_id=r["report_id"],
            report_line_code=r["report_line_code"],
            report_line_name=r["report_line_name"],
            report_section=r["report_section"],
            display_order=r["display_order"],
            source_origin=r["source_origin"],
            source_department=r["source_department"],
            # CRÍTICO: el motor rutea por dept_code. Sin esto, TODA cuenta que mapee
            # a varias líneas según depto (las 17 de planilla, opex compartido) cae en
            # la línea del PRIMER depto → p.ej. la planilla de A&B/Spa/Admin aterriza
            # en OPEX_ROOMS y el overhead queda en 0. Se deriva del nombre del depto.
            dept_code=_dept_code_from_name(r["source_department"]),
            account_code=r["account_code"],
            account_name_example=r["account_name_example"],
            financial_nature=r["financial_nature"],
            rollup_operator=r["rollup_operator"],
            sign_rule=r["sign_rule"],
            notes=r["notes"],
        )
        session.add(obj)
        loaded += 1

    await session.commit()

    return {
        "report_lines_loaded": len(config_rows),
        "mappings_loaded": loaded,
        "mappings_skipped": skipped,
    }


async def _main(path: str):
    print(f"Loading mapping from: {path}")
    async with SessionLocal() as session:
        result = await load_mapping_workbook(session, path)
    print(f"✓ Report lines loaded : {result['report_lines_loaded']}")
    print(f"✓ Mapping rules loaded: {result['mappings_loaded']}")
    if result["mappings_skipped"]:
        print(f"  Skipped (no account/line code): {result['mappings_skipped']}")
    print("\n✅ Mapping loaded successfully.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to the mapping Excel file")
    args = parser.parse_args()
    asyncio.run(_main(args.file))
