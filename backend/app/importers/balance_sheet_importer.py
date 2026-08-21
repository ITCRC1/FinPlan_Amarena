"""Importador del Balance Sheet (Estado de Situación Financiera) desde el
Excel de contabilidad/Integrity ("Month Ending ... CWL Balance & CF.xlsx").

Lee la hoja **Balance Sheet**, bloque **"Balance Sheet Summary"** (versión
gerencial: A/R, Bancos, Inventario, Land, Buildings … Total assets / Total
liabilities / Capital), que cuadra Total assets = Capital & Liab. Sum.

El layout es multi-período: grupos de 3 columnas por mes (CRC, USD, tipo de
cambio). Las fechas de cada período están en la fila 8; los headers
CRC/USD/rate en la fila 9. Devuelve, por (año, mes), las líneas del estado en
orden, con indentación (jerarquía), sección (ASSET/LIABILITY/EQUITY), flag de
subtotal y montos USD + CRC.
"""
from __future__ import annotations
import datetime as _dt
from typing import Optional

SHEET_NAME = "Balance Sheet"
SUMMARY_MARKER = "balance sheet summary"
END_MARKERS = ("capital & liab", "capital and liab")

# Encabezados de sección reconocidos en el bloque Summary (en minúsculas).
SECTION_OF = {
    "assets": "ASSET",
    "liabilities": "LIABILITY",
    "capital": "EQUITY",
}


def _parse_period(v) -> Optional[tuple[int, int]]:
    """Devuelve (año, mes) de una celda de fecha (datetime o texto)."""
    if isinstance(v, _dt.datetime):
        return v.year, v.month
    if isinstance(v, _dt.date):
        return v.year, v.month
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                d = _dt.datetime.strptime(s, fmt)
                return d.year, d.month
            except ValueError:
                continue
    return None


def _num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").strip()
        if s in ("", "-"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _indent(raw: str) -> int:
    """Nivel de jerarquía = sangría (cada 3 espacios ≈ 1 nivel)."""
    n = len(raw) - len(raw.lstrip())
    return n // 3


def parse_balance_sheet(path_or_file) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path_or_file, data_only=True, read_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == SHEET_NAME.lower():
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(f"No se encontró la hoja '{SHEET_NAME}' en el archivo.")

    grid = [list(r) for r in ws.iter_rows(values_only=True)]  # 0-based
    ncols = max((len(r) for r in grid), default=0)
    for r in grid:
        r.extend([None] * (ncols - len(r)))

    def cell(r1: int, c1: int):  # 1-based como Excel
        if 1 <= r1 <= len(grid) and 1 <= c1 <= ncols:
            return grid[r1 - 1][c1 - 1]
        return None

    # ── Detección de formato ─────────────────────────────────────────────────
    # "Clean": una columna USD por mes, con una fila cabecera 'Account' + fechas
    # (export gerencial ya armado por el usuario). "Integrity": grupos de 3
    # columnas por mes (CRC, USD, tipo de cambio), fechas en fila 8.
    account_row = None
    for r in range(1, len(grid) + 1):
        b = cell(r, 2)
        if isinstance(b, str) and b.strip().lower() == "account":
            if any(_parse_period(cell(r, c)) for c in range(3, ncols + 1)):
                account_row = r
                break
    if account_row is not None:
        return _parse_clean(cell, grid, ncols, account_row)
    return _parse_integrity(cell, grid, ncols)


def _emit_lines(cell, grid, ncols, periods, get_vals) -> list:
    """Recorre el bloque Summary y construye las líneas. `periods` = lista de
    dicts con year/month; `get_vals(r)` devuelve {(y,m): (usd, crc)}."""
    start = 1
    for r in range(1, len(grid) + 1):
        b = cell(r, 2)
        if isinstance(b, str) and b.strip().lower().startswith(SUMMARY_MARKER):
            start = r + 1
            break
    lines = []
    section = ""
    order = 0
    last_label = ""
    blanks = 0
    r = start
    while r <= len(grid):
        raw = cell(r, 2)
        if raw is None or str(raw).strip() == "":
            blanks += 1
            if blanks >= 8:
                break
            r += 1
            continue
        blanks = 0
        raw = str(raw)
        label = raw.strip()
        low = label.lower()
        if low == "account":           # filas cabecera repetidas
            r += 1
            continue
        ind = _indent(raw.rstrip())
        if low in SECTION_OF:
            section = SECTION_OF[low]
        is_total = low.startswith("total") or low in SECTION_OF or any(m in low for m in END_MARKERS)
        vals = get_vals(r)
        has_val = any(v[0] is not None or v[1] is not None for v in vals.values())
        # de-duplicar encabezados de sección repetidos y sin valor
        if not has_val and label == last_label:
            r += 1
            continue
        lines.append({"order": order, "label": label, "indent": ind,
                      "section": section, "is_total": is_total, "values": vals})
        order += 1
        last_label = label
        if any(m in low for m in END_MARKERS):
            break
        r += 1
    return lines


def _parse_clean(cell, grid, ncols, account_row) -> dict:
    periods = []
    seen: set[tuple[int, int]] = set()
    for c in range(3, ncols + 1):
        ym = _parse_period(cell(account_row, c))
        if not ym or ym in seen:
            continue
        seen.add(ym)
        periods.append({"col": c, "year": ym[0], "month": ym[1]})
    if not periods:
        raise ValueError("No se detectaron columnas de mes en el Balance Sheet.")

    def get_vals(r):
        return {(p["year"], p["month"]): (_num(cell(r, p["col"])), None) for p in periods}

    lines = _emit_lines(cell, grid, ncols, periods, get_vals)
    return {"periods": [{"year": p["year"], "month": p["month"]} for p in periods], "lines": lines}


def _parse_integrity(cell, grid, ncols) -> dict:
    periods = []
    seen: set[tuple[int, int]] = set()
    for c in range(1, ncols + 1):
        if str(cell(9, c)).strip().upper() != "USD":
            continue
        crc_c = c - 1
        ym = _parse_period(cell(8, crc_c)) or _parse_period(cell(8, c))
        if not ym or ym in seen:
            continue
        seen.add(ym)
        periods.append({"usd_col": c, "crc_col": crc_c, "year": ym[0], "month": ym[1]})
    if not periods:
        raise ValueError("No se detectaron columnas de período (CRC/USD) en el Balance Sheet.")

    def get_vals(r):
        return {(p["year"], p["month"]): (_num(cell(r, p["usd_col"])), _num(cell(r, p["crc_col"])))
                for p in periods}

    lines = _emit_lines(cell, grid, ncols, periods, get_vals)
    return {"periods": [{"year": p["year"], "month": p["month"]} for p in periods], "lines": lines}
