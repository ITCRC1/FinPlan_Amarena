"""
Importador de MÁXIMO DETALLE — GL por cuenta × departamento × mes (Fase detalle).

A diferencia de pl_snapshot_importer (P&L por línea), este lee el archivo "upload"
con detalle de cuenta: col A=clase, col B=departamento (nombre), col C=código de
cuenta (4 díg), col D=nombre de cuenta, y valores mensuales en los 5 bloques
(rotulados en la fila 15, igual que el snapshot).

**Qué clases entran, al 2026-08-14.** Las CINCO financieras: 4xxx ingreso,
5xxx costo, 6xxx planilla, 7xxx opex y 8xxx below-GOP. `consolidate_block()` las
pasa por el mismo motor que los actuals y arma el P&L completo — el below-GOP
(renta, fee, impuesto) sale de las 8xxx reales vía el mapeo, validado al dólar
contra el Dashboard. Por eso «solo el detalle» produce el summary.

⚠️ Este encabezado decía que 4xxx y 8xxx **se ignoraban** «porque el P&L ya los
tiene por el snapshot de línea». Eso dejó de ser cierto cuando el detalle pasó a
ser la única carga, y quedó acá meses diciendo lo contrario de lo que hace el
código. Un comentario viejo miente con la misma cara que uno bueno.

De la clase 9 se leen las tres de siempre (9010/9020/9060) a `stats`, y **el
resto se recoge en `stats_9` pero HOY NADIE LO GUARDA** — ver `filas_clase9()`.

Parser PURO. `parse_gl_detail(bytes)` → bloques con filas {dept_code, account_code,
account_name, months[12]} por clase.
"""
import io
import re
from decimal import Decimal
import openpyxl

from app.importers.verificacion import codigo_de_verificacion as _codigo_de_verificacion

LABEL_COL = 4
DEPT_COL = 2
ACCT_COL = 3
ACCTNAME_COL = 4
FIRST_DATA_COL = 5
BLOCK_WIDTH = 12
VERSION_ROW = 15
MONTH_ROW = 14

_MONTH_NAMES = {
    "january": 1, "enero": 1, "february": 2, "febrero": 2, "march": 3, "marzo": 3,
    "april": 4, "abril": 4, "may": 5, "mayo": 5, "june": 6, "junio": 6,
    "july": 7, "julio": 7, "august": 8, "agosto": 8, "september": 9, "septiembre": 9,
    "october": 10, "octubre": 10, "november": 11, "noviembre": 11, "december": 12, "diciembre": 12,
}
_CLASE_WORDS = {"revenue", "payroll", "opex", "cost", "costs", "operating", "overhead",
                "ingreso", "ingresos", "nomina", "planilla", "gasto", "gastos"}


def _month_idx(v):
    if v is None:
        return None
    if hasattr(v, "month"):
        return v.month
    if isinstance(v, (int, float)):
        iv = int(v)
        return iv if 1 <= iv <= 12 else None
    return _MONTH_NAMES.get(str(v).strip().lower())


# Cuenta 6xxx → columna de concepto en PayrollConceptEntry
CONCEPT_BY_ACCT = {
    "6000": "c6000_sw", "6001": "c6001_overtime", "6002": "c6002_day_off",
    "6003": "c6003_working_holiday", "6004": "c6004_disabilities",
    "6010": "c6010_commissions", "6020": "c6020_ccss", "6021": "c6021_aguinaldo",
    "6022": "c6022_occ_hazard", "6023": "c6023_vacation_prov",
    "6024": "c6024_vacations_taken", "6025": "c6025_cafeteria",
    "6026": "c6026_severance", "6027": "c6027_incentive_bonus",
    "6028": "c6028_housing", "6029": "c6029_transport",
    "6030": "c6030_other", "6031": "c6030_other",
}

# Cuenta clase 9 (estadística) → campo de ScenarioStat. Rooms disponibles/ocupadas y
# huéspedes se suman por mes; ocupación (occ/avail) y ADR (rev rooms/occ) se derivan.
STAT_BY_ACCT = {"9010": "rooms_available", "9020": "rooms_occupied", "9060": "guests"}


#: **Quienes reparten, y en que clases.** Cafeteria (0220) reparte su costo,
#: planilla y opex; Lavanderia (0161) solo planilla y opex — su Laundry Services
#: (ingreso 4xxx + costo de venta 5xxx) es venta real y no se reparte.
#:
#: Este mapa **NO descarta nada**: dice quien es origen de reparto, y con eso el
#: P&L por Departamento le RESTA a cada uno lo que efectivamente repartio, por
#: clase y por mes (ver `scenarios_api`, correccion del 2026-08-27). Lo que sobra
#: queda en su propia fila, que para el 0161 y el 0220 cae en overhead.
DEPTOS_DE_REPARTO = {"0220": {"5", "6", "7"}, "0161": {"6", "7"}}

# VACIO DESDE EL 2026-08-28, POR DECISION DEL OWNER.
#
# «cafeteria y laundry tienen saldo — que salga ese saldo en overhead» · «si
# tiene saldo que lo vea como normal y que aparezca esa diferencia en overhead;
# hasta que se deje en 0, no pasa nada».
#
# Esto es OTRA COSA que `DEPTOS_DE_REPARTO`, y confundirlas fue el bug: aca se
# declara que filas **no llegan a la base**. Antes tenia el mismo contenido, asi
# que las clases 5/6/7 de cafeteria y lavanderia se descartaban en el parser y
# el sobrante no existia en ningun lado — ni para restarlo ni para verlo.
#
# Se vio subiendo los actuales de 2026: marzo y abril entraron (no traian estos
# departamentos) y mayo, junio y julio rebotaron con 409, porque el bloque de
# verificacion del archivo incluia el gasto de cafeteria y el detalle lo habia
# tirado. El archivo tenia razon y el importador no.
#
# Ahora se importa TODO. El neteo lo hace la aritmetica: el motor suma
# `planilla + costo + opex + reparto` por grupo, y `CAFETERIA` y `LAUNDRY_OPS`
# son grupos de OVERHEAD — si el reparto cubre el gasto la linea da cero y ni se
# dibuja; si sobra, el sobrante se ve.
#
# Queda declarado y vacio, no borrado: el dia que haya que volver a descartar
# algo, el lugar existe y esta explicado.
ALLOCATION_EXCLUDE: dict[str, set[str]] = {}

# Sets derivados de QUIEN REPARTE —no de que se descarta—. Los leen el P&L por
# Departamento (para restar lo repartido) y los auxiliares.
ALLOC_EXCL_COST = {d for d, cs in DEPTOS_DE_REPARTO.items() if "5" in cs}
ALLOC_EXCL_PAYROLL = {d for d, cs in DEPTOS_DE_REPARTO.items() if "6" in cs}
ALLOC_EXCL_OPEX = {d for d, cs in DEPTOS_DE_REPARTO.items() if "7" in cs}


def _acct_code(v) -> str | None:
    """Normaliza el código de cuenta venga como número (6020) o texto ('6000')."""
    if isinstance(v, (int, float)):
        s = str(int(v))
    elif isinstance(v, str):
        s = v.strip()
    else:
        return None
    return s if re.fullmatch(r"\d{4}", s) else None


#: Un codigo de departamento al PRINCIPIO de la celda: «0165 · Gift Shop».
#: Tres o cuatro digitos, porque el Club (260), el Area Recreativa (270) y
#: Miscelaneos (280) son de tres.
_CODIGO_AL_INICIO = re.compile(r"^\s*(\d{3,4})\s*(?:[·\-–—:|]|\s|$)")


_POR_PALABRA: list[tuple[str, str]] = [
    ("habitaci", "0110"),
    # Outlets de A&B. Los dos departamentos ya existen y cuelgan del 0120,
    # así que la plata cae en A&B igual que hoy — lo que se gana es que el
    # detalle por outlet NO se pierde. Hasta ahora un GL que dijera
    # «Restaurante» caía en «sin departamento» y se omitía (la Vista previa
    # lo avisa, pero omitido igual).
    #
    # Bar y Room Service NO están acá a propósito: todavía no existen como
    # departamento. Apuntar una palabra clave a un código inexistente sería
    # peor que omitir. Entran cuando se diseñe B2 — y ojo con «bar» como
    # subcadena, que pega en cualquier palabra que la contenga.
    ("restaurante", "0123"), ("restaurant", "0123"),
    ("cocina", "0122"), ("kitchen", "0122"),
    ("a&b", "0120"), ("alimentos", "0120"),
    # ⚠️ El Gift Shop es el 0165 y la Tienda el 0151: son dos locales
    # separados desde el 2026-08-13, con linea propia cada uno. Este
    # importador mandaba «gift» al 0151 y el del resumen al 0165.
    ("spa", "0140"), ("tours", "0150"), ("gift", "0165"), ("tienda", "0151"),
    ("transport", "0152"), ("innocean", "0155"), ("crowther", "0156"),
    ("crowler", "0156"), ("lavander", "0161"), ("administ", "0180"),
    ("ventas", "0190"), ("mercadeo", "0190"), ("marketing", "0190"),
    ("mantenim", "0200"), ("maintenance", "0200"), ("claro", "0205"),
    ("utility", "0210"), ("utilit", "0210"), ("cafeter", "0220"),
    # Employee Benefits es el 0181, no la Cafeteria. Ademas el 0220 se
    # descarta entero en el camino legacy: el gasto desaparecia.
    ("beneficios", "0181"),
    ("property", "0250"), ("propiedad", "0250"),  # Property/below-GOP (8xxx)
    ("miscel", "280"), ("sostenib", "280"),   # Miscelaneos (ingresos 48xx + Sustainability)
    ("madresal", "260"),     # Club Madresal (operativo)
    ("recreativa", "270"),   # Área Recreativa (operativo)
]


#: Los ÚNICOS departamentos que de verdad son de tres dígitos.
#:
#: ⚠️ **No es una lista nueva: se le pregunta al motor.** `_DEPT_TO_GROUP` es la
#: tabla que decide en qué grupo del P&L cae cada departamento; si un código no
#: está ahí, el motor lo manda a `OTHER_OVERHEAD`. Escribir acá una segunda
#: lista de códigos válidos sería exactamente cómo se separan dos reglas que
#: tienen que decir lo mismo.
def _tres_digitos_reales() -> frozenset[str]:
    """Los departamentos que de verdad son de tres dígitos.

    Se conserva porque lo usan las pruebas y `mapping_loader`; la regla vive en
    `app/departamentos.py`, que además aprende del catálogo.
    """
    from app.departamentos import _tres_digitos
    return _tres_digitos()


def _normalizar_dept(code: str) -> str:
    """`110` → `0110`; el Club (260) se queda como está.

    Owner, 2026-09-03: *«el upload tiene mismos departamentos sin 0»*.

    ⚠️ **La cuenta no se hace acá: se delega.** `_CODIGO_AL_INICIO` acepta tres
    o cuatro dígitos —el Club (260), el Área Recreativa (270) y Misceláneos
    (280) son de tres de verdad—, así que un «110 · Habitaciones» pasaba el
    filtro y se guardaba como `110`. Y `110` no está en el catálogo:

        pl_engine.group_for_dept("0110") -> ROOMS
        pl_engine.group_for_dept("110")  -> OTHER_OVERHEAD

    No revienta, no se descarta, no avisa: el gasto de Habitaciones sale como
    Overhead y el P&L cuadra igual.

    La regla vive en `app/departamentos.py` porque este importador **no es la
    única puerta** —hay al menos cuatro caminos que escriben un `dept_code`— y
    dos copias de esta regla es cómo se separan dos cosas que tienen que decir
    lo mismo.
    """
    from app.departamentos import normalizar_dept_code
    return normalizar_dept_code(code)


def dept_code_from_name(name: str) -> str | None:
    """Departamento del archivo → código oficial.

    **Primero busca el CÓDIGO al inicio de la celda.** La plantilla que genera la
    app escribe «0165 · Gift Shop», así que un archivo bajado y vuelto a subir
    resuelve EXACTO, sin adivinar.

    Eso importa por dos motivos. Uno: adivinar por palabra clave es fragil —
    «Restaurante A&B» tiene que dar 0123 y no 0120, y eso depende del orden de
    una tabla. Dos, y peor: un departamento SIN nombre en el catálogo —el 0240,
    por ejemplo— no tiene ninguna palabra que adivinar, así que al volver a subir
    la fila se quedaba sin departamento y se perdía. El código siempre está.

    Si no hay código al inicio, cae al fuzzy de siempre para los archivos que
    escribe el owner a mano.
    """
    m = _CODIGO_AL_INICIO.match(str(name or ""))
    if m:
        return _normalizar_dept(m.group(1))
    n = (name or "").lower()
    # ORDEN IMPORTANTE: gana la PRIMERA que pegue, así que lo específico va
    # antes que lo genérico. «Restaurante A&B» tiene que dar 0123, no 0120.
    table = _POR_PALABRA
    for kw, code in table:
        if kw in n:
            return code
    if re.search(r"\bti\b", n) or "tecnolog" in n:   # Departamento de TI
        return "0230"
    return None


def _num(v) -> Decimal | None:
    if v is None or isinstance(v, str):
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def _parse_version(label: str) -> tuple[str | None, int | None]:
    up = label.upper()
    typ = ("ACTUAL" if "ACTUAL" in up else "FORECAST" if "FORECAST" in up
           else "BUDGET" if "BUDGET" in up else None)
    ym = re.search(r"(20\d{2})", up)
    return typ, (int(ym.group(1)) if ym else None)


def parse_gl_detail(data: bytes) -> list[dict]:
    """Returns [{label, type, year, opex:[...], costs:[...], unmapped:set, skipped:{...}}].

    Cada fila de opex/costs: {dept_code, dept_name, account_code, account_name,
    months: {1..12: float}} (solo meses con valor)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = _pick_gl_sheet(wb)
    # max_col evita materializar miles de columnas fantasma (dimensión inflada).
    rows = list(ws.iter_rows(values_only=True, max_col=300))
    ncols = max((len(r) for r in rows), default=0)

    def cell(r0, c1):
        c = c1 - 1
        return rows[r0][c] if 0 <= r0 < len(rows) and 0 <= c < len(rows[r0]) else None

    # 1) primera columna de datos = la 1ª con MES en fila 14 y rótulo en fila 15
    first_col = None
    for c in range(1, ncols + 1):
        if _month_idx(cell(MONTH_ROW - 1, c)) is not None and str(cell(VERSION_ROW - 1, c) or "").strip():
            first_col = c
            break
    if first_col is None:
        first_col = FIRST_DATA_COL

    # 2) detectar columnas de meta (cuenta/depto/nombre) por CONTENIDO
    acct_col, dept_col, name_col = _detect_gl_columns(cell, len(rows), first_col)
    outlet_col = _detect_outlet_col(cell, first_col)

    # 3) bloques por (rótulo fila 15 + mes fila 14)
    blocks = []
    cur = None
    for c in range(first_col, ncols + 1):
        lbl = cell(VERSION_ROW - 1, c)
        mi = _month_idx(cell(MONTH_ROW - 1, c))
        if not lbl or not str(lbl).strip() or mi is None:
            continue
        s = str(lbl).strip()
        if cur is None or cur["label"] != s:
            typ, year = _parse_version(s)
            cur = {"label": s, "type": typ, "year": year, "colmap": {},
                   "revenue": [], "opex": [], "costs": [], "payroll": [], "belowgop": [],
                   "stats": {}, "unmapped": set(),
                   # Filas que TRAEN MONTO y no traen numero de cuenta. Ver el
                   # comentario del descarte, mas abajo: esto no es estadistica,
                   # es plata que no tiene a donde ir.
                   "sin_cuenta": [],
                   # TODAS las filas de clase 9 con monto, conocidas o no. Antes
                   # las desconocidas se tiraban calladas; ahora se recogen y
                   # quien llama decide.
                   "stats_9": [],
                   # El bloque de VERIFICACION de arriba: {codigo: {mes: monto}}.
                   # Es un CONTROL, no un origen — no suma en ningun total y
                   # nadie lo escribe en la base. Ver `app/importers/verificacion.py`.
                   "verificacion": {},
                   "skipped": {"payroll_noconcept": 0, "nodept": 0}}
            blocks.append(cur)
        cur["colmap"].setdefault(mi, c)

    meta_cols = list(range(1, max(first_col, 2)))
    for r0 in range(len(rows)):
        code = _acct_code(cell(r0, acct_col))
        # ── El bloque de VERIFICACION, arriba del encabezado ─────────────────
        #
        # Owner (2026-08-16): «que el upload tenga la verificacion arriba versus
        # el detalle abajo». Son los totales de control —ingresos, GOP, EBITDA,
        # utilidad neta y el desglose por bucket— contra los que se compara lo
        # que consolide el detalle de abajo.
        #
        # Se leen SOLO arriba del encabezado y SOLO si la fila no trae un codigo
        # de cuenta de cuatro digitos. Las dos condiciones son de seguridad:
        #
        #  · Arriba: un subtotal escrito DENTRO del detalle («TOTAL INGRESOS»)
        #    sigue siendo plata sin cuenta y se sigue rechazando. Si se aceptara
        #    en cualquier lado, una fila de subtotal se convertiria en un control
        #    y su monto dejaria de reclamarse: el agujero de los $40.613 otra vez.
        #  · Sin cuenta: una fila de cuenta nunca puede volverse un control.
        #
        # El monto NO entra a ninguna lista de plata. Se guarda aparte y quien
        # importa lo usa para comparar; nada lo escribe.
        if not code and r0 < VERSION_ROW - 1:
            ctrl = None
            for c in meta_cols:
                ctrl = _codigo_de_verificacion(cell(r0, c))
                if ctrl:
                    break
            if ctrl:
                for blk in blocks:
                    for m, cc in blk["colmap"].items():
                        v = _num(cell(r0, cc))
                        if v is not None:
                            blk["verificacion"].setdefault(ctrl, {})[m] = float(v)
                continue
        if not code:
            # Una fila sin codigo de cuatro digitos suele ser un encabezado, un
            # subtotal o una linea en blanco: saltarla esta bien.
            #
            # **Pero si TRAE MONTO es otra cosa.** Es plata que no tiene a donde
            # ir, y descartarla callado es como se perdieron los $40,613.30 del
            # gasto de Habitaciones en el Actual 2024 (owner, 2026-08-14): dos
            # renglones sin cuenta, en noviembre y diciembre, que el importador
            # se trago sin decir nada. El P&L siguio cuadrando consigo mismo y el
            # descuadre solo aparecio meses despues, comparando contra el
            # auxiliar a ojo.
            #
            # Se registran por bloque —cada version del archivo tiene sus propias
            # columnas de mes— y quien llama decide: hoy, el importador se NIEGA.
            dept_sc = str(cell(r0, dept_col) or "").strip()
            nombre_sc = str(cell(r0, name_col) or "").strip()
            for blk in blocks:
                montos = {}
                for m, c in blk["colmap"].items():
                    v = _num(cell(r0, c))
                    if v is not None and v != 0:
                        montos[m] = float(v)
                if montos:
                    blk["sin_cuenta"].append({
                        "fila": r0 + 1,
                        "departamento": dept_sc,
                        "descripcion": nombre_sc,
                        "meses": montos,
                        "total": sum(montos.values()),
                    })
            continue
        cls = code[0]
        dept_name = cell(r0, dept_col)
        acct_name = str(cell(r0, name_col) or "").strip()
        # clean "X CostCostos Y" style noise
        acct_name = re.sub(r"Costos.*$", "", acct_name).strip() or acct_name

        # Cuenta 4xxx de "Distribución" = allocation interna para dejar el depto en
        # cero; NO es ingreso → fuera del revenue (confirmado con el usuario 2026-06-27).
        if es_contrapartida_de_allocation(code, acct_name):
            continue

        for blk in blocks:
            # Clase 9 = estadísticas. Las tres de siempre (9010/9020/9060)
            # alimentan ScenarioStat, sin depto; ocupación y ADR se derivan al
            # consolidar.
            #
            # ⚠️ Todas las demás se recogen APARTE, en `stats_9`, en vez de
            # tirarse. Hasta 2026-08-14 este bloque hacía `continue` sin más:
            # una cuenta clase 9 que no fuera una de esas tres se descartaba en
            # silencio absoluto —no entraba a `unmapped`, no entraba a
            # `sin_cuenta`, no salía en la vista previa— y el catálogo tiene
            # miles de ellas. Es el espejo del bug de los $40,613: allá la fila
            # no traía cuenta y hoy truena; acá traía cuenta y aun así se perdía.
            #
            # Acá solo se RECOGEN. Quién es conocida y quién no lo decide la capa
            # que tiene la base delante, contra `stat_accounts`.
            if cls == "9":
                smonths = {}
                for m, c in blk["colmap"].items():
                    v = _num(cell(r0, c))
                    if v is not None:
                        smonths[m] = float(v)
                if not smonths:
                    continue
                field = STAT_BY_ACCT.get(code)
                if field:
                    blk["stats"].setdefault(field, {}).update(smonths)
                blk["stats_9"].append({
                    "fila": r0 + 1,
                    "account_code": code,
                    "descripcion": acct_name,
                    "departamento": str(dept_name or "").strip(),
                    "dept_code": dept_code_from_name(str(dept_name or "")) or "",
                    "meses": smonths,
                    "legado": field or "",
                })
                continue
            if cls not in ("4", "5", "6", "7", "8"):
                continue
            dcode = dept_code_from_name(str(dept_name or ""))
            if not dcode and cls == "8":
                # El below-GOP sin departamento va al 0250, que es donde viven
                # sus reglas. El 0240 NO EXISTE en el catalogo ni tiene una sola
                # regla: todo lo suyo resolvia por descarte y funcionaba de
                # casualidad, porque el 0250 es el unico depto con cuentas 8xxx.
                dcode = "0250"
            if not dcode:
                blk["unmapped"].add(str(dept_name)); blk["skipped"]["nodept"] += 1
                continue
            # Deptos de allocation: se excluyen las clases que se reparten (ver
            # ALLOCATION_EXCLUDE). 0220 = todo (5,6,7); 0161 = solo planilla+insumos
            # (6,7), conservando su Laundry Services operativo (ingreso 4xxx + costo 5xxx).
            if cls in ALLOCATION_EXCLUDE.get(dcode, set()):
                blk["skipped"]["allocation"] = blk["skipped"].get("allocation", 0) + 1
                continue
            months = {}
            for m, c in blk["colmap"].items():
                v = _num(cell(r0, c))
                if v is not None and v != 0:
                    months[m] = float(v)
            if not months:
                continue
            if cls == "6":
                concept = CONCEPT_BY_ACCT.get(code)
                if not concept:
                    blk["skipped"]["payroll_noconcept"] += 1; continue
                blk["payroll"].append({"dept_code": dcode, "dept_name": str(dept_name),
                                       "concept": concept, "months": months,
                                       "fila": r0 + 1})
                continue
            # Outlet: el punto de venta dentro del departamento. Solo A&B lo
            # trae; en el resto viene vacío. Se guarda tal cual lo rotula el GL.
            outlet = ""
            if outlet_col:
                outlet = str(cell(r0, outlet_col) or "").strip()
            row = {"dept_code": dcode, "dept_name": str(dept_name),
                   "account_code": code, "account_name": acct_name,
                   "outlet": outlet, "months": months,
                   # La fila del Excel del owner. Se guarda para devolverle la
                   # plantilla EN SU ORDEN: la base no tiene otra forma de saber
                   # como venia ordenado el archivo, y reordenar por criterio
                   # propio le obliga a cruzar dos listas cada vez que compara.
                   "fila": r0 + 1}
            target_list = {"4": "revenue", "5": "costs", "7": "opex", "8": "belowgop"}[cls]
            blk[target_list].append(row)

    for blk in blocks:
        blk.pop("colmap", None)
        blk["unmapped"] = sorted(blk["unmapped"])
    return blocks


def es_contrapartida_de_allocation(account_code, account_name) -> bool:
    """El credito del asiento con que Cafeteria y Lavanderia reparten su costo.

    Son clase 4 pero NO son ingreso: contarlas como tal duplicaria el reparto.
    Por eso el parser las salta.

    ⚠️ **Esta funcion es la unica definicion de la regla.** La importacion la usa
    ademas para NO borrarlas al reemplazar (`scenarios_api`): como el archivo
    nunca las trae, un reemplazo completo se las llevaba — bajar la plantilla,
    corregir una celda y volver a subir borraba $196 mil de allocations, y el P&L
    seguia cuadrando consigo mismo. Si esta regla y la del borrado se separan,
    vuelve el mismo agujero.
    """
    # «distribu» y no «distribuci»: la misma cuenta esta rotulada en los dos
    # idiomas — «Distribuciòn» (4900/4901) y «Expense Distribution» (4999)—, y
    # con la regla en español la inglesa se colaba bajo INGRESO, aparentando ser
    # un ingreso mas. Owner (2026-08-14): «hay 2 cuentas, debe haber una».
    #
    # Se verifico contra el catalogo: ninguna cuenta de ingreso REAL lleva
    # «distribu» en el nombre, asi que ensanchar la regla no se lleva plata por
    # delante — que es el error que seria peor que el original.
    return (str(account_code or "").startswith("4")
            and "distribu" in str(account_name or "").lower())


def filas_sin_cuenta(blocks: list[dict]) -> list[dict]:
    """Todas las filas con monto y sin numero de cuenta, de todos los bloques.

    Se devuelve plano y con el rotulo del bloque, porque el que las tiene que
    corregir abre el Excel una sola vez: separarlas por version lo obligaria a
    cruzar dos listas.
    """
    fuera = []
    for blk in blocks:
        for f in blk.get("sin_cuenta", []):
            fuera.append({**f, "version": blk.get("label", "")})
    return fuera


def filas_clase9(blocks: list[dict]) -> list[dict]:
    """Todas las filas de clase 9 con monto, de todos los bloques.

    Incluye las tres de siempre (llevan `legado` lleno) y las que no. Quien
    llama compara contra `stat_accounts` para saber cuáles reconoce: acá no hay
    base de datos.

    ⚠️ **HOY NADIE LLAMA A ESTA FUNCIÓN.** El parser recoge las clase 9 del
    archivo y el endpoint de importación no las lee, así que se pierden al
    terminar el request. En la práctica: si alguien llena las estadísticas en el
    archivo del GL, **la importación dice que salió bien y las estadísticas
    quedan vacías**. Antes se descartaban en el parser; ahora se descartan una
    capa más adelante. El arreglo quedó a medias.

    La otra puerta —Master Data → Estadísticas, con su propia plantilla— sí las
    guarda. Antes de cablear esta, hay que decidir si se quieren DOS puertas para
    el mismo dato: dos caminos que escriben lo mismo terminan discrepando.
    """
    fuera = []
    for blk in blocks:
        for f in blk.get("stats_9", []):
            fuera.append({**f, "version": blk.get("label", "")})
    return fuera


def consolidate_block(blk: dict, mappings: list[dict], report_lines: list[dict],
                      filas_extra: dict[int, list[dict]] | None = None) -> dict:
    """Consolida un bloque del GL (cuentas 4-8 + planilla) al P&L por línea usando
    el MISMO motor que los actuals (`calculate_pl_from_mapping`), mes a mes, y deriva
    las estadísticas. Función PURA (sin DB) → reutilizable en import y en dry-run.

    Devuelve {'lines': {mes: {line_code: Decimal}}, 'stats': {mes: {campo: Decimal}}}.
    Así "solo el detalle" produce el summary: el below-GOP (renta, fee, impuesto) sale
    de las cuentas 8xxx reales vía el mapeo, tal como se validó contra el Dashboard.

    ⚠️ `filas_extra` — {mes: [{account_code, dept_code, amount}]} — son las filas
    que el archivo NO trae y el reporte SÍ cuenta: las contrapartidas de los
    repartos (`4900`/`4901`/`4999`, «Distribución») las escribe el MOTOR y
    sobreviven al reemplazo justamente porque el archivo no puede traerlas.

    Sin ellas, consolidar «solo lo que subiste» da un número que el reporte
    nunca va a dar, y la verificación fallaría sin que haya un solo error: en
    Rooms del Budget Working 2027 son −92.176,74 y en las dos contrapartidas de
    Lavandería y Cafetería −196.326,17. Medir donde se digita, y no donde
    escribe el motor, ya costó eso dos veces.
    """
    from app.engine import pl_engine

    per_month: dict[int, list[dict]] = {}
    def add(month, code, dcode, amt):
        per_month.setdefault(month, []).append(
            {"account_code": code, "dept_code": dcode, "amount": Decimal(str(amt))})
    for key in ("revenue", "costs", "opex", "belowgop"):
        for r in blk.get(key, []):
            for m, v in r["months"].items():
                add(m, r["account_code"], r["dept_code"], v)
    for r in blk.get("payroll", []):
        acct = pl_engine.payroll_account_for_column(r["concept"])
        if not acct:
            continue
        for m, v in r["months"].items():
            add(m, acct, r["dept_code"], v)
    for m, filas in (filas_extra or {}).items():
        for f in filas:
            add(m, f["account_code"], f["dept_code"], f["amount"])

    lines: dict[int, dict] = {}
    for m, arows in per_month.items():
        res = pl_engine.calculate_pl_from_mapping(arows, mappings, report_lines)
        lines[m] = {L.line_code: L.amount_usd for L in res}

    # Estadísticas: disponibles/ocupadas/huéspedes directos; ocupación y ADR derivados.
    st = blk.get("stats", {})
    stat_months = set()
    for f in ("rooms_available", "rooms_occupied", "guests"):
        stat_months |= set(st.get(f, {}).keys())
    stats: dict[int, dict] = {}
    for m in stat_months:
        avail = st.get("rooms_available", {}).get(m)
        occ = st.get("rooms_occupied", {}).get(m)
        guests = st.get("guests", {}).get(m)
        d: dict[str, Decimal] = {}
        if avail is not None:
            d["rooms_available"] = Decimal(str(avail))
        if occ is not None:
            d["rooms_occupied"] = Decimal(str(occ))
        if guests is not None:
            d["guests"] = Decimal(str(guests))
        if avail and occ:
            d["occupancy_pct"] = Decimal(str(occ)) / Decimal(str(avail))
        if occ:
            # ADR = renta de habitación / ocupadas. REGLA por año (confirmado por el
            # owner): 2026 en adelante = SOLO la cuenta "Rooms" (excluye No Show,
            # Cancellations, otros ingresos del depto). EXCEPCIÓN histórica: ene-2024 a
            # dic-2025 quedó sobre TODO el revenue del depto Rooms — se respeta tal cual.
            year = blk.get("year") or 0
            # Budget también va sobre el total: solo se presupuesta rooms revenue puro
            # (el depto = Rooms puro). 2024-2025 = histórico sobre el depto.
            whole_dept = year <= 2025 or blk.get("type") == "BUDGET"
            if whole_dept:
                rooms_rev = sum(
                    (Decimal(str(v)) for r in blk.get("revenue", [])
                     for mm, v in r["months"].items()
                     if mm == m and pl_engine.group_for_dept(r["dept_code"]) == "ROOMS"),
                    Decimal(0))
            else:
                rooms_rev = sum(
                    (Decimal(str(v)) for r in blk.get("revenue", [])
                     for mm, v in r["months"].items()
                     if mm == m and (r.get("account_name") or "").strip().lower() == "rooms"),
                    Decimal(0))
            if rooms_rev:
                d["adr"] = rooms_rev / Decimal(str(occ))
        stats[m] = d
    return {"lines": lines, "stats": stats}


def _pick_gl_sheet(wb):
    """Elige la hoja con detalle GL: se PRUEBA cada hoja con el parser y se queda
    con la que mapea más cuentas a departamento (menos 'nodept'). Así funciona
    tanto si el detalle está en la hoja del P&L (debajo del resumen) como en una
    hoja aparte ('Data'), sin depender del layout de columnas."""
    best, best_mapped = None, -1
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True, max_col=300))
        ncols = max((len(r) for r in rows), default=0)
        if ncols < FIRST_DATA_COL:
            continue
        mapped = _count_gl_mapped(rows, ncols)
        if mapped > best_mapped:
            best, best_mapped = ws, mapped
    return best or wb.active


def _count_gl_mapped(rows, ncols) -> int:
    """Cuántas filas de cuenta mapean a un departamento con la detección de
    columnas — heurística barata para elegir la hoja de detalle correcta."""
    def cell(r0, c1):
        c = c1 - 1
        return rows[r0][c] if 0 <= r0 < len(rows) and 0 <= c < len(rows[r0]) else None
    first_col = None
    for c in range(1, ncols + 1):
        if _month_idx(cell(MONTH_ROW - 1, c)) is not None and str(cell(VERSION_ROW - 1, c) or "").strip():
            first_col = c
            break
    if first_col is None:
        return 0
    acct_col, dept_col, _ = _detect_gl_columns(cell, len(rows), first_col)
    n = 0
    for r0 in range(len(rows)):
        if _acct_code(cell(r0, acct_col)) and dept_code_from_name(str(cell(r0, dept_col) or "")):
            n += 1
    return n


def _detect_outlet_col(cell, first_col: int):
    """Columna de OUTLET (punto de venta), si el archivo la trae. None si no.

    Va aparte de `_detect_gl_columns` a propósito: esa función la llaman dos
    lugares y cambiarle la firma para una columna opcional es más riesgo que
    beneficio. Acá solo se busca el rótulo, sin heurística por contenido — si
    alguien renombra la columna, se prefiere no capturar outlet antes que
    adivinar mal y meter basura en la llave única.
    """
    for c in range(1, max(first_col, 2)):
        if "outlet" in str(cell(VERSION_ROW - 1, c) or "").strip().lower():
            return c
    return None


def _detect_gl_columns(cell, nrows: int, first_col: int):
    """Detecta qué columna es cuenta / departamento / nombre de cuenta. Primero por
    ENCABEZADO (fila 15 de las columnas meta) si el archivo los trae; si no, por
    CONTENIDO. Soporta el layout reconstruido (dept2·cuenta3·nombre4), el 'Data' del
    owner (cuenta3·nombre5·dept10) y el archivo con encabezados Clase/…/Nombre."""
    meta = list(range(1, max(first_col, 2)))
    # 1) Por encabezado en la fila VERSION_ROW (solo si están los tres rótulos).
    hdr = {c: str(cell(VERSION_ROW - 1, c) or "").strip().lower() for c in meta}
    h_name = next((c for c in meta if "nombre" in hdr[c]), None)
    h_dept = next((c for c in meta if "departamento" in hdr[c] or hdr[c] == "depto"), None)
    h_acct = next((c for c in meta if hdr[c] == "cuenta" or hdr[c].startswith("cuenta")
                   or hdr[c] in ("codigo", "código", "account")), None)
    if h_name and h_dept and h_acct and len({h_name, h_dept, h_acct}) == 3:
        return h_acct, h_dept, h_name
    # 2) Fallback por CONTENIDO (filas de datos).
    score_acct = {c: 0 for c in meta}
    score_dept = {c: 0 for c in meta}
    text_cols = {c: 0 for c in meta}
    for r0 in range(15, min(nrows, 415)):
        for c in meta:
            v = cell(r0, c)
            if v is None:
                continue
            if _acct_code(v):
                s0 = str(int(v)) if isinstance(v, float) else str(v).strip()
                if re.match(r"^\d{4}$", s0):
                    score_acct[c] += 1
            s = str(v).strip()
            if "departamento" in s.lower() or dept_code_from_name(s):
                score_dept[c] += 1
            if isinstance(v, str) and s and not s.replace(".", "").isdigit():
                text_cols[c] += 1
    pick = lambda sc, default: (max(sc, key=sc.get) if sc and max(sc.values()) > 0 else default)
    acct_col = pick(score_acct, ACCT_COL)
    dept_col = pick(score_dept, DEPT_COL)
    for c in (acct_col, dept_col):
        text_cols.pop(c, None)
    name_col = pick(text_cols, ACCTNAME_COL)
    return acct_col, dept_col, name_col
