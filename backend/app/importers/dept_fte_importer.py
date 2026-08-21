# -*- coding: utf-8 -*-
"""Parser del Excel de FTE real por departamento (`export/dept_fte_excel.py`).

Bloques por mes: encabezado con 'Código' en la columna B, filas de depto con
Código (B), Departamento (C), FTE (D), y el nombre del mes en la columna E de
cada fila de datos. Devuelve filas planas {month, dept_code, dept_name, fte}
— solo de bloques con mes detectado (ignora cualquier fila 'TOTAL').
"""
from __future__ import annotations
from io import BytesIO

_MONMAP = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def _month_num(v) -> int | None:
    if not v:
        return None
    s = str(v).strip().lower()
    return _MONMAP.get(s)


def _num(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_dept_fte(file_bytes: bytes, sheet_name: str = "FTE Real") -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        cand = [s for s in wb.sheetnames if "fte" in s.lower()]
        if not cand:
            raise ValueError(f"No se encontró la hoja '{sheet_name}'. Hojas: {wb.sheetnames}")
        sheet_name = cand[0]
    ws = wb[sheet_name]

    out: list[dict] = []
    r = 1
    maxr = ws.max_row
    while r <= maxr:
        b = ws.cell(r, 2).value
        if b and "Código" in str(b):
            rr = r + 1
            rows = []
            while rr <= maxr:
                bb = ws.cell(rr, 2).value
                if bb and (str(bb).strip().upper() == "TOTAL" or "Código" in str(bb)):
                    break
                if bb:
                    rows.append({
                        "dept_code": str(bb).strip(),
                        "dept_name": str(ws.cell(rr, 3).value or "").strip(),
                        "fte": _num(ws.cell(rr, 4).value),
                        "month_raw": ws.cell(rr, 5).value,
                    })
                rr += 1
            for row in rows:
                mn = _month_num(row["month_raw"])
                if mn:
                    out.append({"month": mn, "dept_code": row["dept_code"],
                                "dept_name": row["dept_name"], "fte": row["fte"]})
            r = rr
        else:
            r += 1
    return out
