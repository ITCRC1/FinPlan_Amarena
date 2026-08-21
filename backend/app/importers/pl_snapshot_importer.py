"""
Importador del archivo "upload" — P&L mensual consolidado de varios escenarios.

El archivo (una hoja) trae N bloques de 12 meses, cada uno auto-rotulado en la
fila 15 (ej. "Actual 2024", "Forecast Apr 2026", "Budget 2026"). Cada bloque es
un P&L a nivel de línea (REVENUE/Operating Expenses/... por depto) + estadísticas
de habitaciones arriba.

`parse_pl_snapshot(bytes)` → lista de bloques con stats + líneas por mes, listos
para escribir como ScenarioStat + ActualPLLine. Parser PURO (sin DB).

Layout (1-indexed): col 4 = etiqueta; bloque b (0..N-1) ocupa las 12 columnas
[5 + b*12 .. 16 + b*12]; mes m → columna 5 + b*12 + (m-1). Fila 15 col del bloque
= rótulo de versión. Filas de stats: 3 avail, 4 occ, 6 guests, 7 occ%, 8 ADR.
"""
import io
import re
from decimal import Decimal
import openpyxl

LABEL_COL = 4  # 1-indexed column with the row label
BLOCK_WIDTH = 12
FIRST_DATA_COL = 5
VERSION_ROW = 15
MONTH_ROW = 14   # fila con el nombre/fecha del mes de cada columna

_MONTH_NAMES = {
    "january": 1, "enero": 1, "february": 2, "febrero": 2, "march": 3, "marzo": 3,
    "april": 4, "abril": 4, "may": 5, "mayo": 5, "june": 6, "junio": 6,
    "july": 7, "julio": 7, "august": 8, "agosto": 8, "september": 9, "septiembre": 9,
    "october": 10, "octubre": 10, "november": 11, "noviembre": 11, "december": 12, "diciembre": 12,
}


def _month_idx(v):
    """Devuelve el mes 1..12 de una celda (nombre, fecha o número), o None si no
    es un mes válido (así se saltan columnas 'Full Year', 'VAR', vacías…)."""
    if v is None:
        return None
    if hasattr(v, "month"):          # datetime/date
        return v.month
    if isinstance(v, (int, float)):
        iv = int(v)
        return iv if 1 <= iv <= 12 else None
    s = str(v).strip().lower()
    return _MONTH_NAMES.get(s)


def _detect_blocks(cell, ncols: int):
    """Detecta bloques por (rótulo fila 15 + mes fila 14). Agrupa columnas
    consecutivas del mismo rótulo; cada columna se mapea a su mes. Robusto a
    bloques de ancho variable y columnas 'Full Year' intercaladas."""
    blocks: list[dict] = []
    cur = None
    for col in range(FIRST_DATA_COL, ncols + 1):
        label = cell(VERSION_ROW, col)
        mi = _month_idx(cell(MONTH_ROW, col))
        if not label or not str(label).strip() or mi is None:
            continue   # no es columna de datos de un bloque (spacer / Full Year)
        lab = str(label).strip()
        if cur is None or cur["label"] != lab:
            typ, year = _parse_version(lab)
            cur = {"label": lab, "type": typ, "year": year, "cols": {}}
            blocks.append(cur)
        cur["cols"].setdefault(mi, col)   # 1er col de ese mes gana
    return blocks

# Stats row → ScenarioStat field
STAT_ROWS = {
    3: "rooms_available",
    4: "rooms_occupied",
    6: "guests",
    7: "occupancy_pct",
    8: "adr",
}

# Section headers that switch the active dept-context
_SECTIONS = {
    "REVENUE": "REV",
    "OPERATING EXPENSES": "OPEXP",
    "OPERATING PROFIT": "OPPROFIT",
    "OVERHEAD EXPENSES": "OVH",
}
# Dept label (upper) → code suffix, within REV/OPEXP/OPPROFIT
#
# OJO: este mapa es OPCIONAL fila por fila — solo dice "si esta etiqueta aparece,
# se llama así". Ninguna entrada es obligatoria: un archivo que no traiga la fila
# simplemente no genera esa línea, y la validación de `scenarios_api._pl_block_checks`
# es por prefijo (Σ REV_* = TOTAL_REVENUES), no contra una lista fija. Por eso
# agregar Private Bar NO rompe los archivos viejos que no lo tienen.
_DEPT_SUFFIX = {
    "ROOMS": "ROOMS", "F&B": "FB",
    # Private Bar (depto 0121) — centro propio, va SIEMPRE después de A&B / F&B.
    "PRIVATE BAR": "PRIVATE_BAR", "PRIVATE-BAR": "PRIVATE_BAR",
    "SPA": "SPA", "TOURS": "TOURS",
    "RETAIL-GIFT SHOP": "RETAIL", "TRANSPORTATION": "TRANSPORT",
    "LAUNDRY": "LAUNDRY", "INNOCEANA": "INNOCEANA", "CROWTHER LAB": "CROWTHER",
    "MISCELLANEOUS": "SUSTAINABILITY", "MISCELLANEOS": "SUSTAINABILITY",
}
# Overhead dept label (upper) → full code
_OVH = {
    "ADMINISTRATIONS": "OVH_ADMIN", "SALES & MARKETING": "OVH_SALES",
    "MAINTENANCE": "OVH_MAINTENANCE", "INFORMATION SYSTEM": "OVH_IT",
    "UTILITIES": "OVH_UTILITIES",
}
# Standalone / total labels (upper) → exact code
_EXACT = {
    "TOTAL INCOMES": "TOTAL_REVENUES",
    "TOTAL OPERATIONG EXPENSES": "TOTAL_OPEXP",
    "OPERATING PROFIT": "TOTAL_OP_PROFIT",          # the standalone uppercase total (row 62)
    "TOTAL OVERHEAD EXPENSES": "TOTAL_OVERHEAD",
    "TOTAL GROSS OPERATING PROFIT": "GOP",
    "RENT": "RENT",
    "MANAGEMENT FEES (3%)": "MGMT_FEE",
    "MANAGEMENT FEES (5%) ROYALTIES": "ROYALTIES",
    "TOTAL RENTA AND MANAGEMENT FEES": "TOTAL_MGMT_FEES",
    "PROPERTY INSURANCE": "PROPERTIES_INSURANCE",
    "PROPERTIES INSURANCE": "PROPERTIES_INSURANCE",   # variante real del archivo del owner
    "EBITDA": "EBITDA_BEFORE",
    "FINANCIAL EXPENSES": "FINANCIAL_EXPENSES",
    "DEPRECIATION": "DEPRECIATION",
    "TOTAL DEPRECIATIONS": "TOTAL_DEPRECIATIONS",
    "CAPITAL RESERVE": "CAPITAL_RESERVE",
    "LARGE CAPITAL EXPENDITURE": "LARGE_CAPEX",
    "CAPITAL EXPENSE": "CAPITAL_EXPENSE",
    "EARNINGS BEFORE INCOME TAXES": "EBT",
    "INCOME TAX (30%)": "INCOME_TAXES",
    "EARNINGS AFTER INCOME TAXES": "NET_PROFIT",
}


def _num(v) -> Decimal | None:
    if v is None or isinstance(v, str):
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def _parse_version(label: str) -> tuple[str | None, int | None]:
    """'Forecast Apr 2026' → ('FORECAST', 2026)."""
    up = label.upper()
    typ = ("ACTUAL" if "ACTUAL" in up else
           "FORECAST" if "FORECAST" in up else
           "BUDGET" if "BUDGET" in up else None)
    ym = re.search(r"(20\d{2})", up)
    return typ, (int(ym.group(1)) if ym else None)


def _pick_pl_sheet(wb):
    """Elige la hoja del P&L Summary: la que tiene 'REVENUE' en la columna de
    rótulos (col 4). iter_rows acotado (read_only-friendly). Fallback: activa."""
    for ws in wb.worksheets:
        for r in ws.iter_rows(min_row=10, max_row=42, max_col=LABEL_COL, values_only=True):
            v = r[LABEL_COL - 1] if len(r) >= LABEL_COL else None
            if v and str(v).strip().upper() == "REVENUE":
                return ws
    return wb.active


def parse_pl_snapshot(data: bytes) -> list[dict]:
    """Parse the upload workbook into per-block data.

    Returns [{label, type, year, stats:{month:{field:val}}, lines:{month:{code:amount}}}].
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = _pick_pl_sheet(wb)
    # max_col evita materializar columnas fantasma (dimensión inflada de Excel).
    rows = list(ws.iter_rows(values_only=True, max_col=300))
    ncols = max((len(r) for r in rows), default=0)

    def cell(r1: int, c1: int):
        r, c = r1 - 1, c1 - 1
        if 0 <= r < len(rows) and 0 <= c < len(rows[r]):
            return rows[r][c]
        return None

    # Detectar bloques por (rótulo + mes) → cada bloque trae un mapa {mes: col}
    detected = _detect_blocks(cell, ncols)
    blocks = [{"label": d["label"], "type": d["type"], "year": d["year"],
               "colmap": d["cols"], "stats": {}, "lines": {}, "unmapped": []}
              for d in detected]

    for blk in blocks:
        colmap = blk["colmap"]
        months_here = sorted(colmap.keys())
        stats: dict[int, dict] = {m: {} for m in months_here}
        lines: dict[int, dict] = {m: {} for m in months_here}
        section: str | None = None
        mapped_rows: list[int] = []      # filas que SÍ amarraron (define la zona del summary)
        candidates: list[dict] = []      # filas con datos que NO amarraron

        for r1 in range(1, len(rows) + 1):
            label = cell(r1, LABEL_COL)
            # stats rows
            if r1 in STAT_ROWS:
                field = STAT_ROWS[r1]
                for m in months_here:
                    v = _num(cell(r1, colmap[m]))
                    if v is not None:
                        stats[m][field] = v
                continue
            if label is None:
                continue
            up = str(label).strip().upper()
            if not up:
                continue
            if up in _SECTIONS:
                section = _SECTIONS[up]
                continue
            code = None
            if up in _EXACT:
                code = _EXACT[up]
                section = None  # a total closes the running section
            elif section == "OVH" and up in _OVH:
                code = _OVH[up]
            elif section in ("REV", "OPEXP", "OPPROFIT") and up in _DEPT_SUFFIX:
                code = f"{section}_{_DEPT_SUFFIX[up]}"
            vals = [_num(cell(r1, colmap[m])) if m in colmap else None for m in range(1, 13)]
            if not code:
                # Etiqueta CON datos que no amarró a ninguna línea del summary.
                # Se guarda como candidata; después se filtra a la zona del summary
                # (si no, saldrían cientos de filas de detalle por cuenta).
                con = [v for v in vals if v is not None and v != 0]
                if con:
                    candidates.append({"row": r1, "label": str(label).strip(),
                                       "months_with_data": len(con),
                                       "total": float(sum(con))})
                continue
            mapped_rows.append(r1)
            for m in months_here:
                if vals[m - 1] is not None:
                    lines[m][code] = vals[m - 1]
            # El P&L Summary CIERRA en Net Profit. Todo lo que sigue es detalle por
            # cuenta, y ahí hay etiquetas que repiten códigos del summary (RENT,
            # PROPERTIES_INSURANCE, CAPITAL_RESERVE, LARGE_CAPEX, DEPRECIATION):
            # si se siguieran leyendo, PISARÍAN los valores buenos del summary.
            if code == "NET_PROFIT":
                break

        blk["stats"] = {m: s for m, s in stats.items() if s}
        blk["lines"] = {m: ls for m, ls in lines.items() if ls}
        # Solo reportamos lo no mapeado DENTRO de la zona del summary (hasta la
        # última fila que sí amarró); más abajo es el detalle por cuenta.
        limite = max(mapped_rows) if mapped_rows else 0
        blk["unmapped"] = [c for c in candidates if c["row"] <= limite]
        del blk["colmap"]

    return blocks
