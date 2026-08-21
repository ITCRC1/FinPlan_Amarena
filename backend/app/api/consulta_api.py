# -*- coding: utf-8 -*-
"""Consulta libre — el GL completo, para armar reportes propios.

**Qué pidió el owner (2026-08-14):** «conectá todo el GL, quiero poder generar
reportes por ejemplo horas extras, y jalo todas las cuentas».

Es el equivalente del tab «Power Query» que él mismo hizo en DAILY-OPS: se elige
un conjunto de datos, se filtra, y sale la tabla — para mirarla o bajarla a
Excel y pivotearla.

**No es SQL libre, y es a propósito.** Hay una lista blanca de conjuntos y
columnas. Un armador de consultas abierto sobre una base con planilla y salarios
es una puerta que no se puede cerrar después; con la lista, agregar algo es una
decisión explícita.

**El formato es LARGO, una fila por mes** —y no doce columnas— porque el destino
natural es una tabla dinámica de Excel. Una fila por cuenta × departamento × mes
se pivotea sin tocar nada; en formato ancho hay que despivotar primero, y ahí es
donde la gente se equivoca.

**Los meses en cero no salen.** El GL de un año tiene miles de combinaciones y la
mayoría vacías: devolverlas convierte un archivo de 4.000 filas en uno de 40.000
donde el 90% no dice nada.
"""
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from sqlalchemy import select

from app.errores import ErrorApi
from app.textos import Idioma, t
from app.auth import get_current_user
from app.db import get_session
from app.models.actual_entry import ActualEntry
from app.models.actual_pl_line import ActualPLLine
from app.models.belowgop_account_entry import BelowGopAccountEntry
from app.models.cost_entry import CostEntry
from app.models.opex_entry import OpexEntry
from app.models.payroll_concept_entry import PayrollConceptEntry
from app.models.scenario import Scenario

router = APIRouter()

MESES = ["jan", "feb", "mar", "apr", "may", "jun",
         "jul", "aug", "sep", "oct", "nov", "dec"]
MESES_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Set", "Oct", "Nov", "Dic"]


def _meses(idioma: str) -> list[str]:
    """Las abreviaturas de mes en el idioma de quien pide.

    El orden NO depende de esta columna —para eso está «Mes #»— así que
    traducirla no desordena ninguna dinámica.
    """
    return [t(idioma, f"consulta.mes.{i}") for i in range(1, 13)]


# Los 17 conceptos de planilla, con el número de cuenta que les corresponde. Se
# desarman en filas para que «horas extras» sea un filtro por cuenta como
# cualquier otro, y no una columna que hay que saber que existe.
CONCEPTOS = [
    ("c6000_sw", "6000", "Salary and Wages"),
    ("c6001_overtime", "6001", "Overtime"),
    ("c6002_day_off", "6002", "Day Off"),
    ("c6003_working_holiday", "6003", "Working Holiday"),
    ("c6004_disabilities", "6004", "Disabilities"),
    ("c6010_commissions", "6010", "Commissions"),
    ("c6020_ccss", "6020", "Social Security"),
    ("c6021_aguinaldo", "6021", "Aguinaldo"),
    ("c6022_occ_hazard", "6022", "Work Risk Policy"),
    ("c6023_vacation_prov", "6023", "Vacation Provision"),
    ("c6024_vacations_taken", "6024", "Vacations Taken"),
    ("c6025_cafeteria", "6025", "Cafeteria"),
    ("c6026_severance", "6026", "Severance"),
    ("c6027_incentive_bonus", "6027", "Incentive Bonus"),
    ("c6028_housing", "6028", "Housing"),
    ("c6029_transport", "6029", "Transport"),
    ("c6030_other", "6030", "Other Benefits"),
]

# El ROTULO de cada columna no vive acá: vive en `app/textos.py`, bajo
# `consulta.col.<key>`. La columna «Línea del P&L» tiene que decir lo mismo acá
# que en Admin · Account Mapping —de donde sale el dato—, y eso solo se sostiene
# si las dos leen del mismo catálogo en vez de tener cada una su propia palabra.
COLUMNAS = [
    {"key": "escenario", "tipo": "texto"},
    {"key": "anio", "tipo": "conteo"},
    {"key": "mes", "tipo": "texto"},
    {"key": "mes_num", "tipo": "conteo"},
    {"key": "dept_code", "tipo": "texto"},
    {"key": "dept_name", "tipo": "texto"},
    {"key": "account_code", "tipo": "texto"},
    {"key": "account_name", "tipo": "texto"},
    {"key": "clase", "tipo": "texto"},
    # Lo que el master data sabe de cada fila. Va en el archivo para que la
    # dinamica pueda agrupar por linea del P&L o por grupo de departamento sin
    # que nadie tenga que mantener una tabla de equivalencias aparte en Excel —
    # que es como se desincronizan los reportes (owner, 2026-08-14).
    {"key": "linea_pl", "tipo": "texto"},
    {"key": "linea_pl_nombre", "tipo": "texto"},
    {"key": "seccion", "tipo": "texto"},
    {"key": "grupo", "tipo": "texto"},
    {"key": "dept_padre", "tipo": "texto"},
    {"key": "tipo_dept", "tipo": "texto"},
    {"key": "outlet", "tipo": "texto"},
    # Solo se llenan en planilla. Van en TODOS los conjuntos igual para que el
    # archivo tenga una sola forma: una dinamica que a veces tiene la columna y
    # a veces no obliga a rehacerla cada vez que se cambia de conjunto.
    {"key": "position_code", "tipo": "texto"},
    {"key": "position_name", "tipo": "texto"},
    {"key": "employee", "tipo": "texto"},
    {"key": "detalle", "tipo": "texto"},
    {"key": "monto", "tipo": "moneda"},
]

#: Los conjuntos, en el orden en que se muestran. La nota es opcional: dos de
#: ellos no la tienen y siguen sin tenerla.
CONJUNTOS = ["gl", "opex", "costo", "planilla", "propietario", "pl"]
CONJUNTOS_CON_NOTA = {"gl", "opex", "planilla", "pl"}


def _columnas(idioma: str) -> list[dict]:
    return [{"key": c["key"], "label": t(idioma, f"consulta.col.{c['key']}"),
             "tipo": c["tipo"]} for c in COLUMNAS]


def _conjuntos(idioma: str) -> list[dict]:
    return [{"key": k,
             "label": t(idioma, f"consulta.conjunto.{k}"),
             "nota": (t(idioma, f"consulta.conjunto.{k}_nota")
                      if k in CONJUNTOS_CON_NOTA else "")}
            for k in CONJUNTOS]


def _fila(esc, mes_num, dept, dept_name, cuenta, nombre, monto,
          outlet="", detalle="", pos_code="", pos_name="", empleado="",
          maestro=None, meses=None):
    md = (maestro or {}).get(dept or "", {})
    _lc = (maestro or {}).get("__lineas__", {}).get(((cuenta or ""), (dept or "")), "")
    # En el conjunto de lineas del P&L, la "cuenta" ES la linea.
    if not _lc and (cuenta or "").upper() == (cuenta or "") and not (cuenta or "").isdigit():
        _lc = cuenta or ""
    return {
        "escenario": f"{esc.type} {esc.version} {esc.year}",
        "anio": esc.year,
        "mes": (meses or MESES_ES)[mes_num - 1],
        "mes_num": mes_num,
        "dept_code": dept or "",
        "dept_name": dept_name,
        "account_code": cuenta or "",
        "account_name": nombre or "",
        "clase": (cuenta or " ")[:1],
        "linea_pl": _lc,
        # El codigo solo no le dice nada a nadie: OPEXP_ROOMS contra «Rooms» en
        # la seccion «OPERATING EXPENSES». Con el nombre y la seccion, la
        # dinamica se puede agrupar como el P&L sin mantener equivalencias a mano.
        "linea_pl_nombre": (maestro or {}).get("__nombres_linea__", {}).get(_lc, ""),
        "seccion": (maestro or {}).get("__secciones__", {}).get(_lc, ""),
        "grupo": md.get("grupo", ""),
        "dept_padre": md.get("padre", ""),
        "tipo_dept": md.get("tipo", ""),
        "outlet": outlet or "",
        "position_code": pos_code or "",
        "position_name": pos_name or "",
        "employee": empleado or "",
        "detalle": detalle or "",
        "monto": float(monto),
    }


@router.get("/consulta/conjuntos/")
async def conjuntos(_=Depends(get_current_user), idioma: str = Idioma):
    """Qué se puede consultar, con sus columnas y los escenarios disponibles."""
    async with get_session() as s:
        escs = (await s.execute(select(Scenario).order_by(
            Scenario.year.desc(), Scenario.type))).scalars().all()
    return {
        "conjuntos": _conjuntos(idioma),
        "columnas": _columnas(idioma),
        "escenarios": [{"id": e.id, "label": f"{e.year} · {e.type} {e.version}",
                        "year": e.year} for e in escs],
    }


async def _filas(s, conjunto: str, esc: Scenario, deptos: dict,
                 maestro: dict | None = None,
                 meses: list[str] | None = None) -> list[dict]:
    def nom(d):
        return (maestro or {}).get(d or "", {}).get("nombre", "") or deptos.get(d or "", "")

    if conjunto == "gl":
        filas = []
        for f in (await s.execute(select(ActualEntry).where(
                ActualEntry.scenario_id == esc.id))).scalars().all():
            for i, m in enumerate(MESES):
                v = getattr(f, m) or Decimal("0")
                if v:
                    filas.append(_fila(esc, i + 1, f.dept_code, nom(f.dept_code),
                                       f.account_code, f.account_name, v,
                                       outlet=f.outlet, maestro=maestro,
                                       meses=meses))
        return filas

    if conjunto in ("opex", "costo", "propietario"):
        modelo = {"opex": OpexEntry, "costo": CostEntry,
                  "propietario": BelowGopAccountEntry}[conjunto]
        filas = []
        for f in (await s.execute(select(modelo).where(
                modelo.scenario_id == esc.id))).scalars().all():
            det = f"{getattr(f, 'detail_code', '')} {getattr(f, 'detail_desc', '')}".strip()
            for i, m in enumerate(MESES):
                v = getattr(f, m, None) or Decimal("0")
                if v:
                    filas.append(_fila(esc, i + 1, f.dept_code, nom(f.dept_code),
                                       f.account_code, f.account_name, v, detalle=det,
                                       maestro=maestro, meses=meses))
        return filas

    if conjunto == "planilla":
        # Los conceptos llevan `position_id`, asi que se puede bajar hasta la
        # persona. El CODIGO de posicion es la llave de los reportes de planilla
        # —`0111-01`— y sin el, una fila de horas extras no se puede cruzar con
        # nada (owner, 2026-08-14).
        from app.models.payroll_position import PayrollPosition
        pos = {p.id: p for p in (await s.execute(select(PayrollPosition).where(
            PayrollPosition.scenario_id == esc.id))).scalars().all()}
        filas = []
        for f in (await s.execute(select(PayrollConceptEntry).where(
                PayrollConceptEntry.scenario_id == esc.id))).scalars().all():
            p = pos.get(getattr(f, "position_id", None))
            for campo, cuenta, nombre in CONCEPTOS:
                v = getattr(f, campo, None) or Decimal("0")
                if v:
                    filas.append(_fila(
                        esc, f.month, f.dept_code, nom(f.dept_code), cuenta, nombre, v,
                        pos_code=(p.position_code or "").strip() if p else "",
                        pos_name=(p.position_name or "") if p else "",
                        empleado=(p.employee_name or "") if p else "",
                        maestro=maestro, meses=meses))
        return filas

    if conjunto == "pl":
        filas = []
        for f in (await s.execute(select(ActualPLLine).where(
                ActualPLLine.scenario_id == esc.id))).scalars().all():
            if f.amount_usd:
                filas.append(_fila(esc, f.month, "", "", f.line_code, "", f.amount_usd,
                               maestro=maestro, meses=meses))
        return filas

    raise ErrorApi(422, "consulta.conjunto_desconocido", conjunto=conjunto)


@router.get("/consulta/")
async def consultar(
    conjunto: str = Query("gl"),
    escenarios: str = Query(..., description="ids separados por coma"),
    cuenta: str = Query("", description="prefijo o lista separada por coma: 6001, o 60"),
    dept: str = Query("", description="prefijo o lista"),
    clase: str = Query("", description="digitos de clase: 6,7 — vacio = todas"),
    posicion: str = Query("", description="codigo de posicion, prefijo: 0111 o 0111-01"),
    cuenta_desde: str = Query("", description="rango: desde esta cuenta, inclusive"),
    cuenta_hasta: str = Query("", description="rango: hasta esta cuenta, inclusive"),
    mes_desde: int = Query(1, ge=1, le=12),
    mes_hasta: int = Query(12, ge=1, le=12),
    limite: int = Query(20000, ge=1, le=100000),
    _=Depends(get_current_user),
    idioma: str = Idioma,
):
    """Las filas del conjunto, filtradas. Formato largo: una fila por mes."""
    ids = [x.strip() for x in escenarios.split(",") if x.strip()]
    if not ids:
        raise ErrorApi(422, "escenarios.requerido")
    pref_cta = [x.strip() for x in cuenta.split(",") if x.strip()]
    pref_dep = [x.strip() for x in dept.split(",") if x.strip()]
    # La CLASE es el primer digito. Va aparte del filtro de cuenta —y no
    # metiendola como un prefijo mas— porque se combinan: «clase 6 y 7» + «cuenta
    # 60» tiene que dar la interseccion, no la union. Si compartieran campo, una
    # anularia a la otra sin que se note.
    clases = {c.strip() for c in clase.split(",") if c.strip()}
    pref_pos = [x.strip() for x in posicion.split(",") if x.strip()]

    # Rango de cuentas. Se compara como NUMERO y no como texto: "7100" > "700"
    # es verdad numericamente y falso alfabeticamente, y con codigos de largo
    # distinto el rango de texto se rompe justo en los bordes.
    def _num(c):
        try:
            return int(str(c).strip())
        except (TypeError, ValueError):
            return None
    n_desde, n_hasta = _num(cuenta_desde), _num(cuenta_hasta)

    from app.models.mapping import AccountMapping
    from app.models.department_catalog import DepartmentCatalog

    async with get_session() as s:
        # El maestro se arma UNA vez y se le pasa a todas las filas. Buscarlo por
        # fila serian dos consultas por cada uno de los miles de renglones.
        maestro: dict = {}
        deptos = {}
        for d in (await s.execute(select(DepartmentCatalog))).scalars().all():
            deptos[d.dept_code] = d.dept_name
            maestro[d.dept_code] = {
                "nombre": d.dept_name,
                "grupo": getattr(d, "default_pl_group", "") or "",
                "padre": getattr(d, "parent_dept_code", "") or "",
                "tipo": getattr(d, "pl_kind", "") or "",
            }
        # La linea del P&L depende de la cuenta Y del departamento: la misma
        # cuenta cae en lineas distintas segun donde este.
        lineas: dict[tuple, str] = {}
        for m in (await s.execute(select(AccountMapping))).scalars().all():
            if getattr(m, "active_status", "YES") in ("NO",):
                continue
            lineas[(str(m.account_code or ""), str(m.dept_code or ""))] =                 m.report_line_code or ""
        maestro["__lineas__"] = lineas

        # Nombre y seccion de cada linea del P&L.
        from app.models.mapping import ReportLineConfig
        nombres_l, secciones = {}, {}
        for r in (await s.execute(select(ReportLineConfig))).scalars().all():
            nombres_l[r.line_code] = r.line_name or ""
            secciones[r.line_code] = getattr(r, "section", "") or ""
        maestro["__nombres_linea__"] = nombres_l
        maestro["__secciones__"] = secciones

        salida: list[dict] = []
        for sid in ids:
            esc = await s.get(Scenario, sid)
            if esc is None:
                continue
            salida.extend(await _filas(s, conjunto, esc, deptos, maestro,
                                       _meses(idioma)))

    # El filtro es por PREFIJO, no por igualdad: «60» trae toda la planilla y
    # «6001» solo las horas extras, sin dos campos distintos ni un desplegable
    # con doscientas cuentas.
    def pasa(f):
        if not (mes_desde <= f["mes_num"] <= mes_hasta):
            return False
        if clases and f["clase"] not in clases:
            return False
        if n_desde is not None or n_hasta is not None:
            n = _num(f["account_code"])
            if n is None:
                return False   # una linea del P&L no tiene numero: fuera del rango
            if n_desde is not None and n < n_desde:
                return False
            if n_hasta is not None and n > n_hasta:
                return False
        if pref_cta and not any(f["account_code"].startswith(p) for p in pref_cta):
            return False
        if pref_dep and not any(f["dept_code"].startswith(p) for p in pref_dep):
            return False
        if pref_pos and not any(f["position_code"].startswith(p) for p in pref_pos):
            return False
        return True

    filas = [f for f in salida if pasa(f)]
    filas.sort(key=lambda f: (f["escenario"], f["account_code"], f["dept_code"], f["mes_num"]))
    total = sum(f["monto"] for f in filas)
    return {
        "conjunto": conjunto,
        "filas": filas[:limite],
        "cantidad": len(filas),
        "truncado": len(filas) > limite,
        "total": total,
    }


@router.get("/consulta/excel/")
async def consulta_excel(
    conjunto: str = Query("gl"), escenarios: str = Query(...),
    cuenta: str = Query(""), dept: str = Query(""), clase: str = Query(""),
    posicion: str = Query(""),
    cuenta_desde: str = Query(""), cuenta_hasta: str = Query(""),
    mes_desde: int = Query(1, ge=1, le=12), mes_hasta: int = Query(12, ge=1, le=12),
    _=Depends(get_current_user),
    idioma: str = Idioma,
):
    """Lo mismo, en un .xlsx listo para tabla dinamica.

    Cuatro cosas que lo hacen usable y no solo correcto:

    * **Autofiltro** en la fila 1 — con cuatro versiones y miles de filas, lo
      primero que uno hace es filtrar por escenario.
    * **Fila 1 congelada**, o al segundo scroll no se sabe que columna es cual.
    * **Formato de numero de verdad** en el monto (#,##0.00, negativos en rojo y
      entre parentesis). Sin esto Excel lo trata como numero pero se ve como
      texto plano, y la tabla dinamica hereda ese formato.
    * **El mes va con su numero al lado.** «Ene» ordena alfabetico —Abr, Ago,
      Dic...— que es el clasico desorden de una dinamica por mes. La columna
      «Mes #» es la que se usa para ordenar.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    datos = await consultar(conjunto=conjunto, escenarios=escenarios, cuenta=cuenta,
                            dept=dept, clase=clase, posicion=posicion,
                            cuenta_desde=cuenta_desde, cuenta_hasta=cuenta_hasta,
                            mes_desde=mes_desde, mes_hasta=mes_hasta,
                            limite=100000, _=None, idioma=idioma)
    columnas = _columnas(idioma)
    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta"

    ws.append([c["label"] for c in columnas])
    relleno = PatternFill("solid", fgColor="1F3864")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for f in datos["filas"]:
        ws.append([f[c["key"]] for c in columnas])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{ws.max_row}"

    for i, c in enumerate(columnas, start=1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = {
            "texto": 26, "moneda": 16, "conteo": 9,
        }.get(c["tipo"], 16)
        if c["tipo"] == "moneda":
            for celda in ws[letra][1:]:
                celda.number_format = '#,##0.00;[Red](#,##0.00)'
        elif c["tipo"] == "conteo":
            for celda in ws[letra][1:]:
                celda.number_format = "0"
                celda.alignment = Alignment(horizontal="center")
    # La columna del nombre de cuenta es la que mas texto lleva.
    ws.column_dimensions[get_column_letter(
        [c["key"] for c in columnas].index("account_name") + 1)].width = 34

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Consulta_{conjunto}.xlsx"'},
    )
