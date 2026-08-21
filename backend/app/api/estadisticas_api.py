# -*- coding: utf-8 -*-
"""Estadísticas: el catálogo, el archivo de carga y los valores.

**La puerta.** Vive en Master Data, al lado de Account Mapping, porque las
cuentas estadísticas SON cuentas —clase 9—. Lo que está separado es la tabla, y
por un motivo concreto: una cuenta de dinero solo necesita saber a qué línea del
P&L va; una estadística necesita además **qué unidad mide**, **por qué
dimensiones se abre** y **cómo se acumula el año** —un headcount no se suma entre
meses—. Meter esos campos en las 1,172 reglas contables que no los usan sería
ensuciarlas (owner, 2026-08-14: «prefiero una base de datos separada pero ahí
mismo»).

**El archivo.** Se baja con la grilla YA ARMADA del escenario —sus posiciones,
sus departamentos, sus tipos de habitación—, se llena y se sube. Las columnas se
ubican **por encabezado, nunca por posición**, que es la norma del owner para
todas las plantillas. Y **lo que no se reconoce da error en vez de perderse**: es
la misma regla que se puso el día de los $40,613.
"""
import io
from decimal import Decimal, InvalidOperation

import openpyxl
from fastapi import APIRouter, Depends, File, Query, UploadFile
from app.importers.registro_dep import registro_de_subida
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, select

from app.errores import ErrorApi
from app.auth import get_current_user
from app.db import get_session
from app.engine import estadisticas_grilla as grilla
from app.engine import jornada
from app.export.estadisticas_excel import LLAVE, MESES, construir_libro
from app.models.scenario import Scenario, ScenarioLockedError
from app.models.stat_account import StatAccount
from app.models.statistical_entry import StatisticalEntry

router = APIRouter()


# ─── Catálogo ────────────────────────────────────────────────────────────────

@router.get("/estadisticas/catalogo/")
async def catalogo(_=Depends(get_current_user)):
    """Las cuentas estadísticas, para la pantalla de Master Data."""
    async with get_session() as s:
        filas = (await s.execute(select(StatAccount)
                                 .order_by(StatAccount.grupo, StatAccount.code))
                 ).scalars().all()
    return {"cuentas": [{
        "code": c.code, "grupo": c.grupo, "nombre_es": c.nombre_es,
        "nombre_en": c.nombre_en, "unidad": c.unidad,
        "dims": c.dims_permitidas() and sorted(c.dims_permitidas()) or [],
        "deptos": c.deptos_propios(), "agrega": c.agrega,
        "legado": c.legado, "activa": c.activa,
    } for c in filas],
        # La regla de la jornada, para que la pantalla la muestre en vez de que
        # el owner tenga que acordarse.
        "jornada": {
            "horas_mes": float(jornada.HORAS_MES),
            "horas_dia": float(jornada.HORAS_POR_DIA),
            "dias_base": float(jornada.DIAS_BASE_MES),
            "cierran": list(jornada.CUENTAS_DE_LA_JORNADA),
            "por_encima": list(jornada.CUENTAS_SOBRE_LA_JORNADA),
        }}


@router.post("/estadisticas/catalogo/{code}/")
async def editar_cuenta(code: str, body: dict, _=Depends(get_current_user)):
    """El NOMBRE se puede cambiar; el código no. Misma regla que los tipos de
    habitación: el código es lo que liga el dato y el nombre es la etiqueta.

    Una cuenta NUEVA no se crea acá: va en `seed_data/stats_catalog.json`, para
    que una propiedad nueva nazca con todas."""
    async with get_session() as s:
        c = await s.get(StatAccount, code)
        if c is None:
            raise ErrorApi(404, "cuenta.no_encontrada", cuenta=code)
        if "nombre_es" in body:
            c.nombre_es = str(body["nombre_es"]).strip() or c.nombre_es
        if "nombre_en" in body:
            c.nombre_en = str(body["nombre_en"]).strip()
        if "activa" in body:
            c.activa = bool(body["activa"])
        await s.commit()
    return {"ok": True}


# ─── El archivo ──────────────────────────────────────────────────────────────

#: Las cuentas de segmento que se DERIVAN del Channel Mix, y de qué métrica.
#:
#: «Que las clase 9 se llenen solas del mix» (owner, 18-ago-2026). La 9000 son
#: noches por segmento y la 9001 pax por segmento — exactamente lo que
#: `channel_mix_detail` guarda por market code, que ES el segmento. Cargarlas a
#: mano teniendo el dato al lado es copiar números de una tabla a otra, y cada
#: copia es una oportunidad de que difieran.
DERIVADAS_DEL_MIX = {"9000": "rooms", "9001": "pax"}


#: Por CANAL (el rollup del mix) y por PAÍS (el Country Mix). Misma idea que
#: arriba, una capa más arriba: el dato ya está, copiarlo a mano es invitar a
#: que difiera.
DERIVADAS_POR_CANAL = {"9070": "rooms", "9071": "pax"}
DERIVADAS_POR_PAIS = {"9080": "rooms", "9081": "pax"}


async def _del_channel_mix(session, scenario_id: str) -> dict:
    """`{llave: {mes: valor}}` de todo lo que se deriva de otra tabla.

    Cuatro fuentes, todas exactas —no hay estimación en ninguna—:

    - **9000/9001** ← `channel_mix_detail`, por market code (= SEGMENT).
    - **9070/9071** ← `channel_mix_entries`, por canal.
    - **9080/9081** ← `country_mix_entries`, por país.
    - **9901 FTE**  ← `payroll_positions`, por departamento y posición.

    El FTE es el caso más claro de los cuatro: la planilla ya lo sabe mes por
    mes, y pedirlo otra vez en la grilla de estadísticas es teclear dos veces el
    mismo número para que después no coincidan.
    """
    from app.models.channel_mix import ChannelMixDetail, ChannelMixEntry
    from app.models.country_mix import CountryMixEntry
    from app.models.payroll_position import PayrollPosition

    out: dict = {}

    def sumar(k, mes, v):
        out.setdefault(k, {})[mes] = out.setdefault(k, {}).get(mes, 0.0) + float(v)

    # ── SEGMENT (market code) ────────────────────────────────────────────────
    for f in (await session.execute(select(ChannelMixDetail).where(
            ChannelMixDetail.scenario_id == scenario_id))).scalars().all():
        for cuenta, metrica in DERIVADAS_DEL_MIX.items():
            if f.metric == metrica:
                sumar((cuenta, "", "", "", "SEGMENT", f.market_code), f.month, f.value)

    # ── CHANNEL ──────────────────────────────────────────────────────────────
    for f in (await session.execute(select(ChannelMixEntry).where(
            ChannelMixEntry.scenario_id == scenario_id))).scalars().all():
        for cuenta, metrica in DERIVADAS_POR_CANAL.items():
            if f.metric == metrica:
                sumar((cuenta, "", "", "", "CHANNEL", f.channel), f.month, f.value)

    # ── COUNTRY ──────────────────────────────────────────────────────────────
    for f in (await session.execute(select(CountryMixEntry).where(
            CountryMixEntry.scenario_id == scenario_id))).scalars().all():
        for cuenta, metrica in DERIVADAS_POR_PAIS.items():
            if f.metric == metrica:
                sumar((cuenta, "", "", "", "COUNTRY", f.country), f.month, f.value)

    # ── FTE por posición ─────────────────────────────────────────────────────
    #
    # ⚠️ Se SUMA por (departamento, posición): una misma posición puede tener
    # varias personas —hay tres «AGENTE DE RECEPCION 501» en el mismo depto— y
    # la grilla tiene UNA fila por posición. Quedarse con la última daría un
    # FTE de 1 donde hay 3.
    MESES_FTE = ["fte_jan", "fte_feb", "fte_mar", "fte_apr", "fte_may", "fte_jun",
                 "fte_jul", "fte_aug", "fte_sep", "fte_oct", "fte_nov", "fte_dec"]
    for pos in (await session.execute(select(PayrollPosition).where(
            PayrollPosition.scenario_id == scenario_id))).scalars().all():
        cod = (getattr(pos, "position_code", "") or "").strip()
        if not cod:
            continue          # sin código no hay forma de ubicarla en la grilla
        k = ("9901", (pos.dept_code or "").strip(), cod, "", "", "")
        for i, campo in enumerate(MESES_FTE, start=1):
            v = getattr(pos, campo, None)
            if v:
                sumar(k, i, v)

    return out


async def _cargados(session, scenario_id: str) -> dict:
    """{llave: {mes: monto}} de lo que ya está guardado, más lo DERIVADO.

    ⚠️ Lo cargado a mano MANDA sobre lo derivado. Es la regla de la casa —«lo
    subido no se toca»—: si alguien corrigió un mes en el archivo, el mix no se
    lo pisa. La derivación llena lo que está vacío, no discute lo que ya se
    decidió.
    """
    filas = (await session.execute(select(StatisticalEntry).where(
        StatisticalEntry.scenario_id == scenario_id))).scalars().all()
    out: dict = {}
    for f in filas:
        k = (f.account_code, f.dept_code, f.position_code,
             f.room_type_code, f.dim_type, f.dim_code)
        out.setdefault(k, {})[f.month] = float(f.value)

    for k, meses in (await _del_channel_mix(session, scenario_id)).items():
        destino = out.setdefault(k, {})
        for mes, v in meses.items():
            destino.setdefault(mes, v)      # no pisa lo cargado a mano
    return out


@router.get("/estadisticas/{scenario_id}/plantilla.xlsx")
async def plantilla(scenario_id: str, _=Depends(get_current_user)):
    async with get_session() as s:
        e = await s.get(Scenario, scenario_id)
        if e is None:
            raise ErrorApi(404, "escenario.no_encontrado")
        filas = await grilla.construir(s, e)
        wb = construir_libro(
            f"Estadísticas — {e.type} {e.version} {e.year}",
            f"{len(filas)} filas generadas de este escenario",
            filas, await _cargados(s, scenario_id))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"Estadisticas_{e.type}_{e.year}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'})


def _num(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def leer_libro(data: bytes) -> tuple[list[dict], list[str]]:
    """Lee el archivo. Devuelve (filas, problemas).

    ⚠️ Las columnas se ubican **por encabezado**. Buscarlas por posición es lo
    que rompe una plantilla en cuanto alguien inserta una columna — y el owner
    trabaja sobre estos archivos.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb["Estadisticas"] if "Estadisticas" in wb.sheetnames else wb.worksheets[0]
    filas_xl = list(ws.iter_rows(values_only=True))

    encabezado, fila_enc = None, None
    for i, r in enumerate(filas_xl[:20]):
        vals = [str(x).strip() if x is not None else "" for x in r]
        if "Cuenta" in vals and "Depto" in vals:
            encabezado, fila_enc = vals, i
            break
    if encabezado is None:
        return [], ["No se encontró la fila de encabezados (la que dice «Cuenta» "
                    "y «Depto»). ¿Es el archivo que bajaste de la app?"]

    col = {n: i for i, n in enumerate(encabezado) if n}
    faltan = [n for n in LLAVE if n not in col]
    if faltan:
        return [], [f"Al archivo le faltan columnas: {', '.join(faltan)}"]
    meses_col = {m: col[n] for m, n in enumerate(MESES, start=1) if n in col}
    if not meses_col:
        return [], ["El archivo no tiene ninguna columna de mes."]

    def celda(r, nombre):
        i = col.get(nombre)
        if i is None or i >= len(r) or r[i] is None:
            return ""
        return str(r[i]).strip()

    filas, problemas = [], []
    for n, r in enumerate(filas_xl[fila_enc + 1:], start=fila_enc + 2):
        cuenta = celda(r, "Cuenta")
        if not cuenta:
            continue
        meses = {}
        for m, i in meses_col.items():
            v = _num(r[i]) if i < len(r) else None
            if v is not None and v != 0:
                meses[m] = v
        filas.append({
            "fila": n, "account_code": cuenta,
            "dept_code": celda(r, "Depto"),
            "position_code": celda(r, "Posición"),
            "room_type_code": celda(r, "Tipo hab."),
            "meses": meses,
        })
    return filas, problemas


@router.post("/estadisticas/{scenario_id}/importar/", dependencies=[Depends(registro_de_subida)])
async def importar(scenario_id: str, archivo: UploadFile = File(...),
                   dry_run: bool = Query(False),
                   _=Depends(get_current_user)):
    """Sube el archivo. Con `dry_run` solo dice qué haría.

    **Se NIEGA si algo no se reconoce.** Una fila con dato y con una cuenta o una
    posición que el escenario no tiene es un error de digitación o un archivo
    viejo; guardarla a medias o descartarla en silencio es como se perdieron los
    $40,613 del Actual 2024.
    """
    data = await archivo.read()
    filas, problemas = leer_libro(data)
    if problemas:
        raise ErrorApi(422, "estadisticas.libro_con_problemas",
                       extra={"errores": problemas})

    async with get_session() as s:
        e = await s.get(Scenario, scenario_id)
        if e is None:
            raise ErrorApi(404, "escenario.no_encontrado")

        validas = {f.llave for f in await grilla.construir(s, e)}
        cuentas = {c.code for c in (await s.execute(select(StatAccount))).scalars()}

        desconocidas, con_dato = [], 0
        for f in filas:
            if not f["meses"]:
                continue
            con_dato += 1
            llave = (f["account_code"], f["dept_code"], f["position_code"],
                     f["room_type_code"], "", "")
            if f["account_code"] not in cuentas:
                desconocidas.append(
                    f"fila {f['fila']}: la cuenta {f['account_code']} no existe")
            elif llave not in validas:
                desconocidas.append(
                    f"fila {f['fila']}: {f['account_code']} con departamento "
                    f"«{f['dept_code']}» y posición «{f['position_code']}» no es "
                    "una combinación de este escenario")

        if desconocidas:
            raise ErrorApi(422, "estadisticas.filas_no_reconocidas", extra={
                "errores": desconocidas[:40],
                "total": len(desconocidas),
            })

        if dry_run:
            return {"dry_run": True, "filas_con_dato": con_dato,
                    "filas_en_el_archivo": len(filas)}
        if e.status == "locked":
            raise ScenarioLockedError(scenario_id)

        # Reemplazo total: el archivo es la verdad de este escenario. Es lo que
        # el owner espera («yo bajo, corrijo y subo lo que guardé») y evita que
        # una fila que él borró sobreviva escondida.
        await s.execute(delete(StatisticalEntry).where(
            StatisticalEntry.scenario_id == scenario_id))
        n = 0
        for f in filas:
            for mes, valor in f["meses"].items():
                s.add(StatisticalEntry(
                    scenario_id=scenario_id, account_code=f["account_code"],
                    month=mes, dept_code=f["dept_code"],
                    position_code=f["position_code"],
                    room_type_code=f["room_type_code"],
                    value=valor, origen="ARCHIVO"))
                n += 1
        await s.commit()
    return {"ok": True, "valores_guardados": n, "filas_con_dato": con_dato}


# ─── Los valores ─────────────────────────────────────────────────────────────

@router.get("/estadisticas/{scenario_id}/")
async def valores(scenario_id: str, _=Depends(get_current_user)):
    """Lo cargado, más el control de la jornada por posición."""
    async with get_session() as s:
        e = await s.get(Scenario, scenario_id)
        if e is None:
            raise ErrorApi(404, "escenario.no_encontrado")
        filas = (await s.execute(select(StatisticalEntry).where(
            StatisticalEntry.scenario_id == scenario_id))).scalars().all()
        nombres = {c.code: c.nombre_es for c in
                   (await s.execute(select(StatAccount))).scalars()}

    # El control de la jornada: por posición y por mes, las tres cuentas que
    # cierran tienen que dar 240. Ver `engine/jornada.py`.
    por_pos: dict = {}
    for f in filas:
        if f.account_code in jornada.CUENTAS_DE_LA_JORNADA and f.position_code:
            k = (f.dept_code, f.position_code, f.month)
            por_pos.setdefault(k, {})[f.account_code] = f.value
    descuadres = []
    for (dep, pos, mes), horas in sorted(por_pos.items()):
        cierra, dif = jornada.cierra_la_jornada(horas)
        if not cierra:
            descuadres.append({"dept_code": dep, "position_code": pos,
                               "month": mes, "diferencia": float(dif)})

    return {
        "scenario": {"id": e.id, "type": e.type, "version": e.version,
                     "year": e.year, "status": e.status},
        "nombres": nombres,
        "valores": [{
            "account_code": f.account_code, "month": f.month,
            "dept_code": f.dept_code, "position_code": f.position_code,
            "room_type_code": f.room_type_code, "value": float(f.value),
            "origen": f.origen,
        } for f in filas],
        # Se devuelve SIEMPRE, aunque esté vacío: una lista vacía dice «todo
        # cierra», que es información. Que no aparezca la sección no dice nada.
        "jornada_descuadres": descuadres,
    }
