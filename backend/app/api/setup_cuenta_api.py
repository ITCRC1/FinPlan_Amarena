"""El SETUP DE LA CUENTA — las cinco preguntas, cuenta por cuenta.

Owner (2026-08-16): «lo que sí quiero que se cumpla es que el setup de la cuenta
esté claro: qué es ingreso, qué es costo, qué es gasto y qué es gastos de la
propiedad; qué departamento; y dónde debe aparecer en el P&L, para que se alinee
con los demás años».

Es lo que hace CLONABLE una propiedad: antes de copiar el mapeo a Amarena,
Oxigen y Ojochal hay que poder recorrerlo de una sentada y decir «esto está
bien». Hoy la respuesta existe —el mapeo está sano— pero repartida en once
chequeos de un script, una columna de una plantilla de Excel y un tab que pide
elegir un escenario primero.

Las cinco preguntas que cada cuenta responde acá:

    1. QUÉ ES              ingreso · costo · planilla · gasto · gasto de la
                           propiedad. Sale de la clase USALI del código.
    2. QUÉ DEPARTAMENTO    y si es madre, hijo funcional o set de producto.
    3. EN QUÉ LÍNEA        la línea exacta del P&L, no el grupo.
    4. CÓMO LLEGÓ AHÍ      regla exacta · heredada del padre · sin departamento
                           · POR DESCARTE · SIN REGLA. Las dos últimas son las
                           que hay que ir a arreglar.
    5. ¿SE ALINEA?         la misma cuenta, ¿cae en la misma línea en 2024 que
                           en 2027? Esta es la que hoy no existe en ningún lado.

⚠️ **Esto no es una segunda lista.** No hay ni un dato escrito a mano acá: las
reglas salen de `account_mapping` (que el seed re-afirma desde
`seed_data/mapping_pl.json` en cada deploy), los departamentos del
`department_catalog`, las líneas del `report_line_config` y el ruteo del MISMO
resolvedor del motor (`pl_engine.construir_resolvedor`). Una vista mantenida
aparte se desincroniza —así esta app llegó a mostrar 22 departamentos con 38 en
la base— y un resolvedor propio hace que la herramienta jure una cosa mientras
el P&L hace otra, que ya pasó acá.

⚠️ **Solo lectura.** No escribe en ninguna tabla y no mueve ningún número.
"""
from __future__ import annotations

import io
from collections import defaultdict

from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_db
from app.engine import pl_engine
from app.engine.recalculate import (
    load_active_account_mappings, load_report_line_config,
)
from app.models.department_catalog import DepartmentCatalog
from app.models.mapping import AccountMapping
from app.models.scenario import Scenario
from app.api.audit_api import _sources, es_hijo_funcional
from app.hotel_actual import HOTEL_ID, hotel_slug
from app.i18n import DEFAULT_LOCALE
from app.textos import Idioma, t

router = APIRouter(tags=["setup-cuenta"])


# ── Pregunta 1: qué es ───────────────────────────────────────────────────────
#
# Las cuatro del owner —ingreso, costo, gasto, gasto de la propiedad— son
# exactamente las clases USALI que el sistema ya distingue. La planilla se
# muestra aparte de «gasto» porque tiene su propio auxiliar y sus propias
# cuentas (los 17 conceptos), no porque sea otra cosa: es gasto.
#
# ⚠️ Este español es el que viaja en el catálogo como `setup.clase_<n>`: son las
# palabras del owner y `tests/test_setup_de_la_cuenta.py` las fija. El rótulo
# que sale por la API se pide con `t()` para que el inglés exista; esta tabla
# sigue siendo la que dice QUÉ CLASES hay.
CLASES: dict[str, str] = {
    "4": "Ingreso",
    "5": "Costo",
    "6": "Planilla",
    "7": "Gasto",
    "8": "Gasto de la propiedad",
    "9": "Estadística",
}


def _clase_nombre(clase: str, idioma: str) -> str:
    return t(idioma, f"setup.clase_{clase}" if clase in CLASES
             else "setup.clase_otra")

# ── Pregunta 4: cómo llegó ahí ───────────────────────────────────────────────
#
# Los cinco modos que devuelve `construir_resolvedor`, con el rótulo que ve el
# owner y si cuentan como LIMPIO. `parent` es correcto a propósito: un reparto
# que cae en 0122 Cocina y hereda la regla de 0120 A&B está bien ruteado — es
# el diseño, no un accidente. `dept-agnostic` también: hay cuentas que no
# dependen del departamento y su regla se escribió sin uno.
COMO: dict[str, tuple[str, bool]] = {
    "exact":         ("Regla propia", True),
    "parent":        ("Heredada del padre", True),
    "dept-agnostic": ("Regla sin departamento", True),
    "FALLBACK":      ("Por descarte", False),
    "DROP":          ("Sin regla", False),
    # Los gastos del propietario (`nonop_entries`) NO pasan por el mapeo:
    # siembran su `report_line_code` directo. Marcarlos como «por descarte»
    # —que es lo que da el resolvedor, porque la fila viene sin departamento—
    # sería acusar de sucio algo que está bien y por diseño. Lo único que hay
    # que verificar de esta vía es que la línea sembrada EXISTA en el reporte:
    # si no existe, el monto se pierde sin dar error.
    "siembra":       ("Siembra la línea directo", True),
    "siembra-rota":  ("Siembra una línea que no existe", False),
}

#: El mismo rótulo, en los dos idiomas. La tupla de arriba manda en QUÉ MODOS
#: hay y cuál cuenta como limpio —eso es comportamiento y está probado—; esto
#: solo dice cómo se escribe.
COMO_CLAVE: dict[str, str] = {
    "exact": "setup.como_exact",
    "parent": "setup.como_parent",
    "dept-agnostic": "setup.como_dept_agnostic",
    "FALLBACK": "setup.como_fallback",
    "DROP": "setup.como_drop",
    "siembra": "setup.como_siembra",
    "siembra-rota": "setup.como_siembra_rota",
}

_ORDEN_ESCENARIO = ("FORECAST_current", "FORECAST_working", "ACTUAL",
                    "BUDGET_working", "BUDGET_final")


def _rango_escenario(s: Scenario) -> int:
    """Con cuál de los escenarios del año se mira ese año.

    Es la MISMA regla del owner que ya usa el frontend en
    `lib/escenarioPreferido.ts`: «Budget Working, Forecast Working, Actual».
    Escrita como preferencia por TIPO y no por año, porque acá se aplica a
    todos los años a la vez y la lista de años sale de la base.

    Se elige UNO por año a propósito. Unir los seis presupuestos 2027 daría un
    conjunto de líneas más grande, no más verdadero: la pregunta del owner es
    dónde cae la cuenta en el año, y para eso hace falta un escenario que se
    pueda nombrar. Cuál se usó viaja en la respuesta.
    """
    ver = (s.version or "").strip().lower()
    if s.type == "FORECAST" and getattr(s, "is_current_forecast", False):
        return 0
    if s.type == "FORECAST" and "working" in ver:
        return 1
    if s.type == "ACTUAL":
        return 2
    if s.type == "BUDGET" and "working" in ver:
        return 3
    if s.type == "BUDGET" and "final" in ver:
        return 4
    return 5


def _tipo_de_depto(d: DepartmentCatalog | None, idioma: str = DEFAULT_LOCALE) -> str:
    """Pregunta 2, segunda mitad: madre, hijo funcional o set de producto.

    Usa `es_hijo_funcional` de `audit_api` —la bandera `room_set`, no el
    parentesco— para que no haya dos criterios. `0113` Housekeeping es una
    FUNCIÓN de Habitaciones y su gasto es del padre; `0115` Villas cuelga del
    mismo `0110` y NO lo es: es un set de producto, con gasto propio.
    """
    if d is None:
        return t(idioma, "setup.depto_fuera_catalogo")
    if not (d.parent_dept_code or "").strip():
        return t(idioma, "setup.depto_madre")
    return t(idioma, "setup.depto_hijo_funcional" if es_hijo_funcional(d)
             else "setup.depto_set_producto")


async def _escenario_por_anio(db: AsyncSession) -> list[Scenario]:
    """Un escenario por año, el preferido, y solo los años que tienen alguno."""
    escs = (await db.execute(select(Scenario))).scalars().all()
    por_anio: dict[int, Scenario] = {}
    for s in sorted(escs, key=lambda x: (x.year, _rango_escenario(x), x.version or "")):
        por_anio.setdefault(s.year, s)
    return [por_anio[a] for a in sorted(por_anio)]


async def armar_setup(db: AsyncSession, idioma: str = DEFAULT_LOCALE) -> dict:
    """El cuadro completo. Función aparte del endpoint para que la prueba y el
    Excel llamen a ESTA y no a una copia.

    ⚠️ Los `estado` de la matriz de alineación (`usa`, `no se usó`, `el depto no
    estaba`) NO se traducen: son el CÓDIGO por el que el frontend decide qué
    celda pintar de amarillo. Lo que se traduce es el rótulo, y eso pasa donde
    se pinta — en la pantalla y en el Excel.
    """
    mapeos = await load_active_account_mappings(db)
    resolve = pl_engine.construir_resolvedor(mapeos)
    lineas = {r["line_code"]: r for r in await load_report_line_config(db)}
    deptos = {d.dept_code: d for d in
              (await db.execute(select(DepartmentCatalog))).scalars().all()}

    # Nombre de cuenta. El dict que arma `load_active_account_mappings` para el
    # motor trae solo lo que el motor necesita para RUTEAR (cuenta, depto,
    # línea) — el nombre no está. Se lee aparte de la misma tabla, sin tocar la
    # carga del motor: acá hace falta para que el owner reconozca la cuenta, y
    # ahí sería peso muerto en el camino caliente del P&L.
    nombre_cuenta: dict[str, str] = {}
    for ac, nm in (await db.execute(
            select(AccountMapping.account_code, AccountMapping.account_name_example)
            .where(AccountMapping.active_status == "YES")
            .order_by(AccountMapping.account_code))).all():
        ac = (ac or "").strip()
        if ac and (nm or "").strip():
            nombre_cuenta.setdefault(ac, nm.strip())

    # Qué reglas ACTIVAS existen, por par. Es el universo del «setup»: una
    # cuenta sin movimiento igual tiene que estar bien configurada — la
    # plantilla existe justamente para digitar lo que todavía no está.
    con_regla_propia: set[tuple[str, str]] = set()
    for m in mapeos:
        dc, ac = (m.get("dept_code") or "").strip(), (m.get("account_code") or "").strip()
        if ac:
            con_regla_propia.add((dc, ac))

    # ── La plata, año por año ────────────────────────────────────────────────
    #
    # `_sources` es la MISMA enumeración de fuentes que usa el tab de Control:
    # el auxiliar o el detalle GL según el modo del escenario, más la planilla,
    # los repartos y los gastos del propietario. Reusarla evita la segunda
    # lista, que es el modo de fallar caro de este sistema.
    # La llave es (departamento, cuenta, línea sembrada). El tercer elemento es
    # "" salvo para los gastos del propietario, donde una MISMA cuenta alimenta
    # dos líneas distintas (la 8020 es Reserva de Capital y Capex Grande) y sin
    # él las dos colapsarían en una fila que mentiría sobre las dos.
    escenarios = await _escenario_por_anio(db)
    montos: dict[tuple[str, str, str], dict[int, float]] = defaultdict(dict)
    anios_usados: list[dict] = []
    # Departamentos con alguna actividad en el año. Sirve para la pregunta 5:
    # que un departamento no aparezca en 2024 porque todavía no existía no dice
    # nada sobre el setup de una cuenta.
    deptos_activos: dict[int, set[str]] = {}
    for esc in escenarios:
        acum: dict[tuple[str, str, str], float] = defaultdict(float)
        vivos: set[str] = set()
        for r in await _sources(db, esc, 0):
            ac = (r.get("account_code") or "").strip()
            if not ac:
                continue
            dc = (r.get("dept_code") or "").strip()
            acum[(dc, ac, r.get("seed_line") or "")] += float(r.get("amount") or 0)
            vivos.add(dc)
            # Una cuenta que solo existe en el dato —no en el mapeo— igual
            # tiene que poder reconocerse por el nombre.
            if ac not in nombre_cuenta and (r.get("account_name") or "").strip():
                nombre_cuenta[ac] = r["account_name"].strip()
        if not acum:
            # Un año sin un solo dato (los Working 2028-2035 están vacíos) no
            # dice nada sobre alineación: se saca para que no acuse de
            # «desalineada» a toda cuenta que no exista todavía.
            continue
        for par, v in acum.items():
            if round(v, 2):
                montos[par][esc.year] = round(v, 2)
        deptos_activos[esc.year] = vivos
        anios_usados.append({"anio": esc.year, "escenario_id": esc.id,
                             "escenario": f"{esc.type} {esc.version} {esc.year}"})

    anios = [a["anio"] for a in anios_usados]

    # ── Las filas: mapeo ∪ dato ──────────────────────────────────────────────
    pares = sorted(
        {(d, c, "") for d, c in con_regla_propia} | set(montos))
    filas: list[dict] = []
    for dept, cuenta, sembrada in pares:
        if sembrada:
            # Vía de siembra: la línea la trae la propia fila, no el mapeo.
            lc, como, regla_de = sembrada, "siembra", ""
            if sembrada not in lineas:
                como = "siembra-rota"
        else:
            m, como = resolve(dept, cuenta)
            lc = (m or {}).get("report_line_code") or ""
            # De qué departamento salió la regla, cuando no es el propio. Es lo
            # que separa «heredé del padre» —correcto— de «me resolvieron por
            # descarte», que hay que ir a arreglar en el mapeo.
            regla_de = ""
            if como in ("parent", "FALLBACK"):
                regla_de = (m or {}).get("dept_code") or ""
            elif como == "dept-agnostic":
                regla_de = t(idioma, "setup.sin_departamento")
        meta = lineas.get(lc) or {}
        d = deptos.get(dept)
        limpia = COMO[como][1]
        rotulo = t(idioma, COMO_CLAVE[como])
        clase = cuenta[:1]
        filas.append({
            "dept_code": dept,
            "dept_name": (d.dept_name if d else "") or dept
                         or t(idioma, "setup.sin_departamento"),
            "dept_tipo": _tipo_de_depto(d, idioma) if dept
                         else t(idioma, "setup.depto_nivel_propiedad"),
            "dept_padre": (d.parent_dept_code if d else "") or "",
            "cuenta": cuenta,
            "cuenta_nombre": nombre_cuenta.get(cuenta, ""),
            "clase": clase,
            "clase_nombre": _clase_nombre(clase, idioma),
            "linea_code": lc,
            "linea_nombre": meta.get("line_name") or lc,
            "seccion": meta.get("section") or "",
            "como": como,
            "como_nombre": rotulo,
            "regla_de": regla_de,
            "regla_propia": (dept, cuenta) in con_regla_propia,
            "montos": {str(a): v for a, v
                       in sorted(montos.get((dept, cuenta, sembrada), {}).items())},
            "con_movimiento": bool(montos.get((dept, cuenta, sembrada))),
            "limpia": limpia,
        })

    # ── Pregunta 5: ¿la misma cuenta cae en la misma línea todos los años? ────
    #
    # La resolución depende de (departamento, cuenta), así que un par nunca
    # cambia de línea entre años. Lo que SÍ cambia es bajo qué departamento se
    # contabiliza la cuenta: mover la `7160` del `0240` al `0250` la manda a
    # otra línea sin que nadie toque una regla. Por eso la comparación es por
    # CUENTA, no por par.
    #
    # Solo se comparan años en los que la cuenta tiene plata: una cuenta que
    # aparece en un solo año no está desalineada, todavía no tiene con qué
    # compararse.
    #
    # ⚠️ Y solo cuentan los departamentos VIVOS en los dos años. Sin ese filtro
    # la lista se llena de ruido que no es un problema de setup: el Club (260) y
    # Claro del Bosque (0205) empiezan en 2027, Innoceana (0155) termina en
    # 2026, y con ellos TODA cuenta de planilla y de gasto «cambiaba de línea».
    # Que un departamento no existiera un año no dice nada sobre el setup de una
    # cuenta; que la cuenta se mude de un departamento que sigue vivo, sí.
    por_cuenta: dict[str, dict[int, dict[str, dict]]] = defaultdict(lambda: defaultdict(dict))
    for f in filas:
        for a, v in f["montos"].items():
            anio = int(a)
            if f["dept_code"] and f["dept_code"] not in deptos_activos.get(anio, set()):
                continue
            dest = por_cuenta[f["cuenta"]][anio].setdefault(
                f["linea_code"], {"linea_code": f["linea_code"],
                                  "linea_nombre": f["linea_nombre"],
                                  "monto": 0.0, "deptos": []})
            dest["monto"] = round(dest["monto"] + v, 2)
            if f["dept_code"] not in dest["deptos"]:
                dest["deptos"].append(f["dept_code"])

    # Cada celda (línea × año) dice POR QUÉ está como está. Sin eso la lista es
    # un montón de huecos indistinguibles y no se puede revisar: hay huecos que
    # importan —la cuenta se dejó de usar en un departamento que sigue vivo— y
    # huecos que no —el departamento todavía no existía ese año—.
    HUECO_COMPARABLE = "no se usó"
    desalineadas: list[dict] = []
    for cuenta, por_anio in sorted(por_cuenta.items()):
        if len(por_anio) < 2:
            continue
        # Los departamentos donde vive la cuenta, y en qué años estuvieron
        # vivos. Una línea solo «falta» en un año si TODOS los departamentos
        # que la alimentan estuvieron vivos ese año.
        deptos_de_linea: dict[str, set[str]] = defaultdict(set)
        for dd in por_anio.values():
            for lc, d in dd.items():
                deptos_de_linea[lc].update(dc for dc in d["deptos"] if dc)

        matriz: list[dict] = []
        en_juego = 0.0
        hay_hueco = False
        for lc in sorted(deptos_de_linea):
            celdas = []
            for anio in sorted(por_anio):
                d = por_anio[anio].get(lc)
                if d:
                    estado, monto = "usa", d["monto"]
                elif all(dc in deptos_activos.get(anio, set())
                         for dc in deptos_de_linea[lc]):
                    estado, monto = HUECO_COMPARABLE, 0.0
                    hay_hueco = True
                else:
                    estado, monto = "el depto no estaba", 0.0
                celdas.append({"anio": anio, "estado": estado, "monto": monto})
            if any(c["estado"] == HUECO_COMPARABLE for c in celdas):
                en_juego += sum(c["monto"] for c in celdas)
            matriz.append({
                "linea_code": lc,
                "linea_nombre": next(
                    (d["linea_nombre"] for dd in por_anio.values()
                     if (d := dd.get(lc))), lc),
                "deptos": sorted(deptos_de_linea[lc]),
                "celdas": celdas,
            })
        if not hay_hueco:
            continue
        desalineadas.append({
            "cuenta": cuenta,
            "cuenta_nombre": nombre_cuenta.get(cuenta, ""),
            "clase": cuenta[:1],
            "clase_nombre": _clase_nombre(cuenta[:1], idioma),
            "anios": sorted(por_anio),
            "lineas": matriz,
            # Lo que está en juego: la plata de las líneas que un año usan y
            # otro año, pudiendo usarlas, no.
            "monto_en_juego": round(en_juego, 2),
        })
    desalineadas.sort(key=lambda x: -x["monto_en_juego"])

    cuentas_desalineadas = {d["cuenta"] for d in desalineadas}
    for f in filas:
        f["desalineada"] = f["cuenta"] in cuentas_desalineadas and f["con_movimiento"]
        f["alerta"] = ("" if f["limpia"] else f["como_nombre"].lower())

    por_clase: dict[str, int] = defaultdict(int)
    for f in filas:
        por_clase[f["clase"]] += 1

    sucias = [f for f in filas if not f["limpia"] or f["desalineada"]]
    return {
        "hotel_id": HOTEL_ID,
        "anios": anios_usados,
        "clases": [{"clase": c, "nombre": _clase_nombre(c, idioma), "cuentas": n}
                   for c, n in sorted(por_clase.items())],
        "departamentos": [
            {"dept_code": dc, "dept_name": dn, "dept_tipo": dt}
            for dc, dn, dt in sorted({
                (f["dept_code"], f["dept_name"], f["dept_tipo"]) for f in filas})],
        "resumen": {
            "filas": len(filas),
            "cuentas": len({f["cuenta"] for f in filas}),
            "limpias": sum(1 for f in filas if f["limpia"] and not f["desalineada"]),
            "a_revisar": len(sucias),
            "por_descarte": sum(1 for f in filas if f["como"] == "FALLBACK"),
            "sin_regla": sum(1 for f in filas if f["como"] == "DROP"),
            "desalineadas": len(desalineadas),
            "sin_movimiento": sum(1 for f in filas if not f["con_movimiento"]),
            "anios_comparados": anios,
        },
        "filas": filas,
        "desalineadas": desalineadas,
    }


@router.get("/setup-cuenta/")
async def setup_cuenta(db: AsyncSession = Depends(get_db), idioma: str = Idioma):
    """Las cinco respuestas para cada cuenta. Se filtra en el navegador: son
    ~1.100 filas y el owner necesita cruzar clase con departamento sin esperar
    un viaje al servidor por cada clic."""
    return await armar_setup(db, idioma)


@router.get("/setup-cuenta/excel/")
async def setup_cuenta_excel(
    solo_revisar: bool = Query(False, description="Solo lo que no está limpio"),
    db: AsyncSession = Depends(get_db),
    idioma: str = Idioma,
):
    """El mismo cuadro, para revisar de una sentada y guardarlo como referencia
    de las otras tres propiedades."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    datos = await armar_setup(db, idioma)
    anios = datos["resumen"]["anios_comparados"]

    wb = Workbook()
    HDR = PatternFill("solid", fgColor="16402A")
    WHITE = Font(bold=True, color="FFFFFF", size=10)
    MAL = PatternFill("solid", fgColor="F6E3E3")
    OJO = PatternFill("solid", fgColor="FFF3CD")

    def _encabezado(ws, cols, anchos):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(1, j, c)
            cell.fill = HDR
            cell.font = WHITE
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        from openpyxl.utils import get_column_letter
        for j, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(2, 1)

    # ── Hoja 1: las cinco preguntas, cuenta por cuenta ───────────────────────
    ws = wb.active
    ws.title = t(idioma, "setup.excel_hoja_cuentas")
    cols = [t(idioma, k) for k in (
        "setup.excel_col_que_es", "setup.excel_col_clase",
        "setup.excel_col_cuenta", "setup.excel_col_nombre_cuenta",
        "setup.excel_col_departamento", "setup.excel_col_codigo",
        "setup.excel_col_tipo_depto", "setup.excel_col_depto_padre",
        "setup.excel_col_linea_pl_num", "setup.excel_col_codigo_linea",
        "setup.excel_col_seccion", "setup.excel_col_como",
        "setup.excel_col_regla_de", "setup.excel_col_se_alinea",
        "setup.excel_col_esta_limpia")] + [f"USD {a}" for a in anios]
    _encabezado(ws, cols, [16, 6, 9, 30, 26, 8, 15, 10, 32, 24, 22, 20, 15, 16, 12]
                + [13] * len(anios))
    filas = [f for f in datos["filas"]
             if not solo_revisar or not f["limpia"] or f["desalineada"]]
    for i, f in enumerate(filas, 2):
        alinea = t(idioma, "setup.alinea_revisar" if f["desalineada"]
                   else ("setup.alinea_si" if f["con_movimiento"]
                         else "setup.alinea_sin_dato"))
        vals = [f["clase_nombre"], f["clase"], f["cuenta"], f["cuenta_nombre"],
                f["dept_name"], f["dept_code"], f["dept_tipo"], f["dept_padre"],
                f["linea_nombre"], f["linea_code"], f["seccion"],
                f["como_nombre"], f["regla_de"], alinea,
                t(idioma, "setup.limpia_si"
                  if f["limpia"] and not f["desalineada"] else "setup.limpia_no")]
        for j, v in enumerate(vals, 1):
            ws.cell(i, j, v)
        for j, a in enumerate(anios, len(vals) + 1):
            v = f["montos"].get(str(a))
            if v is not None:
                c = ws.cell(i, j, v)
                c.number_format = "#,##0"
        if not f["limpia"]:
            for j in range(1, len(cols) + 1):
                ws.cell(i, j).fill = MAL
        elif f["desalineada"]:
            for j in range(1, len(cols) + 1):
                ws.cell(i, j).fill = OJO

    # ── Hoja 2: la pregunta 5, cuenta por cuenta y año por año ───────────────
    wa = wb.create_sheet(t(idioma, "setup.excel_hoja_alineacion"))
    cols2 = ([t(idioma, k) for k in (
        "setup.excel_col_cuenta", "setup.excel_col_nombre",
        "setup.excel_col_que_es_corto", "setup.excel_col_en_juego",
        "setup.excel_col_linea_pl", "setup.excel_col_codigo_linea",
        "setup.excel_col_departamentos")]
             + [str(a) for a in anios])
    _encabezado(wa, cols2, [10, 30, 16, 14, 34, 26, 20] + [15] * len(anios))
    r = 2
    for d in datos["desalineadas"]:
        primera = True
        for ln in d["lineas"]:
            wa.cell(r, 1, d["cuenta"])
            wa.cell(r, 2, d["cuenta_nombre"])
            wa.cell(r, 3, d["clase_nombre"])
            if primera:
                wa.cell(r, 4, d["monto_en_juego"]).number_format = "#,##0"
                primera = False
            wa.cell(r, 5, ln["linea_nombre"])
            wa.cell(r, 6, ln["linea_code"])
            wa.cell(r, 7, ", ".join(ln["deptos"]))
            por_anio = {c["anio"]: c for c in ln["celdas"]}
            for j, a in enumerate(anios, 8):
                c = por_anio.get(a)
                if not c:
                    # La cuenta no tuvo un centavo en NINGUNA línea ese año.
                    # Se dice, en vez de dejar la celda en blanco: un blanco se
                    # lee como un error de la herramienta.
                    cell = wa.cell(r, j, t(idioma, "setup.cuenta_no_usada"))
                    cell.font = Font(color="6B7A70", italic=True, size=9)
                    continue
                if c["estado"] == "usa":
                    cell = wa.cell(r, j, c["monto"])
                    cell.number_format = "#,##0"
                else:
                    # `estado` es el código con el que decide el frontend; acá
                    # se escribe, así que acá se traduce.
                    cell = wa.cell(r, j, t(
                        idioma, "setup.estado_no_se_uso"
                        if c["estado"] == "no se usó"
                        else "setup.estado_depto_no_estaba"))
                    cell.font = Font(color="A05000", italic=True, size=9)
                    # El hueco que importa: el departamento estaba vivo y la
                    # cuenta igual no cayó en esta línea ese año.
                    if c["estado"] == "no se usó":
                        cell.fill = MAL
            r += 1
    if r == 2:
        wa.cell(2, 1, t(idioma, "setup.todo_alineado"))

    # ── Hoja 3: cómo se leyó esto ────────────────────────────────────────────
    wi = wb.create_sheet(t(idioma, "setup.excel_hoja_leer"))
    res = datos["resumen"]
    texto = [t(idioma, k) if k else "" for k in (
        "setup.leer_titulo",
        "",
        "setup.leer_p1",
        "setup.leer_p1_clases",
        "",
        "setup.leer_p2",
        "setup.leer_p2_hijo",
        "setup.leer_p2_set",
        "",
        "setup.leer_p3",
        "",
        "setup.leer_p4",
        "setup.leer_p4_exact",
        "setup.leer_p4_parent",
        "setup.leer_p4_agnostic",
        "setup.leer_p4_fallback_1",
        "setup.leer_p4_fallback_2",
        "setup.leer_p4_fallback_3",
        "setup.leer_p4_drop",
        "",
        "setup.leer_p5",
        "setup.leer_p5_b",
        "setup.leer_p5_c",
        "setup.leer_p5_d",
        "setup.leer_p5_e",
        "",
        "setup.leer_anios",
    )]
    texto += [f"   {a['anio']} → {a['escenario']}" for a in datos["anios"]]
    texto += [
        "",
        t(idioma, "setup.leer_fuentes"),
        t(idioma, "setup.leer_fuentes_reglas"),
        t(idioma, "setup.leer_fuentes_deptos"),
        t(idioma, "setup.leer_fuentes_lineas"),
        t(idioma, "setup.leer_fuentes_ruteo"),
        "",
        t(idioma, "setup.leer_resumen"),
        t(idioma, "setup.leer_resumen_filas", n=res["filas"]),
        t(idioma, "setup.leer_resumen_cuentas", n=res["cuentas"]),
        t(idioma, "setup.leer_resumen_limpias", n=res["limpias"]),
        t(idioma, "setup.leer_resumen_a_revisar", n=res["a_revisar"]),
        t(idioma, "setup.leer_resumen_por_descarte", n=res["por_descarte"]),
        t(idioma, "setup.leer_resumen_sin_regla", n=res["sin_regla"]),
        t(idioma, "setup.leer_resumen_desalineadas", n=res["desalineadas"]),
    ]
    for i, linea in enumerate(texto, 1):
        wi.cell(i, 1, linea)
    wi.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    sufijo = "_a_revisar" if solo_revisar else ""
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="Setup_de_la_cuenta_{hotel_slug()}{sufijo}.xlsx"'},
    )
