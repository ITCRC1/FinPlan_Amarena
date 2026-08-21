"""
Control & Auditoría — trazabilidad cuenta×departamento → línea del P&L.

Responde la pregunta operativa: "metí $100 en OPEX de Rooms, ¿dónde cayó?".
Para un escenario devuelve cada fila fuente (auxiliar) con el departamento, la
cuenta, el monto y **la línea exacta del P&L donde aterriza**, más el MODO de
ruteo:

    exact         → hay regla para (depto, cuenta)            ✔ correcto
    dept-agnostic → la regla no distingue depto               ⚠ revisar
    FALLBACK      → cae en la línea de OTRO depto             ✖ misruteo
    DROP          → sin regla: el monto NO llega al P&L       ✖ plata perdida

Además reconcilia: Σ fuentes por línea == P&L calculado por el motor.
"""
from decimal import Decimal

from fastapi import APIRouter, Depends, File, Query, UploadFile
from sqlalchemy import select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.errores import ErrorApi
from app.textos import Idioma, t
from app.db import get_db
from app.engine import pl_engine
from app.engine.recalculate import (
    load_active_account_mappings, load_report_line_config, compute_pl_month,
    PAYROLL_ALL_COLS,
)
from app.models.scenario import Scenario
from app.models.opex_entry import OpexEntry
from app.models.cost_entry import CostEntry
from app.models.payroll_concept_entry import PayrollConceptEntry
from app.models.nonop_entry import NonOpEntry
from app.models.actual_entry import ActualEntry
from app.api._apagados import apagados_por_dimension
from app.models.mapping import AccountMapping
from app.models.department_catalog import DepartmentCatalog
from app.models.allocation_entry import AllocationEntry
from app.engine.allocation_calculator import cuentas_de_reparto
from app.hotel_actual import HOTEL_ID, hotel_slug

router = APIRouter(tags=["control"])
ZERO = Decimal("0")
_M = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _resolver(mappings: list[dict]):
    """El MISMO resolvedor del motor, no una copia.

    Antes esto era una réplica escrita a mano. Una réplica que se queda atrás
    hace que el tab de Control jure una cosa mientras el P&L hace otra, y este
    tab existe precisamente para que el usuario sepa dónde cae cada dato.
    """
    resolve_motor = pl_engine.construir_resolvedor(mappings)

    def resolve(dept: str, acct: str):
        m, como = resolve_motor(dept, acct)
        if not m:
            return None, "DROP", "SUM", ""
        # En 'parent' y 'FALLBACK' el usuario necesita ver de qué regla salió:
        # una dice que heredó del padre —correcto— y la otra que se resolvió
        # por descarte, que hay que ir a arreglar en el mapeo.
        prestado = ""
        if como in ("parent", "FALLBACK"):
            prestado = m.get("dept_code") or m.get("source_department") or ""
        return m.get("report_line_code"), como, m.get("rollup_operator", "SUM"), prestado

    return resolve


def es_hijo_funcional(dept) -> bool:
    """¿Es un hijo que existe como FUNCIÓN de su padre, y no como producto?

    `0111` Front Desk y `0113` Housekeeping son funciones de Habitaciones: su
    gasto es del padre. `0115` Villas y `0116` Residencias cuelgan del mismo
    `0110` y NO lo son — son sets de producto, con su gasto propio.

    Lo que los separa es la bandera `room_set` del catálogo, no el parentesco;
    el mismo criterio que ya usa `revenue_api` («filtra por la bandera, no por
    parentesco»). El `0110` también la tiene en `True` sin colgar de nadie, así
    que hace falta el parentesco Y la ausencia de la bandera.
    """
    return (bool((getattr(dept, "parent_dept_code", None) or "").strip())
            and not bool(getattr(dept, "room_set", False)))


def acepta_gasto_operativo(dept, con_gasto_propio: set[str]) -> bool:
    """¿El selector de Opex/Costos puede ofrecer este departamento?

    Vive acá, a nivel de módulo y no dentro del endpoint, para que la prueba
    llame a ESTA función y no a una copia. Una copia que se queda atrás hace
    que la prueba jure una cosa mientras la pantalla hace otra.

    `con_gasto_propio`: departamentos con al menos una regla de clase 5 o 7 que
    NO sea cuenta de reparto (ver `cuentas_de_reparto`).
    """
    return not es_hijo_funcional(dept) and dept.dept_code in con_gasto_propio


async def _filas_de_ingreso(db: AsyncSession, scenario: Scenario,
                            month: int) -> list[dict]:
    """El ingreso del sub-mayor, con su departamento y su cuenta del mapeo.

    Una fila por línea de ingreso. El monto es de **línea**, no de cuenta: el
    motor calcula «ingreso de Tours», no «ingreso de la 4401». Cuando la línea
    tiene varias cuentas en el mapeo (Tours son la 4400 a la 4403) se usa la
    primera y el nombre dice la línea, para que la fila se explique sola en vez
    de aparentar un detalle por cuenta que nadie calculó.

    ⚠️ **Depende de que el recálculo haya bajado el ingreso al sub-mayor** — ver
    `recalculate.sincronizar_ingreso_al_checkbook`. Antes de ese arreglo, esta
    tabla podía estar meses vieja y acá se vería como el ingreso del escenario.
    """
    from app.engine import pl_engine
    from app.models.revenue_entry import RevenueEntry

    filas = (await db.execute(select(RevenueEntry).where(
        RevenueEntry.scenario_id == scenario.id))).scalars().all()
    if not filas:
        return []

    # `report_line_code` -> (dept_code, account_code), del mapeo activo.
    mappings = await load_active_account_mappings(db)
    destino: dict[str, tuple[str, str]] = {}
    for r in mappings:
        code = str(r.get("report_line_code") or "")
        if not code.startswith("REV_"):
            continue
        par = pl_engine._MOTOR_TO_CANON.get(code)
        code = par[0] if par else code
        cuenta = str(r.get("account_code") or "")
        dept = str(r.get("dept_code") or "")
        actual = destino.get(code)
        if actual is None or cuenta < actual[1]:
            destino[code] = (dept, cuenta)

    out: list[dict] = []
    for e in filas:
        amt = (float(sum(getattr(e, m) or 0 for m in _M)) if not month
               else float(e.get_month(month) or 0))
        if not amt:
            continue
        linea = pl_engine.REVENUE_LINE_TO_REPORT_LINE.get(e.line.lower(), "")
        par = pl_engine._MOTOR_TO_CANON.get(linea)
        linea = par[0] if par else linea
        dept, cuenta = destino.get(linea, ("", ""))
        out.append({"origin": "Ingreso", "dept_code": dept,
                    "account_code": cuenta,
                    "account_name": f"{e.line} · línea {linea or '(sin línea)'}",
                    "amount": round(amt, 2)})
    return out


async def _sources(db: AsyncSession, scenario: Scenario, month: int) -> list[dict]:
    """Filas fuente del escenario: de los auxiliares (checkbook) o del detalle GL."""
    out: list[dict] = []
    cd = pl_engine.consolidate_dept
    imported = getattr(scenario, "source_mode", "imported") != "checkbook"

    if imported:
        rows = (await db.execute(select(ActualEntry).where(
            ActualEntry.scenario_id == scenario.id))).scalars().all()
        for e in rows:
            amt = float(sum(getattr(e, m) or 0 for m in _M)) if not month else float(e.get_month(month) or 0)
            if amt:
                out.append({"origin": "Detalle GL", "dept_code": e.dept_code or "",
                            "account_code": e.account_code or "", "account_name": e.account_name or "",
                            "amount": round(amt, 2)})
        return out

    for Model, origin in ((OpexEntry, "OPEX"), (CostEntry, "Costos")):
        rows = (await db.execute(select(Model).where(Model.scenario_id == scenario.id))).scalars().all()
        for e in rows:
            amt = float(sum(getattr(e, m) or 0 for m in _M)) if not month else float(e.get_month(month) or 0)
            if amt:
                out.append({"origin": origin, "dept_code": cd(e.dept_code or ""),
                            "account_code": e.account_code or "",
                            "account_name": getattr(e, "account_name", "") or "",
                            "amount": round(amt, 2)})
    q = select(PayrollConceptEntry).where(PayrollConceptEntry.scenario_id == scenario.id)
    if month:
        q = q.where(PayrollConceptEntry.month == month)
    for e in (await db.execute(q)).scalars().all():
        for col in PAYROLL_ALL_COLS:
            amt = float(getattr(e, col, None) or 0)
            if amt:
                out.append({"origin": "Planilla", "dept_code": cd(e.dept_code or ""),
                            "account_code": pl_engine.payroll_account_for_column(col),
                            "account_name": col, "amount": round(amt, 2)})
    # Los REPARTOS (cafetería, lavandería, salarios). Van acá porque el P&L los
    # suma como una fuente más: si no se trazan, este tab reconcilia de menos y
    # acusa un descuadre que no existe.
    #
    # Se ven las dos mitades: el CARGO al departamento que recibe (positivo) y
    # el CRÉDITO al que reparte (negativo). Así, filtrando por un par
    # departamento + cuenta, se lee de un vistazo qué entró y qué salió.
    qa = select(AllocationEntry).where(AllocationEntry.scenario_id == scenario.id)
    if month:
        qa = qa.where(AllocationEntry.month == month)
    for e in (await db.execute(qa)).scalars().all():
        amt = float(e.amount_usd or 0)
        if not amt:
            continue
        entra = amt > 0
        out.append({
            "origin": "Reparto",
            "dept_code": cd(e.target_dept or ""),
            "account_code": e.account or "",
            "account_name": (f"{'entra de' if entra else 'sale a'} {e.source_dept}"
                             f" · {e.allocation_type}"),
            "amount": round(amt, 2),
        })

    # ── EL INGRESO ───────────────────────────────────────────────────────────
    #
    # Owner, 2026-08-17: *«debe tomar todas las cuentas»* · *«los canales entran
    # al principio y los resultados que van al GL son el final del proceso»*.
    #
    # Hasta acá esta función era de COSTO —OPEX, Costos, Planilla, Repartos— y el
    # tab de Control lo decía en su propio texto («payroll, OPEX, costs»). Por
    # eso mostraba **«MONTO QUE SE PIERDE $0» sin haber mirado un solo peso de
    # ingreso**: no es que no se perdiera nada, es que no estaba viendo la mitad.
    #
    # ⚠️ Solo hace falta en modo `checkbook`. En `imported` el ingreso ya viene:
    # el detalle GL trae las `4xxx` como cualquier otra fila, y ahí sí se
    # trazaban desde siempre.
    #
    # El (departamento, cuenta) sale del **mismo `account_mapping`** que resuelve
    # el costo — el que el owner edita en Admin · Account Mapping. Las 19 líneas
    # de ingreso ya lo tienen (`REV_ROOMS`→`0110/4000`,
    # `REV_SUSTAINABILITY`→`280/4880`): no se inventa ninguna cuenta acá.
    for fila in await _filas_de_ingreso(db, scenario, month):
        out.append(fila)

    # Gastos del propietario (below-GOP): NO pasan por el mapeo de cuentas — siembran
    # la línea del P&L directamente por su report_line_code. Si ese código no existe
    # en el reporte, el monto se pierde sin aviso: por eso se traza aparte.
    for e in (await db.execute(select(NonOpEntry).where(
            NonOpEntry.scenario_id == scenario.id))).scalars().all():
        amt = float(sum(getattr(e, m) or 0 for m in _M)) if not month else float(e.get_month(month) or 0)
        if amt:
            out.append({"origin": "Gastos Propietario", "dept_code": "",
                        "account_code": getattr(e, "account_code", "") or "",
                        "account_name": getattr(e, "detail_desc", "") or getattr(e, "detail_code", "") or "",
                        "amount": round(amt, 2), "seed_line": e.report_line_code})
    return out


@router.get("/audit/scenario/{scenario_id}/trace/")
async def trace_scenario(
    scenario_id: str,
    month: int = Query(0, description="0 = año completo; 1..12 = un mes"),
    db: AsyncSession = Depends(get_db),
    idioma: str = Idioma,
):
    """Trazabilidad completa: cada fila de los auxiliares → línea del P&L, con el
    modo de ruteo, más la reconciliación contra el P&L que calcula el motor."""
    scen = await db.get(Scenario, scenario_id)
    if not scen:
        raise ErrorApi(404, "escenario.no_encontrado")
    mappings = await load_active_account_mappings(db)
    report_lines = await load_report_line_config(db)
    if not mappings or not report_lines:
        raise ErrorApi(400, "auditoria.falta_configurar_mapeo")
    resolve = _resolver(mappings)
    line_meta = {r["line_code"]: r for r in report_lines}
    dept_names = {d.dept_code: d.dept_name for d in
                  (await db.execute(select(DepartmentCatalog))).scalars().all()}

    rows = await _sources(db, scen, month)
    imported_mode = getattr(scen, "source_mode", "imported") != "checkbook"
    traced, by_line, problems = [], {}, {"DROP": [], "FALLBACK": [], "dept-agnostic": []}
    for r in rows:
        if r.get("seed_line"):
            # Below-GOP: siembra la línea directo; solo se valida que exista.
            lc = r["seed_line"]
            existe = lc in line_meta
            mode, op, src = ("exact" if existe else "DROP"), "SUM", ""
        else:
            lc, mode, op, src = resolve(r["dept_code"], r["account_code"])
        signed = -r["amount"] if op == "SUBTRACT" else r["amount"]
        rec = {**r,
               "dept_name": dept_names.get(r["dept_code"], r["dept_code"]),
               "grupo": pl_engine.group_for_dept(r["dept_code"]),
               "line_code": lc, "line_name": (line_meta.get(lc) or {}).get("line_name", lc or "—"),
               "section": (line_meta.get(lc) or {}).get("section", ""),
               "mode": mode, "rollup": op, "fallback_from": src}
        traced.append(rec)
        if mode in problems:
            problems[mode].append(rec)
        if lc:
            b = by_line.setdefault(lc, {"line_code": lc,
                                        "line_name": rec["line_name"], "section": rec["section"],
                                        "amount_sources": 0.0, "rows": 0, "depts": set()})
            b["amount_sources"] += signed
            b["rows"] += 1
            b["depts"].add(r["dept_code"])

    # P&L real del motor (para reconciliar)
    months = range(1, 13) if not month else [month]
    pl_tot: dict[str, float] = {}
    for m in months:
        for L in await compute_pl_month(db, scen, m):
            pl_tot[L.line_code] = pl_tot.get(L.line_code, 0.0) + float(L.amount_usd)

    lines_out = []
    for lc, b in by_line.items():
        pl_v = pl_tot.get(lc, 0.0)
        dif = round(b["amount_sources"] - pl_v, 2)
        lines_out.append({**b, "depts": sorted(b["depts"]),
                          "amount_sources": round(b["amount_sources"], 2),
                          "amount_pl": round(pl_v, 2), "dif": dif,
                          "ok": abs(dif) < max(1.0, abs(pl_v) * 0.001)})
    lines_out.sort(key=lambda x: (x["section"], -abs(x["amount_sources"])))

    dropped = round(sum(p["amount"] for p in problems["DROP"]), 2)

    # ── Avisos de consistencia (no son errores de ruteo, son riesgos de doble conteo) ──
    from app.importers.gl_detail_importer import ALLOCATION_EXCLUDE
    avisos = []
    if not imported_mode:
        alloc_depts = set(ALLOCATION_EXCLUDE)
        monto_alloc = round(sum(r["amount"] for r in traced
                                if r["dept_code"] in alloc_depts), 2)
        monto_6025 = round(sum(r["amount"] for r in traced
                               if r["account_code"] == "6025"
                               and r["dept_code"] not in alloc_depts), 2)
        # REGLA DEL OWNER: el ingreso del budget sale SOLO del checkbook de ingresos.
        # Si aparece una cuenta 4xxx dentro de OPEX/Costos, o un gasto del propietario
        # apuntando a una línea REV_*, se estaría metiendo ingreso por la puerta de atrás.
        colados = [r for r in traced
                   if str(r["account_code"]).startswith("4") and r["origin"] in ("OPEX", "Costos")]
        colados += [r for r in traced
                    if r["origin"] == "Gastos Propietario"
                    and str(r.get("line_code") or "").startswith("REV")]
        if colados:
            avisos.append({
                "tipo": "ingreso_por_otra_via",
                "titulo": t(idioma, "auditoria.ingreso_por_otra_via"),
                "detalle": t(idioma, "auditoria.ingreso_por_otra_via_detalle",
                             n=len(colados)),
                "monto": round(sum(r["amount"] for r in colados), 2),
            })
        if monto_alloc > 0 and monto_6025 > 0:
            avisos.append({
                "tipo": "posible_doble_conteo",
                "titulo": t(idioma, "auditoria.posible_doble_conteo"),
                "detalle": t(idioma, "auditoria.posible_doble_conteo_detalle",
                             deptos=", ".join(sorted(alloc_depts)),
                             monto_reparto=f"${monto_alloc:,.0f}",
                             monto_6025=f"${monto_6025:,.0f}"),
                "monto_depto_reparto": monto_alloc,
                "monto_concepto_6025": monto_6025,
            })
    return {
        "avisos": avisos,
        "scenario": f"{scen.type} {scen.version} {scen.year}",
        "source_mode": getattr(scen, "source_mode", "imported"),
        "month": month,
        "totales": {
            "filas_fuente": len(traced),
            "monto_fuente": round(sum(r["amount"] for r in traced), 2),
            "monto_perdido_DROP": dropped,
            "filas_DROP": len(problems["DROP"]),
            "filas_FALLBACK": len(problems["FALLBACK"]),
            "filas_dept_agnostic": len(problems["dept-agnostic"]),
            "filas_exact": sum(1 for r in traced if r["mode"] == "exact"),
        },
        "pl_control": {
            "TOTAL_REVENUES": round(pl_tot.get("TOTAL_REVENUES", 0.0), 2),
            "TOTAL_OPERATING_EXPENSES": round(pl_tot.get("TOTAL_OPERATING_EXPENSES", 0.0), 2),
            "TOTAL_OVERHEAD_EXPENSES": round(pl_tot.get("TOTAL_OVERHEAD_EXPENSES", 0.0), 2),
            "TOTAL_GOP": round(pl_tot.get("TOTAL_GOP", 0.0), 2),
            "EBITDA_BEFORE_CAPITAL": round(pl_tot.get("EBITDA_BEFORE_CAPITAL", 0.0), 2),
            "NET_PROFIT": round(pl_tot.get("NET_PROFIT", 0.0), 2),
        },
        "by_line": lines_out,
        "rows": sorted(traced, key=lambda r: (r["mode"] != "DROP", r["mode"] != "FALLBACK",
                                              r["dept_code"], r["account_code"])),
        "problems": {k: v[:200] for k, v in problems.items()},
    }


# ─── Plantilla de carga + validación previa ───────────────────────────────────
_TPL_MONTHS = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
               "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


@router.get("/audit/upload-template/")
async def upload_template(db: AsyncSession = Depends(get_db)):
    """Plantilla de carga de gastos/ingresos con las combinaciones VÁLIDAS de
    departamento × cuenta (las que sí llegan al P&L) y su línea de destino."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from starlette.responses import Response

    mappings = await load_active_account_mappings(db)
    report_lines = {r["line_code"]: r for r in await load_report_line_config(db)}
    dept_names = {d.dept_code: d.dept_name for d in
                  (await db.execute(select(DepartmentCatalog))).scalars().all()}

    wb = Workbook()
    HDR = PatternFill("solid", fgColor="16402A")
    WHITE = Font(bold=True, color="FFFFFF", size=10)

    ws = wb.active
    ws.title = "Carga"
    cols = ["Departamento (código)", "Departamento", "Cuenta", "Nombre de cuenta"] + _TPL_MONTHS
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HDR
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center")
    for w, cl in zip((20, 30, 10, 34), "ABCD"):
        ws.column_dimensions[cl].width = w
    ws.freeze_panes = "E2"
    ws.cell(2, 1, "← poné acá tus filas. Usá SOLO combinaciones de la hoja «Válidas».")

    vs = wb.create_sheet("Válidas")
    vcols = ["Departamento (código)", "Departamento", "Cuenta", "Clase", "→ Línea del P&L", "Sección"]
    for j, c in enumerate(vcols, 1):
        cell = vs.cell(1, j, c)
        cell.fill = HDR
        cell.font = WHITE
    for w, cl in zip((20, 30, 10, 10, 34, 24), "ABCDEF"):
        vs.column_dimensions[cl].width = w
    vs.freeze_panes = "A2"
    CLASE = {"4": "Ingreso", "5": "Costo", "6": "Planilla", "7": "Gasto", "8": "No operativo"}
    r = 2
    seen = set()
    for m in sorted(mappings, key=lambda x: ((x.get("dept_code") or ""), x.get("account_code") or "")):
        dc = (m.get("dept_code") or "").strip()
        ac = (m.get("account_code") or "").strip()
        if not ac or not dc or (dc, ac) in seen:
            continue
        seen.add((dc, ac))
        lc = m.get("report_line_code") or ""
        meta = report_lines.get(lc, {})
        vs.cell(r, 1, dc)
        vs.cell(r, 2, dept_names.get(dc, dc))
        vs.cell(r, 3, ac)
        vs.cell(r, 4, CLASE.get(ac[:1], "—"))
        vs.cell(r, 5, meta.get("line_name", lc))
        vs.cell(r, 6, meta.get("section", ""))
        r += 1

    ins = wb.create_sheet("Instrucciones")
    for i, txt in enumerate([
        "CÓMO USAR ESTA PLANTILLA",
        "",
        "1. Llená la hoja «Carga»: una fila por departamento + cuenta, con los 12 meses.",
        "2. Usá SOLO las combinaciones de la hoja «Válidas»: son las que tienen regla y",
        "   por lo tanto llegan al P&L. Ahí ves a qué línea va cada una.",
        "3. Antes de importar, corré la VALIDACIÓN: te dice fila por fila si el monto",
        "   llega al P&L, si cae en el departamento equivocado o si se perdería.",
        "4. Después de importar, abrí «Control» y verificá el saldo de la línea.",
        "",
        "REGLA DE ORO: el departamento manda. La misma cuenta en otro departamento va",
        "a otra línea del P&L (la planilla de A&B no es la de Rooms).",
    ]):
        ins.cell(i + 1, 1, txt)
    ins.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Plantilla_Carga_{hotel_slug()}.xlsx"'},
    )


@router.post("/audit/validate-upload/")
async def validate_upload(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Validación PREVIA: lee la plantilla y dice, fila por fila, si el monto va a
    llegar al P&L y a qué línea. No escribe nada."""
    import io
    import openpyxl

    mappings = await load_active_account_mappings(db)
    report_lines = {r["line_code"]: r for r in await load_report_line_config(db)}
    resolve = _resolver(mappings)
    dept_names = {d.dept_code: d.dept_name for d in
                  (await db.execute(select(DepartmentCatalog))).scalars().all()}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception as e:
        raise ErrorApi(400, "archivo.ilegible", detalle=str(e))
    ws = wb["Carga"] if "Carga" in wb.sheetnames else wb.active

    filas, problemas, total_ok, total_perdido = [], 0, 0.0, 0.0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row[:4]):
            continue
        dept = str(row[0] or "").strip()
        acct = str(row[2] or "").strip()
        if acct.endswith(".0"):
            acct = acct[:-2]
        if not dept and not acct:
            continue
        monto = sum(float(v) for v in row[4:16] if isinstance(v, (int, float)))
        lc, mode, _op, src = resolve(dept, acct)
        meta = report_lines.get(lc, {})
        ok = mode in ("exact", "dept-agnostic")
        if ok:
            total_ok += monto
        else:
            total_perdido += monto if mode == "DROP" else 0.0
            problemas += 1
        filas.append({
            "fila": i, "dept_code": dept, "dept_name": dept_names.get(dept, dept),
            "account_code": acct, "account_name": str(row[3] or ""), "monto": round(monto, 2),
            "line_code": lc, "line_name": meta.get("line_name", "—"),
            "section": meta.get("section", ""), "mode": mode, "fallback_from": src,
            "ok": ok,
        })
    return {
        "archivo": file.filename,
        "filas": len(filas),
        "filas_ok": len(filas) - problemas,
        "filas_con_problema": problemas,
        "monto_que_llega": round(total_ok, 2),
        "monto_que_se_perderia": round(total_perdido, 2),
        "veredicto": ("✔ Todo viaja al P&L" if problemas == 0 else
                      f"✖ {problemas} fila(s) no llegan bien al P&L"),
        "detalle": filas[:500],
    }


@router.get("/audit/mapping/health/")
async def mapping_health(db: AsyncSession = Depends(get_db),
                         idioma: str = Idioma):
    """Salud del mapeo: reglas sin dept_code (causa de misruteo), cuentas que
    apuntan a varias líneas y pares (depto,cuenta) ambiguos."""
    mappings = await load_active_account_mappings(db)
    report_lines = await load_report_line_config(db)
    line_codes = {r["line_code"] for r in report_lines}
    sin_dept, por_cuenta, por_par = [], {}, {}
    for m in mappings:
        ac = (m.get("account_code") or "").strip()
        dc = (m.get("dept_code") or "").strip()
        if not ac:
            continue
        if not dc:
            sin_dept.append({"account_code": ac, "report_line_code": m.get("report_line_code")})
        por_cuenta.setdefault(ac, set()).add(m.get("report_line_code"))
        por_par.setdefault((dc, ac), set()).add(m.get("report_line_code"))
    multi = {a: sorted(v) for a, v in por_cuenta.items() if len(v) > 1}
    amb = [{"dept_code": d, "account_code": a, "lineas": sorted(v)}
           for (d, a), v in por_par.items() if len(v) > 1]
    huerf = sorted({m.get("report_line_code") for m in mappings
                    if m.get("report_line_code") not in line_codes})
    return {
        "reglas_activas": len(mappings),
        "reglas_sin_dept_code": len(sin_dept),
        "riesgo_misruteo": len(sin_dept) > 0 and len(multi) > 0,
        "cuentas_multi_linea": len(multi),
        "cuentas_multi_linea_detalle": dict(list(multi.items())[:40]),
        "pares_ambiguos": amb[:40],
        "pares_ambiguos_total": len(amb),
        "mapeos_a_linea_inexistente": huerf,
        "veredicto": t(idioma, "auditoria.veredicto_riesgo_misruteo"
                       if (sin_dept and multi) else
                       "auditoria.veredicto_ruteo_ok"),
    }


@router.get("/departments/")
async def listar_departamentos(
    hotel_id: str = HOTEL_ID, db: AsyncSession = Depends(get_db),
):
    """Catálogo de departamentos, para que la pantalla no tenga los nombres a mano.

    El frontend traía una lista escrita en código con 22 departamentos mientras la
    base tiene 36: catorce salían como números pelados (Front Desk, Housekeeping,
    Kitchen, Restaurant, Finance, Club Madresal, Área Recreativa…) y varios nombres
    se habían desviado del catálogo. Con esto hay una sola fuente.

    Viene TAMBIÉN qué escondió el provisionamiento de esta propiedad. El catálogo
    **no se bifurca por hotel** —el universo USALI es uno solo para el grupo—; lo
    que cambia entre propiedades es qué se usa de él, y eso es esta matriz.

    Va acá y no en un endpoint aparte porque este ya es el embudo por donde pasan
    todos los selectores de departamento: si la respuesta trae las dos cosas
    juntas, ninguna pantalla puede pintar la lista antes de saber qué esconder.
    """
    filas = (await db.execute(
        select(DepartmentCatalog).order_by(DepartmentCatalog.dept_code)
    )).scalars().all()
    # `parent_dept_code` viaja porque los selectores lo necesitan para NO ofrecer
    # gasto en un departamento hijo (owner, 2026-08-14: «solo los departamentos
    # padres pueden tener checkbook de gastos»). Sin él, el front no tiene forma
    # de saber quién es hijo y tendría que mantener una segunda lista a mano —
    # que es exactamente cómo esta pantalla terminó mostrando 22 departamentos
    # mientras la base tenía 38.
    #
    # `lleva_gasto` responde la pregunta completa del selector de Opex, y se
    # DERIVA del mapeo en vez de enumerarse: un departamento acepta gasto si no
    # es hijo y si tiene al menos una regla de clase 5 (costo) o 7 (opex).
    #
    # Ser hijo no era el único motivo para no aparecer, y los dos casos que salen
    # son distintos entre sí:
    #
    #   `280` Miscelaneos  — es SOLO ingreso y no tiene costo (owner, 2026-08-13).
    #       Owner, 2026-08-14: «no es gasto, no debe estar acá».
    #   `0250` Gastos de la Propiedad — SÍ es gasto, y eso no se discute. Lo que
    #       pasa es que es clase 8, debajo del GOP, y **tiene checkbook propio
    #       fuera de este**: las pantallas de Owner Expenses y Management Fees.
    #       Owner, 2026-08-14: «0250 tiene su propio checkbook fuera del
    #       checkbook, así que acá no hace nada — pero no se puede quitar de la
    #       naturaleza que es gasto». Sale de ESTE selector, no del sistema.
    #
    # Lo que los une es que ninguno tiene dónde aterrizar un gasto OPERATIVO: lo
    # que se digitara acá caería por descarte en otro departamento, con el GOP
    # cuadrando igual — el modo de falla caro de este sistema.
    #
    # ── Las cuentas de reparto NO cuentan ──────────────────────────────────
    # Una regla de gasto habilita el checkbook porque significa «acá alguien
    # puede DIGITAR un gasto». Las cuentas en las que escribe el motor de
    # repartos no significan eso: existen para recibir un asiento que el motor
    # calcula, no para que nadie las teclee. Es el mismo criterio por el que la
    # `4999` nunca contó — solo que la `4999` es clase 4 y quedaba afuera sola.
    #
    # El caso que lo destapó (2026-08-14): el `0162` Laundry Revenue necesita
    # regla para la `5301`, porque el reparto de lavandería le manda ahí el
    # costo de los kilos de huésped. Sin la regla ese costo caía por descarte
    # en el Spa; con la regla, y sin esta resta, el `0162` volvía a aparecer en
    # el checkbook de Costos, que es de donde el owner lo mandó sacar. Las dos
    # cosas a la vez son ciertas: tiene dónde aterrizar el reparto y no tiene
    # dónde digitar un gasto.
    #
    # La lista sale del motor (`cuentas_de_reparto`), no de una constante acá:
    # un reparto nuevo con una cuenta nueva queda cubierto solo.
    #
    # Ojo: NO alcanza con «no tener ninguna cuenta de reparto». La `7310` y la
    # `7685` también se digitan a mano en medio departamento. Lo que habilita
    # el checkbook es tener al menos una cuenta de gasto que NO sea de reparto,
    # así que Rooms, el Private Bar y los demás no se mueven.
    #
    # ── Ser hijo no basta: hay hijos que SÍ llevan gasto ──────────────────
    # `0115` Villas y `0116` Residencias cuelgan del `0110`, pero no son
    # FUNCIONES de Habitaciones como el Front Desk o el Housekeeping: son SETS
    # DE PRODUCTO, con sus 30 reglas de clase 5/7 propias y su propio gasto.
    # La regla «ningún hijo lleva gasto» los sacó del selector, y `rooms-sets`
    # es un reporte de solo lectura — o sea que se quedaban sin ninguna
    # pantalla donde cargarles gasto. Hoy están en $0,00 porque el owner
    # todavía no cargó los porcentajes; el agujero se destapaba justo el día
    # que los cargara.
    #
    # Lo que distingue a un set de una función es la bandera `room_set` del
    # catálogo, no el parentesco — el mismo criterio que ya usa
    # `revenue_api` («filtra por la bandera, no por parentesco»). Ojo que el
    # `0110` también la tiene en `True` sin ser hijo de nadie, así que la
    # bandera sola tampoco alcanza: es hijo funcional el que tiene padre y NO
    # es set.
    de_reparto = cuentas_de_reparto()
    pares = (await db.execute(
        select(AccountMapping.dept_code, AccountMapping.account_code).where(
            AccountMapping.active_status == "YES",
            or_(AccountMapping.account_code.like("5%"),
                AccountMapping.account_code.like("7%")),
        ).distinct()
    )).all()
    con_gasto = {(d or "").strip() for d, c in pares
                 if (c or "").strip() not in de_reparto}

    return {"departments": [{"dept_code": d.dept_code, "dept_name": d.dept_name,
                             "parent_dept_code": d.parent_dept_code,
                             "lleva_gasto": acepta_gasto_operativo(d, con_gasto)}
                            for d in filas],
            "hotel_id": hotel_id,
            "apagados": await apagados_por_dimension(db, hotel_id)}
