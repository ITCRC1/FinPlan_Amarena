# -*- coding: utf-8 -*-
"""Los 17 conceptos de planilla YA CALCULADOS, un tab por departamento.

Owner, 2026-08-27: *«no hay ningún reporte donde yo veo los beneficios
calculados según las fórmulas»* · *«que a la parte de la descripción venga la
cuenta y driver»* · *«que sea un tab por departamento, todos de un solo»*.

## Por qué existe, si ya había un Excel de planilla

`payroll_excel.py` exporta otra cosa: **las posiciones** —persona, salario, FTE
por mes— y es la mitad de un viaje de ida y vuelta, porque el archivo que baja
es el que `POST /payroll/{id}/import/excel/` vuelve a leer. Por eso está
protegido y sólo deja editar salario y FTE.

Lo que no da —y es lo que se pedía— son los **conceptos derivados**: CCSS,
aguinaldo, provisión de vacaciones, cesantía. Esos no se digitan: los calcula el
motor con los parámetros del escenario, y hasta hoy sólo se veían en pantalla,
un departamento a la vez.

⚠️ **Este archivo NO se sube.** Es un reporte, no una plantilla. El de subir
sigue siendo `payroll_excel.py` — cambiar ése habría roto el import.

## La columna que contesta «¿de dónde salió este número?»

Cada fila trae **cuenta** (6000…6030) y **driver**: la regla con la que el motor
lo calculó, con el valor que tiene ESTE escenario. `CCSS 26.830% sobre BASE` no
es lo mismo que `CCSS` a secas — el primero se puede auditar sin abrir el
código, que es exactamente lo que hace falta en un libro que va a los dueños.

Un concepto que se digita dice **«digitado»**, no un driver inventado. La
diferencia importa: si alguien ve una fórmula donde no la hay, va a buscar el
error en el parámetro en vez de en el dato cargado.

## El orden de las filas es el de la pantalla, a propósito

BASE va después de los siete conceptos que la componen, y TOTAL al final. Quien
compara el Excel contra la pantalla no tiene que reordenar nada mentalmente —y
si algún día no cuadran, se ve en la fila, no en el total.
"""
from __future__ import annotations

from openpyxl import Workbook

from app.export.excel_base import (C, align, border, fill, font, MONTHS_ES,
                                   nombre_de_hoja, set_col_widths,
                                   workbook_to_bytes)

#: (clave del summary, cuenta, rótulo, tipo). El tipo decide el formato de la
#: fila: `sub` es un subtotal (BASE), `tot` el total del departamento.
FILAS = [
    ("c6000", "6000", "Salary and Wages",   "dato"),
    ("c6001", "6001", "Overtime",           "dato"),
    ("c6002", "6002", "Day Off",            "dato"),
    ("c6003", "6003", "Working Holiday",    "dato"),
    ("c6010", "6010", "Commissions",        "dato"),
    ("c6024", "6024", "Vacations Taken",    "dato"),
    ("c6027", "6027", "Incentive Bonus",    "dato"),
    ("base",  "",     "BASE",               "sub"),
    ("c6020", "6020", "Social Security",    "dato"),
    ("c6021", "6021", "Aguinaldo",          "dato"),
    ("c6004", "6004", "Disabilities",       "dato"),
    ("c6022", "6022", "Work Risk Policy",   "dato"),
    ("c6023", "6023", "Vacation Provision", "dato"),
    ("c6025", "6025", "Cafeteria",          "dato"),
    ("c6026", "6026", "Severance",          "dato"),
    ("c6028", "6028", "Housing",            "dato"),
    ("c6029", "6029", "Transport",          "dato"),
    ("c6030", "6030", "Other Benefits",     "dato"),
    ("total", "",     "TOTAL",              "tot"),
]

MONEDA = '#,##0.00;[Red]-#,##0.00;"—"'


def _drivers(p) -> dict[str, str]:
    """La regla de cada concepto, con el valor de ESTE escenario.

    `p` son los `PayrollParams` del escenario, o `None` si no tiene fila — en
    cuyo caso el motor usa sus defaults y acá se dice justamente eso, en vez de
    mostrar un cero que parecería un parámetro puesto en cero a propósito.
    """
    if p is None:
        return {k: "(sin parámetros del escenario — el motor usa sus defaults)"
                for k, _c, _n, t in FILAS if t == "dato"}

    def pct(v) -> str:
        return f"{float(v or 0) * 100:.3f}%"

    def crc(v) -> str:
        return f"₡{float(v or 0):,.2f}"

    def si(cond: bool, texto: str) -> str:
        return texto if cond else "digitado — sin driver configurado"

    ccss = float(p.ccss_rate or 0)
    div = float(p.aguinaldo_divisor or 0)
    return {
        "c6000": "digitado — salario × FTE ÷ TC del mes",
        "c6001": si(float(p.overtime_pct or 0) > 0,
                    f"{pct(p.overtime_pct)} sobre S&W"),
        "c6002": "calendario de días libres del escenario",
        "c6003": "digitado",
        "c6010": "digitado",
        "c6024": "digitado",
        "c6027": si(float(p.bonus_pct or 0) > 0, f"{pct(p.bonus_pct)} sobre S&W"),
        "c6020": f"{pct(ccss)} sobre BASE" if ccss else "sin tasa CCSS configurada",
        "c6021": f"BASE ÷ {div:g}" if div else "sin divisor de aguinaldo configurado",
        "c6004": "digitado",
        "c6022": si(float(p.ins_annual_crc or 0) > 0,
                    f"póliza INS {crc(p.ins_annual_crc)} al año, repartida por FTE"),
        "c6023": si(float(p.vacaciones_rate or 0) > 0,
                    f"{pct(p.vacaciones_rate)} sobre BASE"),
        "c6025": si(float(p.cafeteria_daily_crc or 0) > 0,
                    f"{crc(p.cafeteria_daily_crc)} por día trabajado"),
        "c6026": si(float(p.severance_annual_rate or 0) > 0,
                    f"{pct(p.severance_annual_rate)} anual sobre BASE"),
        "c6028": si(float(p.housing_monthly_crc or 0) > 0,
                    f"{crc(p.housing_monthly_crc)} por mes"),
        "c6029": si(float(p.transport_monthly_crc or 0) > 0,
                    f"{crc(p.transport_monthly_crc)} por mes"),
        "c6030": si(float(p.other_monthly_crc or 0) > 0,
                    f"{crc(p.other_monthly_crc)} por mes"),
    }


def _hoja(wb: Workbook, titulo: str, dept_code: str, dept_name: str,
          meses: list[dict], drv: dict[str, str], etiqueta: str, anio: int,
          usados: set[str]) -> None:
    ws = wb.create_sheet(nombre_de_hoja(f"{dept_code} {dept_name}"[:28], usados))

    ws["A1"] = f"{dept_code} — {dept_name}"
    ws["A1"].font = font(bold=True, size=13, color=C["navy"])
    ws["A2"] = f"{etiqueta} · {anio} · conceptos de planilla ya calculados (USD)"
    ws["A2"].font = font(size=9, color=C["text_mid"])
    ws["A3"] = ("El driver es la regla con la que el motor calculó la fila. "
                "«digitado» = viene de la carga, no de una fórmula.")
    ws["A3"].font = font(size=9, italic=True, color=C["text_mid"])

    ENC = 5
    cabeceras = ["Cuenta", "Concepto", "Driver"] + MONTHS_ES + ["ANUAL"]
    for i, texto in enumerate(cabeceras, start=1):
        c = ws.cell(row=ENC, column=i, value=texto)
        c.fill = fill(C["navy"])
        c.font = font(bold=True, color=C["white"], size=9)
        c.alignment = align("center" if i > 3 else "left", wrap=True)
        c.border = border(C["navy"])

    fila = ENC + 1
    for clave, cuenta, rotulo, tipo in FILAS:
        es_resumen = tipo in ("sub", "tot")
        relleno = (C["blue_light"] if tipo == "sub"
                   else C["blue_header"] if tipo == "tot" else None)

        ws.cell(row=fila, column=1, value=cuenta or "")
        ws.cell(row=fila, column=2, value=rotulo)
        ws.cell(row=fila, column=3, value="" if es_resumen else drv.get(clave, ""))

        anual = 0.0
        for m in range(12):
            v = float(meses[m].get(clave, 0) or 0) if m < len(meses) else 0.0
            anual += v
            c = ws.cell(row=fila, column=4 + m, value=v if v else None)
            c.number_format = MONEDA
        c = ws.cell(row=fila, column=16, value=anual if anual else None)
        c.number_format = MONEDA

        for col in range(1, 17):
            c = ws.cell(row=fila, column=col)
            c.border = border()
            c.font = font(bold=es_resumen, size=9,
                          color=C["navy"] if es_resumen else C["text_dark"])
            if relleno:
                c.fill = fill(relleno)
            if col == 3:
                c.alignment = align("left", wrap=True)
            elif col >= 4:
                c.alignment = align("right")
        fila += 1

    set_col_widths(ws, {1: 9, 2: 24, 3: 42,
                        **{4 + i: 13 for i in range(12)}, 16: 15})
    ws.freeze_panes = "D6"


def export_conceptos_por_depto(
    datos: list[dict],
    params,
    etiqueta: str,
    anio: int,
) -> bytes:
    """Un tab por departamento + un resumen adelante.

    `datos` = [{dept_code, dept_name, monthly: [12 dicts con c6000…total]}],
    ya en el orden en que se quieren las hojas.
    """
    wb = Workbook()
    wb.remove(wb.active)
    drv = _drivers(params)
    usados: set[str] = set()

    # El resumen va PRIMERO: quien abre el archivo tiene que ver el total del
    # hotel sin recorrer veinte pestañas para sumarlas de a una.
    ws = wb.create_sheet(nombre_de_hoja("Resumen", usados))
    ws["A1"] = f"Planilla por departamento — {etiqueta} · {anio}"
    ws["A1"].font = font(bold=True, size=13, color=C["navy"])
    ws["A2"] = "Total de los 17 conceptos, en USD. El detalle por concepto está en la pestaña de cada departamento."
    ws["A2"].font = font(size=9, color=C["text_mid"])

    ENC = 4
    for i, texto in enumerate(["Depto", "Nombre"] + MONTHS_ES + ["ANUAL"], start=1):
        c = ws.cell(row=ENC, column=i, value=texto)
        c.fill = fill(C["navy"])
        c.font = font(bold=True, color=C["white"], size=9)
        c.alignment = align("center" if i > 2 else "left", wrap=True)
        c.border = border(C["navy"])

    fila = ENC + 1
    totales = [0.0] * 12
    for d in datos:
        ws.cell(row=fila, column=1, value=d["dept_code"])
        ws.cell(row=fila, column=2, value=d["dept_name"])
        anual = 0.0
        for m in range(12):
            v = float(d["monthly"][m].get("total", 0) or 0) if m < len(d["monthly"]) else 0.0
            totales[m] += v
            anual += v
            c = ws.cell(row=fila, column=3 + m, value=v if v else None)
            c.number_format = MONEDA
        c = ws.cell(row=fila, column=15, value=anual if anual else None)
        c.number_format = MONEDA
        for col in range(1, 16):
            cc = ws.cell(row=fila, column=col)
            cc.border = border()
            cc.font = font(size=9)
            if col >= 3:
                cc.alignment = align("right")
        fila += 1

    ws.cell(row=fila, column=2, value="TOTAL")
    for m in range(12):
        c = ws.cell(row=fila, column=3 + m, value=totales[m] if totales[m] else None)
        c.number_format = MONEDA
    c = ws.cell(row=fila, column=15, value=sum(totales) if sum(totales) else None)
    c.number_format = MONEDA
    for col in range(1, 16):
        cc = ws.cell(row=fila, column=col)
        cc.fill = fill(C["blue_header"])
        cc.font = font(bold=True, size=9, color=C["navy"])
        cc.border = border()
        if col >= 3:
            cc.alignment = align("right")

    set_col_widths(ws, {1: 9, 2: 30, **{3 + i: 13 for i in range(12)}, 15: 15})
    ws.freeze_panes = "C5"

    for d in datos:
        _hoja(wb, d["dept_name"], d["dept_code"], d["dept_name"],
              d["monthly"], drv, etiqueta, anio, usados)

    return workbook_to_bytes(wb)
