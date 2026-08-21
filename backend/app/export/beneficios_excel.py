# -*- coding: utf-8 -*-
"""Excel de conceptos MANUALES de planilla, por departamento y posición.

No todo beneficio sale de una fórmula. Horas extra, bono, cesantía, transporte,
vivienda y otros se negocian caso por caso, así que el owner los pone a mano.

El archivo trae una hoja por concepto. En cada hoja: una fila por posición,
agrupadas por departamento, y las 12 columnas de meses. Se baja con los datos que
ya hay, se llenan los montos y se vuelve a subir.

Los montos van en DÓLARES, que es como vive la planilla en el P&L. La conversión
de colones ya la hizo el sistema con el TC del mes.
"""
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# Conceptos que el owner llena a mano. La clave es la columna del modelo.
CONCEPTOS_MANUALES = [
    ("c6001_overtime", "6001 Horas extra", "Cotiza CCSS y aguinaldo"),
    ("c6002_day_off", "6002 Dia libre", "Cotiza CCSS y aguinaldo"),
    ("c6010_commissions", "6010 Comisiones", "Cotiza CCSS y aguinaldo"),
    ("c6027_incentive_bonus", "6027 Bono", "Cotiza CCSS y aguinaldo"),
    ("c6024_vacations_taken", "6024 Vac. tomadas", "Cotiza CCSS y aguinaldo"),
    ("c6004_disabilities", "6004 Incapacidades", "No cotiza"),
    ("c6026_severance", "6026 Cesantia", "No cotiza"),
    ("c6029_transport", "6029 Transporte", "No cotiza"),
    ("c6028_housing", "6028 Vivienda", "No cotiza"),
    ("c6030_other", "6030 Otros beneficios", "No cotiza"),
]
# 6022 INS y 6025 Cafetería NO están: vienen del reparto, no de digitación.
# 6000/6020/6021/6023 tampoco: son fórmula.

AZUL = PatternFill("solid", fgColor="2D3A5C")
VERDE = PatternFill("solid", fgColor="1A7F4B")
GRIS = PatternFill("solid", fgColor="EFEFEF")
AMARILLO = PatternFill("solid", fgColor="FFFDE7")
BLANCO = Font(name="Arial", bold=True, color="FFFFFF", size=11)
NEGRO = Font(name="Arial", size=10)
NEGRITA = Font(name="Arial", bold=True, size=10)
AZUL_TXT = Font(name="Arial", bold=True, color="0000FF", size=10)
BORDE = Border(*[Side(style="thin", color="D0D0D0")] * 4)
USD = '#,##0.00;(#,##0.00);"-"'

CABECERA = ["Depto", "Nombre depto", "Codigo", "Puesto", "Empleado"]
COL_MES_0 = len(CABECERA) + 1          # primera columna de mes (F)


def _hoja(wb: Workbook, columna: str, titulo: str, nota: str, posiciones: list[dict]):
    ws = wb.create_sheet(titulo[:31])
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = (f"{nota}. Monto en DOLARES por mes. Deje en blanco o en 0 lo que no "
                f"aplique. Las celdas amarillas son las que usted llena.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")
    ws["A3"] = "NO cambie las columnas Depto / Codigo: son las que amarran cada fila."
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="C0392B")

    fila = 5
    for i, h in enumerate(CABECERA + MESES + ["Total"]):
        c = ws.cell(row=fila, column=i + 1, value=h)
        c.fill = AZUL
        c.font = BLANCO
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDE

    fila += 1
    depto_actual = None
    for p in posiciones:
        # Renglón de departamento, para leerlo agrupado
        if p["dept_code"] != depto_actual:
            depto_actual = p["dept_code"]
            c = ws.cell(row=fila, column=1, value=f"{p['dept_code']} — {p['dept_name']}")
            c.font = NEGRITA
            for k in range(1, len(CABECERA) + 14):
                ws.cell(row=fila, column=k).fill = GRIS
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila, end_column=len(CABECERA))
            fila += 1

        for i, v in enumerate([p["dept_code"], p["dept_name"], p["position_code"],
                               p["position_name"], p["employee_name"]]):
            c = ws.cell(row=fila, column=i + 1, value=v)
            c.font = NEGRO
            c.border = BORDE
        for m in range(12):
            c = ws.cell(row=fila, column=COL_MES_0 + m, value=float(p["meses"][m]))
            c.font = AZUL_TXT
            c.fill = AMARILLO
            c.number_format = USD
            c.border = BORDE
        ini, fin = get_column_letter(COL_MES_0), get_column_letter(COL_MES_0 + 11)
        t = ws.cell(row=fila, column=COL_MES_0 + 12, value=f"=SUM({ini}{fila}:{fin}{fila})")
        t.font = NEGRITA
        t.number_format = USD
        t.border = BORDE
        fila += 1

    # Total de la hoja
    c = ws.cell(row=fila + 1, column=len(CABECERA), value="TOTAL")
    c.font = BLANCO
    c.fill = VERDE
    for m in range(13):
        col = get_column_letter(COL_MES_0 + m)
        t = ws.cell(row=fila + 1, column=COL_MES_0 + m,
                    value=f"=SUM({col}6:{col}{fila - 1})")
        t.font = BLANCO
        t.fill = VERDE
        t.number_format = USD

    anchos = [9, 26, 11, 34, 30] + [12] * 13
    for i, w in enumerate(anchos):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = ws.cell(row=6, column=COL_MES_0)
    return ws


def export_beneficios(posiciones_por_concepto: dict[str, list[dict]],
                      etiqueta: str, year: int) -> bytes:
    """posiciones_por_concepto[columna] = [{dept_code, dept_name, position_code,
    position_name, employee_name, meses: [12 Decimals]}, …]"""
    wb = Workbook()
    wb.remove(wb.active)

    guia = wb.create_sheet("Guia")
    guia["A1"] = f"Conceptos manuales de planilla — {etiqueta} {year}"
    guia["A1"].font = Font(name="Arial", bold=True, size=14)
    for i, texto in enumerate([
        "",
        "Una hoja por concepto. Llene el monto en dolares, mes por mes, en las celdas amarillas.",
        "Vuelva a subir el mismo archivo: se identifican las filas por Depto + Codigo de puesto.",
        "",
        "Lo que NO esta aqui, porque no se digita:",
        "   6000 Salario        = salario x FTE / TC",
        "   6020 CCSS           = BASE x tasa",
        "   6021 Aguinaldo      = BASE / divisor",
        "   6023 Vac. provision = BASE x tasa   (si la tasa esta en cero, si se puede digitar)",
        "   6003 Feriados       = salario / dias del mes x feriados",
        "   6022 INS            = REPARTO del monto de la poliza, por FTE",
        "   6025 Cafeteria      = REPARTO del costo del depto 0220, por FTE (tab Allocations)",
        "",
        "Ojo: si en Parametros de Planilla le pone una tasa a un concepto, la formula manda",
        "y pisa lo que usted suba aqui. Deje la tasa en CERO para que mande el dato manual.",
    ], start=2):
        guia[f"A{i}"] = texto
        guia[f"A{i}"].font = Font(name="Arial", size=10)
    guia.column_dimensions["A"].width = 100

    for columna, titulo, nota in CONCEPTOS_MANUALES:
        _hoja(wb, columna, titulo, nota, posiciones_por_concepto.get(columna, []))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_beneficios(contenido: bytes) -> tuple[dict[str, list[dict]], list[str]]:
    """Lee el archivo de vuelta.

    Devuelve ({columna: [{dept_code, position_code, meses[12]}, …]}, avisos).
    Solo se leen las hojas que corresponden a un concepto conocido; una hoja de
    mas se ignora en vez de tumbar toda la carga.
    """
    wb = load_workbook(BytesIO(contenido), data_only=True)
    por_titulo = {t: c for c, t, _ in CONCEPTOS_MANUALES}
    out: dict[str, list[dict]] = {}
    avisos: list[str] = []

    for hoja in wb.sheetnames:
        columna = por_titulo.get(hoja)
        if columna is None:
            if hoja != "Guia":
                avisos.append(f"Hoja '{hoja}' ignorada: no corresponde a ningun concepto.")
            continue
        ws = wb[hoja]
        filas = []
        for r in ws.iter_rows(min_row=6, values_only=True):
            if not r or r[0] is None:
                continue
            dept = str(r[0]).strip()
            if "—" in dept or not dept:      # renglón de departamento
                continue
            code = str(r[2]).strip() if r[2] is not None else ""
            if not code:
                avisos.append(f"{hoja}: fila sin codigo de puesto ({dept}) — se ignora.")
                continue
            meses = []
            for m in range(12):
                v = r[COL_MES_0 - 1 + m] if len(r) > COL_MES_0 - 1 + m else None
                try:
                    meses.append(Decimal(str(v)) if v not in (None, "") else Decimal("0"))
                except (ValueError, TypeError):
                    meses.append(Decimal("0"))
                    avisos.append(f"{hoja}: valor no numerico en {dept}/{code}, mes {m + 1}.")
            if any(v != 0 for v in meses):
                filas.append({"dept_code": dept, "position_code": code, "meses": meses})
        out[columna] = filas
    return out, avisos
