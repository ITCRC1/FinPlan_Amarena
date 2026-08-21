"""P&L Full Detail → Excel.

El reporte existe en pantalla desde la Fase 2, pero no se podía mandar por
correo. Esto lo cierra: el mismo armado, el mismo cuadre, en un `.xlsx` que se
abre en cualquier lado.

## Qué se respeta del Excel original

- **Los negativos entre paréntesis y en rojo**, y **el cero no se imprime**
  (`#,##0;[Red](#,##0);""`). Es la convención del archivo del owner y lo que
  deja respirar una grilla de 700 filas.
- **Un bloque por departamento**, con la plantilla que los 16 bloques del Excel
  repiten igual: INGRESOS → COSTO DE VENTAS → NÓMINA → GASTOS OPERATIVOS →
  UTILIDAD NETA, más los dos ratios.
- **La jerarquía por color**, no por sangría (en el original las 1,007 filas
  tienen `indent = 0` y lo único que distingue un total de un detalle es el
  relleno). Acá va el color **y** la sangría, que es lo legible.

## Qué se corrige

- Los ratios «% de Ingresos del Depto.» y «% Utilidad» tenían **formato de
  moneda** en el original: se leían `$0.35` en vez de `35.0%`. Acá van como
  porcentaje.
- Los números iban **centrados**. Acá van a la derecha, que es como se leen las
  columnas de cifras.
- El total anual es **siempre** la suma de los doce meses. En el original la
  fila 76 sacaba los meses de un lado y el anual de otro, y era la única fila
  del archivo con dos fuentes.

La hoja **CUADRE** va primero a propósito: si el detalle no amarra con el
resumen, eso es lo primero que hay que ver, no algo escondido al final.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment

from app.export.excel_base import (
    MONTHS_ES, align, border, fill, font, set_col_widths, workbook_to_bytes,
)

# Negativos entre paréntesis y en rojo; el cero NO se imprime.
FMT_USD = '#,##0;[Red](#,##0);""'
FMT_PCT = '0.0%;[Red](0.0%);""'
FMT_NUM = '#,##0;[Red](#,##0);""'

COL_CUENTA, COL_ETIQUETA, COL_M1 = 1, 2, 3
COL_TOTAL = COL_M1 + 12          # columna 15
ANCHO = {COL_CUENTA: 9, COL_ETIQUETA: 46, COL_TOTAL: 15,
         **{COL_M1 + i: 12 for i in range(12)}}

# Un color por tipo de fila. Reemplaza los 13 estilos del original, que decían
# lo mismo con más ruido.
ESTILO = {
    "seccion":  {"bg": "1A3A5C", "fg": "FFFFFF", "bold": True},
    "total":    {"bg": "E8EDF3", "fg": "1A1A2E", "bold": True},
    "subtotal": {"bg": "F8F9FA", "fg": "1A1A2E", "bold": True},
    "pct":      {"bg": None,     "fg": "4A5568", "bold": False, "italic": True},
    "stat":     {"bg": None,     "fg": "4A5568", "bold": False},
    "detalle":  {"bg": None,     "fg": "1A1A2E", "bold": False},
}

# Noches y unidades no son dólares. Con formato de moneda, «12,410 noches» se
# lee «$12,410» — el mismo error que el original cometía con los ratios.
FMT_STAT = '#,##0.#;[Red](#,##0.#);""'



def _encabezado(ws, row: int, titulo: str) -> int:
    ws.cell(row=row, column=COL_CUENTA, value="Cta")
    ws.cell(row=row, column=COL_ETIQUETA, value=titulo)
    for i, m in enumerate(MONTHS_ES):
        ws.cell(row=row, column=COL_M1 + i, value=m)
    ws.cell(row=row, column=COL_TOTAL, value="TOTAL AÑO")
    for c in range(COL_CUENTA, COL_TOTAL + 1):
        cel = ws.cell(row=row, column=c)
        cel.fill = fill("2D5A9E")
        cel.font = font(bold=True, color="FFFFFF", size=10)
        cel.alignment = align("center" if c >= COL_M1 else "left")
        cel.border = border()
    return row + 1


def _fila(ws, row: int, f: dict) -> int:
    est = ESTILO.get(f["tipo"], ESTILO["detalle"])
    es_pct = f["tipo"] == "pct"
    fmt = FMT_PCT if es_pct else (FMT_STAT if f["tipo"] == "stat" else FMT_USD)

    cta = ws.cell(row=row, column=COL_CUENTA, value=f.get("cuenta") or None)
    cta.font = font(color="4A5568", size=9)
    cta.alignment = align("left")

    # La sangría dice el nivel. En el original las 1,007 filas tenían
    # `indent = 0` y la jerarquía vivía solo en el color de relleno; acá van las
    # dos, que es lo que se lee de un vistazo.
    eti = ws.cell(row=row, column=COL_ETIQUETA, value=f["etiqueta"])
    eti.font = font(bold=est["bold"], color=est["fg"], size=10,
                    italic=est.get("italic", False))
    eti.alignment = Alignment(horizontal="left", vertical="center",
                              indent=int(f.get("nivel") or 0))

    for i, v in enumerate(f["meses"]):
        c = ws.cell(row=row, column=COL_M1 + i, value=(v if v else None))
        c.number_format = fmt
        c.font = font(bold=est["bold"], color=est["fg"], size=10,
                      italic=est.get("italic", False))
        c.alignment = align("right")
    t = ws.cell(row=row, column=COL_TOTAL, value=(f["total"] if f["total"] else None))
    t.number_format = fmt
    t.font = font(bold=True, color=est["fg"], size=10, italic=est.get("italic", False))
    t.alignment = align("right")
    t.border = border(sides="all")

    if est["bg"]:
        for c in range(COL_CUENTA, COL_TOTAL + 1):
            ws.cell(row=row, column=c).fill = fill(est["bg"])
    if f["tipo"] in ("total", "subtotal"):
        for c in range(COL_CUENTA, COL_TOTAL + 1):
            ws.cell(row=row, column=c).border = border(sides="top_bottom")
    return row + 1


def _hoja(wb: Workbook, nombre: str) -> "Worksheet":  # noqa: F821
    # Excel no admite nombres de hoja de más de 31 caracteres ni con : \\ / ? * [ ]
    limpio = "".join(ch for ch in nombre if ch not in ':\\/?*[]')[:31]
    return wb.create_sheet(limpio)


def _cuadre(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "CUADRE"
    set_col_widths(ws, {1: 38, 2: 20, 3: 20, 4: 18})

    ws.cell(row=1, column=1, value="P&L FULL DETAIL").font = font(bold=True, size=16)
    ws.cell(row=2, column=1, value=data["scenario"]).font = font(size=12, color="4A5568")
    ws.cell(row=3, column=1,
            value=f"Moneda {data['moneda']} · fuente del P&L: {data['source_mode']}"
            ).font = font(size=10, color="4A5568")

    c = data["cuadre"]
    r = 5
    ws.cell(row=r, column=1, value="¿EL DETALLE AMARRA CON EL RESUMEN?").font = font(bold=True, size=11)
    r += 1
    veredicto = "SÍ — cuadra" if c["ok"] else "NO — hay diferencia"
    cel = ws.cell(row=r, column=1, value=veredicto)
    cel.font = font(bold=True, size=12, color=("1A5C3A" if c["ok"] else "C0392B"))
    r += 2

    for h, col in (("Concepto", 1), ("Detalle", 2), ("Resumen (motor)", 3), ("Diferencia", 4)):
        x = ws.cell(row=r, column=col, value=h)
        x.fill = fill("2D5A9E"); x.font = font(bold=True, color="FFFFFF")
        x.alignment = align("center" if col > 1 else "left")
    r += 1
    filas = [("Ingresos", c["ingresos_detalle"], c["ingresos_pl"], c["dif_ingresos"]),
             ("Gastos (operativos + overhead)", c["gastos_detalle"], c["gastos_pl"], c["dif_gastos"])]
    for nombre, det, pl, dif in filas:
        ws.cell(row=r, column=1, value=nombre)
        for col, v in ((2, det), (3, pl), (4, dif)):
            x = ws.cell(row=r, column=col, value=v)
            x.number_format = FMT_USD
            x.alignment = align("right")
        r += 1
    r += 1
    for nombre, v in (("GOP", c["gop_pl"]), ("Utilidad neta", c["net_pl"])):
        ws.cell(row=r, column=1, value=nombre).font = font(bold=True)
        x = ws.cell(row=r, column=2, value=v)
        x.number_format = FMT_USD; x.font = font(bold=True); x.alignment = align("right")
        r += 1

    if not c["ingreso_por_cuenta"]:
        r += 1
        ws.cell(row=r, column=1,
                value="El ingreso de este escenario NO está abierto por cuenta.").font = font(bold=True)
        r += 1

    for aviso in data.get("avisos", []):
        r += 1
        cel = ws.cell(row=r, column=1, value=aviso)
        cel.font = font(size=10, color="92400E")
        cel.alignment = align("left", wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 46


def _kpis(wb: Workbook, data: dict) -> None:
    k = data.get("kpis") or {}
    sets = k.get("sets") or []
    if not sets:
        return
    ws = _hoja(wb, "HABITACIONES POR SET")
    set_col_widths(ws, {1: 26, **{i: 15 for i in range(2, 9)}})
    ws.cell(row=1, column=1, value="Habitaciones — abierto por set").font = font(bold=True, size=13)

    cols = ["Set", "Unidades", "Noches disp.", "Noches ocup.",
            "% Ocupación", "ADR", "RevPAR", "Ingreso"]
    for i, h in enumerate(cols):
        x = ws.cell(row=3, column=1 + i, value=h)
        x.fill = fill("2D5A9E"); x.font = font(bold=True, color="FFFFFF")
        x.alignment = align("center" if i else "left")

    filas = list(sets) + ([k["consolidado"]] if k.get("consolidado") else [])
    r = 4
    for i, s in enumerate(filas):
        es_total = k.get("consolidado") and i == len(filas) - 1
        nombre = s["nombre"] + ("   (sin ocupación cargada)" if s.get("sin_ocupacion") else "")
        ws.cell(row=r, column=1, value=nombre).font = font(bold=es_total)
        vals = [(2, s["unidades"], FMT_NUM),
                (3, sum(s["noches_disponibles"]), FMT_NUM),
                (4, sum(s["noches_ocupadas"]), FMT_NUM),
                (5, s["ocupacion_anual"], FMT_PCT),
                (6, s["adr_anual"], FMT_USD),
                (7, s["revpar_anual"], FMT_USD),
                (8, s["revenue_anual"], FMT_USD)]
        for col, v, fmt in vals:
            x = ws.cell(row=r, column=col, value=(v or None))
            x.number_format = fmt; x.alignment = align("right"); x.font = font(bold=es_total)
        if es_total:
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = fill("E8EDF3")
        r += 1

    if k.get("diluyen"):
        r += 1
        cel = ws.cell(row=r, column=1, value=(
            "Sin ocupación cargada: " + ", ".join(k["diluyen"]) +
            ". Sus unidades SÍ suman noches disponibles al consolidado, así que "
            "diluyen la ocupación general y el RevPAR del hotel."))
        cel.font = font(size=10, color="92400E")
        cel.alignment = align("left", wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 32


def _bloque_de_filas(wb: Workbook, nombre: str, titulo: str, filas: list[dict]) -> None:
    if not filas:
        return
    ws = _hoja(wb, nombre)
    set_col_widths(ws, ANCHO)
    ws.cell(row=1, column=COL_ETIQUETA, value=titulo).font = font(bold=True, size=13)
    r = _encabezado(ws, 3, "Concepto")
    ws.freeze_panes = ws.cell(row=r, column=COL_M1)
    for f in filas:
        r = _fila(ws, r, f)


def build_pl_full_detail_workbook(data: dict) -> bytes:
    """`data` es exactamente lo que devuelve `GET /reports/pl-full-detail/{id}/`.

    El exportador NO recalcula nada: si dibujara sus propios números, el Excel y
    la pantalla podrían decir cosas distintas y no habría forma de saber cuál
    creer.
    """
    wb = Workbook()
    _cuadre(wb, data)
    _kpis(wb, data)
    _bloque_de_filas(wb, "RESUMEN", f"Resumen — {data['scenario']}", data.get("resumen") or [])

    usados: set[str] = set()
    for b in data.get("bloques") or []:
        base = f"{b['dept_code']} {b['titulo']}"
        nombre = "".join(ch for ch in base if ch not in ':\\/?*[]')[:31]
        # Dos departamentos con nombre largo pueden colapsar en el mismo rótulo
        # de 31 caracteres; Excel rechaza el libro entero si se repite.
        if nombre in usados:
            nombre = f"{nombre[:27]} {b['dept_code']}"[:31]
        usados.add(nombre)
        # Los sets de Rooms llevan el aviso en el título de su hoja: en un libro
        # de veinte pestañas iguales, nada distingue una apertura de un
        # departamento, y sumar las cuatro contaría Rooms dos veces.
        if b.get("es_apertura"):
            etiqueta = f"apertura de {b.get('apertura_de') or ''} · NO se suma aparte"
        else:
            etiqueta = "overhead" if b["tipo"] == "OVERHEAD" else "operativo"
        _bloque_de_filas(wb, nombre, f"{b['titulo']} ({b['dept_code']} · {etiqueta})",
                         b["filas"])

    _bloque_de_filas(wb, "GASTOS DE PROPIEDAD", "Gastos de Propiedad (below-GOP)",
                     data.get("propiedad") or [])
    return workbook_to_bytes(wb)
