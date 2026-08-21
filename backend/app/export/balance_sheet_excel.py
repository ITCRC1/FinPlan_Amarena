"""Export del Balance Sheet a una **plantilla Excel bloqueada** y round-trip
con el importador (formato Summary limpio: fila 'Account' + fechas, una
columna USD por mes).

El histórico queda bloqueado; SOLO la columna del mes nuevo (corriente) es
editable. El usuario llena esa columna y vuelve a subir el archivo por
`import-balance-sheet` (merge agrega ese mes).
"""
from __future__ import annotations
import datetime as _dt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.export.excel_base import (
    font, fill, border, align, set_col_widths, workbook_to_bytes,
    unlock, protect_sheet, C,
)

MONEY = '#,##0.00;(#,##0.00)'
DATEFMT = "mmm-yyyy"


def export_balance_sheet_template(columns: list[dict], rows: list[dict]) -> bytes:
    """`columns` = [{year, month, editable}] de izq. a der.
    `rows` = [{label, indent, is_total, section, values:{(y,m): usd|None}}]."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # Título (lo usa el importador como marcador de bloque)
    t = ws.cell(row=1, column=2, value="Balance Sheet Summary")
    t.font = font(bold=True, color=C["navy"], size=14)

    HDR = 3
    FIRST = HDR + 1
    # Cabecera 'Account' + fechas (datetime para que el importador lea el mes)
    h = ws.cell(row=HDR, column=2, value="Account")
    h.fill = fill(C["navy"]); h.font = font(bold=True, color="FFFFFF", size=10)
    h.alignment = align("left"); h.border = border()
    for i, col in enumerate(columns):
        cc = 3 + i
        cell = ws.cell(row=HDR, column=cc, value=_dt.datetime(col["year"], col["month"], 1))
        cell.number_format = DATEFMT
        cell.fill = fill(C["navy_mid"] if col["editable"] else C["navy"])
        cell.font = font(bold=True, color="FFFFFF", size=10)
        cell.alignment = align("center"); cell.border = border()

    # Filas del estado
    r = FIRST
    for row in rows:
        label = ("   " * int(row.get("indent", 0))) + row["label"]
        is_total = bool(row.get("is_total"))
        section = row.get("section", "")
        is_section = is_total and row["label"].strip().lower() in ("assets", "liabilities", "capital", "equity")
        lc = ws.cell(row=r, column=2, value=label)
        lc.alignment = align("left")
        if is_section:
            lc.font = font(bold=True, color="FFFFFF", size=11)
            lc.fill = fill(C["navy_mid"])
        else:
            lc.font = font(bold=is_total, color=C["text_dark"], size=10)
        for i, col in enumerate(columns):
            cc = 3 + i
            v = row["values"].get((col["year"], col["month"]))
            cell = ws.cell(row=r, column=cc)
            cell.number_format = MONEY
            cell.alignment = align("right")
            cell.font = font(bold=is_total, color=C["text_dark"], size=10)
            if col["editable"]:
                unlock(cell)                       # columna del mes a llenar: editable
                cell.fill = fill(C["amber_light"])
                # si ese mes ya tiene datos (re-carga/corrección), pre-llenar
                if v is not None:
                    cell.value = round(float(v), 2)
            else:
                cell.value = None if v is None else round(float(v), 2)
            if is_section:
                cell.fill = fill(C["navy_mid"])
        r += 1
    last = r - 1

    # Nota de uso (debajo)
    note = ws.cell(row=last + 2, column=2,
                   value="Complete SOLO la columna resaltada (mes nuevo) y vuelva a subir este archivo. "
                         "El resto está bloqueado. Total assets debe ser igual a Capital & Liab. Sum.")
    note.font = font(italic=True, color=C["text_mid"], size=9)

    set_col_widths(ws, {1: 2, 2: 46, **{3 + i: 16 for i in range(len(columns))}})
    ws.freeze_panes = "C4"

    # Bloqueo: todo locked salvo lo que marcamos unlock()
    protect_sheet(ws)
    return workbook_to_bytes(wb)
