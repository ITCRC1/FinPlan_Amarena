# -*- coding: utf-8 -*-
"""El archivo de carga de estadísticas: se baja con la grilla, se llena, se sube.

Mismas mañas que la plantilla de actuales, que es la norma del owner: «yo bajo,
corrijo y subo lo que guardé». Concretamente:

* **Las columnas se ubican por ENCABEZADO, nunca por posición.** Si alguien mueve
  una columna o agrega la suya, la carga la sigue encontrando.
* Viene con la grilla YA ARMADA a partir del escenario: sus posiciones, sus
  departamentos, sus tipos de habitación. Lo que se deja en blanco es cero.
* Trae los valores que ya estén cargados, para poder corregir sin volver a
  digitar todo.

Las columnas de identidad —cuenta, departamento, posición, tipo de habitación—
son la LLAVE. Si se tocan, la fila deja de ser la misma; por eso van bloqueadas
visualmente (fondo gris) y la carga avisa cuando no reconoce una.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL = "2D3A5C"
GRIS = "EDEFF2"
VERDE = "1A7F4B"

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# El orden y el nombre EXACTO de cada columna. La carga busca por estos
# encabezados, así que cambiarlos acá cambia el contrato del archivo.
LLAVE = ["Cuenta", "Concepto", "Unidad", "Depto", "Departamento",
         "Posición", "Nombre posición", "Tipo hab.", "Tipo de habitación"]

_borde = Side(style="thin", color="D0D5DD")
BORDE = Border(_borde, _borde, _borde, _borde)


def construir_libro(titulo: str, subtitulo: str, filas, valores=None):
    """`filas` son `FilaGrilla`; `valores` es {llave: {mes: monto}}."""
    valores = valores or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadisticas"

    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", bold=True, size=15, color=AZUL)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="667085")
    ws["A3"] = ("Llená los meses. Lo que dejes en blanco es cero. NO cambies las "
                "columnas grises: son la llave de cada fila.")
    ws["A3"].font = Font(name="Calibri", size=10, color="B54708")

    cols = LLAVE + MESES
    for i, t in enumerate(cols, start=1):
        c = ws.cell(row=5, column=i, value=t)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL if i > len(LLAVE) else VERDE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE
    ws.row_dimensions[5].height = 26

    gris = PatternFill("solid", fgColor=GRIS)
    for r, f in enumerate(filas, start=6):
        datos = [f.account_code, f.account_name, f.unidad,
                 f.dept_code, f.dept_name,
                 f.position_code, f.position_name,
                 f.room_type_code, f.room_type_name]
        for i, v in enumerate(datos, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.fill = gris
            c.border = BORDE
            c.font = Font(name="Calibri", size=9.5,
                          bold=(i == 1), color="344054")
        serie = valores.get(f.llave, {})
        for m in range(1, 13):
            c = ws.cell(row=r, column=len(LLAVE) + m, value=serie.get(m))
            c.border = BORDE
            c.number_format = '#,##0.00;[Red](#,##0.00);""'
            c.font = Font(name="Calibri", size=10)

    for i, ancho in enumerate([9, 30, 12, 8, 24, 11, 26, 10, 24] + [11] * 12, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "J6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(cols))}{5 + len(filas)}"

    # Segunda hoja: el catálogo, para saber qué mide cada cuenta sin adivinar.
    ws2 = wb.create_sheet("Cuentas")
    ws2["A1"] = "Qué mide cada cuenta"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color=AZUL)
    for i, t in enumerate(["Cuenta", "Concepto", "Unidad", "Se abre por",
                           "El año es"], start=1):
        c = ws2.cell(row=3, column=i, value=t)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.border = BORDE
    vistas, r = set(), 4
    for f in filas:
        if f.account_code in vistas:
            continue
        vistas.add(f.account_code)
        for i, v in enumerate([f.account_code, f.account_name, f.unidad], start=1):
            ws2.cell(row=r, column=i, value=v).border = BORDE
        r += 1
    for i, ancho in enumerate([10, 34, 14, 26, 24], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = ancho
    return wb
