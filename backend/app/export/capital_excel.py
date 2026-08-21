"""
Capital Project — export / import de Excel.

Una sola hoja, en el mismo orden en que se presenta: las áreas separadas por su
encabezado y su subtotal, y abajo el total general.

Todo lo que se puede sumar va como FÓRMULA, no como número quemado: el total de
cada renglón, el subtotal de cada área y el total general. Así el dueño edita un
mes en Excel y ve el efecto sin volver a subir el archivo, que es justo para lo
que abre el Excel.

La hoja va protegida dejando abiertas SOLO las celdas de captura (nombre, notas
y los doce meses). No es para esconder nada — es para que no se borre por
accidente una fórmula y el archivo vuelva descuadrado.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app.export.excel_base import (
    C, fill, font, border, align,
    merged_header, set_col_widths, workbook_to_bytes,
    unlock, protect_sheet,
    MONTHS_ES, MONTH_ATTRS,
)

HOJA = "Capital Project"

COL_AREA = 1    # A
COL_NAME = 2    # B
COL_JAN = 3     # C
COL_DEC = 14    # N
COL_TOTAL = 15  # O
COL_NOTES = 16  # P

FILA_ENCABEZADO = 4
PRIMERA_FILA = FILA_ENCABEZADO + 1

# Dólares sin centavos: la lista se arma en miles y los centavos solo agregan
# ruido a doce columnas. El cero se muestra vacío para que se lea de un vistazo
# en qué meses cae el gasto.
USD = '$#,##0;($#,##0);""'
USD_TOTAL = '$#,##0;($#,##0);"$0"'

# Amarillo pálido = celda que se puede escribir. Es la misma convención que ya
# usan los otros Excel del sistema (beneficios, planilla), así que quien abre
# este archivo no tiene que aprender un código nuevo. Los montos además van en
# azul: en un mar de doce columnas, el color de fondo solo no alcanza cuando la
# celda está vacía.
EDITABLE = "FFFDE7"
AZUL_INPUT = "2D5A9E"


def export_capital_to_excel(rows: list[dict], titulo: str) -> bytes:
    """`rows` = renglones ya ordenados (area, name, notes, jan..dec)."""
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA

    merged_header(ws, 1, COL_AREA, COL_NOTES, "CAPITAL PROJECT", C["navy"], sz=14)
    merged_header(ws, 2, COL_AREA, COL_NOTES, titulo, C["navy_mid"], sz=10)

    # Leyenda del color, arriba y no al pie: se tiene que ver antes de empezar a
    # escribir, no después.
    leyenda = ws.cell(row=3, column=COL_AREA,
                      value="Las celdas en amarillo se pueden editar. El resto son fórmulas y están bloqueadas.")
    leyenda.font = font(bold=True, color=C["text_mid"], size=9)
    ws.cell(row=3, column=COL_TOTAL, value="editable →").font = font(color=C["text_mid"], size=8, italic=True)
    muestra = ws.cell(row=3, column=COL_NOTES)
    muestra.fill = fill(EDITABLE)
    muestra.border = border()

    enc = ["ÁREA", "PROYECTO", *MONTHS_ES, "TOTAL", "NOTAS"]
    for i, texto in enumerate(enc, start=COL_AREA):
        c = ws.cell(row=FILA_ENCABEZADO, column=i, value=texto)
        c.fill = fill(C["navy"])
        c.font = font(bold=True, color=C["white"], size=9)
        c.alignment = align("center" if i >= COL_JAN else "left")
        c.border = border()

    # Agrupadas por área conservando el orden de captura.
    #
    # Los renglones vacíos (sin nombre y sin montos) NO se escriben. Son los que
    # deja el botón «Agregar» en la pantalla mientras se llenan, no llevan
    # información, y en el archivo serían indistinguibles de la banda del área
    # —misma columna A, todo lo demás en blanco—, así que al subirlo de vuelta
    # el lector no sabría cuál es cuál. Mejor que no existan que arriesgar que
    # una banda se lea como proyecto.
    orden: list[str] = []
    grupos: dict[str, list[dict]] = {}
    for r in rows:
        if not (r.get("name") or "").strip() and not any(float(r.get(m) or 0) for m in MONTH_ATTRS):
            continue
        a = r.get("area") or "Sin área"
        if a not in grupos:
            grupos[a] = []
            orden.append(a)
        grupos[a].append(r)

    fila = PRIMERA_FILA
    filas_subtotal: list[int] = []

    for area in orden:
        # Sin .upper(): el encabezado y la columna del renglón tienen que decir
        # exactamente lo mismo. Si difieren y alguien borra la columna A de una
        # fila, al subirlo aparecería un área duplicada con otra grafía.
        #
        # La banda va en azul y NO en ámbar: el ámbar se parecía demasiado al
        # amarillo de las celdas de captura y hacía dudar si el encabezado del
        # área se podía escribir. Acá el amarillo significa una sola cosa.
        c = ws.cell(row=fila, column=COL_AREA, value=area)
        c.font = font(bold=True, color=C["white"], size=10)
        for col in range(COL_AREA, COL_NOTES + 1):
            ws.cell(row=fila, column=col).fill = fill(C["navy_mid"])
        fila += 1

        primera = fila
        for r in grupos[area]:
            # El área se repite en cada renglón para que el archivo se pueda leer
            # de vuelta sin depender de los encabezados de color.
            ca = ws.cell(row=fila, column=COL_AREA, value=area)
            ca.font = font(color=C["text_mid"], size=9)
            ca.border = border()
            ca.fill = fill(EDITABLE)
            unlock(ca)

            cn = ws.cell(row=fila, column=COL_NAME, value=r.get("name") or "")
            cn.font = font(size=10)
            cn.border = border()
            cn.fill = fill(EDITABLE)
            unlock(cn)

            for j, m in enumerate(MONTH_ATTRS):
                valor = float(r.get(m) or 0)
                cm = ws.cell(row=fila, column=COL_JAN + j, value=valor or None)
                cm.number_format = USD
                cm.border = border()
                cm.alignment = align("right")
                cm.fill = fill(EDITABLE)
                cm.font = font(bold=True, color=AZUL_INPUT, size=10)
                unlock(cm)

            ct = ws.cell(
                row=fila, column=COL_TOTAL,
                value=f"=SUM({get_column_letter(COL_JAN)}{fila}:{get_column_letter(COL_DEC)}{fila})",
            )
            ct.number_format = USD_TOTAL
            ct.font = font(bold=True)
            ct.border = border()

            cnot = ws.cell(row=fila, column=COL_NOTES, value=r.get("notes") or "")
            cnot.font = font(color=C["text_mid"], size=9, italic=True)
            cnot.border = border()
            cnot.fill = fill(EDITABLE)
            unlock(cnot)
            fila += 1

        # Subtotal del área. Si el área quedó sin renglones, igual se escribe en
        # cero: una lista con un área vacía es información, no un hueco.
        cs = ws.cell(row=fila, column=COL_NAME, value=f"Subtotal {area}")
        cs.font = font(bold=True, color=C["navy"])
        for col in range(COL_JAN, COL_TOTAL + 1):
            letra = get_column_letter(col)
            cc = ws.cell(row=fila, column=col,
                         value=f"=SUM({letra}{primera}:{letra}{fila - 1})" if fila > primera else 0)
            cc.number_format = USD_TOTAL
            cc.font = font(bold=True, color=C["navy"])
        for col in range(COL_AREA, COL_NOTES + 1):
            ws.cell(row=fila, column=col).fill = fill(C["blue_light"])
            ws.cell(row=fila, column=col).border = border()
        filas_subtotal.append(fila)
        fila += 2   # renglón en blanco: separa las áreas a la vista

    # Total general = suma de los subtotales (no de los renglones): así no se
    # duplica si alguien agrega filas dentro de un área.
    ct = ws.cell(row=fila, column=COL_NAME, value="TOTAL INVERSIÓN DE CAPITAL")
    ct.font = font(bold=True, color=C["white"], size=11)
    for col in range(COL_JAN, COL_TOTAL + 1):
        letra = get_column_letter(col)
        suma = "+".join(f"{letra}{f}" for f in filas_subtotal) or "0"
        cc = ws.cell(row=fila, column=col, value=f"={suma}")
        cc.number_format = USD_TOTAL
        cc.font = font(bold=True, color=C["white"], size=11)
    for col in range(COL_AREA, COL_NOTES + 1):
        ws.cell(row=fila, column=col).fill = fill(C["navy"])
        ws.cell(row=fila, column=col).border = border()

    ws.cell(row=fila + 2, column=COL_AREA,
            value="Se pueden editar: área, proyecto, los doce meses y las notas. "
                  "Los totales son fórmulas y están bloqueados. Para volver a subirlo, "
                  "no cambie el orden de las columnas.").font = font(color=C["text_mid"], size=8, italic=True)

    set_col_widths(ws, {COL_AREA: 24, COL_NAME: 38, COL_TOTAL: 14, COL_NOTES: 26,
                        **{c: 11 for c in range(COL_JAN, COL_DEC + 1)}})
    ws.freeze_panes = ws.cell(row=PRIMERA_FILA, column=COL_JAN)
    protect_sheet(ws)
    return workbook_to_bytes(wb)


def _num(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    try:
        return Decimal(str(v).replace("$", "").replace(",", "").replace(" ", "").strip() or "0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


def import_capital_from_excel(contenido: bytes) -> list[dict]:
    """Lee el archivo exportado y devuelve los renglones.

    Se descartan los subtotales, el total general y las filas en blanco: los
    primeros vuelven a calcularse y las últimas son los separadores que el
    propio export escribe entre áreas.
    """
    wb = load_workbook(filename=__import__("io").BytesIO(contenido), data_only=True)
    ws = wb[HOJA] if HOJA in wb.sheetnames else wb.active

    filas: list[dict] = []
    area_actual = ""
    for fila in ws.iter_rows(min_row=PRIMERA_FILA, values_only=True):
        celdas = list(fila) + [None] * (COL_NOTES - len(fila))
        area = str(celdas[COL_AREA - 1] or "").strip()
        nombre = str(celdas[COL_NAME - 1] or "").strip()
        meses = [_num(celdas[COL_JAN - 1 + i]) for i in range(12)]

        if nombre.lower().startswith("subtotal") or nombre.upper().startswith("TOTAL"):
            continue
        # Encabezado de área: solo la columna A, en mayúsculas y sin montos.
        if area and not nombre and not any(meses):
            area_actual = area
            continue
        if not nombre and not any(meses):
            continue
        if area:
            area_actual = area

        filas.append({
            "area": area_actual, "name": nombre,
            "notes": str(celdas[COL_NOTES - 1] or "").strip(),
            "sort_order": len(filas),
            **{m: meses[i] for i, m in enumerate(MONTH_ATTRS)},
        })
    return filas
