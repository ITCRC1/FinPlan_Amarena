# -*- coding: utf-8 -*-
"""Excel de FTE real por departamento — ida y vuelta.

Mismo espíritu que `room_stats_importer.py`: un bloque por mes, un
departamento por fila, columna de mes para que el importador sepa a cuál va
cada bloque. Acá el archivo lo generamos NOSOTROS (no viene de Opera), así
que el formato es el que más simple le resulta al importador: sin ambigüedad
de columnas fijas, sin bloque YTD que descartar.
"""
from __future__ import annotations
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.hotel_actual import HOTEL_ID

MONTHS_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

NAVY = "1F3A6E"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color=NAVY, size=13)
TOTAL_FILL = PatternFill("solid", fgColor="EDEDED")
TOTAL_FONT = Font(bold=True, size=10.5)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INPUT_FILL = PatternFill("solid", fgColor="FFFDE7")  # ya tiene dato, editable
BLANK_FILL = PatternFill("solid", fgColor="FFF3E0")  # todavía sin llenar

COL_DEPT_CODE = 2   # B
COL_DEPT_NAME = 3   # C
COL_FTE = 4         # D
COL_MONTH = 5       # E


def export_dept_fte_template(
    depts: list[tuple[str, str]],
    prellenado: dict[tuple[str, int], float],
    scenario_label: str,
) -> bytes:
    """`depts` = [(dept_code, dept_name), ...]. `prellenado` = {(dept_code,
    month): fte} — lo que ya está guardado en `actual_dept_fte`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FTE Real"

    ws["B1"] = f"FinPlan {HOTEL_ID} — FTE Real por Departamento · {scenario_label}"
    ws["B1"].font = TITLE_FONT
    ws["B2"] = ("Un bloque por mes. Completar D (FTE) para los departamentos y meses donde "
                "el FTE calculado esté en 0 pero el costo sea real. Al subir, cada mes "
                "reemplaza SOLO ese mes en esta versión — no toca los demás.")
    ws["B2"].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 28
    ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")

    row = 4
    for m_idx in range(1, 13):
        mes_nombre = MONTHS_ES[m_idx - 1]
        for j, label in enumerate(["Código", "Departamento", "FTE", "Month"]):
            c = ws.cell(row=row, column=2 + j, value=label)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        row += 1
        first_data_row = row
        for code, name in depts:
            ws.cell(row=row, column=COL_DEPT_CODE, value=code).border = BORDER
            ws.cell(row=row, column=COL_DEPT_NAME, value=name).border = BORDER
            valor = prellenado.get((code, m_idx))
            fte_cell = ws.cell(row=row, column=COL_FTE, value=valor)
            fte_cell.fill = INPUT_FILL if valor is not None else BLANK_FILL
            fte_cell.border = BORDER
            fte_cell.number_format = "#,##0.00"
            mj = ws.cell(row=row, column=COL_MONTH, value=mes_nombre)
            mj.border = BORDER; mj.alignment = Alignment(horizontal="center")
            mj.font = Font(color="999999", size=9)
            row += 1
        last_data_row = row - 1
        ws.cell(row=row, column=COL_DEPT_CODE, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=row, column=COL_DEPT_CODE).fill = TOTAL_FILL
        ws.cell(row=row, column=COL_DEPT_NAME).fill = TOTAL_FILL
        ws.cell(row=row, column=COL_FTE,
                value=f"=SUM(D{first_data_row}:D{last_data_row})")
        ws.cell(row=row, column=COL_FTE).number_format = "#,##0.00"
        for col in (COL_DEPT_CODE, COL_DEPT_NAME, COL_FTE, COL_MONTH):
            cc = ws.cell(row=row, column=col)
            cc.fill = TOTAL_FILL; cc.font = TOTAL_FONT; cc.border = BORDER
        row += 2

    widths = {COL_DEPT_CODE: 10, COL_DEPT_NAME: 32, COL_FTE: 12, COL_MONTH: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
