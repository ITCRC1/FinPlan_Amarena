"""
Cost of Sales (Clase 5) Excel export / import.

Format:
  One sheet per dept_code.
  Columns: A=account_code  B=account_name  C=driver_type  D..O=ENE..DIC  P=TOTAL
  MANUAL rows are editable. DRIVER rows show computed values (import ignores them).
"""
from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app.export.excel_base import (
    C, fill, font, border, align,
    merged_header, month_header_row, set_col_widths, workbook_to_bytes,
    unlock, protect_sheet, add_dropdown, nombre_de_hoja,
    MONTHS_ES, MONTH_ATTRS,
)
from app.hotel_actual import HOTEL_ID

# Nombres de departamento. **NO se escriben acá.**
#
# Los tres exportadores (OPEX, costos, planilla) tenían cada uno su propia lista
# de ~16 departamentos mientras `department_catalog` tiene 38: el 260 y el 270
# —y cualquiera que se agregue— salían con la pestaña rotulada «260 260», el
# código repetido en vez del nombre. Tres copias parciales de la misma verdad.
#
# Ahora el nombre lo pasa quien exporta, tomado del catálogo. Este diccionario
# queda SOLO como respaldo para que un llamado viejo no rompa, y si un código no
# está, la pestaña sale con el código: nunca inventado.
DEPT_NAMES: dict[str, str] = {
    "0110": "Rooms & Housekeeping",
    "0113": "Housekeeping",
    "0120": "Food & Beverage",
    "0140": "Spa",
    "0150": "Tour Activities",
    "0151": "Transport",
    "0152": "Corcovado Bosque",
    "0160": "Laundry",
    "0161": "Laundry (ext)",
    "0180": "Administration",
    "0190": "Sales & Marketing",
    "0191": "Marketing",
    "0200": "Maintenance",
    "0210": "Energy / Utilities",
    "0220": "Cafeteria",
    "0230": "IT Systems",
    "0240": "Property",
}

COL_ACCT   = 1   # A
COL_NAME   = 2   # B
COL_DRIVER = 3   # C
COL_JAN    = 4   # D
COL_TOTAL  = 16  # P
# Al FINAL a proposito: meterla en medio correria las columnas que el
# importador ya asume por posicion.
COL_MONEDA = 17  # Q
HEADER_ROW = 5


def export_costs_to_excel(
    entries_by_dept: dict[str, list[dict]],
    scenario_label: str,
    year: int,
    dept_names: dict[str, str] | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    usados: set[str] = set()
    for dept_code, entries in sorted(entries_by_dept.items()):
        dept_name = (dept_names or {}).get(dept_code) or DEPT_NAMES.get(dept_code, dept_code)
        sheet_name = nombre_de_hoja(f"{dept_code} {dept_name}", usados, sufijo=dept_code)
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: title
        merged_header(ws, 1, COL_ACCT, COL_TOTAL,
                      f"FinPlan {HOTEL_ID} — Cost of Sales (Clase 5) | {scenario_label} {year}",
                      C["teal_dark"], sz=13)
        # Row 2: dept
        merged_header(ws, 2, COL_ACCT, COL_TOTAL,
                      f"{dept_code} — {dept_name}", C["navy_mid"], sz=11)
        # Row 3: note
        inst = ws.cell(row=3, column=COL_ACCT,
                       value="Filas MANUAL son editables. Filas DRIVER se recalculan automáticamente.")
        inst.font = font(italic=True, color=C["text_mid"], size=9)
        ws.merge_cells(start_row=3, start_column=COL_ACCT,
                       end_row=3, end_column=COL_TOTAL)
        # Row 4: spacer
        # Row 5: headers
        # "Modo", no "Driver": la columna trae MANUAL/DRIVER (el modo de la
        # fila), no cuál driver específico la calcula. Decía "Driver" y hacía
        # pensar que ahí estaba el % o la tarifa — no está, esas siguen sin
        # plantilla (B7 residual, docs/PENDIENTES.md).
        for label, col in [("Cuenta", COL_ACCT), ("Nombre", COL_NAME),
                           ("Modo", COL_DRIVER), ("Moneda", COL_MONEDA)]:
            c = ws.cell(row=HEADER_ROW, column=col, value=label)
            c.fill = fill(C["blue_header"])
            c.font = font(bold=True, color=C["navy"], size=10)
            c.alignment = align("center")
            c.border = border()
        month_header_row(ws, HEADER_ROW, COL_JAN, COL_TOTAL)

        row = HEADER_ROW + 1
        for e in entries:
            is_manual = (e.get("driver_type", "MANUAL") == "MANUAL")
            row_bg = C["white"] if is_manual else C["gray_light"]

            ws.cell(row=row, column=COL_ACCT,
                    value=e.get("account_code", "")).border = border()
            nm = ws.cell(row=row, column=COL_NAME, value=e.get("account_name", ""))
            nm.font = font(size=10)
            nm.border = border()
            dr = ws.cell(row=row, column=COL_DRIVER, value=e.get("driver_type", "MANUAL"))
            dr.font = font(size=9, color=C["text_mid"])
            dr.alignment = align("center")
            dr.border = border()

            es_crc = (e.get("currency") or "USD").upper() == "CRC"
            for i, mk in enumerate(MONTH_ATTRS):
                # Una linea en colones exporta COLONES: es lo que el owner edita
                # y lo que tiene que volver al subirlo.
                v = (e.get(f"crc_{mk}", 0) if es_crc else e.get(mk, 0)) or 0
                cell = ws.cell(row=row, column=COL_JAN + i, value=float(v))
                cell.number_format = "#,##0.00;(#,##0.00)"
                cell.alignment = align("right")
                cell.fill = fill(row_bg)
                cell.border = border()
                if is_manual:
                    unlock(cell)   # ← editable only for MANUAL rows
                else:
                    cell.font = font(color=C["text_mid"], size=9)

            # TOTAL
            first = get_column_letter(COL_JAN)
            last  = get_column_letter(COL_JAN + 11)
            tc = ws.cell(row=row, column=COL_TOTAL,
                         value=f"=SUM({first}{row}:{last}{row})")
            tc.number_format = "#,##0.00;(#,##0.00)"
            tc.fill = fill(C["blue_light"])
            tc.font = font(bold=True, color=C["navy"], size=10)
            tc.alignment = align("right")
            tc.border = border()

            # Moneda de la linea. Si dice CRC, los meses de esta fila estan en
            # COLONES: son el dato maestro y el dolar lo deriva el sistema.
            mon = (e.get("currency") or "USD").upper()
            mc = ws.cell(row=row, column=COL_MONEDA, value=mon)
            mc.alignment = align("center")
            mc.font = font(size=9, bold=(mon == "CRC"),
                           color=C["navy"] if mon == "CRC" else C["text_mid"])
            mc.border = border()
            unlock(mc)
            row += 1

        # Grand total
        data_start = HEADER_ROW + 1
        data_end   = row - 1
        gt = ws.cell(row=row, column=COL_ACCT, value="TOTAL COGS")
        gt.fill = fill(C["teal_dark"])
        gt.font = font(bold=True, color="FFFFFF", size=11)
        gt.border = border()
        ws.merge_cells(start_row=row, start_column=COL_ACCT,
                       end_row=row, end_column=COL_DRIVER)

        for i in range(12):
            cl = get_column_letter(COL_JAN + i)
            tc = ws.cell(row=row, column=COL_JAN + i,
                         value=f"=SUM({cl}{data_start}:{cl}{data_end})")
            tc.number_format = "#,##0.00;(#,##0.00)"
            tc.fill = fill(C["teal_dark"])
            tc.font = font(bold=True, color="FFFFFF", size=10)
            tc.alignment = align("right")
            tc.border = border()

        tl = get_column_letter(COL_TOTAL)
        tt = ws.cell(row=row, column=COL_TOTAL,
                     value=f"=SUM({tl}{data_start}:{tl}{data_end})")
        tt.number_format = "#,##0.00;(#,##0.00)"
        tt.fill = fill(C["teal_dark"])
        tt.font = font(bold=True, color="FFFFFF", size=11)
        tt.alignment = align("right")
        tt.border = border()

        protect_sheet(ws)
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=COL_JAN)
        set_col_widths(ws, {
            COL_ACCT: 10, COL_NAME: 28, COL_DRIVER: 14,
            **{COL_JAN + i: 11 for i in range(12)},
            COL_TOTAL: 13,
        })

    return workbook_to_bytes(wb)


def import_costs_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Mismo cuidado que en OPEX: `[:4]` sobre «260 Club Madresal» devolvía
        # «260 » —con el espacio— y no calzaba con ningún departamento.
        m = re.match(r"^(\d{3,4})(?!\d)", sheet_name.strip())
        if not m:
            continue
        dept_code = m.group(1)

        header_row = None
        for r in range(1, 10):
            val = str(ws.cell(row=r, column=COL_ACCT).value or "").upper()
            if "CUENTA" in val or "ENE" in val:
                header_row = r
                break
        if header_row is None:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            acct_raw = ws.cell(row=r, column=COL_ACCT).value
            name_raw = ws.cell(row=r, column=COL_NAME).value
            driver   = str(ws.cell(row=r, column=COL_DRIVER).value or "MANUAL").strip()

            if not acct_raw:
                continue
            acct = str(acct_raw).split(".")[0].strip()
            if not re.match(r"^\d{4,5}$", acct):
                continue
            name = str(name_raw or "").strip()
            if name.upper() in ("TOTAL COGS", "TOTAL", ""):
                continue
            if driver != "MANUAL":
                continue  # skip driver rows on import — they get recalculated

            moneda = str(ws.cell(row=r, column=COL_MONEDA).value or "USD").strip().upper()
            if moneda not in ("USD", "CRC"):
                moneda = "USD"

            row_dict: dict = {
                "currency": moneda,
                "dept_code": dept_code,
                "account_code": acct,
                "account_name": name,
                "driver_type": "MANUAL",
            }
            # En una linea CRC los montos del Excel son COLONES: van a crc_*, que
            # es el dato maestro. Meterlos en las columnas de dolares seria
            # multiplicar el presupuesto por el tipo de cambio.
            destino = (lambda mk: f"crc_{mk}") if moneda == "CRC" else (lambda mk: mk)
            for i, mk in enumerate(MONTH_ATTRS):
                raw = ws.cell(row=r, column=COL_JAN + i).value
                try:
                    v = Decimal(str(raw)) if raw not in (None, "") else Decimal("0")
                except InvalidOperation:
                    v = Decimal("0")
                row_dict[destino(mk)] = v

            results.append(row_dict)

    return results
