"""
Payroll positions Excel export / import.

Format: ONE sheet ("Planilla") for all departments. 'Departamento' is repeated
on every row so the data is pivot-table friendly. A blank row separates depts.

Columns:
    A  = dept_code
    B  = dept_name
    C  = position_code
    D  = position_name
    E  = employee_name
    F  = employee_type
    G  = salary_amount        (EDITABLE)
    H  = salary_currency      (CRC / USD)
    I..T = FTE_ENE..FTE_DIC   (EDITABLE)
    U  = salary_annual_usd    (computed reference, not imported)

Only G (salary) and I-T (FTE) are unlocked; the rest is protected.
Import reads A..T (dept_code from col A), ignores U and blank separator rows.
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from app.hotel_actual import HOTEL_ID

from app.export.excel_base import (
    C, fill, font, border, align,
    merged_header, set_col_widths, workbook_to_bytes,
    unlock, protect_sheet,
    MONTHS_ES,
)

EMPLOYEE_TYPES = ["1-Permanente", "2-Temporal", "3-Estacional", "4-Contrato"]
CURRENCIES     = ["CRC", "USD"]

# Nombres de departamento. **NO se escriben acá.**
#
# Los tres exportadores (OPEX, costos, planilla) tenían cada uno su propia lista
# de ~16 departamentos mientras `department_catalog` tiene 38: el 260 y el 270
# —y cualquiera que se agregue— salían con la pestaña rotulada «260 260», el
# código repetido en vez del nombre. Tres copias parciales de la misma verdad.
#
# Ahora el nombre lo pasa quien exporta, tomado del catálogo. Este diccionario
# queda SOLO como respaldo para que un llamado viejo no rompa, y si un código no
# está, la pestaña sale con el código: nunca inventado.
DEPT_NAMES: dict[str, str] = {
    "0110": "Rooms & Housekeeping",
    "0113": "Housekeeping",
    "0120": "Food & Beverage",
    "0140": "Spa",
    "0150": "Tour Activities",
    "0151": "Transport",
    "0152": "Corcovado Bosque",
    "0160": "Laundry",
    "0161": "Laundry (ext)",
    "0180": "Administration",
    "0190": "Sales & Marketing",
    "0191": "Marketing",
    "0200": "Maintenance",
    "0210": "Energy / Utilities",
    "0220": "Cafeteria",
    "0230": "IT Systems",
    "0240": "Property",
}

FTE_ATTRS = [f"fte_{m}" for m in
             ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]]

# Single sheet for ALL departments. 'Departamento' repeated on every row (pivot-friendly).
COL_DEPT     = 1   # A  dept_code
COL_DEPTNAME = 2   # B  dept_name
COL_PCODE    = 3   # C
COL_PNAME    = 4   # D
COL_ENAME    = 5   # E
COL_ETYPE    = 6   # F
COL_SALARY   = 7   # G  (editable)
COL_CUR      = 8   # H
COL_FTE_JAN  = 9   # I
COL_REF      = 21  # U  (reference-only annual, not imported)
HEADER_ROW   = 5


def export_payroll_to_excel(
    positions_by_dept: dict[str, list[dict]],
    scenario_label: str,
    year: int,
    tc_usd: float = 640.0,
    dept_names: dict[str, str] | None = None,
    tc_por_mes: list[float] | None = None,
) -> bytes:
    """
    ⚠️ `tc_por_mes` (12 valores, enero primero) es lo que hace que «SW Anual
    USD*» diga lo mismo que la pantalla. `tc_usd` queda solo para la nota del
    encabezado y como respaldo si no llega la lista — el modelo dice
    `SW_mes = salary × FTE_mes / TC_mes` (`payroll_position.py`), un TC por
    mes, no uno para el año entero. La fórmula vieja usaba SOLO el de enero
    para los doce meses: con TC móvil, el número no amarraba con la pantalla.
    """
    if not tc_por_mes or len(tc_por_mes) != 12:
        tc_por_mes = [tc_usd] * 12
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla"

    # Title / subtitle / note
    merged_header(ws, 1, COL_DEPT, COL_REF,
                  f"FinPlan {HOTEL_ID} — Planilla de Personal | {scenario_label} {year}",
                  C["navy"], sz=13)
    merged_header(ws, 2, COL_DEPT, COL_REF, "Todos los departamentos", C["navy_mid"], sz=11)
    # El TC ya no es uno solo: "SW Anual USD*" usa el de CADA mes (ver
    # `tc_por_mes`), así que un rango dice más que un promedio que nadie usa
    # para calcular nada.
    tc_min, tc_max = min(tc_por_mes), max(tc_por_mes)
    tc_nota = (f"₡{tc_min:,.0f}/USD" if tc_min == tc_max
              else f"₡{tc_min:,.0f}–{tc_max:,.0f}/USD según el mes")
    note = ws.cell(row=3, column=COL_DEPT,
                   value=(f"TC: {tc_nota}  |  SOLO editar: G (salario) y "
                          "I-T (FTE por mes). El resto está bloqueado. La columna "
                          "'Departamento' se repite en cada fila para tablas dinámicas."))
    note.font = font(italic=True, color=C["text_mid"], size=9)
    ws.merge_cells(start_row=3, start_column=COL_DEPT, end_row=3, end_column=COL_REF)

    # Headers (row 5)
    headers = [
        (COL_DEPT,     "Departamento"),
        (COL_DEPTNAME, "Nombre Depto"),
        (COL_PCODE,    "Cód.Pos."),
        (COL_PNAME,    "Puesto"),
        (COL_ENAME,    "Colaborador"),
        (COL_ETYPE,    "Tipo"),
        (COL_SALARY,   "Salario"),
        (COL_CUR,      "Moneda"),
    ]
    for col, label in headers:
        c = ws.cell(row=HEADER_ROW, column=col, value=label)
        c.fill = fill(C["blue_header"]); c.font = font(bold=True, color=C["navy"], size=10)
        c.alignment = align("center"); c.border = border()
    for i, m in enumerate(MONTHS_ES):
        c = ws.cell(row=HEADER_ROW, column=COL_FTE_JAN + i, value=f"FTE {m}")
        c.fill = fill(C["amber"]); c.font = font(bold=True, color="FFFFFF", size=9)
        c.alignment = align("center"); c.border = border()
    ref = ws.cell(row=HEADER_ROW, column=COL_REF, value="SW Anual USD*")
    ref.fill = fill(C["blue_header"]); ref.font = font(bold=True, color=C["navy"], size=9)
    ref.alignment = align("center"); ref.border = border()

    row = HEADER_ROW + 1
    for dept_code in sorted(positions_by_dept):
        positions = positions_by_dept[dept_code]
        if not positions:
            continue
        # Prefer the real dept_name stored on the position; fall back to the
        # static map, then to the code itself.
        dept_name = (str(positions[0].get("dept_name") or "").strip()
                     or (dept_names or {}).get(dept_code)
                     or DEPT_NAMES.get(dept_code, dept_code))

        for p in positions:
            sal = float(p.get("salary_amount", 0) or 0)
            cur = p.get("salary_currency", "CRC")

            def wc(col, val, **kw):
                c = ws.cell(row=row, column=col, value=val)
                c.border = border()
                for k, v in kw.items():
                    setattr(c, k, v)
                return c

            # locked identity columns (repeated dept on every row)
            wc(COL_DEPT, dept_code, font=font(size=9, color=C["text_mid"]), alignment=align("center"))
            wc(COL_DEPTNAME, dept_name, font=font(size=9, color=C["text_mid"]))
            wc(COL_PCODE, p.get("position_code", ""),
               font=font(size=9, color=C["text_mid"]), alignment=align("center"))
            wc(COL_PNAME, p.get("position_name", ""), font=font(size=10))
            wc(COL_ENAME, p.get("employee_name", "VACANTE"), font=font(size=10, color=C["text_mid"]))
            wc(COL_ETYPE, p.get("employee_type", "1-Permanente"),
               font=font(size=9, color=C["text_mid"]), alignment=align("center"))

            # editable: salary (G)
            sal_c = wc(COL_SALARY, sal)
            sal_c.number_format = "#,##0;(#,##0)"; sal_c.alignment = align("right")
            sal_c.fill = fill(C["amber_light"]); sal_c.font = font(bold=True, color=C["amber"], size=10)
            unlock(sal_c)

            # locked: currency (H)
            wc(COL_CUR, cur, alignment=align("center"), font=font(size=9, color=C["text_mid"]))

            # editable: FTE (I-T)
            for i, attr in enumerate(FTE_ATTRS):
                fte_val = float(p.get(attr, 0) or 0)
                fc = ws.cell(row=row, column=COL_FTE_JAN + i, value=fte_val)
                fc.number_format = "0.00"; fc.alignment = align("center")
                fc.fill = fill(C["amber_light"])
                # Mismo criterio que la pantalla (payroll/fte/page.tsx): 0 es
                # gris —una posición sin persona ese mes, no un error—, rojo
                # es solo 0<FTE<0.5 (media jornada rara), ámbar es 0.5–1 y
                # negro-negrita es 1 o más. Antes el exportador pintaba TODO
                # cero de rojo, que se leía como alarma donde la pantalla no
                # muestra ninguna.
                if fte_val == 0:
                    fte_color, fte_bold = C["text_mid"], False
                elif fte_val < 0.5:
                    fte_color, fte_bold = "C0392B", False
                elif fte_val < 1:
                    fte_color, fte_bold = C["amber"], False
                else:
                    fte_color, fte_bold = C["text_dark"], True
                fc.font = font(bold=fte_bold, color=fte_color, size=10)
                fc.border = border()
                unlock(fc)

            # reference: annual SW in USD — mismo criterio que `calc_sw`
            # (payroll_calculator.py): SW_mes = salario × FTE_mes / TC_mes, y
            # el anual es la SUMA de los doce, cada uno con SU TC. Antes esto
            # dividía por 12 (de más: `salary_amount` YA es mensual, no
            # necesita anualizarse) y usaba el TC de un solo mes para los
            # doce — con TC móvil, el resultado no se parecía al de la
            # pantalla. El SUMPRODUCT reemplaza esa suma mes a mes.
            fs = get_column_letter(COL_FTE_JAN); fe = get_column_letter(COL_FTE_JAN + 11)
            sl = get_column_letter(COL_SALARY);  cl = get_column_letter(COL_CUR)
            coef_1_tc = ",".join(f"{(1 / t if t else 0):.10f}" for t in tc_por_mes)
            formula = (
                f'=IF({cl}{row}="CRC",'
                f'SUMPRODUCT({fs}{row}:{fe}{row},{{{coef_1_tc}}})*{sl}{row},'
                f'SUM({fs}{row}:{fe}{row})*{sl}{row})'
            )
            rc = ws.cell(row=row, column=COL_REF, value=formula)
            rc.number_format = "#,##0;(#,##0)"; rc.fill = fill(C["blue_light"])
            rc.font = font(bold=True, color=C["navy"], size=10); rc.alignment = align("right")
            rc.border = border()

            row += 1

        row += 1  # blank separator row between departments

    protect_sheet(ws)
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=COL_FTE_JAN)
    set_col_widths(ws, {
        COL_DEPT: 12, COL_DEPTNAME: 22, COL_PCODE: 8, COL_PNAME: 30, COL_ENAME: 22,
        COL_ETYPE: 12, COL_SALARY: 13, COL_CUR: 8,
        **{COL_FTE_JAN + i: 8 for i in range(12)},
        COL_REF: 14,
    })
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    return workbook_to_bytes(wb)


def import_payroll_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Locate header row (the one whose Puesto column says PUESTO/POSICION)
    header_row = None
    for r in range(1, 12):
        val = str(ws.cell(row=r, column=COL_PNAME).value or "").upper()
        if "PUESTO" in val or "POSICION" in val or "POSICIÓN" in val:
            header_row = r
            break
    if header_row is None:
        return []

    results: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        pname = str(ws.cell(row=r, column=COL_PNAME).value or "").strip()
        if not pname or pname.upper() in ("TOTAL DEPT", "TOTAL"):
            continue  # blank separator or total row
        dept_code = str(ws.cell(row=r, column=COL_DEPT).value or "").strip()
        if not dept_code:
            continue

        sal_raw = ws.cell(row=r, column=COL_SALARY).value
        try:
            salary = Decimal(str(sal_raw)) if sal_raw not in (None, "") else Decimal("0")
        except InvalidOperation:
            salary = Decimal("0")

        cur = str(ws.cell(row=r, column=COL_CUR).value or "CRC").strip().upper()
        if cur not in ("CRC", "USD"):
            cur = "CRC"

        dept_name = str(ws.cell(row=r, column=COL_DEPTNAME).value or "").strip() \
            or DEPT_NAMES.get(dept_code, "")

        row_dict: dict = {
            "dept_code": dept_code,
            "dept_name": dept_name,
            "position_code": str(ws.cell(row=r, column=COL_PCODE).value or "").strip(),
            "position_name": pname,
            "employee_name": str(ws.cell(row=r, column=COL_ENAME).value or "VACANTE").strip(),
            "employee_type": str(ws.cell(row=r, column=COL_ETYPE).value or "1-Permanente").strip(),
            "salary_amount": salary,
            "salary_currency": cur,
        }
        for i, attr in enumerate(FTE_ATTRS):
            raw = ws.cell(row=r, column=COL_FTE_JAN + i).value
            try:
                row_dict[attr] = Decimal(str(raw)) if raw not in (None, "") else Decimal("0")
            except InvalidOperation:
                row_dict[attr] = Decimal("0")

        results.append(row_dict)

    return results
