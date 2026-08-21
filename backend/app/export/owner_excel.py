"""
Owner Report Excel export.

Single sheet "Reporte Propietarios" that mirrors the on-screen owner report:
  - Title block (hotel + scenario)
  - Room KPIs (Ocupación, ADR, RevPAR, Total Pax)
  - P&L summary (Revenue → GOP → EBITDA → Net Profit)
  - "Explicación de variaciones" narrative grouped by section

The frontend computes the displayed numbers (KPIs come from drivers, which the
P&L does not carry in checkbook mode) and POSTs them here, so the spreadsheet
matches the screen exactly.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.hotel_actual import HOTEL_NAME

from app.export.excel_base import (
    C, fill, font, border, align,
    merged_header, set_col_widths, workbook_to_bytes,
)

# Layout: A = label, B..C = value (merged for narrative readability)
COL_A = 1
COL_B = 2
COL_C = 3
LAST_COL = 4   # D — keeps the title band a comfortable width

USD0 = '"$"#,##0;("$"#,##0);"—"'
PCT1 = '0.0"%"'
NUM0 = "#,##0"


def export_owner_report_to_excel(
    scenario_label: str,
    kpis: dict,
    pl_rows: list[dict],
    notes: list[dict],
) -> bytes:
    """
    kpis:    {"occ_pct", "adr", "revpar", "pax"}
    pl_rows: [{"label", "value", "strong"}]  strong → bold (totals/expense tag)
    notes:   [{"section", "ref", "month_name", "body"}]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Propietarios"

    # ── Title band ──────────────────────────────────────────────────────────
    merged_header(ws, 1, COL_A, LAST_COL,
                  f"{HOTEL_NAME} — {scenario_label}",
                  C["navy"], sz=14)
    merged_header(ws, 2, COL_A, LAST_COL,
                  "Resumen ejecutivo del presupuesto",
                  C["navy_mid"], sz=10)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    row = 4

    # ── KPIs ────────────────────────────────────────────────────────────────
    _section_title(ws, row, "INDICADORES")
    row += 1
    kpi_rows = [
        ("Ocupación", kpis.get("occ_pct", 0), PCT1),
        ("ADR", kpis.get("adr", 0), USD0),
        ("RevPAR", kpis.get("revpar", 0), USD0),
        ("Total Pax", kpis.get("pax", 0), NUM0),
    ]
    for label, value, fmt in kpi_rows:
        lc = ws.cell(row=row, column=COL_A, value=label)
        lc.font = font(color=C["text_mid"], size=10)
        lc.border = border()
        vc = ws.cell(row=row, column=COL_B, value=float(value or 0))
        vc.number_format = fmt
        vc.font = font(bold=True, color=C["navy"], size=11)
        vc.alignment = align("right")
        vc.border = border()
        ws.merge_cells(start_row=row, start_column=COL_B, end_row=row, end_column=COL_C)
        row += 1

    row += 1

    # ── P&L summary ─────────────────────────────────────────────────────────
    _section_title(ws, row, "RESUMEN DE RESULTADOS (USD)")
    row += 1
    totals = {"Total Revenue", "GOP", "EBITDA", "Net Profit"}
    for r in pl_rows:
        label = r["label"]
        value = float(r.get("value", 0) or 0)
        is_total = label in totals
        lc = ws.cell(row=row, column=COL_A, value=label)
        lc.font = font(bold=is_total, color=C["text_dark"] if not r.get("strong") else C["text_mid"], size=10)
        lc.border = border(sides="top_bottom") if is_total else border()
        # El valor va CON su signo. Antes se escribía el valor absoluto y se le
        # ponía un formato con paréntesis literales para que *pareciera*
        # negativo: la celda de un gasto de $50,000 tenía +50000, así que un
        # SUM() sobre la columna sumaba los gastos en vez de restarlos. El
        # formato de tres secciones ya muestra los negativos entre paréntesis
        # sin tocar el número.
        vc = ws.cell(row=row, column=COL_B, value=value)
        vc.number_format = USD0
        vc.font = font(bold=is_total, color=C["navy"] if is_total else C["text_dark"], size=10)
        vc.alignment = align("right")
        vc.border = border(sides="top_bottom") if is_total else border()
        ws.merge_cells(start_row=row, start_column=COL_B, end_row=row, end_column=COL_C)
        if is_total:
            for col in (COL_A, COL_B):
                ws.cell(row=row, column=col).fill = fill(C["blue_header"])
        row += 1

    row += 1

    # ── Narrative ───────────────────────────────────────────────────────────
    _section_title(ws, row, "EXPLICACIÓN DE VARIACIONES")
    row += 1
    if not notes:
        nc = ws.cell(row=row, column=COL_A, value="Sin comentarios cargados todavía.")
        nc.font = font(italic=True, color=C["text_mid"], size=10)
        ws.merge_cells(start_row=row, start_column=COL_A, end_row=row, end_column=LAST_COL)
        row += 1
    else:
        # group by section, preserving first-seen order
        grouped: dict[str, list[dict]] = {}
        for n in notes:
            grouped.setdefault(n.get("section", ""), []).append(n)
        for section, items in grouped.items():
            sc = ws.cell(row=row, column=COL_A, value=section or "General")
            sc.font = font(bold=True, color=C["navy_mid"], size=11)
            ws.merge_cells(start_row=row, start_column=COL_A, end_row=row, end_column=LAST_COL)
            row += 1
            for n in items:
                prefix_parts = [p for p in (n.get("ref"), n.get("month_name")) if p]
                prefix = (" · ".join(prefix_parts) + ": ") if prefix_parts else ""
                bc = ws.cell(row=row, column=COL_A, value=f"• {prefix}{n.get('body', '')}")
                bc.font = font(color=C["text_dark"], size=10)
                bc.alignment = align("left", "top", wrap=True)
                ws.merge_cells(start_row=row, start_column=COL_A, end_row=row, end_column=LAST_COL)
                ws.row_dimensions[row].height = 30
                row += 1
            row += 1

    set_col_widths(ws, {COL_A: 34, COL_B: 16, COL_C: 16, LAST_COL: 16})
    ws.sheet_view.showGridLines = False
    return workbook_to_bytes(wb)


def _section_title(ws, row: int, text: str) -> None:
    merged_header(ws, row, COL_A, LAST_COL, text, C["navy_mid"], sz=11)
