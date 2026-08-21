# -*- coding: utf-8 -*-
"""Exporta el archivo ÚNICO de Detalle (mismo formato que se sube) desde los datos
del sistema: una hoja 'Detalle' con todas las cuentas (clases 4-8) + estadísticas
(clase 9), ordenadas por depto y clase. El owner lo baja, edita solo las filas que
necesita y lo vuelve a subir — round-trip controlado."""
import io
import json
import pathlib
from functools import lru_cache

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.engine.pl_engine import group_for_dept
from app.hotel_actual import HOTEL_ID

MES = ["January", "February", "March", "April", "May", "June", "July", "August",
       "September", "October", "November", "December"]
CLASE_BY_PREFIX = {"4": "Revenue", "5": "Cost", "6": "Payroll", "7": "Opex",
                   "8": "BelowGOP", "9": "Stat"}
#: El orden de las SECCIONES de la hoja. Owner (2026-08-14): «quiero que diga
#: ingreso, payroll, cost and opex». Cost iba antes que Payroll; se invirtio
#: porque asi lee el owner sus estructuras.
CLASE_ORD = {"Revenue": 0, "Payroll": 1, "Cost": 2, "Opex": 3,
             "Distribucion": 4, "BelowGOP": 5, "Stat": 9}

#: Como se titula cada seccion en la hoja.
CLASE_TITULO = {"Revenue": "INGRESO", "Payroll": "PLANILLA", "Cost": "COSTO",
                "Opex": "GASTO OPERATIVO (OPEX)", "BelowGOP": "BELOW GOP",
                # Las contrapartidas de allocation son clase 4 pero NO son
                # ingreso: en el archivo del owner van en su propio bloque, al
                # final del departamento. Sin titulo propio salian rotuladas
                # «INGRESO» una segunda vez y parecia un error de la hoja.
                "Distribucion": "DISTRIBUCIÓN"}
CANON = ["ROOMS", "FB", "PRIVATE_BAR", "SPA", "TOURS", "RETAIL", "TRANSPORT",
         "LAUNDRY", "LAUNDRY_OPS", "INNOCEANA", "CROWTHER", "CLUB",
         "SUSTAINABILITY", "MISC_OTHER",
         "ADMIN", "SALES", "MAINTENANCE", "IT", "UTILITIES", "CAFETERIA",
         "AREC", "OTHER_OVERHEAD"]  # AREC = centro de costo (overhead), ver pl_engine
# PRIVATE_BAR va pegado a FB, no al final: un grupo que no esté en CANON cae al
# fondo de la hoja (grank = len(CANON)), detrás del overhead. El Private Bar es
# centro de utilidad propio (dept 0121) y se lee junto a A&B, no de último.
VERSION_ROW, MONTH_ROW = 15, 14
STAT_ROWS = [("9010", "Rooms disponibles", "rooms_available"),
             ("9020", "Rooms ocupadas", "rooms_occupied"),
             ("9060", "Huéspedes", "guests")]




#: El orden que mando el owner (`ORDEN PARA EL UPLOAD.xlsx`, 2026-08-14),
#: convertido por `scripts/generar_orden_plantilla.py`.
#:
#: **Es una LISTA y no una regla.** El orden de las clases cambia segun el
#: departamento: Rooms no tiene costo; en Tours el costo va DESPUES de planilla.
#: No hay regla que produzca eso, y tres intentos de inventarla fallaron.
_ORDEN_JSON = pathlib.Path(__file__).resolve().parents[1] / "seed_data" / "orden_plantilla.json"


@lru_cache(maxsize=1)
def orden_canonico() -> dict:
    """{(dept_code, cuenta): posicion}. Vacio si el archivo no esta."""
    try:
        datos = json.loads(_ORDEN_JSON.read_text(encoding="utf-8"))["orden"]
    except Exception:
        # Sin el archivo la plantilla sigue saliendo, con el orden por defecto.
        # Que falte no puede tumbar una descarga.
        return {}
    return {(f["dept_code"], str(f["cuenta"])): i for i, f in enumerate(datos)}


@lru_cache(maxsize=1)
def orden_departamentos() -> dict:
    """{dept_code: posicion}, en el orden en que el owner los lista."""
    vistos: dict = {}
    for (dept, _cta), i in sorted(orden_canonico().items(), key=lambda kv: kv[1]):
        vistos.setdefault(dept, len(vistos))
    return vistos



@lru_cache(maxsize=1)
def orden_clase_en_depto() -> dict:
    """{(dept_code, clase): posicion} — donde empieza esa clase en ese depto.

    Sirve para las cuentas que el owner NO listo: sin esto se iban todas al final
    del departamento y las secciones se repetian (A&B mostraba INGRESO, COSTO,
    PLANILLA, OPEX y otra vez INGRESO). Con esto caen dentro de su seccion.
    """
    fuera: dict = {}
    for (dept, cta), i in sorted(orden_canonico().items(), key=lambda kv: kv[1]):
        clase = CLASE_BY_PREFIX.get(str(cta)[:1], "?")
        fuera.setdefault((dept, clase), i)
    return fuera


def rotulo_depto(dept_code: str, dept_names: dict) -> str:
    """Siempre «codigo · nombre». Nunca uno de los dos solo.

    Owner (2026-08-14), viendo la plantilla: «por que Tienda no tiene numero de
    departamento». La columna mostraba el NOMBRE cuando el depto estaba en el
    catalogo y el CODIGO PELADO cuando no — asi que `0240` salia como numero y
    `Tienda` como texto, en la misma columna. Leyendo la hoja no habia forma de
    saber si «0240» era un codigo sin nombre o un depto que se llama asi.

    El codigo es lo canonico —es lo que liga el dato entre escenarios y
    propiedades— y el nombre es la etiqueta. Se muestran los dos, y el que no
    tiene nombre se ve que NO LO TIENE en vez de disfrazarse de nombre.
    """
    nombre = (dept_names or {}).get(dept_code)
    if not nombre:
        return f"{dept_code} · (sin nombre en el catalogo)"
    return f"{dept_code} · {nombre}"


def _escribir_verificacion(det, bmc: dict, verificacion: dict, nmeta: int) -> None:
    """El bloque de control, ARRIBA — filas 1 a 12.

    Owner (2026-08-16): «que el upload tenga la verificación arriba versus el
    detalle abajo […] así el sistema consolida el detalle y valida que estos
    resultados hagan match».

    Entra justo en el hueco que ya existía sobre el encabezado del Detalle: la
    fila 13 es el título, la 14 los meses y la 15 las versiones, y esas dos
    últimas están cableadas en el parser y en todos los archivos que el owner ya
    tiene. Por eso son ONCE controles y no doce — ver `CONTROLES`.

    La fila 1 hace doble trabajo: título a la izquierda y nombre del mes sobre
    cada columna, para que el bloque se lea sin bajar hasta la fila 14.

    ⚠️ Lo que se escribe acá es un CONTROL. Al volver a subir el archivo se
    compara y se descarta: no suma en ningún total ni se guarda en ninguna
    tabla. Una segunda fuente de plata es exactamente el problema que este
    bloque existe para cerrar.
    """
    from app.importers.verificacion import CONTROLES

    TITULO = Font(bold=True, color="16402A", size=11)
    HDRV = PatternFill("solid", fgColor="FFF3CD")
    BLOQ = PatternFill("solid", fgColor="F6E3E3")
    det.cell(1, 1, "VERIFICACIÓN — totales de control").font = TITULO
    det.cell(1, 2, "Bloquea").font = Font(bold=True, size=9)
    det.cell(1, 3, "Sección").font = Font(bold=True, size=9)
    det.cell(1, 4, "Código").font = Font(bold=True, size=9)
    det.cell(1, 5, "Concepto").font = Font(bold=True, size=9)
    for (blk, m), c in bmc.items():
        det.cell(1, c, MES[m - 1]).font = Font(color="6B7A70", size=9)

    r = 2
    for ctrl in CONTROLES:
        det.cell(r, 1, "VERIF")
        det.cell(r, 2, "SÍ" if ctrl.bloquea else "aviso")
        det.cell(r, 3, ctrl.seccion)
        det.cell(r, 4, ctrl.codigo)
        det.cell(r, 5, ctrl.etiqueta)
        for j in range(1, nmeta + 1):
            det.cell(r, j).fill = BLOQ if ctrl.bloquea else HDRV
            if ctrl.bloquea:
                det.cell(r, j).font = Font(bold=True, size=10)
        for blk, porcod in (verificacion or {}).items():
            for m, v in (porcod.get(ctrl.codigo) or {}).items():
                c = bmc.get((blk, int(m)))
                if c:
                    cc = det.cell(r, c, round(float(v), 2))
                    cc.number_format = "#,##0"
        r += 1


def build_detail_workbook(dest_labels: list[str], accts: list[dict],
                          stats: dict[tuple, dict], dept_names: dict[str, str],
                          verificacion: dict | None = None) -> bytes:
    """dest_labels: rótulos de bloque en orden. accts: [{clase,grupo,dept_code,cuenta,
    nombre, vals:{(label,mes):monto}}]. stats: {(label,mes):{campo:valor}}.
    verificacion: {label: {codigo_control: {mes: monto}}} — el bloque de control
    de arriba. Si viene vacío el bloque sale igual, pero en blanco: la plantilla
    tiene que ofrecer dónde escribirlo aunque el sistema todavía no tenga el
    número (que es el caso de una propiedad que arranca desde cero)."""
    wb = Workbook(); det = wb.active; det.title = "Detalle"
    HDR = PatternFill("solid", fgColor="16402A"); WHITE = Font(bold=True, color="FFFFFF", size=10)
    STATF = PatternFill("solid", fgColor="FFF3CD"); DEPTF = PatternFill("solid", fgColor="EAF1EC")
    GREY = Font(color="6B7A70", size=10)
    # La 2a columna decia «GrupoPL» y mostraba `group_for_dept`, que para un
    # departamento sin grupo cae por descarte a OTHER_OVERHEAD. Owner
    # (2026-08-14) sobre Misceláneos: un departamento de INGRESO rotulado como
    # overhead. No afectaba el P&L —que rutea por el mapeo, no por el grupo—
    # pero decia algo falso, y ahora se ve la linea de verdad.
    META = ["Clase", "Línea del P&L", "Departamento", "Cuenta", "Nombre de cuenta"]
    NMETA = len(META)
    for j, t in enumerate(META, 1):
        c = det.cell(VERSION_ROW, j, t); c.fill = HDR; c.font = WHITE
    bmc = {}; col = NMETA + 1
    for blk in dest_labels:
        for m in range(1, 13):
            det.cell(MONTH_ROW, col, MES[m - 1]).font = GREY
            c = det.cell(VERSION_ROW, col, blk); c.fill = HDR; c.font = WHITE
            c.alignment = Alignment(horizontal="center")
            bmc[(blk, m)] = col; col += 1
    _escribir_verificacion(det, bmc, verificacion or {}, NMETA)
    det.cell(13, 1, f"{HOTEL_ID} · DETALLE ÚNICO — bajar, editar solo lo necesario y volver a subir").font = Font(bold=True, color="16402A", size=11)
    det.cell(13, 3, "Arriba van los totales de control; el sistema consolida esto y los compara.").font = Font(color="6B7A70", size=9)

    r = VERSION_ROW + 1
    # 1) estadísticas (clase 9)
    for code, label, field in STAT_ROWS:
        det.cell(r, 1, "Stat"); det.cell(r, 2, "KPI"); det.cell(r, 3, "Estadísticas")
        det.cell(r, 4, code); det.cell(r, 5, label)
        for blk in dest_labels:
            for m in range(1, 13):
                v = stats.get((blk, m), {}).get(field)
                if v is not None:
                    cc = det.cell(r, bmc[(blk, m)], round(float(v), 2)); cc.number_format = "#,##0"
        for j in range(1, NMETA + 1):
            det.cell(r, j).fill = STATF
        r += 1
    r += 1

    # 2) Las cuentas.
    #
    # **El orden es el del archivo que subio el owner** (`orden_archivo`, mig
    # 111). Antes se ordenaba por grupo del P&L y clase: determinista y estable
    # entre descargas, pero distinto del archivo historico del owner, que tenia
    # que cruzar dos listas cada vez que comparaba.
    #
    # Las filas SIN orden —las que no vinieron de un archivo, o son anteriores a
    # la migracion— van al final con el orden de siempre. Asi un escenario que
    # nunca se importo se ve exactamente como antes.
    def grank(g): return CANON.index(g) if g in CANON else len(CANON)
    # DEPARTAMENTO afuera, clase adentro — y las dos cosas EN EL ORDEN DEL OWNER.
    #
    # Owner (2026-08-14): «solo acomoda el archivo de upload en este formato.
    # Mismo orden de los departamentos, mismo orden interno. Para Rooms sale
    # ingreso y sus cuentas, costo y sus cuentas, planilla y sus cuentas y opex y
    # sus cuentas. Y despues sigues con F&B, Spa, y asi hasta terminar.»
    #
    # ⚠️ **El orden de las clases NO es el mismo en todos los departamentos**:
    # Rooms no tiene costo, y en Tours el costo va DESPUES de planilla. No hay
    # regla que produzca eso — hay un archivo, y esta en `orden_plantilla.json`.
    # Tres intentos anteriores fallaron buscando la regla.
    #
    # Lo que no este en esa lista sale detras de su departamento, por clase y
    # cuenta. Y un departamento que el owner no listo va al final, para que una
    # cuenta nueva nunca desaparezca de la hoja.
    GRANDE = 10 ** 9
    pos_cuenta = orden_canonico()
    pos_dept = orden_departamentos()
    pos_clase = orden_clase_en_depto()

    def clave(x):
        """(seccion del depto, posicion dentro de la seccion).

        ⚠️ La SECCION depende solo de (departamento, clase) — nunca de la cuenta.
        Esa es la garantia de que cada clase salga en UN solo bloque. Cuando la
        seccion salia de la posicion de la CUENTA, las cuentas listadas y las no
        listadas de la misma clase caian en dos bloques distintos y el
        departamento mostraba «OPEX … DISTRIBUCION … OPEX» otra vez.
        """
        d, cta, clase = x["dept_code"], str(x["cuenta"]), x["clase"]
        sec = pos_clase.get((d, clase))
        if sec is None:
            # El owner no listo esta clase para este departamento: se INTERCALA
            # entre las que si listo, segun el orden estandar. (Owner viendo
            # Utilities: «no se que paso aca» — su archivo lista solo Opex ahi,
            # asi que la planilla quedaba al final.)
            mio = CLASE_ORD.get(clase, 9)
            despues = [p for (dd, cl), p in pos_clase.items()
                       if dd == d and CLASE_ORD.get(cl, 9) > mio]
            antes = [p for (dd, cl), p in pos_clase.items()
                     if dd == d and CLASE_ORD.get(cl, 9) < mio]
            if despues:
                sec = min(despues) - 0.5
            elif antes:
                sec = max(antes) + 0.5
            else:
                sec = GRANDE + mio          # el depto entero no esta listado

        p = pos_cuenta.get((d, cta))
        if p is not None:
            dentro = (0, p)                 # tal cual lo puso el owner
        else:
            # Detras de las listadas de su clase, y entre ellas manda el orden
            # del archivo que se subio, si lo hay.
            dentro = (1, x["orden"] if x.get("orden") is not None else GRANDE, cta)

        # El grupo del P&L solo ubica un departamento que el owner no listo: un
        # depto puede tener dos grupos (Lavanderia: LAUNDRY y LAUNDRY_OPS) y
        # ordenar por grupo lo partiria en dos.
        return (pos_dept.get(d, GRANDE + grank(x["grupo"])), d, sec, dentro)

    prev_clase = None
    prev_dept = None
    for a in sorted(accts, key=clave):

        # El DEPARTAMENTO manda: encabezado fuerte y reinicio de las clases.
        if a["dept_code"] != prev_dept:
            det.cell(r, 1, "")
            det.cell(r, 3, f"── {rotulo_depto(a['dept_code'], dept_names)} ──")
            for j in range(1, NMETA + 1):
                det.cell(r, j).fill = HDR
                det.cell(r, j).font = WHITE
            r += 1
            prev_dept = a["dept_code"]
            prev_clase = None      # cada depto vuelve a anunciar sus clases

        # Y adentro, el subtitulo de la clase: ingreso, costo, planilla, opex.
        if a["clase"] != prev_clase:
            det.cell(r, 1, CLASE_TITULO.get(a["clase"], a["clase"]))
            for j in range(1, NMETA + 1):
                det.cell(r, j).fill = DEPTF
                det.cell(r, j).font = Font(bold=True)
            r += 1
            prev_clase = a["clase"]

        det.cell(r, 1, a["clase"])
        det.cell(r, 2, a.get("linea_pl") or a["grupo"])
        det.cell(r, 3, rotulo_depto(a["dept_code"], dept_names))
        det.cell(r, 4, a["cuenta"]); det.cell(r, 5, a["nombre"])
        for (blk, m), v in a["vals"].items():
            cc = det.cell(r, bmc[(blk, m)], round(float(v), 2)); cc.number_format = "#,##0"
        r += 1
    for cl, w in [("A", 9), ("B", 13), ("C", 30), ("D", 9), ("E", 32)]:
        det.column_dimensions[cl].width = w
    for c in range(NMETA + 1, col):
        det.column_dimensions[get_column_letter(c)].width = 11
    det.freeze_panes = det.cell(VERSION_ROW + 1, NMETA + 1)

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
