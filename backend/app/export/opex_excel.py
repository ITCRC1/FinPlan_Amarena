"""
OPEX Excel export / import — Fase 10.

Export: genera un .xlsx con una hoja por departamento.
        Filas de detalle editables + subtotales por cuenta + grand total.
Import: lee ese mismo formato y devuelve lista de dicts listos para upsert.

Formato de columnas (fijo):
    A  = account_code   (4 dígitos, ej. "7065")
    B  = account_name
    C  = detail_desc
    D  = ENE  … O = DIC   (12 meses)
    P  = TOTAL  (fórmula SUM D:O en export; ignorada en import)
"""
from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from app.export.excel_base import nombre_de_hoja, protect_sheet, unlock
from app.hotel_actual import HOTEL_ID

# ── Paleta ───────────────────────────────────────────────────────────────────

_C = {
    "navy":        "1A3A5C",
    "navy_mid":    "2D5A9E",
    "blue_light":  "EBF3FB",
    "blue_header": "F0F4F8",
    "white":       "FFFFFF",
    "text_dark":   "1A1A2E",
    "text_mid":    "4A5568",
    "gray_light":  "F8F9FA",
    "border":      "CBD5E0",
    "green_dark":  "1A5C3A",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="1A1A2E", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _border(color="CBD5E0", sides="all") -> Border:
    s = Side(style="thin", color=color)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    if sides == "top":
        return Border(top=Side(style="medium", color=color))
    if sides == "top_bottom":
        return Border(top=s, bottom=s)
    return Border()

_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

MONTHS_ES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
             "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
MONTH_ATTRS = ["jan", "feb", "mar", "apr", "may", "jun",
               "jul", "aug", "sep", "oct", "nov", "dec"]

# Col letters: A=acct, B=name, C=detail, D-O=months, P=total
COL_ACCT   = 1   # A
COL_NAME   = 2   # B
COL_DETAIL = 3   # C
COL_JAN    = 4   # D
COL_DEC    = 15  # O
COL_TOTAL  = 16  # P
# Al FINAL a proposito: en medio correria las columnas que el importador
# ya asume por posicion — y correria tambien los archivos que la gente ya
# tenga bajados, que se suben dias despues.
COL_MONEDA = 17  # Q
# El «#» de la pantalla. Sin el, el viaje de ida y vuelta lo renumeraba: la
# importacion asignaba un correlativo nuevo por posicion, asi que reordenar el
# Excel le cambiaba el codigo a una fila. Lo lee `len(row) >= COL_NUM`, igual
# que la moneda, para que un archivo viejo —que no lo trae— siga subiendo.
COL_NUM    = 18  # R — detail_code

# Row where column headers live (1-based)
HEADER_ROW = 5

# Dept display names
# Nombres de departamento. **NO se escriben acá.**
#
# Los tres exportadores (OPEX, costos, planilla) tenían cada uno su propia lista
# de ~16 departamentos mientras `department_catalog` tiene 38: el 260 y el 270
# —y cualquiera que se agregue— salían con la pestaña rotulada «260 260», el
# código repetido en vez del nombre. Tres copias parciales de la misma verdad.
#
# Ahora el nombre lo pasa quien exporta, tomado del catálogo. Este diccionario
# queda SOLO como respaldo para que un llamado viejo no rompa, y si un código no
# está, la pestaña sale con el código: nunca inventado.
DEPT_NAMES: dict[str, str] = {
    "0110": "Rooms & Housekeeping",
    "0120": "Food & Beverage",
    "0140": "Spa",
    "0150": "Tours",
    "0151": "Tienda",
    "0152": "Transportation",
    "0155": "Innoceanna",
    "0156": "Crowler Lab",
    "0161": "Laundry",
    "0180": "Administration",
    "0190": "Sales & Marketing",
    "0200": "Maintenance",
    "0205": "Claro del Bosque",
    "0210": "Utilities",
    "0220": "Employee Dining",
    "0230": "IT",
}


# ── Export ────────────────────────────────────────────────────────────────────

def export_opex_to_excel(
    entries_by_dept: dict[str, list[dict]],
    scenario_label: str,
    year: int,
    dept_names: dict[str, str] | None = None,
) -> bytes:
    """
    entries_by_dept: {dept_code: [entry_dict, ...]}
      entry_dict keys: account_code, account_name, detail_desc, jan..dec

    Returns raw bytes of the .xlsx file.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    usados: set[str] = set()
    for dept_code in sorted(entries_by_dept):
        entries = entries_by_dept[dept_code]
        dept_name = (dept_names or {}).get(dept_code) or DEPT_NAMES.get(dept_code, dept_code)
        sheet_name = nombre_de_hoja(f"{dept_code} {dept_name}", usados, sufijo=dept_code)
        ws = wb.create_sheet(title=sheet_name)
        _write_dept_sheet(ws, dept_code, dept_name, entries, scenario_label, year)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_dept_sheet(
    ws,
    dept_code: str,
    dept_name: str,
    entries: list[dict],
    scenario_label: str,
    year: int,
) -> None:
    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10   # account
    ws.column_dimensions["B"].width = 28   # account name
    ws.column_dimensions["C"].width = 32   # detail
    for col in range(COL_JAN, COL_DEC + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 13

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(COL_TOTAL)}1")
    c = ws.cell(1, 1, f"FinPlan {HOTEL_ID} — OPEX {year}  ·  {scenario_label}")
    c.font = _font(bold=True, color=_C["white"], size=13)
    c.fill = _fill(_C["navy"])
    c.alignment = _CENTER

    # ── Row 2: Dept name ──────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    ws.merge_cells(f"A2:{get_column_letter(COL_TOTAL)}2")
    c = ws.cell(2, 1, f"{dept_name}  ({dept_code})")
    c.font = _font(bold=True, color=_C["white"], size=11)
    c.fill = _fill(_C["navy_mid"])
    c.alignment = _CENTER

    # ── Row 3: subtitle ───────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"A3:{get_column_letter(COL_TOTAL)}3")
    c = ws.cell(3, 1, "Importar: no modificar columnas A-C · solo editar celdas de meses (D-O)")
    c.font = _font(italic=True, color=_C["text_mid"], size=8)
    c.fill = _fill(_C["blue_header"])
    c.alignment = _CENTER

    # ── Row 4: empty spacer ───────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6
    for col in range(1, COL_TOTAL + 1):
        ws.cell(4, col).fill = _fill(_C["white"])

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[5].height = 20
    headers = ["Cuenta", "Nombre de Cuenta", "Descripción / Detalle"] + MONTHS_ES + ["TOTAL"]
    aligns = [_CENTER, _LEFT, _LEFT] + [_CENTER] * 12 + [_CENTER]
    for col, (h, al) in enumerate(zip(headers, aligns), start=1):
        c = ws.cell(5, col, h)
        c.font = _font(bold=True, color=_C["navy"], size=9)
        c.fill = _fill(_C["blue_header"])
        c.alignment = al
        c.border = _border(_C["border"])

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Group entries by account_code
    by_account: dict[str, list[dict]] = {}
    for e in entries:
        by_account.setdefault(e["account_code"], []).append(e)

    row = HEADER_ROW + 1
    subtotal_rows_by_account: dict[str, int] = {}

    for acct_code in sorted(by_account):
        acct_entries = by_account[acct_code]
        acct_name = acct_entries[0]["account_name"]
        detail_rows: list[int] = []

        for entry in acct_entries:
            ws.row_dimensions[row].height = 16
            # A: account code (gray, not editable intent)
            c = ws.cell(row, COL_ACCT, acct_code)
            c.font = _font(color=_C["text_mid"], size=9)
            c.alignment = _CENTER
            c.border = _border(_C["border"])

            # B: account name
            c = ws.cell(row, COL_NAME, acct_name)
            c.font = _font(color=_C["text_mid"], size=9)
            c.alignment = _LEFT
            c.border = _border(_C["border"])

            # C: detail desc  ← EDITABLE
            c = ws.cell(row, COL_DETAIL, entry.get("detail_desc", ""))
            c.font = _font(color=_C["text_dark"], size=9)
            c.alignment = _LEFT
            c.border = _border(_C["border"])
            unlock(c)

            # D-O: month values  ← EDITABLE
            # Una línea en COLONES exporta colones: es lo que el owner edita y lo
            # que tiene que volver al subir el archivo. Exportar los dólares haría
            # que al reimportarlos se tomaran como colones.
            es_crc = str(entry.get("currency") or "USD").upper() == "CRC"
            for mi, attr in enumerate(MONTH_ATTRS):
                col = COL_JAN + mi
                val = float((entry.get(f"crc_{attr}", 0) if es_crc
                             else entry.get(attr, 0)) or 0)
                c = ws.cell(row, col, val if val else None)
                c.font = _font(color=_C["text_dark"], size=9)
                c.alignment = _RIGHT
                c.border = _border(_C["border"])
                c.number_format = '#,##0.00;(#,##0.00);"-"'
                c.fill = _fill(_C["white"])
                unlock(c)

            # P: row total formula
            d_letter = get_column_letter(COL_JAN)
            o_letter = get_column_letter(COL_DEC)
            c = ws.cell(row, COL_TOTAL, f"=SUM({d_letter}{row}:{o_letter}{row})")
            c.font = _font(bold=True, color=_C["navy"], size=9)
            c.alignment = _RIGHT
            c.border = _border(_C["border"])
            c.number_format = '#,##0.00;(#,##0.00);"-"'
            c.fill = _fill(_C["gray_light"])

            # R: el «#» de la pantalla. Viaja para que al volver a subir el
            # archivo la fila conserve SU código y no uno nuevo por posición.
            cn = ws.cell(row, COL_NUM, entry.get("detail_code", ""))
            cn.font = _font(size=9, color=_C["text_mid"])
            cn.alignment = _RIGHT
            cn.border = _border(_C["border"])

            # Q: moneda de la línea ← EDITABLE. Si dice CRC, los montos de esta
            # fila están en colones.
            c = ws.cell(row, COL_MONEDA, "CRC" if es_crc else "USD")
            c.font = _font(bold=es_crc, size=9,
                           color=_C["navy"] if es_crc else _C["text_mid"])
            c.alignment = _RIGHT
            c.border = _border(_C["border"])
            unlock(c)

            detail_rows.append(row)
            row += 1

        # Subtotal row per account
        ws.row_dimensions[row].height = 16
        subtotal_rows_by_account[acct_code] = row

        c = ws.cell(row, COL_ACCT, acct_code)
        c.font = _font(bold=True, color=_C["navy"], size=9)
        c.fill = _fill(_C["blue_light"])
        c.alignment = _CENTER
        c.border = _border(_C["border"])

        c = ws.cell(row, COL_NAME, acct_name)
        c.font = _font(bold=True, color=_C["navy"], size=9)
        c.fill = _fill(_C["blue_light"])
        c.alignment = _LEFT
        c.border = _border(_C["border"])

        c = ws.cell(row, COL_DETAIL, "SUBTOTAL")
        c.font = _font(bold=True, color=_C["navy"], size=9)
        c.fill = _fill(_C["blue_light"])
        c.alignment = _LEFT
        c.border = _border(_C["border"])

        for mi in range(12):
            col = COL_JAN + mi
            letter = get_column_letter(col)
            refs = "+".join(f"{letter}{r}" for r in detail_rows)
            c = ws.cell(row, col, f"={refs}")
            c.font = _font(bold=True, color=_C["navy"], size=9)
            c.fill = _fill(_C["blue_light"])
            c.alignment = _RIGHT
            c.border = _border(_C["border"])
            c.number_format = '#,##0.00;(#,##0.00);"-"'

        p_letter = get_column_letter(COL_TOTAL)
        d_letter = get_column_letter(COL_JAN)
        o_letter = get_column_letter(COL_DEC)
        c = ws.cell(row, COL_TOTAL, f"=SUM({d_letter}{row}:{o_letter}{row})")
        c.font = _font(bold=True, color=_C["navy"], size=9)
        c.fill = _fill(_C["blue_light"])
        c.alignment = _RIGHT
        c.border = _border(_C["border"])
        c.number_format = '#,##0.00;(#,##0.00);"-"'

        row += 1

    # ── Grand total row ───────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20

    c = ws.cell(row, COL_ACCT, "")
    c.fill = _fill(_C["navy"])
    c.border = _border(_C["navy"])

    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row, COL_NAME, f"TOTAL  {dept_name.upper()}")
    c.font = _font(bold=True, color=_C["white"], size=10)
    c.fill = _fill(_C["navy"])
    c.alignment = _LEFT

    for mi in range(12):
        col = COL_JAN + mi
        letter = get_column_letter(col)
        refs = "+".join(f"{letter}{r}" for r in subtotal_rows_by_account.values())
        c = ws.cell(row, col, f"={refs}" if refs else 0)
        c.font = _font(bold=True, color=_C["white"], size=10)
        c.fill = _fill(_C["navy"])
        c.alignment = _RIGHT
        c.border = _border(_C["navy"])
        c.number_format = '#,##0.00;(#,##0.00);"-"'

    d_letter = get_column_letter(COL_JAN)
    o_letter = get_column_letter(COL_DEC)
    c = ws.cell(row, COL_TOTAL, f"=SUM({d_letter}{row}:{o_letter}{row})")
    c.font = _font(bold=True, color=_C["white"], size=10)
    c.fill = _fill(_C["navy"])
    c.alignment = _RIGHT
    c.border = _border(_C["navy"])
    c.number_format = '#,##0.00;(#,##0.00);"-"'

    # freeze panes below headers, right of detail cols
    ws.freeze_panes = ws.cell(HEADER_ROW + 1, COL_JAN)
    protect_sheet(ws)


# ── Import ────────────────────────────────────────────────────────────────────

def import_opex_from_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse an OPEX Excel file generated by export_opex_to_excel().
    Returns list of dicts with keys:
        dept_code, account_code, account_name, detail_desc,
        jan, feb, mar, apr, may, jun, jul, aug, sep, oct, nov, dec

    Only reads rows where:
        - Column A has a 4-digit numeric account code
        - Column C is not empty / not "SUBTOTAL" / not "TOTAL"
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    results: list[dict] = []

    for ws in wb.worksheets:
        # El código del departamento es lo primero del título de la hoja.
        #
        # ⚠️ Pedía EXACTAMENTE cuatro dígitos, y el Club Madresal es `260` y el
        # Área Recreativa `270`: de tres. Sus hojas no calzaban y se saltaban
        # SIN DECIR NADA — se podía llenar el OPEX del Club en el Excel, subirlo,
        # ver «importado correctamente» y que no hubiera entrado una sola línea.
        # Ahora acepta tres o cuatro, prefiriendo cuatro cuando los hay.
        sheet_name = ws.title.strip()
        dept_code_match = re.match(r"^(\d{3,4})(?!\d)", sheet_name)
        if not dept_code_match:
            continue
        dept_code = dept_code_match.group(1)

        # Find the header row (contains "Cuenta" or "ENE" in col D)
        header_row = None
        for r in ws.iter_rows(min_row=1, max_row=10):
            for cell in r:
                if cell.value and str(cell.value).strip().upper() in ("CUENTA", "ENE"):
                    header_row = cell.row
                    break
            if header_row:
                break

        if not header_row:
            continue

        # Read data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            acct_raw = row[COL_ACCT - 1]   # col A (0-indexed)
            name_raw = row[COL_NAME - 1]   # col B
            detail_raw = row[COL_DETAIL - 1]  # col C

            # Skip blank or structural rows
            if not acct_raw:
                continue
            acct_str = str(acct_raw).strip().split(".")[0]  # handle float "7065.0"
            if not re.match(r"^\d{3,5}$", acct_str):
                continue
            detail_str = str(detail_raw).strip() if detail_raw else ""
            if detail_str.upper() in ("SUBTOTAL", "TOTAL"):
                continue

            detail_code = ""
            if len(row) >= COL_NUM and row[COL_NUM - 1] not in (None, ""):
                detail_code = str(row[COL_NUM - 1]).strip().split(".")[0]

            moneda = "USD"
            if len(row) >= COL_MONEDA:
                crudo = row[COL_MONEDA - 1]
                if crudo and str(crudo).strip().upper() == "CRC":
                    moneda = "CRC"

            # En una linea CRC los montos del Excel son COLONES: van a crc_*, que
            # es el dato maestro. Meterlos en las columnas de dolares seria
            # multiplicar el presupuesto por el tipo de cambio.
            prefijo = "crc_" if moneda == "CRC" else ""
            months: dict[str, Decimal] = {}
            for mi, attr in enumerate(MONTH_ATTRS):
                raw = row[COL_JAN - 1 + mi]
                try:
                    v = Decimal(str(raw)).quantize(Decimal("0.01")) if raw is not None else Decimal("0")
                except (InvalidOperation, TypeError):
                    v = Decimal("0")
                months[prefijo + attr] = v

            results.append({
                "dept_code": dept_code,
                "account_code": acct_str.zfill(4),
                "account_name": str(name_raw).strip() if name_raw else "",
                "detail_desc": detail_str,
                "detail_code": detail_code,
                "currency": moneda,
                **months,
            })

    return results
