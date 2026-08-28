# -*- coding: utf-8 -*-
"""El P&L Detail del owner, en Excel y con su forma.

Owner, 2026-08-27: *«que se baje a excel super nítido y profesional»*.

Dos hojas en un archivo, porque son dos lecturas distintas del mismo dato:

* **Cierre** — los tres cortes lado a lado (el mes, el acumulado y el año), cada
  uno con sus versiones y la variación. Es el cuadro que arma cada mes.
* **Cascada** — los doce meses abiertos, con el detalle por línea.

## Por qué un exportador propio y no el genérico

`export/cuadro_excel.py` arma una tabla plana: un encabezado y filas. Este cuadro
tiene **encabezados de dos pisos** —el bloque (`Mayo`, `YTD Mayo`, `Full Year`) y
adentro las columnas de cada versión— y ese segundo piso es justamente lo que lo
hace leerse. Aplanarlo daría doce columnas seguidas sin decir a qué bloque
pertenece cada una.

## Detalles que hacen que se lea

* **Negativos entre paréntesis y en rojo**, que es como los lee un contador.
* **Cero se muestra como `—`**, no como `$0.00`: una línea sin movimiento no
  compite visualmente con las que sí lo tuvieron.
* **Panel congelado** en la primera columna: al desplazarse a la derecha, el
  rótulo de la fila sigue ahí. Sin eso, treinta filas adentro nadie sabe qué
  está mirando.
* **Una línea vertical gruesa entre bloques.** Es lo único que separa `Full
  Year` de `YTD` cuando las dos traen los mismos rótulos.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.export.excel_base import C, align, fill, font, set_col_widths, workbook_to_bytes

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
MES_CORTO = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

#: Contable: negativo entre paréntesis y en rojo, cero como raya.
MONEDA = '#,##0.00;[Red](#,##0.00);"—"'
CONTEO = '#,##0;[Red](#,##0);"—"'
PORC = '0.0%;[Red](0.0%);"—"'

_FINA = Side(style="thin", color=C["border"])
_GRUESA = Side(style="medium", color=C["navy"])


def _borde(izq_gruesa=False) -> Border:
    return Border(left=_GRUESA if izq_gruesa else _FINA,
                  right=_FINA, top=_FINA, bottom=_FINA)


def _titulo(ws, fila: int, texto: str, sub: str = "") -> int:
    ws.cell(row=fila, column=1, value=texto).font = font(
        bold=True, size=14, color=C["navy"])
    if sub:
        ws.cell(row=fila + 1, column=1, value=sub).font = font(
            size=9.5, color=C["text_mid"])
        return fila + 3
    return fila + 2


def _celda(ws, r, c, valor, formato, *, negrita=False, izq_gruesa=False,
           relleno=None, texto=False):
    x = ws.cell(row=r, column=c, value=valor)
    if not texto:
        x.number_format = formato
        x.alignment = align("right")
    else:
        x.alignment = align("left", wrap=False)
    x.font = font(bold=negrita, size=9.5,
                  color=C["navy"] if negrita else C["text_dark"])
    x.border = _borde(izq_gruesa)
    if relleno:
        x.fill = fill(relleno)
    return x


def _hoja_cierre(wb: Workbook, d: dict, mes: int) -> None:
    """Los tres cortes lado a lado, cada uno con sus versiones."""
    ws = wb.create_sheet("Cierre")
    vs = d["versiones"]
    n = len(vs)
    hay_var = n >= 2

    cortes = [
        (MESES[mes - 1], [mes - 1]),
        (f"YTD {MESES[mes - 1]}", list(range(mes))),
        ("Full Year", list(range(12))),
    ]
    # Cada bloque: una columna por versión, y si hay con qué comparar, la
    # variación en plata y en porcentaje.
    por_bloque = n + (2 if hay_var else 0)

    f = _titulo(ws, 1, f"P&L Detail — {d['titulo_ambito']}",
                f"{d['escenario']} · {d['nota_ambito']} · USD · cierre de "
                f"{MESES[mes - 1]} {d['year']}")

    # ── Encabezado de dos pisos ──────────────────────────────────────────────
    ws.cell(row=f, column=1, value="ACCOUNT DESCRIPTION")
    for c in (f, f + 1):
        x = ws.cell(row=c, column=1)
        x.fill = fill(C["navy"])
        x.font = font(bold=True, color=C["white"], size=9.5)
        x.border = _borde()
        x.alignment = align("left")
    ws.merge_cells(start_row=f, start_column=1, end_row=f + 1, end_column=1)

    col = 2
    for rotulo, _idx in cortes:
        ws.merge_cells(start_row=f, start_column=col,
                       end_row=f, end_column=col + por_bloque - 1)
        x = ws.cell(row=f, column=col, value=rotulo)
        x.fill = fill(C["navy"])
        x.font = font(bold=True, color=C["white"], size=10)
        x.alignment = align("center")
        for k in range(por_bloque):
            ws.cell(row=f, column=col + k).border = _borde(k == 0)
        sub = [v["escenario"] for v in vs] + (["Var $", "Var %"] if hay_var else [])
        for k, s in enumerate(sub):
            y = ws.cell(row=f + 1, column=col + k, value=s)
            y.fill = fill(C["navy_mid"])
            y.font = font(bold=True, color=C["white"], size=8.5)
            y.alignment = align("center", wrap=True)
            y.border = _borde(k == 0)
        col += por_bloque

    fila = f + 2

    def escribir(rotulo, series, formato, negrita=False, relleno=None,
                 razon=None):
        """`razon(idx, kpis)` para las estadísticas, que NO se suman."""
        nonlocal fila
        _celda(ws, fila, 1, rotulo, None, negrita=negrita, relleno=relleno,
               texto=True)
        c = 2
        for _r, idx in cortes:
            vals = []
            for i in range(n):
                if razon is not None:
                    vals.append(razon(idx, vs[i]["kpis"]))
                else:
                    s = series[i]
                    vals.append(sum(s[j] for j in idx) if s else None)
            for k, v in enumerate(vals):
                _celda(ws, fila, c + k, v, formato, negrita=negrita,
                       izq_gruesa=(k == 0), relleno=relleno)
            if hay_var:
                a, b = vals[0], vals[1]
                dif = (a - b) if (a is not None and b is not None) else None
                _celda(ws, fila, c + n, dif, formato, negrita=negrita,
                       relleno=relleno)
                # Sobre base cero no hay porcentaje: un 0% o un ∞ serían una
                # lectura inventada.
                p = (dif / abs(b)) if (dif is not None and b) else None
                _celda(ws, fila, c + n + 1, p, PORC, negrita=negrita,
                       relleno=relleno)
            c += por_bloque
        fila += 1

    def suma(idx, k, campo):
        return sum(k[campo][j] for j in idx)

    escribir("Total available Rooms", None, CONTEO,
             razon=lambda i, k: suma(i, k, "rooms_available"))
    escribir("Total Rooms Occupied", None, CONTEO,
             razon=lambda i, k: suma(i, k, "rooms_occupied"))
    escribir("Total Guests", None, CONTEO,
             razon=lambda i, k: suma(i, k, "guests"))
    escribir("% Occupancy", None, PORC, razon=lambda i, k: (
        suma(i, k, "rooms_occupied") / suma(i, k, "rooms_available")
        if suma(i, k, "rooms_available") else 0))
    escribir("Average Daily Room Only", None, MONEDA, razon=lambda i, k: (
        suma(i, k, "rooms_revenue") / suma(i, k, "rooms_occupied")
        if suma(i, k, "rooms_occupied") else 0))
    escribir("Total RevPAR", None, MONEDA, razon=lambda i, k: (
        suma(i, k, "rooms_revenue") / suma(i, k, "rooms_available")
        if suma(i, k, "rooms_available") else 0))

    fila += 1
    porRotulo = {x["rotulo"]: x for x in d["filas"]}
    for r in d["clave"]:
        x = porRotulo.get(r)
        if x:
            escribir(r, x["series"], MONEDA, negrita=True,
                     relleno=C["blue_header"])

    fila += 1
    for k, rot in d["clases_rotulos"]:
        escribir(rot, [c[k] for c in d["clases"]], MONEDA)
    total = [[sum(c[k][m] for k, _ in d["clases_rotulos"]) for m in range(12)]
             for c in d["clases"]]
    escribir("Total Operating and Property Expenses", total, MONEDA,
             negrita=True, relleno=C["blue_light"])

    anchos = {1: 42}
    for c in range(2, 2 + len(cortes) * por_bloque):
        anchos[c] = 15
    set_col_widths(ws, anchos)
    ws.freeze_panes = ws.cell(row=f + 2, column=2)


def _hoja_cascada(wb: Workbook, d: dict) -> None:
    """Los doce meses abiertos, con el detalle por línea, de la versión principal."""
    ws = wb.create_sheet("Cascada")
    f = _titulo(ws, 1, f"P&L Detail — {d['titulo_ambito']}",
                f"{d['escenario']} · {d['nota_ambito']} · USD")

    cab = ["ACCOUNT DESCRIPTION"] + MES_CORTO + ["Full Year"]
    for i, t in enumerate(cab, start=1):
        x = ws.cell(row=f, column=i, value=t)
        x.fill = fill(C["navy"])
        x.font = font(bold=True, color=C["white"], size=9.5)
        x.alignment = align("center" if i > 1 else "left", wrap=True)
        x.border = _borde()

    fila = f + 1
    for x in d["filas"]:
        if x["tipo"] == "esp":
            fila += 1
            continue
        fuerte = x["tipo"] in ("tot", "sec")
        relleno = (C["blue_header"] if x["tipo"] == "tot"
                   else C["gray_light"] if x["tipo"] == "sec" else None)
        rot = ("    " if x["tipo"] == "det" else "") + x["rotulo"]
        _celda(ws, fila, 1, rot, None, negrita=fuerte, relleno=relleno, texto=True)
        serie = x["series"][0]
        for m in range(12):
            _celda(ws, fila, 2 + m, serie[m] if serie else None, MONEDA,
                   negrita=fuerte, relleno=relleno)
        _celda(ws, fila, 14, sum(serie) if serie else None, MONEDA,
               negrita=True, relleno=relleno)
        fila += 1

    fila += 1
    ctl = d["control"]
    cuadra = abs(ctl["diferencia"]) < 0.01
    x = ws.cell(row=fila, column=1, value=(
        f"CONTROL · Ingresos {ctl['ingresos']:,.2f} · Gastos {ctl['gastos']:,.2f} · "
        f"Utilidad {ctl['utilidad']:,.2f} · Diferencia {ctl['diferencia']:,.2f} "
        f"{'✓' if cuadra else '⚠ NO CUADRA'}"))
    x.font = font(bold=True, size=10,
                  color=C["green_dark"] if cuadra else "B00020")

    set_col_widths(ws, {1: 44, **{2 + i: 14 for i in range(12)}, 14: 16})
    ws.freeze_panes = ws.cell(row=f + 1, column=2)


def export_pl_detail(d: dict, mes: int) -> bytes:
    """El archivo completo: Cierre primero, Cascada después.

    Cierre va primero a propósito: es el que se manda, y el que abre el archivo
    tiene que ver el cuadro del mes sin buscar pestaña.
    """
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_cierre(wb, d, mes)
    _hoja_cascada(wb, d)
    return workbook_to_bytes(wb)
