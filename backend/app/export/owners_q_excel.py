"""Export de `Owners Q` — el SCP Monthly P&L Report (POR/PAR).

TRES HOJAS:

  · `SCP<entidad>` — el archivo que recibe SCP, REPLICADO del suyo. Todo el
    formato salió de mirar el `.xlsx` que el owner manda: el bloque de
    encabezado con `As of Date` y `Location`, las fechas de las filas 6-7,
    Helvetica 12 en negrita, los dos azules de subtotal y total, los bordes,
    los anchos columna por columna y la sangría de dos espacios por nivel.
    El nombre interno "Owners Q" no viaja al entregable.

  · `Meses` y `Trimestres y Año` — los 12 meses, Q1..Q4 y Full Year, cada
    período con el set de seis columnas: Actual / Budget / Last Year y el
    acumulado de los tres (pedido del owner, 2026-08-17). No es lo que SCP
    pide; es para mirar el año sin abrir doce archivos.

El formato no es decoración: la sangría y los resaltes son CÓMO SCP LEE el
reporte, y lo consolida por posición de fila. Por eso el estilo de cada fila
vive en el catálogo (`report_lines.estilo`, leído del archivo del owner) y no
en este código: no se deriva de nada — la fila 49 lleva línea arriba sin ser
subtotal, y `TOTAL DEPARTMENTAL PROFIT` se pinta como subtotal mientras `GROSS
OPERATING PROFIT`, del mismo tipo, se pinta como total.
"""
from __future__ import annotations

import calendar as _cal
import datetime as _dt
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string as _col, get_column_letter

from app.export.excel_base import (
    C, align, border, fill, font, workbook_to_bytes,
)

# ── Paleta propia del reporte, derivada de la del proyecto ────────────────────
AZUL_TITULO = C["navy"]          # cabecera principal
AZUL_BLOQUE = C["navy_mid"]      # cabecera de cada bloque de 4 columnas
AZUL_CLARO = C["blue_light"]     # fondo de las filas de estadística
GRIS_CABEZA = C["blue_header"]   # fondo de la cabecera de columnas
VERDE = C["green_dark"]          # subtotales y utilidades
AMBAR = C["amber_light"]         # bloques de variación
ROJO = "C0392B"                  # negativos
BORDE = C["border"]

# ── Los formatos EXACTOS del archivo del owner (hoja `SCPCWL`) ───────────────
# Con signo de dólar y el negativo entre paréntesis en rojo, que es como Excel
# lo escribe: `_)` reserva el ancho del paréntesis para que la columna alinee.
MONEDA = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
PORCENTAJE = '0.0%_);[Red]\\(0.0%\\)'
ENTERO = '#,##0_);[Red](#,##0)'
FECHA = "mm-dd-yy"
RATIO = MONEDA          # ADR, RevPar y POR/PAR van con $ en el original

FUENTE = "Helvetica"
TAM = 12
#: El archivo del owner tiene TODA la grilla en negrita. No es un descuido: es
#: como SCP lo lee.
ALTO_FILA = 15.75
ALTO_FILA_TOTAL = 16.5

#: Los dos azules del original.
AZUL_SUBTOTAL = "FFDDEBF7"
AZUL_TOTAL = "FFBDD7EE"
BLANCO = "FFFFFFFF"

#: Anchos de columna del archivo del owner, columna por columna.
ANCHOS = {
    "A": 40.9, "B": 14.0, "C": 12.9, "D": 11.0, "E": 16.7, "F": 14.0,
    "G": 12.1, "H": 11.0, "I": 19.1, "J": 22.1, "K": 17.1, "L": 14.0,
    "M": 12.9, "N": 11.0, "O": 14.3, "P": 16.9, "Q": 59.3, "R": 16.7,
    "S": 14.0, "T": 12.1, "U": 10.3, "V": 16.7, "W": 14.0, "X": 12.1,
    "Y": 10.3, "Z": 19.0, "AA": 22.0, "AB": 21.3, "AC": 14.0, "AD": 11.0,
    "AE": 10.3, "AF": 16.7, "AG": 13.0,
}

#: Las columnas que SIEMPRE llevan su propio formato, sea cual sea la fila.
COL_PORCENTAJE = {"B", "F", "J", "L", "P", "S", "W", "AA", "AC", "AG"}
COL_RATIO = {"C", "D", "G", "H", "M", "N", "T", "U", "X", "Y", "AD", "AE"}
#: Las de valor y diferencia toman el formato DE LA FILA (las noches son
#: enteros, la ocupación es porcentaje, el resto es plata).
COL_DE_LA_FILA = {"A", "E", "K", "R", "V", "AB", "I", "O", "Z", "AF"}

#: Los tres bloques de `Month Ending` y los tres de acumulado, con el rango que
#: abarca cada rótulo en las filas 6 y 7 (fusionado, como en el original).
BLOQUES_CABECERA = [
    ("A", "B", "D", "Month Ending", "actual"),
    ("E", "F", "J", "Month Ending", "budget"),
    ("K", "L", "N", "Month Ending", "py"),
    ("R", "S", "U", "Year To Date", "actual"),
    ("V", "W", "AA", "Year To Date", "budget"),
    ("AB", "AC", "AE", "Prior Year To Date", "py"),
]

#: Las 33 columnas de la hoja de SCP, en orden. `Q` es la etiqueta.
CABECERAS = [
    ("A", "PTD Actual"), ("B", "% Revenue"), ("C", "POR"), ("D", "PAR"),
    ("E", "PTD Budget"), ("F", "% Revenue"), ("G", "POR"), ("H", "PAR"),
    ("I", "PTD Budget Diff"), ("J", "PTD Budget % Var"),
    ("K", "PTD PY Actual"), ("L", "% Revenue"), ("M", "POR"), ("N", "PAR"),
    ("O", "PTD PY Diff"), ("P", "PTD PY % Var"),
    ("Q", " "),
    ("R", "YTD Actual"), ("S", "% Revenue"), ("T", "POR"), ("U", "PAR"),
    ("V", "YTD Budget"), ("W", "% Revenue"), ("X", "POR"), ("Y", "PAR"),
    ("Z", "YTD Budget Diff"), ("AA", "YTD Budget % Var"),
    ("AB", "YTD PY Actual"), ("AC", "% Revenue"), ("AD", "POR"), ("AE", "PAR"),
    ("AF", "YTD PY Diff"), ("AG", "YTD PY % Var"),
]

#: Cómo se llama cada posición cuando hay que explicarla en el archivo.
POSICION = {"actual": "Actual", "budget": "columna Budget", "py": "columna Año Anterior"}

MESES_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _num(v):
    """El API manda strings (Decimal serializado) o None. None = celda vacía."""
    if v is None or v == "":
        return None
    try:
        return float(Decimal(str(v)))
    except (InvalidOperation, ValueError):
        return None


def _estilo_fila(fila: dict) -> tuple[bool, str, str | None]:
    """(negrita, color de letra, relleno) según el tipo de fila."""
    lt = fila["line_type"]
    if lt == "HEADER":
        return True, AZUL_TITULO, GRIS_CABEZA
    if lt in ("SUBTOTAL", "CALC"):
        return True, VERDE if fila["nature"] == "profit" else AZUL_TITULO, GRIS_CABEZA
    if lt == "STAT":
        return False, C["text_mid"], AZUL_CLARO
    return False, C["text_dark"], None


def _escribir_celda(ws, r, c, valor, tipo, negrita, color_letra, relleno):
    cell = ws.cell(row=r, column=c)
    v = _num(valor)
    if v is not None:
        cell.value = v
        cell.number_format = {"money": MONEDA, "pct": PORCENTAJE,
                              "ratio": RATIO}.get(tipo, MONEDA)
        # El rojo del negativo le gana al color del tipo de fila: en un P&L el
        # signo es lo primero que se mira.
        cell.font = font(bold=negrita, color=ROJO if v < 0 else color_letra)
    else:
        cell.font = font(bold=negrita, color=color_letra)
    cell.alignment = align("right")
    cell.border = border(BORDE)
    if relleno:
        cell.fill = fill(relleno)
    return cell


def _hoja_scp(wb, datos: dict) -> None:
    """La hoja que recibe SCP, replicada de su propio archivo.

    Todo lo de acá salió de mirar el `.xlsx` que el owner manda: el bloque de
    encabezado, las fechas de las filas 6-7, la fuente Helvetica 12 en negrita,
    los dos azules, los bordes, los anchos y la sangría de dos espacios por
    nivel. Nada es criterio propio — el reporte tiene que verse como el que ya
    están acostumbrados a leer.
    """
    ws = wb.active
    # El nombre de la hoja también es el del owner. `SCPCWL` para Corcovado;
    # sale de la entidad para que la próxima propiedad no necesite código.
    ws.title = f"SCP{datos['entidad']}"[:31]

    entidad, anio, mes = datos["entidad"], datos["anio"], datos["mes"]
    etiqueta_entidad = f"SCP {entidad}"

    # `As of` es el ÚLTIMO día del mes, como en el original (30-06-2026).
    def fin_de(a: int, m: int) -> _dt.date:
        return _dt.date(a, m, _cal.monthrange(a, m)[1])

    fin_mes = fin_de(anio, mes)
    es_un_mes = datos.get("es_un_mes", True)
    # Con un trimestre o el año, «Month Ending» sería mentira. El rótulo sigue
    # al período; con un mes simple queda EXACTAMENTE como el archivo del owner.
    rotulo_periodo = {"Month Ending": "Month Ending" if es_un_mes else "Period Ending",
                      "Year To Date": "Year To Date",
                      "Prior Year To Date": "Prior Year To Date"}
    # La fecha de CADA bloque sale del escenario que quedó en esa posición. Si
    # el owner puso un Forecast 2026 donde va el año anterior, el encabezado
    # dice 2026 — el archivo no puede mentir sobre qué está comparando.
    bloques = datos.get("bloques") or {}

    def fecha_bloque(cual: str) -> _dt.date:
        b = bloques.get(cual)
        if not b:
            return fin_de(anio - 1, mes) if cual == "py" else fin_mes
        return fin_de(b["anio"], b["mes"])

    def celda(fila, col_letra, valor=None, *, negrita=True, fmt=None,
              relleno=BLANCO, h="right", top="", bottom=""):
        c = ws.cell(row=fila, column=_col(col_letra))
        if valor is not None:
            c.value = valor
        c.font = Font(name=FUENTE, size=TAM, bold=negrita)
        c.alignment = Alignment(horizontal=h, vertical="bottom")
        if fmt:
            c.number_format = fmt
        if relleno:
            c.fill = PatternFill("solid", fgColor=relleno)
        if top or bottom:
            c.border = Border(top=Side(style=top or None),
                              bottom=Side(style=bottom or None))
        return c

    # ── Filas 1-5: identificación ────────────────────────────────────────────
    celda(1, "A", etiqueta_entidad, h="left")
    celda(2, "A", "Statement of Income", h="left")
    celda(3, "A", "As of Date:", negrita=False, h="left")
    celda(3, "B", fin_mes, negrita=False, fmt=FECHA, h="left")
    celda(4, "A", "Location:", negrita=False, h="left")
    celda(4, "B", etiqueta_entidad, negrita=False, h="left")
    # La fila 5 está en blanco en el original y así se queda cuando la
    # comparación es la estándar. Si NO lo es, ahí va el aviso: el archivo tiene
    # que decir en su cara qué se puso en cada columna, porque quien lo reciba
    # no tiene forma de saberlo mirando los números.
    if datos.get("es_estandar", True):
        celda(5, "A", " ", negrita=False, h="left")
    else:
        partes = []
        if not es_un_mes:
            partes.append(f"período {datos.get('periodo_etiqueta', '')}"
                          f" (SCP pide UN mes)")
        partes += [f"{POSICION[k]}: {bloques[k]['etiqueta']}"
                   f"{'' if bloques[k].get('periodo_etiqueta') in (None, datos.get('periodo_etiqueta')) else ' / ' + bloques[k]['periodo_etiqueta']}"
                   for k in ("actual", "budget", "py")
                   if bloques.get(k) and not bloques[k].get("por_defecto", True)]
        celda(5, "A", "⚠ NO es el reporte estándar de SCP — " + " · ".join(partes),
              negrita=True, h="left")

    # ── Filas 6-7: el bloque y su fecha, fusionados sobre sus ratios ─────────
    for ancla, m_ini, m_fin, titulo, cual in BLOQUES_CABECERA:
        celda(6, ancla, rotulo_periodo.get(titulo, titulo))
        celda(7, ancla, fecha_bloque(cual), negrita=False, fmt=FECHA)
        for fila in (6, 7):
            ws.merge_cells(f"{m_ini}{fila}:{m_fin}{fila}")
            celda(fila, m_ini, " ", negrita=(fila == 6))

    # ── Fila 8: los 33 rótulos de columna ───────────────────────────────────
    for letra, titulo in CABECERAS:
        celda(8, letra, titulo or " ", h="left" if letra == "Q" else "right")

    # ── Filas 9-56: las 48 del reporte ──────────────────────────────────────
    for fila in datos["filas"]:
        r = fila["row_no"]
        est = fila.get("estilo") or {}
        relleno = {"subtotal": AZUL_SUBTOTAL, "total": AZUL_TOTAL}.get(
            est.get("resalte", ""), BLANCO)
        top, bottom = est.get("top", ""), est.get("bottom", "")
        # La sangría es DATO: SCP lee el reporte por nivel, y son espacios
        # dentro del texto, no `indent` de Excel — así viene el original.
        espacios = est.get("sangria_espacios", fila["indent"] * 2)

        celda(r, "Q", " " * espacios + fila["label"], h="left",
              relleno=relleno, top=top, bottom=bottom)

        fmt_fila = {"int": ENTERO, "pct": PORCENTAJE}.get(
            est.get("formato", "money"), MONEDA)
        for letra, _t in CABECERAS:
            if letra == "Q":
                continue
            fmt = (PORCENTAJE if letra in COL_PORCENTAJE
                   else RATIO if letra in COL_RATIO else fmt_fila)
            c = celda(r, letra, fmt=fmt, relleno=relleno, top=top, bottom=bottom)
            v = _num(fila["celdas"].get(letra)) if fila["line_type"] != "HEADER" else None
            if v is not None:
                c.value = v

        ws.row_dimensions[r].height = est.get("alto") or ALTO_FILA

    for letra, ancho in ANCHOS.items():
        ws.column_dimensions[letra].width = ancho
    # Congela en B9 como el original: deja la columna A —el actual del mes— a la
    # vista mientras se recorren las otras 32.
    ws.freeze_panes = "B9"


#: El SET que pidió el owner (2026-08-17): seis columnas por período — el
#: período con sus tres datasets, y su acumulado con los mismos tres.
#: (clave_dataset, etiqueta, usa_acumulado)
SET_COLUMNAS = [
    ("actual", "ACTUAL", False),
    ("budget", "BUDGET", False),
    ("py", "LAST YEAR", False),
    ("actual", "ACUM ACT", True),
    ("budget", "ACUM BUD", True),
    ("py", "ACUM LY", True),
]


def _hoja_periodos(wb, anio_datos: dict, bases: list[str], titulo: str) -> None:
    """Una hoja con el set de 6 columnas por cada período de `bases`.

    `bases` son las claves SIN el sufijo `_ACUM` (`M01`…`M12`, o `Q1`…`Q4`,`FY`);
    el acumulado de cada una se busca como `<base>_ACUM`.
    """
    datasets = anio_datos["datasets"]
    if not datasets:
        return
    ws = wb.create_sheet(titulo[:31])
    filas = anio_datos["filas"]

    def valor(base: str, ds: str, acum: bool, code: str):
        bloque = datasets.get(ds)
        if bloque is None:
            return None
        clave = f"{base}_ACUM" if acum else base
        # El año completo no tiene «acumulado» aparte: el acumulado ES el año.
        # Sin este respaldo las tres columnas de acumulado de FY saldrían vacías.
        p = bloque["periodos"].get(clave) or bloque["periodos"].get(base)
        return None if p is None else p["valores"].get(code)

    def etiqueta_base(base: str) -> str:
        for bloque in datasets.values():
            p = bloque["periodos"].get(base)
            if p:
                return p["etiqueta"]
        return base

    ancho = len(SET_COLUMNAS)
    total_cols = 1 + ancho * len(bases)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1,
                value=f"{titulo}  ·  {anio_datos['entidad']} {anio_datos['anio']}  ·  USD"
                      f"  ·  cada período con Actual / Budget / Last Year, y su acumulado")
    t.font = font(bold=True, color="FFFFFF", size=14)
    t.fill = fill(AZUL_TITULO)
    t.alignment = align("center")

    # Fila 2: la banda del período, sobre sus seis columnas.
    ws.cell(row=2, column=1, value="").fill = fill(GRIS_CABEZA)
    ws.cell(row=3, column=1, value="").fill = fill(GRIS_CABEZA)
    for b, base in enumerate(bases):
        c0 = 2 + b * ancho
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + ancho - 1)
        h = ws.cell(row=2, column=c0, value=etiqueta_base(base))
        h.fill = fill(VERDE if base.startswith(("Q", "FY")) else AZUL_TITULO)
        h.font = font(bold=True, color="FFFFFF", size=11)
        h.alignment = align("center")
        h.border = border(BORDE)

        for k, (_ds, et, acum) in enumerate(SET_COLUMNAS):
            s = ws.cell(row=3, column=c0 + k, value=et)
            # El acumulado va en claro: separa de un vistazo el dato del arrastre.
            s.fill = fill(AZUL_CLARO if acum else AZUL_BLOQUE)
            s.font = font(bold=True, color=AZUL_TITULO if acum else "FFFFFF", size=9)
            s.alignment = align("center", wrap=True)
            s.border = border(BORDE)

    PRIMERA = 5
    for i, fila in enumerate(filas, start=PRIMERA):
        negrita, color_letra, relleno = _estilo_fila(fila)
        et = ws.cell(row=i, column=1,
                     value=("    " * (fila["indent"] - 1)) + fila["label"])
        et.font = font(bold=negrita, color=color_letra)
        et.border = border(BORDE)
        if relleno:
            et.fill = fill(relleno)

        tipo = {"STAT_OCC": "pct", "STAT_ADR": "ratio", "STAT_REVPAR": "ratio",
                "STAT_TOTAL_REVPAR": "ratio", "STAT_ROOMS_AVAILABLE": "ratio",
                "STAT_ROOMS_OCCUPIED": "ratio"}.get(fila["report_code"], "money")

        for b, base in enumerate(bases):
            c0 = 2 + b * ancho
            for k, (ds, _et, acum) in enumerate(SET_COLUMNAS):
                if fila["line_type"] == "HEADER":
                    c = ws.cell(row=i, column=c0 + k)
                    c.border = border(BORDE)
                    if relleno:
                        c.fill = fill(relleno)
                    continue
                _escribir_celda(
                    ws, i, c0 + k, valor(base, ds, acum, fila["report_code"]), tipo,
                    negrita, color_letra,
                    relleno or (AZUL_CLARO if acum else None))

    ws.column_dimensions["A"].width = 46
    for j in range(2, 1 + total_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.freeze_panes = f"B{PRIMERA}"
    ws.sheet_view.showGridLines = False


def _hoja_anio(wb, anio_datos: dict, dataset: str, titulo: str) -> None:
    """12 meses + Q1..Q4 + Full Year, para un dataset."""
    bloque = anio_datos["datasets"].get(dataset)
    if bloque is None:
        return
    ws = wb.create_sheet(titulo[:31])
    orden = anio_datos["orden_periodos"]
    periodos = bloque["periodos"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(orden))
    t = ws.cell(row=1, column=1,
                value=f"{titulo}  ·  {anio_datos['entidad']} {anio_datos['anio']}  ·  USD"
                      f"  ·  cada período con su mes y su acumulado")
    t.font = font(bold=True, color="FFFFFF", size=14)
    t.fill = fill(AZUL_TITULO)
    t.alignment = align("center")

    # Fila 2: la banda que agrupa cada par (el período y su acumulado), para
    # que se lea de un golpe qué columna es el mes y cuál viene arrastrando.
    ws.cell(row=2, column=1, value="").fill = fill(GRIS_CABEZA)
    j = 2
    while j < 2 + len(orden):
        clave = orden[j - 2]
        pareja = 2 if (j - 1) < len(orden) and orden[j - 1].endswith("_ACUM") else 1
        base = clave.replace("_ACUM", "")
        ws.merge_cells(start_row=2, start_column=j, end_row=2, end_column=j + pareja - 1)
        h = ws.cell(row=2, column=j,
                    value=periodos[base]["etiqueta"] if base in periodos else base)
        h.fill = fill(VERDE if base.startswith(("Q", "FY")) else AZUL_TITULO)
        h.font = font(bold=True, color="FFFFFF", size=10)
        h.alignment = align("center")
        h.border = border(BORDE)
        j += pareja

    ws.cell(row=3, column=1, value="").fill = fill(GRIS_CABEZA)
    for j, clave in enumerate(orden, start=2):
        es_acum = periodos[clave].get("acumulado", False)
        es_resumen = clave.startswith(("Q", "FY"))
        h = ws.cell(row=3, column=j,
                    value="ACUM" if es_acum else periodos[clave]["etiqueta"])
        # Verde = trimestre o año; azul = mes. El acumulado va más claro que su
        # período, así la lectura horizontal distingue el dato del arrastre.
        h.fill = fill(AZUL_CLARO if es_acum else (VERDE if es_resumen else AZUL_BLOQUE))
        h.font = font(bold=True, color=AZUL_TITULO if es_acum else "FFFFFF", size=9)
        h.alignment = align("center")
        h.border = border(BORDE)

    PRIMERA = 5
    for i, fila in enumerate(anio_datos["filas"], start=PRIMERA):
        negrita, color_letra, relleno = _estilo_fila(fila)
        et = ws.cell(row=i, column=1,
                     value=("    " * (fila["indent"] - 1)) + fila["label"])
        et.font = font(bold=negrita, color=color_letra)
        et.border = border(BORDE)
        if relleno:
            et.fill = fill(relleno)

        # El formato depende de la FILA, no de la columna: Occ% es porcentaje,
        # ADR y RevPar son ratios, el resto es plata.
        tipo = {"STAT_OCC": "pct", "STAT_ADR": "ratio", "STAT_REVPAR": "ratio",
                "STAT_TOTAL_REVPAR": "ratio",
                "STAT_ROOMS_AVAILABLE": "ratio",
                "STAT_ROOMS_OCCUPIED": "ratio"}.get(fila["report_code"], "money")
        for j, clave in enumerate(orden, start=2):
            if fila["line_type"] == "HEADER":
                c = ws.cell(row=i, column=j)
                c.border = border(BORDE)
                if relleno:
                    c.fill = fill(relleno)
                continue
            es_resumen = clave.startswith(("Q", "FY"))
            es_acum = periodos[clave].get("acumulado", False)
            _escribir_celda(ws, i, j, periodos[clave]["valores"].get(fila["report_code"]),
                            tipo, negrita or es_resumen, color_letra,
                            relleno or (AZUL_CLARO if (es_resumen or es_acum) else None))

    ws.column_dimensions["A"].width = 46
    for j in range(2, 2 + len(orden)):
        ws.column_dimensions[get_column_letter(j)].width = 15
    ws.freeze_panes = f"B{PRIMERA}"
    ws.sheet_view.showGridLines = False


def export_owners_q(datos: dict, anio_datos: dict | None = None) -> bytes:
    """`datos` = salida de `/reports/owners-q/`; `anio_datos` la de `/anio/`.

    Tres hojas: la de SCP (el mes que se manda), los doce meses, y los
    trimestres con el año. Las dos últimas con el mismo set de seis columnas
    —Actual / Budget / Last Year, y el acumulado de los tres— para que se lean
    igual sin volver a aprender el formato.
    """
    wb = Workbook()
    _hoja_scp(wb, datos)
    if anio_datos:
        _hoja_periodos(wb, anio_datos, [f"M{m:02d}" for m in range(1, 13)], "Meses")
        _hoja_periodos(wb, anio_datos, ["Q1", "Q2", "Q3", "Q4", "FY"],
                       "Trimestres y Año")
    return workbook_to_bytes(wb)


def nombre_archivo(entidad: str, anio: int, datos: dict | int) -> str:
    """`SCP_CWL_JUN26_Statement_of_Income.xlsx` — sin "Owners Q" adentro.

    Con un trimestre o el año el nombre lo dice (`Q2_26`, `FY26`): dos archivos
    del mismo año que se llamaran igual y dijeran cosas distintas serían una
    trampa esperando.
    """
    # `datos` puede ser el mes suelto (firma vieja) o el reporte entero.
    if isinstance(datos, int):
        clave, mes = f"M{datos:02d}", datos
    else:
        clave, mes = datos.get("periodo") or f"M{datos['mes']:02d}", datos["mes"]

    aa = f"{anio % 100:02d}"
    if clave.startswith("M"):
        abrev = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                 "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"][mes - 1]
        tramo = f"{abrev}{aa}"
    elif clave == "FY":
        tramo = f"FY{aa}"
    else:
        tramo = f"{clave}_{aa}"
    return f"SCP_{entidad}_{tramo}_Statement_of_Income.xlsx"
