"""
Shared styles and helpers for all FinPlan CWL Excel exporters.
"""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MONTHS_ES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
             "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
MONTH_ATTRS = ["jan", "feb", "mar", "apr", "may", "jun",
               "jul", "aug", "sep", "oct", "nov", "dec"]

# ── Shared color palette ───────────────────────────────────────────────────────

C = {
    "navy":        "1A3A5C",
    "navy_mid":    "2D5A9E",
    "blue_light":  "EBF3FB",
    "blue_header": "F0F4F8",
    "white":       "FFFFFF",
    "text_dark":   "1A1A2E",
    "text_mid":    "4A5568",
    "gray_light":  "F8F9FA",
    "border":      "CBD5E0",
    "green_dark":  "1A5C3A",
    "amber":       "92400E",
    "amber_light": "FEF3C7",
    "teal_dark":   "0D5E5E",
    "teal_light":  "E0F7F7",
}


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color="1A1A2E", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def border(color="CBD5E0", sides="all") -> Border:
    s = Side(style="thin", color=color)
    m = Side(style="medium", color=color)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    if sides == "top":
        return Border(top=m)
    if sides == "top_bottom":
        return Border(top=s, bottom=s)
    return Border()


def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def merged_header(ws, row: int, col_start: int, col_end: int,
                  text: str, bg: str, fg: str = "FFFFFF", sz: int = 12) -> None:
    """Write a merged cell spanning col_start..col_end on given row."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end,
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=sz)
    cell.alignment = align("center")


def month_header_row(ws, row: int, first_col: int, total_col: int | None = None,
                     bg: str = "2D5A9E", fg: str = "FFFFFF") -> None:
    """Write ENE..DIC headers starting at first_col; optionally write TOTAL."""
    for i, m in enumerate(MONTHS_ES):
        c = ws.cell(row=row, column=first_col + i, value=m)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=10)
        c.alignment = align("center")
        c.border = border()
    if total_col:
        t = ws.cell(row=row, column=total_col, value="TOTAL")
        t.fill = fill(bg)
        t.font = font(bold=True, color=fg, size=10)
        t.alignment = align("center")
        t.border = border()


def set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Protection helpers ─────────────────────────────────────────────────────────

SHEET_PASSWORD = "cwl2026"

_UNLOCKED = Protection(locked=False)
_LOCKED   = Protection(locked=True)


def unlock(cell) -> None:
    """Mark a cell as editable when the sheet is protected."""
    cell.protection = _UNLOCKED


def protect_sheet(ws, password: str = SHEET_PASSWORD) -> None:
    """
    Enable sheet protection. All cells are locked by default; call unlock()
    on data-entry cells before calling this.
    Allows: selecting any cell, using AutoFilter, formatting unlocked cells.
    """
    ws.protection.sheet          = True
    ws.protection.password       = password
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False


def add_dropdown(ws, col_letter: str, first_data_row: int, last_row: int,
                 options: list[str], prompt: str = "") -> None:
    """Add an in-cell dropdown for col_letter from first_data_row to last_row."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,   # False = show the arrow
        prompt=prompt,
        promptTitle="Opciones",
    )
    dv.sqref = f"{col_letter}{first_data_row}:{col_letter}{last_row}"
    ws.add_data_validation(dv)


# ── Nombre de pestaña ─────────────────────────────────────────────────────────

# Excel prohíbe estos caracteres en el título de una hoja y el libro entero se
# cae al guardarlo. No es hipotético: los tres exportadores pasaron a rotular las
# pestañas con el `department_catalog` —para que el 260 dejara de salir «260
# 260»— y el catálogo dice «Rooms / Habitaciones». La barra reventó el export
# completo en el primer intento.
_PROHIBIDOS = r':\/?*[]'
LARGO_MAX_HOJA = 31


def nombre_de_hoja(base: str, usados: set[str], *, sufijo: str = "") -> str:
    """Un título de hoja válido y ÚNICO, y lo registra en `usados`.

    Dos cosas que Excel no perdona y que se descubren tarde, al abrir el archivo:

    * los caracteres de arriba;
    * y **el nombre repetido** — dos departamentos con nombre largo colapsan en
      los mismos 31 caracteres y el libro no abre. Por eso hay `sufijo`: algo
      corto y único (el código del departamento) con lo que desempatar.
    """
    limpio = "".join(ch for ch in (base or "") if ch not in _PROHIBIDOS).strip()
    nombre = (limpio or "Hoja")[:LARGO_MAX_HOJA]
    if nombre in usados and sufijo:
        nombre = f"{nombre[:LARGO_MAX_HOJA - len(sufijo) - 1]} {sufijo}"[:LARGO_MAX_HOJA]
    n = 2
    while nombre in usados:                      # último recurso: numerar
        cola = f" ({n})"
        nombre = f"{nombre[:LARGO_MAX_HOJA - len(cola)]}{cola}"
        n += 1
    usados.add(nombre)
    return nombre
