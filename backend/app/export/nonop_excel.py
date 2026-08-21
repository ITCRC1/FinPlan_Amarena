"""
NonOp (below-GOP / owner) Excel export / import.

Format:
  One sheet per report_line_code (Rent, Insurance, CapEx, etc.).
  Columns: A=detail_code  B=account_code  C=detail_desc  D..O=ENE..DIC  P=TOTAL
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from app.export.excel_base import (
    C, fill, font, border, align,
    merged_header, month_header_row, set_col_widths, workbook_to_bytes,
    unlock, protect_sheet,
    MONTHS_ES, MONTH_ATTRS,
)

# La plantilla tiene que devolver EXACTAMENTE lo que se subió: se baja, se
# corrige y se vuelve a subir. Todo lo que la plantilla NO lleve se pierde,
# porque la importación borra las filas antes de escribir las nuevas.
#
# `account_code` no estaba, y no se restauraba: bajar y subir el archivo dejaba
# en blanco la cuenta 8xxx de TODA línea below-GOP. No daba error — la línea
# del P&L la decide `report_line_code`, no la cuenta, así que los números
# seguían cuadrando y el dato de referencia desaparecía sin más.
COL_CODE  = 1   # A — detail_code (correlativo)
COL_ACCT  = 2   # B — account_code (8015, 8020…), referencia
COL_DESC  = 3   # C
COL_JAN   = 4   # D
COL_TOTAL = 16  # P

HEADER_ROW = 4

# Standard below-GOP manual lines (mirror the owner screen). Used to build a
# fillable template when the scenario has no entries yet.
DEFAULT_NONOP_LINES: dict[str, str] = {
    "RENT": "Rent",
    "PROPERTY_INSURANCE": "Properties Insurance",
    "OTHER_EXPENSES": "Other Expenses",
    "CAPITAL_RESERVE": "Capital Reserve",
    "LARGE_CAPEX": "Large Capital Expenditure",
    "BANK_INTEREST": "Bank Interest Charges",
    "LEASINGS_RENTS": "Leasings/Rents",
    "FINANCIAL_LOSSES": "Financial Losses",
    "DEPRECIATION": "Depreciation",
    "ASSET_LOSS": "Asset Loss",
}


def export_nonop_to_excel(
    entries_by_line: dict[str, list[dict]],
    scenario_label: str,
    year: int,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    # Empty scenario → emit a fillable template (one sheet per standard line).
    if not entries_by_line:
        entries_by_line = {
            lc: [{"detail_code": "", "detail_desc": desc,
                  **{mk: 0 for mk in MONTH_ATTRS}}]
            for lc, desc in DEFAULT_NONOP_LINES.items()
        }

    for line_code, entries in sorted(entries_by_line.items()):
        sheet_name = line_code[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: system title
        merged_header(ws, 1, COL_CODE, COL_TOTAL,
                      f"FinPlan {HOTEL_ID} — Non-Operating Expenses | {scenario_label} {year}",
                      C["navy"], sz=13)

        # Row 2: line code
        merged_header(ws, 2, COL_CODE, COL_TOTAL,
                      line_code, C["navy_mid"], sz=11)

        # Row 3: instructions
        texto_inst = ("Editar Cuenta, Descripción y ENE-DIC. La columna A (Código) la "
                     "asigna el sistema: no la toque.")
        # ⚠️ Capital Reserve tiene un driver de % de ingreso en Management Fees
        # (pl_engine.py: `seeds["CAPITAL_RESERVE"] = total_rev × pct`) que, si
        # está configurado, REEMPLAZA lo que se tipee acá — no se suman, y no
        # hay error ni aviso en ningún lado más que este. Antes esta hoja se
        # veía igual de editable que las otras nueve y no decía que podía no
        # servir de nada.
        if line_code == "CAPITAL_RESERVE":
            texto_inst += (" ⚠ Si el % de Capital Reserve está configurado en "
                           "Management Fees, ese % MANDA y lo que se tipee acá "
                           "se ignora — no se suma. Cargá acá solo si ese % "
                           "está en 0.")
        inst = ws.cell(row=3, column=COL_CODE, value=texto_inst)
        inst.font = font(italic=True, color=C["text_mid"], size=9)
        ws.merge_cells(start_row=3, start_column=COL_CODE,
                       end_row=3, end_column=COL_TOTAL)

        # Row 4: column headers
        for label, col in [("Código", COL_CODE), ("Cuenta", COL_ACCT),
                           ("Descripción", COL_DESC)]:
            c = ws.cell(row=HEADER_ROW, column=col, value=label)
            c.fill = fill(C["blue_header"])
            c.font = font(bold=True, color=C["navy"], size=10)
            c.alignment = align("center")
            c.border = border()
        month_header_row(ws, HEADER_ROW, COL_JAN, COL_TOTAL)

        # Data rows
        row = HEADER_ROW + 1
        for e in entries:
            vals = [e.get(mk, 0) or 0 for mk in MONTH_ATTRS]

            ws.cell(row=row, column=COL_CODE, value=e.get("detail_code", "")).border = border()
            ac = ws.cell(row=row, column=COL_ACCT, value=e.get("account_code", ""))
            ac.font = font(size=10)
            ac.alignment = align("left")
            ac.border = border()
            unlock(ac)   # ← editable: es dato del usuario, no un correlativo
            dc = ws.cell(row=row, column=COL_DESC, value=e.get("detail_desc", ""))
            dc.font = font(size=10)
            dc.alignment = align("left")
            dc.border = border()
            unlock(dc)   # ← editable

            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=COL_JAN + i, value=float(v))
                cell.number_format = "#,##0.00;(#,##0.00)"
                cell.alignment = align("right")
                cell.fill = fill(C["white"])
                cell.border = border()
                unlock(cell)   # ← editable

            # TOTAL formula
            first = get_column_letter(COL_JAN)
            last  = get_column_letter(COL_JAN + 11)
            tc = ws.cell(row=row, column=COL_TOTAL,
                         value=f"=SUM({first}{row}:{last}{row})")
            tc.number_format = "#,##0.00;(#,##0.00)"
            tc.fill = fill(C["blue_light"])
            tc.font = font(bold=True, color=C["navy"], size=10)
            tc.alignment = align("right")
            tc.border = border()

            row += 1

        # Grand total row
        data_start = HEADER_ROW + 1
        data_end   = row - 1
        gt = ws.cell(row=row, column=COL_CODE, value="TOTAL")
        gt.fill = fill(C["navy"])
        gt.font = font(bold=True, color="FFFFFF", size=11)
        gt.border = border()
        ws.merge_cells(start_row=row, start_column=COL_CODE,
                       end_row=row, end_column=COL_DESC)

        for i in range(12):
            col_letter = get_column_letter(COL_JAN + i)
            tc = ws.cell(row=row, column=COL_JAN + i,
                         value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
            tc.number_format = "#,##0.00;(#,##0.00)"
            tc.fill = fill(C["navy"])
            tc.font = font(bold=True, color="FFFFFF", size=10)
            tc.alignment = align("right")
            tc.border = border()

        tot_letter = get_column_letter(COL_TOTAL)
        tt = ws.cell(row=row, column=COL_TOTAL,
                     value=f"=SUM({tot_letter}{data_start}:{tot_letter}{data_end})")
        tt.number_format = "#,##0.00;(#,##0.00)"
        tt.fill = fill(C["navy"])
        tt.font = font(bold=True, color="FFFFFF", size=11)
        tt.alignment = align("right")
        tt.border = border()

        protect_sheet(ws)
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=COL_JAN)
        set_col_widths(ws, {
            COL_CODE: 10, COL_ACCT: 10, COL_DESC: 32,
            **{COL_JAN + i: 11 for i in range(12)},
            COL_TOTAL: 13,
        })
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

    return workbook_to_bytes(wb)


def get_column_letter(col: int) -> str:
    from openpyxl.utils import get_column_letter as gcl
    return gcl(col)


def _columnas(ws, header_row: int) -> dict[str, int]:
    """Ubica las columnas por su encabezado, con la posición vieja de respaldo.

    Un archivo bajado antes de que existiera la columna «Cuenta» sigue
    subiendo bien: no la encuentra, `cuenta` queda en `None` y esa fila entra
    sin cuenta en vez de entrar con los datos corridos.
    """
    ALIAS = {
        "codigo": ("CÓDIGO", "CODIGO"),
        "cuenta": ("CUENTA",),
        "desc":   ("DESCRIPCIÓN", "DESCRIPCION"),
        "ene":    ("ENE", "ENERO", "JAN"),
    }
    encontrado: dict[str, int] = {}
    for c in range(1, (ws.max_column or 20) + 1):
        titulo = str(ws.cell(row=header_row, column=c).value or "").strip().upper()
        if not titulo:
            continue
        for llave, nombres in ALIAS.items():
            if llave not in encontrado and titulo in nombres:
                encontrado[llave] = c
    # Respaldo: un archivo sin encabezados reconocibles se lee como siempre.
    encontrado.setdefault("codigo", COL_CODE)
    encontrado.setdefault("desc", COL_DESC)
    encontrado.setdefault("ene", COL_JAN)
    return encontrado


def import_nonop_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        line_code = sheet_name.strip()

        # La fila de encabezados se reconoce por IGUALDAD, no por «contiene».
        # La fila de instrucciones nombra «Código» y «ENE-DIC» dentro de una
        # frase: con `in` se la tragaba como cabecera y leía los encabezados
        # de verdad como si fueran una fila de datos.
        header_row = None
        for r in range(1, 10):
            val = str(ws.cell(row=r, column=COL_CODE).value or "").strip().upper()
            if val in ("CÓDIGO", "CODIGO"):
                header_row = r
                break
        if header_row is None:
            continue

        # Las columnas se ubican por su ENCABEZADO, no por su posición fija.
        #
        # Si se leyeran por posición, agregar una columna rompería todos los
        # archivos que la gente ya tenga bajados: los baja, los corrige durante
        # días y los sube — y de golpe la subida cargaría los datos corridos una
        # celda, sin dar error. Leyendo el encabezado, el archivo viejo y el
        # nuevo entran los dos bien, y la columna que falte simplemente no viene.
        col = _columnas(ws, header_row)

        for r in range(header_row + 1, ws.max_row + 1):
            code_raw = ws.cell(row=r, column=col["codigo"]).value
            acct_raw = ws.cell(row=r, column=col["cuenta"]).value if col.get("cuenta") else None
            desc_raw = ws.cell(row=r, column=col["desc"]).value
            if not code_raw and not desc_raw:
                continue
            code = str(code_raw or "").strip()
            desc = str(desc_raw or "").strip()
            # Los renglones de totales que el propio exportador escribe llevan
            # «TOTAL» en la columna del CÓDIGO. Antes se miraban por la
            # descripción, que en esa fila viene vacía porque está combinada:
            # desde que la descripción vacía dejó de descartarse —la pantalla
            # permite dejarla en blanco— el TOTAL se colaba como una fila más.
            if code.upper() in ("TOTAL", "SUBTOTAL") or desc.upper() in ("TOTAL", "SUBTOTAL"):
                continue

            row_dict: dict = {
                "report_line_code": line_code,
                "detail_code": code,
                # Sin esto, el viaje de ida y vuelta borraba la cuenta.
                "account_code": str(acct_raw or "").strip(),
                "detail_desc": desc,
            }
            for i, mk in enumerate(MONTH_ATTRS):
                raw = ws.cell(row=r, column=col["ene"] + i).value
                try:
                    row_dict[mk] = Decimal(str(raw)) if raw not in (None, "") else Decimal("0")
                except InvalidOperation:
                    row_dict[mk] = Decimal("0")

            results.append(row_dict)

    return results


import io
from app.hotel_actual import HOTEL_ID
