"""
Scenario Stats (KPIs) Excel export / import.

Format: single sheet, rows = KPI metrics, columns = ENE..DIC.
Metrics: rooms_available, rooms_occupied, guests, occupancy_pct, adr
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from app.hotel_actual import HOTEL_ID

from app.export.excel_base import (
    C, fill, font, border, align,
    merged_header, month_header_row, set_col_widths, workbook_to_bytes,
    unlock, protect_sheet,
    MONTHS_ES, MONTH_ATTRS,
)

METRICS = [
    ("rooms_available", "Habitaciones Disponibles", "# hab", "#,##0"),
    ("rooms_occupied",  "Habitaciones Ocupadas",    "# hab", "#,##0.0"),
    ("guests",          "Huéspedes",                "# pax", "#,##0.0"),
    ("occupancy_pct",   "Ocupación %",              "%",     "0.0%"),
    ("adr",             "ADR (Tarifa Promedio)",    "USD",   "$#,##0.00"),
]

COL_LABEL  = 1   # A
COL_UNIT   = 2   # B
COL_JAN    = 3   # C
COL_TOTAL  = 15  # O
HEADER_ROW = 4


def export_stats_to_excel(
    stats_by_month: dict[int, dict],  # month 1-12 → dict with metric fields
    scenario_label: str,
    year: int,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]
    ws = wb.create_sheet(title="KPIs")

    merged_header(ws, 1, COL_LABEL, COL_TOTAL,
                  f"FinPlan {HOTEL_ID} — Estadísticas de Ocupación | {scenario_label} {year}",
                  C["teal_dark"], sz=13)
    merged_header(ws, 2, COL_LABEL, COL_TOTAL,
                  "KPIs — Habitaciones · Huéspedes · Ocupación · ADR",
                  C["navy_mid"], sz=11)

    inst = ws.cell(row=3, column=COL_LABEL,
                   value="Editar columnas C-N (meses). Ocupación % se puede dejar en 0 — el motor la recalcula.")
    inst.font = font(italic=True, color=C["text_mid"], size=9)
    ws.merge_cells(start_row=3, start_column=COL_LABEL,
                   end_row=3, end_column=COL_TOTAL)

    # Headers
    for label, col in [("Indicador", COL_LABEL), ("Unidad", COL_UNIT)]:
        c = ws.cell(row=HEADER_ROW, column=col, value=label)
        c.fill = fill(C["blue_header"])
        c.font = font(bold=True, color=C["navy"], size=10)
        c.alignment = align("center")
        c.border = border()
    month_header_row(ws, HEADER_ROW, COL_JAN, total_col=COL_TOTAL,
                     bg=C["teal_dark"])

    # Data rows
    for row_idx, (field, label, unit, fmt) in enumerate(METRICS):
        row = HEADER_ROW + 1 + row_idx
        lc = ws.cell(row=row, column=COL_LABEL, value=label)
        lc.font = font(bold=True, size=10)
        lc.fill = fill(C["blue_header"])
        lc.border = border()

        uc = ws.cell(row=row, column=COL_UNIT, value=unit)
        uc.font = font(size=9, color=C["text_mid"])
        uc.alignment = align("center")
        uc.fill = fill(C["blue_header"])
        uc.border = border()

        for m in range(1, 13):
            stats = stats_by_month.get(m, {})
            raw = stats.get(field, 0) or 0
            # occupancy_pct stored as fraction (e.g. 0.82) or percent? Check both
            if field == "occupancy_pct" and float(raw) > 1:
                raw = float(raw) / 100
            val = float(raw)
            col = COL_JAN + (m - 1)
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = fmt
            cell.alignment = align("right")
            cell.fill = fill(C["teal_light"] if row_idx % 2 == 0 else C["white"])
            cell.border = border()
            unlock(cell)   # ← editable

        # Total / avg in last col
        first = get_column_letter(COL_JAN)
        last  = get_column_letter(COL_JAN + 11)
        if field in ("occupancy_pct", "adr"):
            formula = f"=AVERAGE({first}{row}:{last}{row})"
        else:
            formula = f"=SUM({first}{row}:{last}{row})"
        tc = ws.cell(row=row, column=COL_TOTAL, value=formula)
        tc.number_format = fmt
        tc.fill = fill(C["teal_dark"])
        tc.font = font(bold=True, color="FFFFFF", size=10)
        tc.alignment = align("right")
        tc.border = border()

    protect_sheet(ws)
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=COL_JAN)
    set_col_widths(ws, {
        COL_LABEL: 28, COL_UNIT: 8,
        **{COL_JAN + i: 11 for i in range(12)},
        COL_TOTAL: 13,
    })
    ws.row_dimensions[1].height = 22

    return workbook_to_bytes(wb)


def import_stats_from_excel(file_bytes: bytes) -> list[dict]:
    """Return list of dicts: {month: 1-12, field: value}"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active or wb.worksheets[0]

    # Find header row
    header_row = None
    for r in range(1, 10):
        val = str(ws.cell(row=r, column=COL_LABEL).value or "").upper()
        if "INDICADOR" in val or "KPI" in val:
            header_row = r
            break
    if header_row is None:
        return []

    # Build month→col mapping from header
    month_cols: dict[int, int] = {}
    for col in range(COL_JAN, COL_JAN + 12):
        val = str(ws.cell(row=header_row, column=col).value or "").upper()
        for i, m in enumerate(MONTHS_ES):
            if m in val:
                month_cols[i + 1] = col
                break

    if not month_cols:
        return []

    # Read metric rows
    field_map: dict[str, str] = {label.upper(): field for field, label, *_ in METRICS}
    results_by_month: dict[int, dict] = {m: {"month": m} for m in range(1, 13)}

    for r in range(header_row + 1, ws.max_row + 1):
        label_raw = str(ws.cell(row=r, column=COL_LABEL).value or "").strip().upper()
        field = field_map.get(label_raw)
        if not field:
            continue
        for month, col in month_cols.items():
            raw = ws.cell(row=r, column=col).value
            try:
                val = Decimal(str(raw)) if raw not in (None, "") else Decimal("0")
            except InvalidOperation:
                val = Decimal("0")
            results_by_month[month][field] = val

    return list(results_by_month.values())
