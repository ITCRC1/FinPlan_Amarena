"""Plantilla de ida y vuelta del Country Mix: bajo, corrijo, subo.

Norma del owner (18-ago-2026, sobre el mix por país): «Debe haber la opción de
editar por si acaso un ajuste; se podría quizás bajar a Excel controlado y
después subir con el cambio.»

⚠️ **UNA HOJA POR MÉTRICA.** «Se me hace confuso subir si pax y rooms; ocupo que
haya 2 tabs para esto» (owner). La primera versión ponía las dos métricas en la
misma hoja, separadas por una columna «Métrica»: para corregir las noches de un
país había que encontrar SU fila entre el doble de filas, y una fila mal
rotulada cambiaba la métrica equivocada sin que se notara. Ahora la métrica es
la PESTAÑA, y dentro de cada una hay una fila por país y nada más:

    hoja «Habitaciones»          hoja «Pax»
    País          | Ene | Feb …  País          | Ene | Feb …
    United States |  322|  300   United States |  612|  557

⚠️ **Las columnas se ubican por ENCABEZADO, nunca por posición.** Es la norma
del proyecto y tiene razón de ser: el owner trabaja sobre estos archivos, y
buscar por posición rompe la plantilla en cuanto alguien inserta una columna.
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
COL_PAIS = "País"
#: Columna de la primera versión de la plantilla. Ya no se escribe, pero se
#: sigue LEYENDO: alguien puede tener bajado un archivo del formato viejo.
COL_METRICA = "Métrica"

#: Métrica ↔ pestaña.
HOJA_DE = {"rooms": "Habitaciones", "pax": "Pax"}
METRICA_DE = {"habitaciones": "rooms", "rooms": "rooms", "noches": "rooms",
              "pax": "pax", "huespedes": "pax", "huéspedes": "pax"}


def construir_libro(titulo: str, filas: list[dict]) -> openpyxl.Workbook:
    """`filas` = [{pais, metric, values: [12]}]. Una hoja por métrica."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for metric, hoja in HOJA_DE.items():
        ws = wb.create_sheet(hoja)
        ws["A1"] = f"{titulo} — {hoja}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = ("Corregí los números y volvé a subir este mismo archivo. Cada "
                    "pestaña es una métrica: no las mezcles. Las columnas se "
                    "ubican por su encabezado, así que podés mover o insertar "
                    "columnas sin romper nada. Un país que no exista se crea; "
                    "para que deje de existir, borrá su fila.")
        ws["A2"].font = Font(size=9, italic=True, color="FF808080")

        for j, nombre in enumerate([COL_PAIS] + MESES, start=1):
            c = ws.cell(row=4, column=j, value=nombre)
            c.font = Font(bold=True, color="FFFFFFFF")
            c.fill = PatternFill("solid", fgColor="FF2D3A5C")
            c.alignment = Alignment(horizontal="center")

        i = 5
        for f in filas:
            if f["metric"] != metric:
                continue
            ws.cell(row=i, column=1, value=f["pais"])
            for m in range(12):
                v = f["values"][m] if m < len(f["values"]) else 0
                celda = ws.cell(row=i, column=2 + m, value=float(v or 0))
                celda.number_format = "#,##0"
            i += 1

        ws.column_dimensions["A"].width = 28
        for j in range(2, 14):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 9
        ws.freeze_panes = "B5"
    return wb


def _num(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _norm(s: str) -> str:
    t = (s or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        t = t.replace(a, b)
    return t


def leer_libro(data: bytes) -> tuple[list[dict], list[str]]:
    """Lee la plantilla. Devuelve `(filas, problemas)`.

    La métrica sale del NOMBRE DE LA HOJA. Si la hoja además trae una columna
    «Métrica» —el formato viejo, de una sola hoja— se respeta esa, para que un
    archivo bajado antes del cambio siga subiendo.

    Un problema NO es una excepción: se juntan todos y se devuelven, para que
    el que subió el archivo vea de una vez todo lo que hay que arreglar y no de
    a uno por intento.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"No se pudo abrir el archivo: {e}"]

    filas: list[dict] = []
    problemas: list[str] = []
    vistas: set[tuple[str, str]] = set()
    hojas_leidas = 0

    for ws in wb.worksheets:
        met_hoja = METRICA_DE.get(_norm(ws.title))
        f, p, leyo = _leer_hoja(ws, met_hoja, vistas)
        filas.extend(f)
        problemas.extend(p)
        hojas_leidas += 1 if leyo else 0

    if not hojas_leidas:
        return [], ["No se encontró ninguna hoja con la columna «País». "
                    "¿Es el archivo que bajaste de la app?"]
    if not filas and not problemas:
        problemas.append("El archivo no trae ninguna fila con país.")
    return filas, problemas


def _leer_hoja(ws, met_hoja: str | None, vistas: set) -> tuple[list[dict], list[str], bool]:
    filas_xl = list(ws.iter_rows(values_only=True))

    # ⚠️ Por ENCABEZADO, no por posición.
    encabezado = fila_enc = None
    for i, r in enumerate(filas_xl[:20]):
        vals = [_norm(str(x)) if x is not None else "" for x in r]
        if _norm(COL_PAIS) in vals:
            encabezado, fila_enc = vals, i
            break
    if encabezado is None:
        return [], [], False

    col = {n: i for i, n in enumerate(encabezado) if n}
    i_pais = col[_norm(COL_PAIS)]
    i_met = col.get(_norm(COL_METRICA))      # formato viejo, opcional
    meses_col = {m: col[_norm(n)] for m, n in enumerate(MESES, start=1) if _norm(n) in col}
    if not meses_col:
        return [], [f"La hoja «{ws.title}» no tiene ninguna columna de mes (Ene … Dic)."], True
    if i_met is None and met_hoja is None:
        return [], [f"No se sabe qué métrica es la hoja «{ws.title}». Renombrala a "
                    f"«Habitaciones» o «Pax»."], True

    filas, problemas = [], []
    for n, r in enumerate(filas_xl[fila_enc + 1:], start=fila_enc + 2):
        pais = str(r[i_pais]).strip() if i_pais < len(r) and r[i_pais] is not None else ""
        if not pais:
            continue

        metric = met_hoja
        if i_met is not None:
            cruda = str(r[i_met]).strip() if i_met < len(r) and r[i_met] is not None else ""
            metric = METRICA_DE.get(_norm(cruda))
            if metric is None:
                problemas.append(
                    f"«{ws.title}» fila {n}: «{cruda or '(vacío)'}» no es una "
                    f"métrica. Tiene que decir «Habitaciones» o «Pax».")
                continue

        if (pais, metric) in vistas:
            # Dos filas del mismo país y métrica: una pisaría a la otra sin
            # avisar, y el total quedaría mal sin que nada lo dijera.
            problemas.append(f"«{ws.title}» fila {n}: «{pais}» está repetido.")
            continue
        vistas.add((pais, metric))

        vals = [0.0] * 12
        for m, i in meses_col.items():
            v = _num(r[i]) if i < len(r) else None
            if v is None:
                continue
            if v < 0:
                problemas.append(
                    f"«{ws.title}» fila {n}: «{pais}» tiene {v} en {MESES[m - 1]}. "
                    f"No puede haber noches ni pax negativos.")
                continue
            vals[m - 1] = float(v)
        filas.append({"pais": pais, "metric": metric, "values": vals})

    return filas, problemas, True
