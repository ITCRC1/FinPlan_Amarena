"""Exportador GENÉRICO de cuadros a Excel, con el formato de la casa.

**Por qué existe.** El owner pidió que todos los cuadros de todos los tabs se
puedan bajar a Excel con formato profesional. Son ~47 pantallas sin exportación
y ~11 más que exportan mal. Escribir 47 exportadores a mano no es viable: cada
uno son 200 líneas y todos se desincronizan del estilo con el tiempo.

**Y hay una razón técnica que cierra la discusión.** Las 10 pantallas que hoy
bajan Excel lo hacen desde el navegador con `xlsx` (SheetJS Community), que **no
escribe estilos de celda**: negrita, relleno, bordes y formato de moneda son de
la edición paga. Con esa librería, «formato profesional» es imposible por más
código que se escriba. Por eso esto vive en el servidor, con `openpyxl`.

**El contrato.** La pantalla manda lo que YA tiene renderizado:

    {
      "titulo": "Big Picture — Budget 2027",
      "subtitulo": "Corcovado · USD",          # opcional
      "columnas": [
        {"label": "Concepto", "ancho": 42, "formato": "texto"},
        {"label": "2026",     "ancho": 14, "formato": "usd"},
        {"label": "Var %",    "ancho": 10, "formato": "pct"},
      ],
      "filas": [
        {"label": "Ingresos",       "nivel": 0, "es_total": True,  "valores": [1000, 0.12]},
        {"label": "  Habitaciones", "nivel": 1, "es_total": False, "valores": [800, 0.10]},
      ],
    }

Los valores van como NÚMERO, nunca como texto ya formateado. Es la diferencia
entre un Excel que se puede sumar y uno que no — hoy `/reports/summary` manda
`"$1,234.00"` como cadena y el archivo resultante no sirve para nada.

Un libro puede llevar varios cuadros: cada uno es su hoja. Las pantallas con
tabs (allocations tiene 12 cuadros, cash flow directo 6) bajan todo de una.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from app.export.excel_base import (
    C, align, border, fill, font, merged_header, nombre_de_hoja,
    set_col_widths, workbook_to_bytes,
)

# Mismos formatos que `pl_full_detail_excel`, que es la referencia del repo:
# negativo en rojo y entre paréntesis, y el cero NO se imprime — una grilla
# llena de ceros esconde las cifras que sí importan.
FORMATOS = {
    "usd":   '#,##0;[Red](#,##0);""',
    "usd2":  '#,##0.00;[Red](#,##0.00);""',
    "pct":   '0.0%;[Red](0.0%);""',
    "num":   '#,##0;[Red](#,##0);""',
    "num1":  '#,##0.0;[Red](#,##0.0);""',
    "texto": None,
}

FILA_TITULO = 1
FILA_SUBTITULO = 2
FILA_CABECERA = 4
PRIMERA_FILA = 5


def _kpis(ws, cuadro: dict, desde: int) -> int:
    """La franja de estadísticas, arriba del cuadro. Devuelve la fila siguiente.

    Owner, 2026-09-03: *«no están saliendo las estadísticas en cada tab»*.

    ⚠️ En la pantalla la franja se dibuja UNA vez arriba de los sub-tabs, así
    que se ve en todos. Acá **cada hoja se lee sola** —se imprime, se manda
    suelta— y sin las estadísticas al lado los montos no tienen contra qué
    leerse: 56.001 de ingreso con 132 noches vendidas dice algo muy distinto
    que con 400.

    Va en gris y compacta: es contexto, no el cuadro.
    """
    filas = cuadro.get("kpis") or []
    columnas = cuadro.get("kpis_columnas") or []
    if not filas or not columnas:
        return desde

    fila = desde
    c = ws.cell(fila, 1, "ESTADÍSTICAS")
    c.font = font(bold=True, size=9, color=C["navy_mid"])
    for i, col in enumerate(columnas, start=2):
        c = ws.cell(fila, i, col)
        c.font = font(bold=True, size=9, color=C["navy_mid"])
        c.alignment = align("right")
    fila += 1

    for f in filas:
        rot = str(f.get("label") or "")
        ws.cell(fila, 1, rot).font = font(size=9)
        # El formato lo decide el rótulo: la ocupación es un porcentaje y la
        # tarifa son dólares. Mandarlo por fila desde la pantalla sería una
        # tercera copia de la misma decisión.
        bajo = rot.lower()
        fmt = ("pct" if "%" in rot else
               "usd2" if ("adr" in bajo or "daily" in bajo or "revpar" in bajo
                          or "cuota" in bajo) else "num")
        for i, v in enumerate(f.get("valores") or [], start=2):
            celda = ws.cell(fila, i, v)
            celda.number_format = FORMATOS.get(fmt, FORMATOS["usd"])
            celda.alignment = align("right")
            celda.font = font(size=9)
        fila += 1
    return fila + 1          # una en blanco antes del cuadro


def _hoja(wb: Workbook, cuadro: dict, usados: set[str]):
    columnas = cuadro.get("columnas") or []
    filas = cuadro.get("filas") or []
    titulo = (cuadro.get("titulo") or "Cuadro").strip()
    n_col = max(1, len(columnas))

    ws = wb.create_sheet(nombre_de_hoja(cuadro.get("hoja") or titulo, usados))

    merged_header(ws, FILA_TITULO, 1, n_col, titulo, C["navy"], sz=13)
    if cuadro.get("subtitulo"):
        merged_header(ws, FILA_SUBTITULO, 1, n_col, cuadro["subtitulo"],
                      C["navy_mid"], sz=10)

    # ⚠️ La cabecera del cuadro se corre hacia abajo lo que ocupe la franja.
    # Las constantes de fila eran fijas; con la franja delante, escribir la
    # tabla en la fila 4 la pisaría.
    FILA_CABECERA = _kpis(ws, cuadro, FILA_SUBTITULO + 2)
    PRIMERA_FILA = FILA_CABECERA + 1

    for i, col in enumerate(columnas, start=1):
        c = ws.cell(FILA_CABECERA, i, col.get("label", ""))
        c.fill = fill(C["navy_mid"])
        c.font = font(bold=True, color=C["white"], size=10)
        # La primera columna es la etiqueta de la fila; el resto son números.
        c.alignment = align("left" if i == 1 else "center", wrap=True)
        c.border = border()

    for j, f in enumerate(filas):
        fila = PRIMERA_FILA + j
        es_total = bool(f.get("es_total"))
        nivel = int(f.get("nivel") or 0)

        # La jerarquía va con SANGRÍA de Excel, no con espacios dentro del texto.
        # Con espacios, ordenar la columna o copiarla a otro lado se lleva la
        # sangría puesta y el nivel deja de significar nada. Es lo que hace hoy
        # `/reports/expenses`, que simula la jerarquía con espacios.
        etiqueta = ws.cell(fila, 1, f.get("label", ""))
        etiqueta.font = font(bold=es_total)
        etiqueta.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=min(nivel, 8))
        etiqueta.border = border()
        if es_total:
            etiqueta.fill = fill(C["blue_header"])

        # La fila puede pisar el formato de la columna. Hace falta cuando un mismo
        # cuadro mezcla unidades en la misma columna — el bloque de drivers del
        # Big Picture tiene noches, ocupación % y ADR en dólares, una debajo de
        # otra. Sin esto habría que partirlo en tres cuadros.
        fmt_fila = f.get("formato")

        for i, valor in enumerate(f.get("valores") or [], start=2):
            if i > n_col:
                break
            celda = ws.cell(fila, i, valor)
            fmt = FORMATOS.get(fmt_fila or columnas[i - 1].get("formato") or "usd",
                               FORMATOS["usd"])
            if fmt:
                celda.number_format = fmt
            # El texto se alinea a la izquierda: una columna de nombres de cuenta
            # alineada a la derecha es ilegible. Pasa en las pantallas de mapeo,
            # que son casi todas de texto (cuenta · departamento · línea del P&L).
            celda.alignment = align("left" if isinstance(valor, str) else "right")
            celda.font = font(bold=es_total)
            celda.border = border()
            if es_total:
                celda.fill = fill(C["blue_header"])

    set_col_widths(ws, {i: (col.get("ancho") or (38 if i == 1 else 14))
                        for i, col in enumerate(columnas, start=1)})
    # Congelar la cabecera y la columna de etiquetas: sin esto, un cuadro de 12
    # meses obliga a adivinar qué fila se está mirando al llegar a diciembre.
    ws.freeze_panes = ws.cell(PRIMERA_FILA, 2)

    # ── Que imprima en UNA hoja ──────────────────────────────────────────────
    #
    # Owner, 2026-08-27: «el Excel debe ser en una sola página sin separar». Un
    # cuadro de 12 meses son 14 columnas: en vertical y sin ajuste, Excel lo
    # parte en tres o cuatro hojas y los meses quedan repartidos entre papeles
    # distintos. Un reporte partido no se puede leer ni mandar.
    #
    # `fitToPage` en `sheet_properties.pageSetUpPr` es OBLIGATORIO: sin él,
    # `fitToWidth`/`fitToHeight` quedan escritos en el archivo y Excel los
    # ignora — se ve bien en el XML y sale partido igual.
    #
    # `fitToHeight = 0` es «las hojas de alto que haga falta». Se usa 1 porque
    # el pedido es una sola hoja; un cuadro larguísimo sale con letra chica,
    # que es preferible a que se parta.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    # El área de impresión se acota a lo escrito: sin esto, una celda tocada
    # por accidente lejos de la tabla arrastra hojas en blanco.
    ultima = PRIMERA_FILA + max(0, len(filas)) - 1
    if ultima >= FILA_TITULO:
        ws.print_area = f"A{FILA_TITULO}:{get_column_letter(n_col)}{ultima}"
    return ws


def _indice(wb: Workbook, cuadros: list[dict], nombres: list[str]) -> None:
    """La portada del libro: qué trae y en qué hoja está cada cosa.

    Owner, 2026-09-03: *«que baje bien profesional y claro»*, pidiendo que el
    Excel traiga todos los sub-tabs «tal como Word».

    ⚠️ El Word tiene su página de CONTENIDO; un libro de doce hojas sin índice
    obliga a recorrer las pestañas de abajo una por una, y los nombres van
    cortados a 31 caracteres —«Profit & Loss Statement YTD JU»—, así que ni
    siquiera se leen enteros. El índice es donde el título completo cabe.

    Va PRIMERO y con los nombres tal como quedaron, no como se pidieron: si dos
    cuadros se llamaban parecido, el libro los desambiguó y el índice tiene que
    mostrar el nombre real de la pestaña o no sirve para encontrarla.
    """
    ws = wb.create_sheet("Índice", 0)
    merged_header(ws, 1, 1, 3, "CONTENIDO", C["navy"], sz=13)
    for i, rotulo in enumerate(("#", "Hoja", "Cuadro"), start=1):
        c = ws.cell(3, i, rotulo)
        c.fill = fill(C["navy_mid"])
        c.font = font(bold=True, color=C["white"], size=10)
        c.alignment = align("left")
        c.border = border()
    for j, (cuadro, hoja) in enumerate(zip(cuadros, nombres)):
        fila = 4 + j
        titulo = (cuadro.get("titulo") or "Cuadro").strip()
        sub = (cuadro.get("subtitulo") or "").strip()
        for i, valor in enumerate((j + 1, hoja, titulo + (f"  ·  {sub}" if sub else "")),
                                  start=1):
            c = ws.cell(fila, i, valor)
            c.alignment = align("left")
            c.border = border()
    set_col_widths(ws, {1: 5, 2: 34, 3: 88})
    ws.freeze_panes = ws.cell(4, 1)


def build_cuadros_workbook(cuadros: list[dict]) -> bytes:
    """Un libro con una hoja por cuadro, y un índice adelante."""
    wb = Workbook()
    wb.remove(wb.active)
    usados: set[str] = set()
    nombres: list[str] = []
    for cuadro in cuadros or []:
        nombres.append(_hoja(wb, cuadro, usados).title)
    # ⚠️ El índice sólo cuando hay VARIAS hojas. En un libro de una, una portada
    # que dice «1. esa hoja» es un clic de más para llegar al único cuadro.
    if len(nombres) > 1:
        _indice(wb, cuadros or [], nombres)
    if not wb.sheetnames:            # nunca devolver un libro sin hojas
        wb.create_sheet("Sin datos")
    return workbook_to_bytes(wb)
