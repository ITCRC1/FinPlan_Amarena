"""Excel del Channel Mix: reporte de 4 pestañas y plantilla de ida y vuelta.

**El pedido (owner, 2026-08-18).** «Quiero tener las estadísticas mes por mes
por market codes y channel, total rooms y total pax. Quisiera bajar el Excel y
ver todos los números de ambas cosas por mes y en diferente tab. Igual quisiera
bajar y subir el archivo, tal como country.»

Son DOS archivos distintos y conviene no confundirlos:

**El reporte** (`⬇ Excel`) — cuatro pestañas, solo para mirar:

    Canal · Habitaciones   Canal · Pax
    Market code · Habitaciones   Market code · Pax

**La plantilla** (`⬇ Plantilla` / `⬆ Plantilla`) — dos pestañas, editable:

    Habitaciones   Pax        (filas = market code)

⚠️ **La plantilla es SOLO por market code, a propósito.** El canal se deriva del
código con la tabla `market_codes`: si además se pudiera editar el canal, el
resumen podría terminar contradiciendo a su propio detalle y ninguno de los dos
sería la verdad. Se corrige el átomo y el canal se recalcula.

⚠️ Las columnas se ubican **por encabezado, nunca por posición** — la norma del
proyecto: el owner trabaja sobre estos archivos.
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
COL_CODE = "Market code"
COL_CANAL = "Canal"

HOJA_DE = {"rooms": "Habitaciones", "pax": "Pax"}
METRICA_DE = {"habitaciones": "rooms", "rooms": "rooms", "noches": "rooms",
              "pax": "pax", "huespedes": "pax", "huéspedes": "pax"}

_AZUL = "FF2D3A5C"


def _encabezado(ws, cols: list[str], fila: int = 4) -> None:
    for j, nombre in enumerate(cols, start=1):
        c = ws.cell(row=fila, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=_AZUL)
        c.alignment = Alignment(horizontal="center")


def _hoja(wb, titulo: str, subtitulo: str, primera: str, filas: list[dict],
          con_canal: bool) -> None:
    """`filas` = [{label, canal, values:[12]}]. Agrega TOTAL al final."""
    ws = wb.create_sheet(titulo[:31])
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(size=9, italic=True, color="FF808080")

    cols = [primera] + ([COL_CANAL] if con_canal else []) + MESES + ["Total"]
    _encabezado(ws, cols)

    i = 5
    for f in filas:
        ws.cell(row=i, column=1, value=f["label"])
        c0 = 2
        if con_canal:
            ws.cell(row=i, column=2, value=f.get("canal") or "— sin canal —")
            c0 = 3
        for m in range(12):
            v = f["values"][m] if m < len(f["values"]) else 0
            celda = ws.cell(row=i, column=c0 + m, value=float(v or 0))
            celda.number_format = "#,##0"
        t = ws.cell(row=i, column=c0 + 12, value=float(sum(f["values"])))
        t.number_format = "#,##0"
        t.font = Font(bold=True)
        i += 1

    # El TOTAL de la hoja. Sin esto hay que sumarlo a mano para cruzarlo contra
    # el On the Books, que es justo el control que hace falta.
    ws.cell(row=i, column=1, value="TOTAL").font = Font(bold=True)
    c0 = 3 if con_canal else 2
    for m in range(12):
        s = sum((f["values"][m] if m < len(f["values"]) else 0) for f in filas)
        c = ws.cell(row=i, column=c0 + m, value=float(s))
        c.number_format = "#,##0"
        c.font = Font(bold=True)
    c = ws.cell(row=i, column=c0 + 12, value=float(sum(sum(f["values"]) for f in filas)))
    c.number_format = "#,##0"
    c.font = Font(bold=True)

    ws.column_dimensions["A"].width = 26
    if con_canal:
        ws.column_dimensions["B"].width = 18
    for j in range(c0, c0 + 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 9
    ws.freeze_panes = ws.cell(row=5, column=c0).coordinate


def construir_reporte(titulo: str, por_canal: dict, por_code: dict,
                      canal_de: dict) -> openpyxl.Workbook:
    """Las CUATRO pestañas del reporte. `por_*` = {(clave, metric): [12]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for metric, etiqueta in HOJA_DE.items():
        canales = sorted({k for k, m in por_canal if m == metric})
        _hoja(wb, f"Canal · {etiqueta}", titulo,
              "Canal",
              [{"label": c, "values": por_canal[(c, metric)]} for c in canales],
              con_canal=False)
    for metric, etiqueta in HOJA_DE.items():
        codes = sorted({k for k, m in por_code if m == metric})
        _hoja(wb, f"Market code · {etiqueta}", titulo,
              COL_CODE,
              [{"label": c, "canal": canal_de.get(c, ""), "values": por_code[(c, metric)]}
               for c in codes],
              con_canal=True)
    if not wb.sheetnames:
        wb.create_sheet("Sin datos")
    return wb


def construir_plantilla(titulo: str, filas: list[dict], canal_de: dict) -> openpyxl.Workbook:
    """La plantilla EDITABLE: una hoja por métrica, filas = market code.

    `filas` = [{code, metric, values:[12]}].
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for metric, hoja in HOJA_DE.items():
        ws = wb.create_sheet(hoja)
        ws["A1"] = f"{titulo} — {hoja}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = ("Corregí los números y volvé a subir este mismo archivo. Cada "
                    "pestaña es una métrica. La columna «Canal» es informativa: "
                    "sale de Master Data → Market Codes y no se lee al subir — el "
                    "canal se recalcula del código, para que el resumen no pueda "
                    "contradecir a su propio detalle.")
        ws["A2"].font = Font(size=9, italic=True, color="FF808080")
        _encabezado(ws, [COL_CODE, COL_CANAL] + MESES)

        i = 5
        for f in filas:
            if f["metric"] != metric:
                continue
            ws.cell(row=i, column=1, value=f["code"])
            c = ws.cell(row=i, column=2, value=canal_de.get(f["code"], "") or "— sin canal —")
            c.font = Font(color="FF808080")
            for m in range(12):
                v = f["values"][m] if m < len(f["values"]) else 0
                celda = ws.cell(row=i, column=3 + m, value=float(v or 0))
                celda.number_format = "#,##0"
            i += 1

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 18
        for j in range(3, 15):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 9
        ws.freeze_panes = "C5"
    return wb


def _num(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _norm(s: str) -> str:
    t = (s or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        t = t.replace(a, b)
    return t


def leer_plantilla(data: bytes) -> tuple[list[dict], list[str]]:
    """Lee la plantilla corregida. Devuelve `({code, metric, values}, problemas)`.

    La métrica sale del NOMBRE DE LA HOJA. La columna «Canal» se ignora a
    propósito: es informativa, y el canal se recalcula del código.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"No se pudo abrir el archivo: {e}"]

    filas: list[dict] = []
    problemas: list[str] = []
    vistas: set[tuple[str, str]] = set()
    hojas = 0

    for ws in wb.worksheets:
        metric = METRICA_DE.get(_norm(ws.title))
        filas_xl = list(ws.iter_rows(values_only=True))

        # ⚠️ Por ENCABEZADO, no por posición.
        encabezado = fila_enc = None
        for i, r in enumerate(filas_xl[:20]):
            vals = [_norm(str(x)) if x is not None else "" for x in r]
            if _norm(COL_CODE) in vals:
                encabezado, fila_enc = vals, i
                break
        if encabezado is None:
            continue
        hojas += 1
        if metric is None:
            problemas.append(f"No se sabe qué métrica es la hoja «{ws.title}». "
                             f"Renombrala a «Habitaciones» o «Pax».")
            continue

        col = {n: i for i, n in enumerate(encabezado) if n}
        i_code = col[_norm(COL_CODE)]
        meses_col = {m: col[_norm(n)] for m, n in enumerate(MESES, start=1) if _norm(n) in col}
        if not meses_col:
            problemas.append(f"La hoja «{ws.title}» no tiene ninguna columna de mes.")
            continue

        for n, r in enumerate(filas_xl[fila_enc + 1:], start=fila_enc + 2):
            code = str(r[i_code]).strip().upper() if i_code < len(r) and r[i_code] is not None else ""
            if not code or code == "TOTAL":
                continue
            if (code, metric) in vistas:
                problemas.append(f"«{ws.title}» fila {n}: «{code}» está repetido.")
                continue
            vistas.add((code, metric))
            vals = [0.0] * 12
            for m, i in meses_col.items():
                v = _num(r[i]) if i < len(r) else None
                if v is None:
                    continue
                if v < 0:
                    problemas.append(
                        f"«{ws.title}» fila {n}: «{code}» tiene {v} en {MESES[m - 1]}. "
                        f"No puede haber noches ni pax negativos.")
                    continue
                vals[m - 1] = float(v)
            filas.append({"code": code, "metric": metric, "values": vals})

    if not hojas:
        return [], ["No se encontró ninguna hoja con la columna «Market code». "
                    "¿Es el archivo que bajaste de la app?"]
    if not filas and not problemas:
        problemas.append("El archivo no trae ninguna fila con market code.")
    return filas, problemas
