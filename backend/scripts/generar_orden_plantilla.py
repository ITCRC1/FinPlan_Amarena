# -*- coding: utf-8 -*-
"""Convierte el Excel del owner en el ORDEN CANONICO de la plantilla.

Owner (2026-08-14): «solo acomoda el archivo de upload en este formato. Mismo
orden de los departamentos, mismo orden interno. Para Rooms sale ingreso y sus
cuentas, costo y sus cuentas, planilla y sus cuentas y opex y sus cuentas. Y
despues sigues con F&B, Spa, y asi hasta terminar.»

**Por que es una LISTA y no una regla.** El orden de las clases CAMBIA segun el
departamento:

    Rooms  -> Ingreso, Planilla, Opex          (no tiene costo)
    A&B    -> Ingreso, Costo, Planilla, Opex
    Tours  -> Ingreso, Planilla, Costo, Opex   (el costo va DESPUES de planilla)

No hay regla que produzca eso. Hay un archivo. Tres intentos anteriores fallaron
buscando la regla; este lo lee.

    python -m scripts.generar_orden_plantilla "C:/ruta/ORDEN PARA EL UPLOAD.xlsx"

Escribe `app/seed_data/orden_plantilla.json`. Se corre a mano cuando el owner
cambie su estructura, no en cada arranque: el orden es un dato del negocio y
tiene que quedar versionado en git para poder ver que cambio.
"""
import json
import pathlib
import re
import sys
import unicodedata

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

RAIZ = pathlib.Path(__file__).resolve().parents[1]
SALIDA = RAIZ / "app" / "seed_data" / "orden_plantilla.json"

#: Confirmado por el owner (2026-08-14): las primeras filas del archivo van
#: ANTES del primer encabezado y son el ingreso de habitaciones.
DEPTO_INICIAL = "0110"

#: Confirmado por el owner: RENT, PROPERTY INSURANCE, EXCHANGE GAIN/LOSSES e
#: INCOME TAX no son departamentos sino lineas de below-GOP, y en la app viven
#: todas en Property Expenses.
SECCIONES_BELOW_GOP = {
    "rent": "0250",
    "property insurance1": "0250",
    "property insurance": "0250",
    "exchange gain/losses": "0250",
    "income tax": "0250",
}

CLASE_POR_PREFIJO = {"4": "Revenue", "5": "Cost", "6": "Payroll",
                     "7": "Opex", "8": "BelowGOP", "9": "Stat"}


def norm(t) -> str:
    t = unicodedata.normalize("NFD", str(t or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def leer(ruta: str, mapa_nombres: dict) -> list[dict]:
    import openpyxl

    ws = openpyxl.load_workbook(ruta, data_only=True).worksheets[0]
    fuera, actual, sin_resolver = [], DEPTO_INICIAL, {}
    for r in range(1, ws.max_row + 1):
        a, d, c = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if d is None or c is None:
            t = str(a or "").strip()
            if not t or t.upper().startswith("TOTAL"):
                continue
            n = norm(t)
            dc = mapa_nombres.get(n) or SECCIONES_BELOW_GOP.get(n)
            if dc:
                actual = dc
            else:
                sin_resolver[t] = sin_resolver.get(t, 0) + 1
            continue
        cta = str(c).strip()
        if not re.fullmatch(r"\d{4}", cta):
            continue
        fuera.append({"dept_code": actual, "cuenta": cta,
                      "clase": CLASE_POR_PREFIJO.get(cta[0], "?")})
    if sin_resolver:
        print("  OJO: encabezados sin departamento (se heredo el anterior):")
        for t, n in sin_resolver.items():
            print(f"       {t}  x{n}")
    return fuera


async def main(ruta: str):
    from scripts._prodenv import usar_produccion
    usar_produccion()
    from sqlalchemy import text
    from app.db import get_session

    async with get_session() as s:
        mapa = {}
        for sd, dc in (await s.execute(text(
            "SELECT DISTINCT source_department, dept_code FROM account_mapping "
            "WHERE active_status='YES' AND dept_code IS NOT NULL"))).all():
            mapa.setdefault(norm(sd), dc)

    filas = leer(ruta, mapa)
    # (depto, cuenta) repetido: manda la PRIMERA aparicion, que es donde el owner
    # la puso. Repetirla mas abajo no la mueve.
    visto, unicas = set(), []
    for f in filas:
        k = (f["dept_code"], f["cuenta"])
        if k in visto:
            continue
        visto.add(k)
        unicas.append(f)

    SALIDA.write_text(json.dumps({
        "_nota": [
            "El ORDEN de la plantilla del Detalle, tal como lo mando el owner",
            "(ORDEN PARA EL UPLOAD.xlsx, 2026-08-14).",
            "",
            "Es una LISTA y no una regla: el orden de las clases cambia segun el",
            "departamento (Rooms no tiene costo; en Tours el costo va DESPUES de",
            "planilla). No hay regla que produzca eso.",
            "",
            "Se regenera con `python -m scripts.generar_orden_plantilla <xlsx>`.",
            "Lo que no este en esta lista sale despues, por departamento y cuenta.",
        ],
        "orden": unicas,
    }, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")

    deptos = list(dict.fromkeys(f["dept_code"] for f in unicas))
    print(f"\n  {len(unicas)} pares (depto, cuenta) en {len(deptos)} departamentos")
    print(f"  -> {SALIDA.relative_to(RAIZ)}")
    print("\n  orden de departamentos:")
    print("   ", " ".join(deptos))


if __name__ == "__main__":
    import asyncio
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    asyncio.run(main(sys.argv[1]))
