# -*- coding: utf-8 -*-
"""Baja el catálogo de cuentas estadísticas a Excel para revisarlo.

El owner las revisa en Excel, no en una pantalla (2026-08-14). Se genera desde
`seed_data/stats_catalog.json`, que es la lista de verdad — no desde la base, así
que sirve también para revisar un cambio antes de desplegarlo.

    python -m scripts.excel_catalogo_estadisticas [salida.xlsx]
"""
import sys
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from app.seed_stats import leer_catalogo, ARCHIVO  # noqa: E402

AZUL = "2D3A5C"
VERDE = "1A7F4B"
ROJO = "C0392B"
GRIS = "F2F4F7"

DIM_ES = {
    "DEPT": "Departamento",
    "POSITION": "Posición",
    "ROOMTYPE": "Tipo de habitación",
    "CHANNEL": "Canal",
    "COUNTRY": "País",
    "SEGMENT": "Segmento de mercado",
    "OUTLET": "Punto de venta",
}

UNIDAD_ES = {
    "rooms": "habitaciones", "nights": "noches", "pax": "personas",
    "covers": "covers", "treatments": "tratamientos", "kilos": "kilos",
    "hours": "horas", "count": "cantidad", "trips": "viajes",
    "fte": "FTE", "usd": "dólares",
}

GRUPO_ES = {
    "9000": "Habitaciones y comercial",
    "9110": "Alimentos y bebidas",
    "9201": "Spa",
    "9400": "Tours y actividades",
    "9500": "Transporte y bote",
    "9600": "Innoceana",
    "9700": "Lavandería",
    "9900": "Planilla — personas",
    "9980": "Planilla — horas",
}

_borde = Border(*[Side(style="thin", color="D0D5DD")] * 4)


def _encabezado(ws, fila, titulos, color=AZUL):
    for i, t in enumerate(titulos, start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde
    ws.row_dimensions[fila].height = 32


def _hoja_cuentas(wb, cuentas):
    ws = wb.create_sheet("Cuentas")
    ws["A1"] = "Cuentas estadísticas — clase 9"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=AZUL)
    ws["A2"] = (f"{len(cuentas)} cuentas. Los rangos siguen CLAUDE.md §18.1. "
                "Los nombres se pueden cambiar; los códigos no se mueven.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="667085")

    # Sin columnas de dinero: acá no entra plata (owner, 2026-08-14).
    cols = ["Cuenta", "Grupo", "Nombre", "Name (EN)", "Qué cuenta",
            "Se abre por", "El año es", "Ya existía en"]
    _encabezado(ws, 4, cols)

    fila = 5
    grupo_actual = None
    for c in cuentas:
        if c["grupo"] != grupo_actual:
            grupo_actual = c["grupo"]
            ws.cell(row=fila, column=1,
                    value=f"{grupo_actual} · {GRUPO_ES.get(grupo_actual, '')}")
            for i in range(1, len(cols) + 1):
                cel = ws.cell(row=fila, column=i)
                cel.fill = PatternFill("solid", fgColor=VERDE)
                cel.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                cel.border = _borde
            fila += 1

        dims = ", ".join(DIM_ES.get(d, d) for d in c.get("dims", []))
        vals = [
            c["code"],
            GRUPO_ES.get(c["grupo"], c["grupo"]),
            c["nombre_es"],
            c.get("nombre_en", ""),
            UNIDAD_ES.get(c["unidad"], c["unidad"]),
            dims or "— (total del hotel)",
            "la suma de los 12 meses" if c.get("agrega", "SUM") == "SUM"
            else "el saldo de diciembre",
            c.get("legado", ""),
        ]
        for i, v in enumerate(vals, start=1):
            cel = ws.cell(row=fila, column=i, value=v)
            cel.border = _borde
            cel.alignment = Alignment(vertical="center", wrap_text=(i in (3, 4, 6, 7)))
            cel.font = Font(name="Calibri", size=10)
            if i == 1:
                cel.font = Font(name="Calibri", size=11, bold=True)
                cel.alignment = Alignment(horizontal="center", vertical="center")
            if c.get("legado") and i == 8:
                cel.font = Font(name="Calibri", size=10, bold=True, color=VERDE)
        fila += 1

    for col, ancho in zip("ABCDEFGH", (10, 24, 46, 36, 16, 32, 26, 18)):
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{fila - 1}"
    return ws


def _hoja_dimensiones(wb):
    ws = wb.create_sheet("Dimensiones")
    ws["A1"] = "Dimensiones — por qué se puede abrir una estadística"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=AZUL)
    ws["A2"] = ("Cada cuenta declara cuáles acepta. La carga RECHAZA cualquier "
                "otra: una cuenta de kilos con un código de canal adentro es un "
                "error de digitación, no un dato.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="667085")

    _encabezado(ws, 4, ["Dimensión", "Qué es", "De dónde salen los códigos"])
    detalle = [
        ("Departamento", "0110 Habitaciones, 0120 A&B, 0140 Spa…",
         "El catálogo de departamentos que ya usa todo el sistema"),
        ("Posición", "El código de posición de planilla, tipo 0111-01",
         "Las posiciones del escenario"),
        ("Tipo de habitación", "BL01, BI02, PO03…",
         "Código FIJO de Master Data. No cambia aunque se renombre la villa"),
        ("Canal", "Canal de venta",
         "⚠️ Hoy hay DOS listas distintas conviviendo. Falta decidir cuál manda"),
        ("País", "País de procedencia del huésped",
         "Lista abierta, la que ya usa el mix por país"),
        ("Segmento de mercado", "Market code: 001 Retail, 005 Wholesale, 017 House Use…",
         "CLAUDE.md §18.2 — 13 segmentos. Todavía no está construido"),
        ("Punto de venta", "El outlet de A&B",
         "La columna que ya existe en el detalle del mayor"),
    ]
    for i, (a, b, c) in enumerate(detalle, start=5):
        for j, v in enumerate((a, b, c), start=1):
            cel = ws.cell(row=i, column=j, value=v)
            cel.border = _borde
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            cel.font = Font(name="Calibri", size=10,
                            bold=(j == 1), color=(ROJO if "⚠️" in c and j == 3 else "000000"))
        ws.row_dimensions[i].height = 30
    for col, ancho in zip("ABC", (24, 52, 62)):
        ws.column_dimensions[col].width = ancho
    return ws


def _hoja_notas(wb, cuentas):
    ws = wb.create_sheet("Léame")
    ws.column_dimensions["A"].width = 110
    lineas = [
        ("Qué es este archivo", "t"),
        ("El catálogo de cuentas estadísticas que quedó sembrado en producción "
         "el 2026-08-14. Son cuentas de CANTIDADES —noches, covers, kilos, "
         "horas, personas— no de dinero.", "p"),
        ("", "p"),
        ("Los códigos no se mueven", "t"),
        ("El número de cuenta es la llave. El nombre se puede cambiar cuando "
         "quieras; el código, no. Es la misma regla de los códigos de tipo de "
         "habitación.", "p"),
        ("", "p"),
        ("Acá no entra dinero", "t"),
        ("Decisión del owner (2026-08-14): solo cantidades. La primera versión "
         "traía tres cuentas de venta de habitaciones —por canal, por país y "
         "por segmento— y se quitaron.", "p"),
        ("Es lo correcto. Una venta abierta por canal es la MISMA plata que el "
         "P&L ya reporta, partida de otra forma; el día que la suma no diera "
         "igual habría dos verdades sobre el mismo número y ninguna avisaría. "
         "Cuadrarlas con una prueba mitigaba el riesgo; no tenerlas lo elimina.", "p"),
        ("El sistema ahora RECHAZA cualquier cuenta estadística en dólares.", "p"),
        ("", "p"),
        ("Qué falta decidir (owner)", "t"),
        ("1. Cuál de las dos listas de canales manda. Hoy conviven "
         "«Travel Agency / Directo + Web / OTA / Otros» y «TA / OTA / Directo». "
         "Aplica a 9070 y 9071 (habitaciones y pax por canal).", "p"),
        ("2. Los 13 segmentos de mercado del §18.2 no están construidos.", "p"),
        ("3. Si algún nombre de cuenta no es el que usás en la operación, "
         "cambialo acá y se ajusta.", "p"),
        ("", "p"),
        ("Estado", "t"),
        (f"{len(cuentas)} cuentas creadas y vivas en producción. Las tablas "
         "nacen VACÍAS: todavía no se ha cargado un solo dato. El archivo para "
         "subirlos es el siguiente paso.", "p"),
        (f"Fuente: {ARCHIVO.name}", "p"),
    ]
    for i, (txt, tipo) in enumerate(lineas, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if tipo == "t":
            c.font = Font(name="Calibri", bold=True, size=13,
                          color=(ROJO if txt.startswith("⚠️") else AZUL))
        else:
            c.font = Font(name="Calibri", size=11)
        ws.row_dimensions[i].height = 18 if tipo == "t" else 32
    return ws


def main(salida: str):
    cuentas = leer_catalogo()
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_notas(wb, cuentas)
    _hoja_cuentas(wb, cuentas)
    _hoja_dimensiones(wb)
    wb.save(salida)
    # Texto plano a proposito: la consola de Windows es cp1252 y un caracter
    # bonito acá tumba el script DESPUES de haber hecho el trabajo (ya pasó con
    # la carga del 8060, que dejó el escenario a medias por un emoji).
    print(f"{len(cuentas)} cuentas -> {salida}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "CATALOGO_ESTADISTICAS.xlsx")
