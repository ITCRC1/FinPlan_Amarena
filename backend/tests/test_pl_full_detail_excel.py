# -*- coding: utf-8 -*-
"""El export a Excel del P&L Full Detail.

Lo que vigilan estas pruebas es que el archivo diga lo MISMO que la pantalla —
el exportador no recalcula nada— y que se respete el formato del Excel del
owner: negativos entre paréntesis y en rojo, el cero sin imprimir, y los ratios
como porcentaje y no como plata (en el original tenían formato de moneda y se
leían `$0.35` en vez de `35.0%`).
"""
import io

import openpyxl
import pytest

from app.export.pl_full_detail_excel import (
    FMT_PCT, FMT_USD, build_pl_full_detail_workbook,
)


def _fila(tipo, etiqueta, total, cuenta="", nivel=1, meses=None):
    meses = meses if meses is not None else [total / 12] * 12
    return {"tipo": tipo, "nivel": nivel, "cuenta": cuenta, "etiqueta": etiqueta,
            "clave": f"{tipo}|{cuenta}|{etiqueta}", "meses": meses, "total": total}


@pytest.fixture
def payload():
    return {
        "scenario_id": "x", "scenario": "BUDGET Working 2027", "year": 2027,
        "source_mode": "checkbook", "moneda": "USD",
        "avisos": ["Este escenario no trae el ingreso abierto por cuenta."],
        "kpis": {
            "disponible": True, "diluyen": ["Villas"],
            "sets": [{
                "clave": "STANDARD", "nombre": "Rooms Standard", "unidades": 30,
                "noches_disponibles": [300] * 12, "noches_ocupadas": [150] * 12,
                "revenue": [90000.0] * 12, "ocupacion": [0.5] * 12,
                "adr": [600.0] * 12, "revpar": [300.0] * 12,
                "ocupacion_anual": 0.5, "adr_anual": 600.0, "revpar_anual": 300.0,
                "revenue_anual": 1080000.0, "sin_ocupacion": False,
            }, {
                "clave": "0115", "nombre": "Villas", "unidades": 2,
                "noches_disponibles": [20] * 12, "noches_ocupadas": [0] * 12,
                "revenue": [0.0] * 12, "ocupacion": [0.0] * 12,
                "adr": [0.0] * 12, "revpar": [0.0] * 12,
                "ocupacion_anual": 0.0, "adr_anual": 0.0, "revpar_anual": 0.0,
                "revenue_anual": 0.0, "sin_ocupacion": True,
            }],
            "consolidado": {
                "clave": "TOTAL", "nombre": "Rooms (consolidado)", "unidades": 32,
                "noches_disponibles": [320] * 12, "noches_ocupadas": [150] * 12,
                "revenue": [90000.0] * 12, "ocupacion": [0.47] * 12,
                "adr": [600.0] * 12, "revpar": [281.0] * 12,
                "ocupacion_anual": 0.47, "adr_anual": 600.0, "revpar_anual": 281.0,
                "revenue_anual": 1080000.0,
            },
        },
        "resumen": [_fila("total", "TOTAL REVENUES", 1080000.0, nivel=0)],
        "bloques": [{
            "clave": "0110", "dept_code": "0110", "titulo": "Rooms / Habitaciones",
            "titulo_en": "Rooms", "tipo": "OPERATIVO",
            "ingreso_anual": 1080000.0, "gasto_anual": 480000.0, "utilidad_anual": 600000.0,
            "filas": [
                _fila("seccion", "INGRESOS", 0.0, nivel=0, meses=[0.0] * 12),
                _fila("detalle", "Room Revenue", 1080000.0, cuenta="4000"),
                _fila("subtotal", "Total Ingresos", 1080000.0),
                _fila("total", "UTILIDAD NETA", 600000.0, nivel=0),
                {"tipo": "pct", "nivel": 2, "cuenta": "", "etiqueta": "% Utilidad",
                 "clave": "pct|utilidad", "meses": [0.5555] * 12, "total": 0.5555},
                _fila("detalle", "Una pérdida", -1200.0, cuenta="7065"),
            ],
        }],
        "propiedad": [_fila("detalle", "Depreciation", 312000.0, cuenta="8040")],
        "cuadre": {
            "ingresos_detalle": 0.0, "ingresos_pl": 5997345.87,
            "dif_ingresos": -5997345.87, "gastos_detalle": 3650126.54,
            "gastos_pl": 3650126.54, "dif_gastos": 0.0,
            "gop_pl": 2347219.33, "net_pl": 833494.39,
            "ingreso_por_cuenta": False, "ok": True,
        },
    }


@pytest.fixture
def wb(payload):
    return openpyxl.load_workbook(io.BytesIO(build_pl_full_detail_workbook(payload)))


# ── Estructura ────────────────────────────────────────────────────────────────

def test_el_cuadre_va_primero(wb):
    """Si el detalle no amarra, eso es lo primero que hay que ver — no algo
    escondido en la última hoja."""
    assert wb.sheetnames[0] == "CUADRE"


def test_hay_una_hoja_por_bloque(wb):
    assert "RESUMEN" in wb.sheetnames
    assert "HABITACIONES POR SET" in wb.sheetnames
    assert "GASTOS DE PROPIEDAD" in wb.sheetnames
    assert any(n.startswith("0110") for n in wb.sheetnames)


def test_ningun_nombre_de_hoja_pasa_de_31_caracteres(wb):
    """Excel rechaza el libro entero si se pasa, o si dos hojas se llaman igual."""
    assert all(len(n) <= 31 for n in wb.sheetnames)
    assert len(set(wb.sheetnames)) == len(wb.sheetnames)


def test_dos_departamentos_de_nombre_largo_no_colapsan_en_la_misma_hoja():
    largo = "Departamento de Alimentos y Bebidas del Restaurante Principal"
    data = {
        "scenario": "X", "moneda": "USD", "source_mode": "checkbook", "avisos": [],
        "kpis": {}, "resumen": [], "propiedad": [],
        "cuadre": {"ingresos_detalle": 0, "ingresos_pl": 0, "dif_ingresos": 0,
                   "gastos_detalle": 0, "gastos_pl": 0, "dif_gastos": 0,
                   "gop_pl": 0, "net_pl": 0, "ingreso_por_cuenta": True, "ok": True},
        "bloques": [
            {"clave": d, "dept_code": d, "titulo": largo, "titulo_en": "", "tipo": "OPERATIVO",
             "ingreso_anual": 0, "gasto_anual": 0, "utilidad_anual": 0,
             "filas": [_fila("detalle", "x", 1.0)]}
            for d in ("0120", "0121")
        ],
    }
    libro = openpyxl.load_workbook(io.BytesIO(build_pl_full_detail_workbook(data)))
    assert len(set(libro.sheetnames)) == len(libro.sheetnames)


# ── Formato: lo que se corrige del Excel original ────────────────────────────

def test_los_ratios_van_como_porcentaje_no_como_plata(wb):
    """En el Excel del owner «% Utilidad» tenía formato de MONEDA: se leía
    `$0.56` en vez de `55.6%`."""
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("0110"))
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "% Utilidad")
    assert ws.cell(row=fila, column=15).number_format == FMT_PCT
    assert ws.cell(row=fila, column=3).number_format == FMT_PCT


def test_la_plata_lleva_parentesis_rojo_y_no_imprime_el_cero(wb):
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("0110"))
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "Room Revenue")
    fmt = ws.cell(row=fila, column=15).number_format
    assert fmt == FMT_USD
    assert "[Red]" in fmt and "(" in fmt and fmt.endswith('""')


def test_el_cero_se_escribe_vacio_no_como_cero(wb):
    """Una grilla de 700 filas llena de ceros no se lee. Es la convención del
    Excel original y acá se respeta."""
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("0110"))
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "INGRESOS")
    assert ws.cell(row=fila, column=3).value is None


def test_los_numeros_van_a_la_derecha(wb):
    """En el original estaban centrados. Una columna de cifras se lee alineada
    a la derecha."""
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("0110"))
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "Room Revenue")
    assert ws.cell(row=fila, column=3).alignment.horizontal == "right"


def test_el_negativo_llega_negativo(wb):
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("0110"))
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "Una pérdida")
    assert ws.cell(row=fila, column=15).value == -1200.0


# ── El Excel dice lo mismo que la pantalla ───────────────────────────────────

def test_el_cuadre_copia_los_numeros_del_payload(wb, payload):
    ws = wb["CUADRE"]
    c = payload["cuadre"]
    vistos = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
              for r in range(1, ws.max_row + 1)}
    assert vistos["Ingresos"] == c["ingresos_detalle"]
    assert vistos["Gastos (operativos + overhead)"] == c["gastos_detalle"]
    assert vistos["GOP"] == c["gop_pl"]
    assert vistos["Utilidad neta"] == c["net_pl"]


def test_el_aviso_del_ingreso_sin_apertura_sale_en_el_archivo(wb):
    ws = wb["CUADRE"]
    texto = " ".join(str(ws.cell(row=r, column=1).value or "")
                     for r in range(1, ws.max_row + 1))
    assert "NO está abierto por cuenta" in texto
    assert "ingreso abierto por cuenta" in texto   # el aviso del payload


def test_el_set_sin_ocupacion_queda_marcado(wb):
    ws = wb["HABITACIONES POR SET"]
    texto = " ".join(str(ws.cell(row=r, column=1).value or "")
                     for r in range(1, ws.max_row + 1))
    assert "sin ocupación cargada" in texto
    assert "diluyen la ocupación general" in texto


def test_el_total_anual_del_archivo_es_la_suma_de_los_doce_meses(wb):
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("0110"))
    fila = next(r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=2).value == "Room Revenue")
    meses = sum(ws.cell(row=fila, column=3 + i).value or 0 for i in range(12))
    assert round(meses, 2) == round(ws.cell(row=fila, column=15).value, 2)
