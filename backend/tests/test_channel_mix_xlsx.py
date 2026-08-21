# -*- coding: utf-8 -*-
"""El Excel del Channel Mix: 4 pestañas para mirar, 2 para editar.

**El pedido (owner, 2026-08-18).** «Quiero tener las estadísticas mes por mes
por market codes y channel, total rooms y total pax. Quisiera bajar el Excel y
ver todos los números de ambas cosas por mes y en diferente tab. Igual quisiera
bajar y subir el archivo, tal como country.»

Son dos archivos distintos:

- **Reporte** (`⬇ Excel`): 4 pestañas — Canal · Habitaciones, Canal · Pax,
  Market code · Habitaciones, Market code · Pax. Solo para mirar.
- **Plantilla** (`⬇`/`⬆ Plantilla`): 2 pestañas, filas = market code. Editable.

⚠️ **La plantilla es SOLO por market code, a propósito.** El canal se deriva del
código con `market_codes`; si además se pudiera editar el canal, el resumen
podría terminar contradiciendo a su propio detalle y ninguno de los dos sería la
verdad. Se corrige el átomo y el canal se recalcula.
"""
import inspect
import io

import openpyxl
import pytest

from app.export.channel_mix_xlsx import (
    MESES, construir_plantilla, construir_reporte, leer_plantilla,
)


def _bytes(wb) -> bytes:
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def _clave(filas):
    """El contenido, sin depender del orden (sale agrupado por pestaña)."""
    return sorted((f["code"], f["metric"], tuple(float(x) for x in f["values"]))
                  for f in filas)


CANAL_DE = {"TA": "Travel Agent", "TAFIT": "Travel Agent", "WEB": "Website",
            "CORP": ""}
POR_CODE = {
    ("TA", "rooms"): [530, 538, 531, 211, 0, 87, 0, 0, 0, 0, 0, 0],
    ("TA", "pax"): [975, 1000, 1007, 438, 0, 153, 0, 0, 0, 0, 0, 0],
    ("WEB", "rooms"): [59, 37, 73, 56, 0, 26, 0, 0, 0, 0, 0, 0],
}
POR_CANAL = {
    ("Travel Agent", "rooms"): [530, 538, 531, 211, 0, 87, 0, 0, 0, 0, 0, 0],
    ("Website", "rooms"): [59, 37, 73, 56, 0, 26, 0, 0, 0, 0, 0, 0],
}
FILAS = [{"code": c, "metric": m, "values": v} for (c, m), v in POR_CODE.items()]


def test_el_reporte_trae_las_CUATRO_pestanas():
    wb = construir_reporte("t", POR_CANAL, POR_CODE, CANAL_DE)
    assert wb.sheetnames == [
        "Canal · Habitaciones", "Canal · Pax",
        "Market code · Habitaciones", "Market code · Pax",
    ]


def test_el_reporte_trae_los_doce_meses_y_el_total():
    ws = construir_reporte("t", POR_CANAL, POR_CODE, CANAL_DE)["Canal · Habitaciones"]
    enc = [ws.cell(row=4, column=j).value for j in range(1, 15)]
    assert enc[0] == "Canal"
    assert enc[1:13] == MESES
    assert enc[13] == "Total"


def test_la_hoja_de_market_code_dice_a_que_canal_pertenece():
    """Es lo que el owner pidió ver: el código y su canal."""
    ws = construir_reporte("t", POR_CANAL, POR_CODE, CANAL_DE)["Market code · Habitaciones"]
    assert ws.cell(row=4, column=2).value == "Canal"
    filas = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
             for r in range(5, ws.max_row)}
    assert filas["TA"] == "Travel Agent"
    assert filas["WEB"] == "Website"


def test_el_total_de_la_hoja_suma_las_filas():
    """Sin el TOTAL hay que sumarlo a mano para cruzarlo contra el On the Books,
    que es justo el control que hace falta."""
    ws = construir_reporte("t", POR_CANAL, POR_CODE, CANAL_DE)["Canal · Habitaciones"]
    fila_total = ws.max_row
    assert ws.cell(row=fila_total, column=1).value == "TOTAL"
    assert ws.cell(row=fila_total, column=2).value == 530 + 59      # enero
    assert ws.cell(row=fila_total, column=14).value == sum(
        sum(v) for (c, m), v in POR_CANAL.items() if m == "rooms")


# ── La plantilla ─────────────────────────────────────────────────────────────

def test_la_plantilla_va_en_dos_pestanas():
    wb = construir_plantilla("t", FILAS, CANAL_DE)
    assert wb.sheetnames == ["Habitaciones", "Pax"]


def test_lo_que_bajo_es_lo_que_vuelve():
    filas, problemas = leer_plantilla(_bytes(construir_plantilla("t", FILAS, CANAL_DE)))
    assert problemas == []
    assert _clave(filas) == _clave(FILAS)


def test_un_ajuste_a_mano_entra():
    wb = construir_plantilla("t", FILAS, CANAL_DE)
    wb["Habitaciones"].cell(row=5, column=3, value=999)     # primera fila, Ene
    filas, problemas = leer_plantilla(_bytes(wb))
    assert problemas == []
    assert any(f["values"][0] == 999 for f in filas if f["metric"] == "rooms")


def test_mover_una_columna_NO_rompe_la_carga():
    """⚠️ La norma del proyecto: por encabezado, nunca por posición."""
    wb = construir_plantilla("t", FILAS, CANAL_DE)
    wb["Habitaciones"].insert_cols(1)
    filas, problemas = leer_plantilla(_bytes(wb))
    assert problemas == []
    assert _clave(filas) == _clave(FILAS)


def test_la_columna_canal_es_INFORMATIVA_y_no_se_lee():
    """⚠️ Si el canal se leyera del archivo, el resumen podría contradecir a su
    propio detalle. Se corrige el código y el canal se recalcula."""
    wb = construir_plantilla("t", FILAS, CANAL_DE)
    wb["Habitaciones"].cell(row=5, column=2, value="Un canal inventado")
    filas, problemas = leer_plantilla(_bytes(wb))
    assert problemas == []
    assert all("canal" not in f for f in filas), "la plantilla no devuelve canal"


def test_la_fila_TOTAL_no_se_toma_como_market_code():
    wb = construir_plantilla("t", FILAS, CANAL_DE)
    ws = wb["Habitaciones"]
    ws.cell(row=ws.max_row + 1, column=1, value="TOTAL")
    ws.cell(row=ws.max_row, column=3, value=99999)
    filas, _ = leer_plantilla(_bytes(wb))
    assert not any(f["code"] == "TOTAL" for f in filas)


def test_un_negativo_se_avisa():
    wb = construir_plantilla("t", FILAS, CANAL_DE)
    wb["Habitaciones"].cell(row=5, column=3, value=-5)
    _filas, problemas = leer_plantilla(_bytes(wb))
    assert any("negativos" in p for p in problemas)


def test_un_archivo_que_no_es_la_plantilla_lo_dice():
    wb = openpyxl.Workbook()
    wb.active["A1"] = "cualquier cosa"
    filas, problemas = leer_plantilla(_bytes(wb))
    assert filas == []
    assert any("Market code" in p for p in problemas)


# ── El endpoint ──────────────────────────────────────────────────────────────

def test_subir_sin_confirmar_no_guarda():
    from app.api import revenue_api

    src = inspect.getsource(revenue_api.channel_mix_subir_plantilla)
    assert "if not confirmar:" in src
    assert src.index("if not confirmar:") < src.index("delete(ChannelMixDetail)")


def test_el_resumen_se_RECALCULA_del_detalle_al_subir():
    """⚠️ El canal nunca sale del archivo. Si saliera, las dos capas podrían
    quedar diciendo cosas distintas sobre el mismo mes."""
    from app.api import revenue_api

    src = inspect.getsource(revenue_api.channel_mix_subir_plantilla)
    assert "canal_de.get(code" in src, "el canal se deriva del código"
    assert "roll" in src and "ChannelMixEntry(" in src
