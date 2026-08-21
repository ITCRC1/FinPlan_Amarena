# -*- coding: utf-8 -*-
"""Lo que se sube es lo que queda. El viaje redondo del Detalle.

**Por que existe (owner, 2026-08-14).** «Yo los resultados de Corcovado no los
voy a volver a subir porque ya los tengo. Solo quiero asegurarme que el upload
funciona bien. Para los otros hoteles que estaran en blanco.»

Corcovado se cargo hace meses y su historia esta revisada. Los tres hoteles
nuevos se van a cargar DESDE CERO, y ahi no hay con que comparar: si el
importador pierde una fila, nadie tiene el numero anterior para notarlo. El P&L
va a cuadrar consigo mismo igual.

Asi que esto prueba el camino completo con numeros inventados y conocidos:

    plantilla generada  ->  se llena  ->  parser  ->  ¿salio lo mismo?

No usa la base: es el contrato entre el exportador y el importador, que son las
dos puntas por donde se pierde la plata.
"""
import io

import openpyxl
import pytest

from app.export.detail_excel import build_detail_workbook
from app.importers.gl_detail_importer import parse_gl_detail

BLOQUE = "Actual actual 2025"
DEPTOS = {"0110": "Habitaciones", "0120": "A&B", "0140": "Spa",
          "0180": "Administracion", "0250": "Property"}

#: Una cuenta de cada clase, con un monto distinto por mes para que un
#: corrimiento de columnas se note.
CASOS = [
    ("0110", "4000", "Rooms", "Revenue"),
    ("0120", "5100", "Food Cost", "Cost"),
    ("0110", "6000", "Salary and Wages", "Payroll"),
    ("0140", "7065", "Cleaning Supplies", "Opex"),
    ("0250", "8015", "Property Insurance", "BelowGOP"),
]


def _plantilla_llena(casos=CASOS, meses=(1, 6, 12)) -> tuple[bytes, dict]:
    """Genera la plantilla y le pone montos. Devuelve (xlsx, esperado)."""
    esperado = {}
    accts = []
    for i, (dept, cta, nombre, clase) in enumerate(casos, start=1):
        vals = {}
        for m in meses:
            monto = i * 1000 + m          # distinto por cuenta Y por mes
            vals[(BLOQUE, m)] = float(monto)
            esperado[(dept, cta, m)] = float(monto)
        accts.append({"clase": clase, "grupo": "ROOMS", "dept_code": dept,
                      "cuenta": cta, "nombre": nombre, "vals": vals,
                      "orden": None})
    return build_detail_workbook([BLOQUE], accts, {}, DEPTOS), esperado


def _leer_parseado(xls: bytes) -> dict:
    """{(dept, cuenta, mes): monto} segun el PARSER."""
    fuera = {}
    for blk in parse_gl_detail(xls):
        for key in ("revenue", "costs", "opex", "belowgop"):
            for r in blk.get(key, []):
                for m, v in r["months"].items():
                    fuera[(r["dept_code"], r["account_code"], m)] = float(v)
        for r in blk.get("payroll", []):
            # La planilla viaja por CONCEPTO; se re-expande a su cuenta 6xxx.
            from app.engine import pl_engine
            acct = pl_engine.payroll_account_for_column(r["concept"])
            for m, v in r["months"].items():
                fuera[(r["dept_code"], acct, m)] = float(v)
    return fuera


def test_no_se_pierde_ninguna_fila():
    xls, esperado = _plantilla_llena()
    salio = _leer_parseado(xls)
    faltan = {k: v for k, v in esperado.items() if k not in salio}
    assert not faltan, (
        "El importador NO leyo estas filas de la plantilla que genera la propia "
        f"app: {sorted(faltan)}. En un hotel nuevo esa plata desaparece y no hay "
        "contra que compararla.")


def test_los_montos_llegan_intactos():
    xls, esperado = _plantilla_llena()
    salio = _leer_parseado(xls)
    malos = {k: (v, salio[k]) for k, v in esperado.items()
             if k in salio and abs(salio[k] - v) > 0.005}
    assert not malos, f"Montos cambiados en el viaje: {malos}"


def test_cada_monto_cae_en_SU_mes():
    """Un corrimiento de una columna daria numeros perfectamente creibles del
    mes equivocado, y eso no lo delata nada."""
    xls, esperado = _plantilla_llena(meses=(1, 6, 12))
    salio = _leer_parseado(xls)
    for (dept, cta, mes), monto in esperado.items():
        assert salio.get((dept, cta, mes)) == monto, (
            f"{dept}/{cta}: el mes {mes} no trae {monto}. "
            f"Lo que llego: { {k[2]: v for k, v in salio.items() if k[:2] == (dept, cta)} }")


def test_el_departamento_se_resuelve_por_codigo():
    """La plantilla escribe «0140 · Spa». El parser tiene que leer 0140 EXACTO,
    sin adivinar por la palabra — un departamento sin nombre en el catalogo no
    tiene palabra que adivinar."""
    xls, _ = _plantilla_llena(casos=[("0250", "8015", "Property Insurance", "BelowGOP")])
    salio = _leer_parseado(xls)
    assert any(k[0] == "0250" for k in salio), (
        "El parser no reconocio el departamento que escribio el exportador.")


def test_las_doce_columnas_de_mes_viajan():
    xls, esperado = _plantilla_llena(meses=tuple(range(1, 13)))
    salio = _leer_parseado(xls)
    for m in range(1, 13):
        assert any(k[2] == m for k in salio), f"el mes {m} no llego"
    assert len(salio) >= len(esperado)


def test_el_separador_de_departamento_no_entra_como_dato():
    """La hoja lleva filas de titulo y separador. Ninguna puede convertirse en
    una fila de datos ni en «plata sin cuenta», o el importador se negaria a
    cargar el archivo que el mismo genero."""
    xls, _ = _plantilla_llena()
    for blk in parse_gl_detail(xls):
        assert not blk.get("sin_cuenta"), (
            f"El importador vio plata sin cuenta en su propia plantilla: "
            f"{blk['sin_cuenta']}")


def test_un_hotel_en_blanco_no_rompe_el_parser():
    """Los tres hoteles nuevos arrancan sin una sola cifra."""
    accts = [{"clase": "Revenue", "grupo": "ROOMS", "dept_code": "0110",
              "cuenta": "4000", "nombre": "Rooms", "vals": {}, "orden": None}]
    xls = build_detail_workbook([BLOQUE], accts, {}, DEPTOS)
    blks = parse_gl_detail(xls)
    assert blks, "una plantilla vacia tiene que parsear igual"
    assert not any(b.get("sin_cuenta") for b in blks)


@pytest.mark.parametrize("clase,dept,cta", [
    ("Revenue", "0110", "4000"), ("Cost", "0120", "5100"),
    ("Payroll", "0110", "6000"), ("Opex", "0140", "7065"),
    ("BelowGOP", "0250", "8015"),
])
def test_las_cinco_clases_sobreviven(clase, dept, cta):
    """Si una clase entera se cae, en un hotel nuevo se nota tarde y feo."""
    xls, esperado = _plantilla_llena(casos=[(dept, cta, "X", clase)])
    salio = _leer_parseado(xls)
    assert salio, f"la clase {clase} no sobrevivio el viaje"
    assert set(esperado) <= set(salio)


# ── La otra mitad: que el RESUMEN cuadre con el DETALLE ─────────────────────
#
# Owner (2026-08-14): «como valido que el resumen pega con el detalle».
#
# El P&L se construye DESDE el detalle, asi que deberian cuadrar por definicion.
# Lo que rompe esa igualdad es lo que se cae en el camino: una cuenta sin regla
# de mapeo, una fila sin departamento, una clase que nadie consolida. Cada una de
# esas es plata que sale del detalle y NO llega al resumen — sin que nada falle.

def _mapeo_del_repo():
    """El mapeo real, leido del JSON del repositorio (no de la base)."""
    import json
    import pathlib

    arch = (pathlib.Path(__file__).resolve().parents[1]
            / "app" / "seed_data" / "mapping_pl.json")
    d = json.loads(arch.read_text(encoding="utf-8"))
    maps = [{"account_code": r["account_code"],
             "dept_code": r.get("dept_code") or "",
             "report_line_code": r["report_line_code"],
             "active_status": r["active_status"],
             "rollup_operator": r.get("rollup_operator", "SUM")}
            for r in d["account_mapping"] if r.get("active_status") == "YES"]
    cfg = sorted([r for r in d["report_line_config"] if r.get("active", True)],
                 key=lambda r: r["display_order"])
    return maps, cfg


def test_el_ingreso_del_detalle_llega_entero_al_resumen():
    """Lo que se sube como ingreso tiene que aparecer en TOTAL_REVENUES.

    Si una cuenta no tiene regla de mapeo, su plata se evapora entre el detalle y
    el P&L: el reporte cuadra consigo mismo y el total es menor. En un hotel
    nuevo no hay con que compararlo.
    """
    from decimal import Decimal

    from app.importers.gl_detail_importer import consolidate_block

    maps, cfg = _mapeo_del_repo()
    # Cuentas de ingreso que SI estan en el mapeo del repo.
    reales = [(m["dept_code"], m["account_code"]) for m in maps
              if str(m["account_code"]).startswith("4") and m["dept_code"]][:6]
    assert reales, "el mapeo del repo no trae cuentas de ingreso"

    casos = [(d, c, "X", "Revenue") for d, c in reales]
    xls, esperado = _plantilla_llena(casos=casos, meses=(3,))
    puesto = sum(esperado.values())

    total = Decimal("0")
    for blk in parse_gl_detail(xls):
        r = consolidate_block(blk, maps, cfg)
        for _mes, lineas in r["lines"].items():
            total += lineas.get("TOTAL_REVENUES", Decimal("0"))

    assert abs(float(total) - puesto) < 0.01, (
        f"Se subieron {puesto:,.2f} de ingreso y el resumen muestra "
        f"{float(total):,.2f}. La diferencia es plata que se cae entre el "
        f"detalle y el P&L.")


def test_una_cuenta_sin_regla_no_desaparece_callada():
    """El caso peligroso: una cuenta que el mapeo no conoce.

    No se le pide al motor que la adivine — se le pide que NO diga que todo
    cuadra. Hoy la fila se parsea (llega al detalle) y no suma al P&L; lo que
    esta prueba fija es que el detalle SI la trae, para que la comparacion
    detalle-vs-resumen pueda delatarla.
    """
    xls, esperado = _plantilla_llena(
        casos=[("0110", "4989", "Cuenta inventada", "Revenue")], meses=(5,))
    salio = _leer_parseado(xls)
    assert set(esperado) <= set(salio), (
        "Una cuenta sin regla de mapeo tampoco llega al detalle: entonces no hay "
        "forma de detectarla comparando, y desaparece del todo.")


# ── La verificacion que viaja ARRIBA del archivo ────────────────────────────
#
# Owner (2026-08-16): «necesito que el upload de los resultados tenga la
# verificacion arriba versus el detalle abajo […] para las nuevas propiedades,
# que debo empezar desde cero». Es el requisito que destraba clonar propiedades,
# y se cuelga de esta prueba porque esta es la que protege el camino por donde
# van a entrar los hoteles que no tienen historia contra que comparar.

from app.importers.verificacion import (  # noqa: E402
    CONTROLES, LOS_CUATRO_DEL_OWNER, comparar, meses_comparables)


def _con_verificacion(verif, casos=CASOS, meses=(1, 6, 12)):
    """La plantilla llena, con el bloque de control de arriba puesto."""
    _xls, esperado = _plantilla_llena(casos=casos, meses=meses)
    accts = []
    for i, (dept, cta, nombre, clase) in enumerate(casos, start=1):
        vals = {(BLOQUE, m): float(i * 1000 + m) for m in meses}
        accts.append({"clase": clase, "grupo": "ROOMS", "dept_code": dept,
                      "cuenta": cta, "nombre": nombre, "vals": vals, "orden": None})
    xls = build_detail_workbook([BLOQUE], accts, {}, DEPTOS,
                                verificacion={BLOQUE: verif})
    return xls, esperado


def test_el_bloque_de_verificacion_viaja_en_la_bajada_y_vuelve_igual():
    """Bajo, corrijo, subo: el control tiene que sobrevivir el viaje redondo.

    Si solo existiera en la subida habria que teclearlo de memoria cada vez, y
    un control que hay que escribir a mano es un control que nadie llena.
    """
    puesto = {"VER_INGRESOS": {1: 111.0, 6: 222.0},
              "VER_GOP": {1: -50.0},
              "VER_UTILIDAD_NETA": {12: 9.5}}
    xls, _ = _con_verificacion(puesto)
    blks = parse_gl_detail(xls)
    assert len(blks) == 1
    salio = blks[0]["verificacion"]
    for cod, meses in puesto.items():
        for m, v in meses.items():
            assert salio.get(cod, {}).get(m) == v, (
                f"{cod} mes {m}: se escribio {v} y volvio {salio.get(cod, {}).get(m)}")


def test_la_verificacion_no_es_una_segunda_fuente_de_plata():
    """Lo de arriba es un CONTROL, no un origen.

    Si un centavo del bloque de verificacion terminara sumando en el detalle,
    seria exactamente el problema que este bloque existe para cerrar: dos
    fuentes que se pisan y un total que cuadra consigo mismo.
    """
    enorme = {c.codigo: {m: 9_999_999.0 for m in range(1, 13)} for c in CONTROLES}
    xls, esperado = _con_verificacion(enorme)
    salio = _leer_parseado(xls)
    assert salio == esperado, (
        "El bloque de verificacion se colo en el detalle. Eso es plata inventada.")


def test_la_verificacion_no_se_confunde_con_plata_sin_cuenta():
    """El importador se NIEGA ante una fila con monto y sin cuenta (los $40.613
    del Actual 2024). El bloque de control tiene monto y no tiene cuenta: si no
    se distinguiera, la app no podria subir su propia plantilla."""
    verif = {c.codigo: {1: 1000.0} for c in CONTROLES}
    xls, _ = _con_verificacion(verif)
    for blk in parse_gl_detail(xls):
        assert not blk.get("sin_cuenta"), (
            f"La verificacion se leyo como plata sin cuenta: {blk['sin_cuenta']}")


def test_un_subtotal_metido_en_el_detalle_se_sigue_rechazando():
    """La verificacion se lee SOLO arriba del encabezado.

    Una fila «TOTAL INGRESOS» escrita en medio del detalle sigue siendo plata
    que no tiene a donde ir, y tiene que seguir tronando. Aceptarla como control
    la convertiria en un numero que nadie reclama.
    """
    from openpyxl import load_workbook
    import io as _io
    xls, _ = _plantilla_llena()
    wb = load_workbook(_io.BytesIO(xls))
    ws = wb["Detalle"]
    fila = ws.max_row + 2
    ws.cell(fila, 5, "Ingresos totales")
    ws.cell(fila, 6, 12345.0)
    buf = _io.BytesIO(); wb.save(buf)
    blks = parse_gl_detail(buf.getvalue())
    assert any(b.get("sin_cuenta") for b in blks), (
        "Un subtotal dentro del detalle dejo de reclamarse: se leyo como control.")


def test_los_cuatro_del_owner_bloquean_y_el_resto_avisa():
    """«Ingresos, GOP, EBITDA y net profit» — esos cuatro frenan la carga.

    El desglose por bucket explica DONDE esta la diferencia, pero una diferencia
    de presentacion adentro de un bucket que cuadra no tiene por que frenar
    nada: por eso avisa.
    """
    assert LOS_CUATRO_DEL_OWNER == (
        "VER_INGRESOS", "VER_GOP", "VER_EBITDA", "VER_UTILIDAD_NETA")
    otros = [c.codigo for c in CONTROLES if not c.bloquea]
    assert "VER_IMPUESTO" in otros and "VER_OVERHEAD" in otros


def test_cuando_cuadra_no_molesta():
    from decimal import Decimal
    verif = {"VER_INGRESOS": {1: 100.0}, "VER_GOP": {1: 40.0}}
    consolidado = {1: {"TOTAL_REVENUES": Decimal("100"), "TOTAL_GOP": Decimal("40.4")}}
    rep = comparar(verif, consolidado, list(range(1, 13)))
    assert rep["cuadra"] and not rep["bloquea"], rep


def test_cuando_no_cuadra_dice_cuanto_y_en_que_mes():
    """No se rechaza a secas: se muestra la comparacion. El owner puede tener
    una razon legitima para seguir, pero tiene que VERLA."""
    from decimal import Decimal
    verif = {"VER_INGRESOS": {1: 100.0, 2: 100.0}}
    consolidado = {1: {"TOTAL_REVENUES": Decimal("100")},
                   2: {"TOTAL_REVENUES": Decimal("70")}}
    rep = comparar(verif, consolidado, list(range(1, 13)))
    assert rep["bloquea"]
    linea = next(L for L in rep["lineas"] if L["codigo"] == "VER_INGRESOS")
    assert linea["dif"] == -30.0
    assert [d["mes"] for d in linea["meses_que_no_cuadran"]] == [2]


def test_el_impuesto_solo_no_hace_ver_roto_al_archivo():
    """La unica linea donde una diferencia puede significar «se subio una
    provision que los libros todavia no tienen». Los dos Forecast 2026 reportan
    $17.881,10 y $2.473,73 de diferencia neta, y toda es del impuesto: mezclarlo
    los hace ver rotos cuando su GOP y su EBT cuadran exactos."""
    from decimal import Decimal
    verif = {"VER_IMPUESTO": {1: 1000.0}, "VER_UTILIDAD_NETA": {1: 5000.0}}
    consolidado = {1: {"INCOME_TAXES": Decimal("1500"), "NET_PROFIT": Decimal("4500")}}
    rep = comparar(verif, consolidado, list(range(1, 13)))
    assert not rep["bloquea"], "el impuesto solo no puede frenar la carga"
    neta = next(L for L in rep["lineas"] if L["codigo"] == "VER_UTILIDAD_NETA")
    assert not neta["cuadra"], "la diferencia se dice, no se esconde"
    assert "impuesto" in neta["nota"].lower()


def test_una_utilidad_neta_mal_cargada_si_bloquea():
    """El descuento del impuesto no puede volverse una puerta de atras."""
    from decimal import Decimal
    verif = {"VER_IMPUESTO": {1: 1000.0}, "VER_UTILIDAD_NETA": {1: 5000.0}}
    consolidado = {1: {"INCOME_TAXES": Decimal("1000"), "NET_PROFIT": Decimal("4000")}}
    rep = comparar(verif, consolidado, list(range(1, 13)))
    assert rep["bloquea"] and "VER_UTILIDAD_NETA" in rep["bloqueantes"]


def test_un_forecast_no_se_mide_sobre_meses_que_no_usa():
    """Un forecast toma sus meses cerrados del Actual enlazado, no de si mismo.
    Comparar los doce es lo que hizo creer que el Working 2026 estaba
    desalineado cuando no lo estaba."""
    from decimal import Decimal
    comparables, cerrados = meses_comparables("FORECAST", 4)
    assert comparables == [5, 6, 7, 8, 9, 10, 11, 12] and cerrados == [1, 2, 3, 4]
    verif = {"VER_INGRESOS": {1: 999.0, 5: 100.0}}       # enero es mes cerrado
    consolidado = {1: {"TOTAL_REVENUES": Decimal("1")},   # y no coincide
                   5: {"TOTAL_REVENUES": Decimal("100")}}
    rep = comparar(verif, consolidado, comparables, cerrados)
    assert rep["cuadra"], rep
    assert rep["meses_no_comparados"] == [1, 2, 3, 4]


def test_una_celda_vacia_no_se_compara():
    """Es lo que deja subir una propiedad nueva mes a mes: lo que no se declara
    no se controla, y lo que se declara se controla al dolar."""
    from decimal import Decimal
    rep = comparar({}, {1: {"TOTAL_REVENUES": Decimal("100")}}, list(range(1, 13)))
    assert not rep["hay_verificacion"] and not rep["bloquea"]


def test_la_plantilla_en_blanco_ofrece_donde_escribir_el_control():
    """Una propiedad que arranca desde cero no tiene numeros que bajar, pero
    tiene que ver el bloque: una casilla que no existe es una que nadie llena."""
    from openpyxl import load_workbook
    import io as _io
    xls, _ = _con_verificacion({})
    ws = load_workbook(_io.BytesIO(xls))["Detalle"]
    codigos = {ws.cell(r, 4).value for r in range(1, 13)}
    for c in CONTROLES:
        assert c.codigo in codigos, f"{c.codigo} no esta en la plantilla vacia"


def test_el_bloque_de_control_no_pisa_el_encabezado_del_detalle():
    """Las filas 14 y 15 estan cableadas en el parser y en todos los archivos
    que el owner ya tiene. El control tiene que caber ARRIBA sin moverlas."""
    from openpyxl import load_workbook
    import io as _io
    from app.importers.gl_detail_importer import MONTH_ROW, VERSION_ROW
    xls, _ = _con_verificacion({c.codigo: {1: 1.0} for c in CONTROLES})
    ws = load_workbook(_io.BytesIO(xls))["Detalle"]
    assert ws.cell(VERSION_ROW, 4).value == "Cuenta"
    assert ws.cell(MONTH_ROW, 6).value == "January"
    assert len(CONTROLES) <= MONTH_ROW - 3, (
        "no caben los controles arriba del encabezado sin mover MONTH_ROW")


# ── Las tres garantias del ENDPOINT ─────────────────────────────────────────
#
# Se miran sobre el codigo y no sobre la base a proposito: son propiedades de
# ORDEN y de AUSENCIA (que nada se escriba antes, que el control no se guarde
# nunca), y eso una prueba de datos no lo puede demostrar — solo puede no
# encontrarlo.

def _fuente_del_endpoint() -> str:
    import io as _io
    import pathlib as _pl
    arch = _pl.Path(__file__).resolve().parents[1] / "app" / "api" / "scenarios_api.py"
    src = _io.open(arch, encoding="utf-8").read()
    return src.split("async def import_gl_detail(")[1].split("\n@router.")[0]


def test_la_verificacion_corre_antes_de_escribir_una_sola_fila():
    """Si bloqueara despues de escribir, quedaria media importacion adentro y un
    mensaje de error diciendo que no se cargo nada. Es el mismo cuidado con que
    `import_all` valida el GL antes de que el snapshot haga commit."""
    src = _fuente_del_endpoint()
    # El 409 se levanta con `ErrorApi(409, ...)` — el mensaje vive en el
    # catalogo bilingue (`app/errores.py`), pero la PUERTA sigue estando aca.
    puerta = src.index('ErrorApi(409, "gl.verificacion_no_cuadra"')
    primera_escritura = min(
        src.index("await _write_accounts(RevenueAccountEntry"),
        src.index("sa_delete(ActualEntry)"))
    assert puerta < primera_escritura, (
        "la verificacion bloquea DESPUES de escribir: quedaria media carga adentro")


def test_el_bloque_de_control_no_se_guarda_en_ninguna_tabla():
    """Es un CONTROL, no un origen. Si algo de lo que trae terminara en la base
    seria una segunda fuente de plata — exactamente el problema que cierra."""
    src = _fuente_del_endpoint()
    for linea in src.splitlines():
        if 'blk["verificacion"]' in linea or "blk.get(\"verificacion\")" in linea:
            assert "db.add" not in linea and "Model(" not in linea, linea


def test_se_consolida_con_lo_que_escribe_el_motor_no_solo_con_lo_que_se_digita():
    """Las contrapartidas de reparto (4900/4901/4999) las escribe el motor y el
    archivo no puede traerlas: sobreviven al reemplazo justamente por eso. Sin
    ellas la comparacion fallaria sin que haya un solo error — son −196.326,17
    entre Lavanderia y Cafeteria y −92.176,74 en Rooms del Working 2027."""
    src = _fuente_del_endpoint()
    assert "_filas_que_sobreviven" in src
    assert "filas_extra=" in src


def test_la_puerta_no_se_abre_sola():
    """Seguir con una diferencia tiene que ser una decision explicita del owner,
    no un default. Un `confirmar_diferencias=True` por defecto convertiria todo
    esto en un aviso que nadie lee."""
    import inspect as _insp
    from app.api.scenarios_api import import_gl_detail
    p = _insp.signature(import_gl_detail).parameters["confirmar_diferencias"]
    assert p.default.default is False


# ── Subir UN MES no puede llevarse el reparto de ese mes ────────────────────
#
# Owner (2026-08-16), describiendo el ciclo real: «Yo subo julio, se actualiza el
# ACTUAL 2026, el Forecast Working se actualiza automaticamente». O sea que el
# camino de TODOS los meses, de CADA hotel, es el merge — la pantalla de carga
# manda `merge=true`.
#
# Y ese camino tenia el mismo agujero que el reemplazo total ya tenia tapado: las
# contrapartidas de reparto (`4900`/`4901`/`4999`, «Distribucion») se ponian en
# cero para los meses del archivo. El archivo no puede reponerlas —el parser las
# excluye a proposito, porque son el credito del asiento y no ingreso— y nada
# las regenera: un ACTUAL no recalcula repartos.
#
# Medido sobre el Actual 2025 en produccion: subir SOLO julio se llevaba
# −5.007,57 de Lavanderia (0161/4900) y −12.537,17 de Cafeteria (0220/4901) =
# **−17.544,74** de credito. El overhead lo mostraba de mas y el P&L seguia
# cuadrando consigo mismo. En el año: −196.326,17.

class _FakeScalars:
    def __init__(self, filas): self._filas = filas
    def all(self): return self._filas


class _FakeResult:
    def __init__(self, filas): self._filas = filas
    def scalars(self): return _FakeScalars(self._filas)


class _FakeDB:
    """Devuelve siempre las mismas filas: `_filas_que_sobreviven` hace UNA query."""
    def __init__(self, filas): self._filas = filas
    async def execute(self, *_a, **_k): return _FakeResult(self._filas)


class _Destino:
    id = "esc-1"


def _fila(dept, code, name, montos):
    from app.models.actual_entry import ActualEntry
    e = ActualEntry(scenario_id="esc-1", hotel_id="CWL", dept_code=dept,
                    account_code=code, account_name=name, outlet="")
    for m, v in montos.items():
        e.set_month(m, Decimal(str(v)))
    return e


from decimal import Decimal  # noqa: E402


def _sobrevivientes(merge, meses_archivo):
    import asyncio

    from app.api.scenarios_api import _filas_que_sobreviven
    filas = [
        # Las dos contrapartidas reales del Actual 2025, con julio de verdad.
        _fila("0161", "4900", "Distribución", {7: -5007.57, 8: -4000.00}),
        _fila("0220", "4901", "Distribución", {7: -12537.17, 8: -11000.00}),
        # Un gasto normal: ESTE si lo pisa el archivo.
        _fila("0110", "7065", "Cleaning Supplies", {7: 900.00, 8: 800.00}),
    ]
    db = _FakeDB(filas)
    return asyncio.run(_filas_que_sobreviven(db, _Destino(), merge, meses_archivo))


def test_subir_un_mes_no_borra_la_contrapartida_de_ese_mes():
    """El defecto, en el camino que se recorre todos los meses.

    Se mide sobre `_filas_que_sobreviven` porque es la funcion que DECLARA que
    va a seguir ahi despues de la carga — y es la que alimenta la verificacion
    de la puerta. Si dijera que la contrapartida se pierde, la puerta compararia
    contra un consolidado que el reporte nunca va a dar.
    """
    vivos = _sobrevivientes(merge=True, meses_archivo=[7])
    julio = {(f["dept_code"], f["account_code"]): f["amount"]
             for f in vivos.get(7, [])}
    assert (("0161", "4900")) in julio, (
        "La contrapartida de Lavanderia NO sobrevivio a subir julio: son "
        "−5.007,57 de credito que el archivo no puede reponer.")
    assert ("0220", "4901") in julio, (
        "La contrapartida de Cafeteria NO sobrevivio a subir julio: −12.537,17.")
    assert float(julio[("0161", "4900")]) == pytest.approx(-5007.57)
    assert float(julio[("0220", "4901")]) == pytest.approx(-12537.17)
    perdido = -17544.74
    assert float(sum(julio.values())) == pytest.approx(perdido, abs=0.01), (
        f"El credito de reparto de julio tiene que sobrevivir entero ({perdido:,.2f}).")


def test_el_gasto_normal_del_mes_subido_SI_lo_pisa_el_archivo():
    """El arreglo no puede volverse «no se pisa nada»: el archivo manda sobre el
    mes que trae. Solo se protege lo que el archivo no puede traer."""
    vivos = _sobrevivientes(merge=True, meses_archivo=[7])
    julio = {(f["dept_code"], f["account_code"]) for f in vivos.get(7, [])}
    assert ("0110", "7065") not in julio, (
        "Un gasto normal de julio sobrevivio: entonces el archivo no manda "
        "sobre el mes que sube, y subir dos veces sumaria.")


def test_los_meses_que_el_archivo_NO_trae_quedan_intactos():
    """Propiedad 2 del contrato con el modelo real: subir julio no mueve junio."""
    vivos = _sobrevivientes(merge=True, meses_archivo=[7])
    agosto = {(f["dept_code"], f["account_code"]): f["amount"]
              for f in vivos.get(8, [])}
    assert len(agosto) == 3, "un mes no subido tiene que conservar TODAS sus filas"
    assert float(agosto[("0110", "7065")]) == pytest.approx(800.00)


def test_el_escritor_del_merge_saltea_las_contrapartidas():
    """Que la proteccion este REALMENTE en el que escribe.

    `_filas_que_sobreviven` solo DECLARA. Si el escritor no hiciera lo mismo, la
    puerta diria que la plata sigue ahi y el merge la borraria igual — que es
    peor que el bug original, porque ahora habria un control afirmandolo.
    """
    src = _fuente_del_endpoint()
    # El merge de ActualEntry, acotado: arranca donde lee las filas existentes y
    # termina donde empieza la rama de reemplazo. Anclar en `if merge:` a secas
    # agarraba el de la planilla, que va antes.
    bloque = src[src.index("existing_ae = "):src.index("sa_delete(ActualEntry)")]
    assert "es_contrapartida_de_allocation" in bloque, (
        "La rama de merge ya no protege las contrapartidas: subir un mes vuelve "
        "a borrar el credito de reparto de ese mes.")
    assert "continue" in bloque, (
        "Se nombra la regla pero no se saltea la fila: la proteccion no aplica.")


def test_el_alcance_por_defecto_es_el_MES_y_no_el_ANO():
    """El ciclo del owner es mensual; el default tiene que serlo tambien.

    Con `merge=False` por defecto, subir julio borraba y refabricaba el escenario
    entero — y solo estaba a salvo porque la pantalla manda `merge=true` a mano.
    Cualquier llamada por fuera (curl, script) se llevaba los meses ya cerrados.
    """
    import inspect as _insp
    from app.api.scenarios_api import import_gl_detail
    p = _insp.signature(import_gl_detail).parameters["merge"]
    assert p.default.default is True, (
        "El default volvio a ser el reemplazo total del escenario.")


# ── El viaje redondo por las CARGAS MASIVAS: la moneda tiene que volver ──────
#
# Norma de trabajo del owner: **bajo, corrijo, subo**. Las dos puntas de ese
# viaje en las cargas masivas de checkbook son el serializador de lectura
# (`_entry_to_dict`, lo que se BAJA) y el schema de escritura (`*BulkRow`, lo
# que se SUBE). Entre las dos se perdia la moneda:
#
#   · `bulk_replace_opex` DECLARABA `currency` y los `crc_*` en su schema y no
#     los copiaba al INSERT. La linea entraba siempre como USD y con los
#     colones en cero.
#   · `bulk_replace_costs` ni siquiera los declaraba: no habia donde recibirlos.
#
# Y los dos endpoints BORRAN todo el escenario antes de escribir, asi que el
# dato maestro no quedaba en ningun lado. No revienta —los dolares que venia
# trayendo el archivo quedan, y el P&L cuadra consigo mismo— pero la linea deja
# de acompanar al tipo de cambio: la proxima vez que el TC se mueve, esa cuenta
# se queda con el dolar viejo y nada lo avisa.

_MK = ["jan", "feb", "mar", "apr", "may", "jun",
       "jul", "aug", "sep", "oct", "nov", "dec"]


class _DBQueEscribe:
    """DB de mentira que anota lo que se INSERTA. Sin tipos de cambio cargados,
    asi que la derivacion a dolares no corre y se ve lo que se guardo tal cual."""

    def __init__(self, escenario):
        self._esc = escenario
        self.agregados: list = []

    async def get(self, _model, _id):
        return self._esc

    async def execute(self, *_a, **_k):
        return _FakeResult([])

    def add(self, obj):
        self.agregados.append(obj)

    async def flush(self):
        pass

    async def commit(self):
        pass


def _escenario_editable():
    from app.models.scenario import Scenario
    return Scenario(id="esc-1", hotel_id="CWL", year=2027, type="BUDGET",
                    version="Working", status="draft")


#: Un mes de electricidad de verdad en colones, y su dolar derivado.
CRC_JULIO = Decimal("1750000.00")
USD_JULIO = Decimal("3370.1900")


def test_opex_la_moneda_sobrevive_a_bajar_y_subir():
    import asyncio

    from app.api import opex_api
    from app.models.opex_entry import OpexEntry

    linea = OpexEntry(scenario_id="esc-1", hotel_id="CWL", dept_code="0210",
                      account_code="7160", account_name="Electricity",
                      detail_code="001", detail_desc="CNFL",
                      currency="CRC",
                      **{m: Decimal("0") for m in _MK},
                      **{f"crc_{m}": Decimal("0") for m in _MK})
    linea.set_crc(7, CRC_JULIO)
    linea.set_month(7, USD_JULIO)

    # BAJO: exactamente lo que la pantalla recibe al leer la linea.
    bajado = opex_api._entry_to_dict(linea)
    assert bajado["currency"] == "CRC"

    # SUBO: ese mismo contenido, por la carga masiva.
    fila = opex_api.OpexBulkRow(
        dept_code=linea.dept_code, account_code=linea.account_code,
        account_name=linea.account_name, detail_code=linea.detail_code,
        detail_desc=linea.detail_desc,
        currency=bajado["currency"],
        **{f"crc_{m}": Decimal(bajado["crc_months"][m]) for m in _MK},
        **{m: Decimal(bajado["months"][m]) for m in _MK},
    )
    db = _DBQueEscribe(_escenario_editable())
    asyncio.run(opex_api.bulk_replace_opex("esc-1", [fila], db))

    guardada = db.agregados[0]
    assert guardada.currency == "CRC", (
        "La linea volvio marcada en DOLARES: bajar y subir le borro la moneda, "
        "y este endpoint borra el escenario entero antes de escribir, asi que "
        "los colones no quedaron en ningun lado.")
    assert guardada.get_crc(7) == CRC_JULIO, (
        f"Los colones de julio no volvieron ({CRC_JULIO:,.2f}): quedo el dolar "
        "como unico dato y la linea deja de acompanar al tipo de cambio.")
    assert guardada.get_month(7) == USD_JULIO


def test_costos_la_moneda_sobrevive_a_bajar_y_subir():
    import asyncio

    from app.api import costs_api
    from app.models.cost_entry import CostEntry

    linea = CostEntry(scenario_id="esc-1", hotel_id="CWL", dept_code="0120",
                      account_code="5100", account_name="Food Cost",
                      calc_mode="MANUAL", currency="CRC",
                      **{m: Decimal("0") for m in _MK},
                      **{f"crc_{m}": Decimal("0") for m in _MK})
    linea.set_crc(7, CRC_JULIO)
    linea.set_month(7, USD_JULIO)

    bajado = costs_api._entry_to_dict(linea)
    assert bajado["currency"] == "CRC"

    fila = costs_api.CostBulkRow(
        dept_code=linea.dept_code, account_code=linea.account_code,
        account_name=linea.account_name, calc_mode="MANUAL",
        currency=bajado["currency"],
        **{f"crc_{m}": Decimal(bajado[f"crc_{m}"]) for m in _MK},
        **{m: Decimal(bajado["months"][m]) for m in _MK},
    )
    db = _DBQueEscribe(_escenario_editable())
    asyncio.run(costs_api.bulk_replace_costs("esc-1", [fila], db))

    guardada = db.agregados[0]
    assert guardada.currency == "CRC", (
        "El schema de costos no recibia la moneda: la linea planificada en "
        "colones volvia en dolares.")
    assert guardada.get_crc(7) == CRC_JULIO
    assert guardada.get_month(7) == USD_JULIO


@pytest.mark.parametrize("modulo,schema", [
    ("opex_api", "OpexBulkRow"),
    ("costs_api", "CostBulkRow"),
])
def test_los_dos_schemas_masivos_declaran_la_moneda(modulo, schema):
    """Gemelos: si uno gana un campo de moneda y el otro no, el que quede atras
    pierde el dato en silencio. Ya paso una vez."""
    import importlib
    campos = getattr(importlib.import_module(f"app.api.{modulo}"), schema).model_fields
    assert "currency" in campos
    for m in _MK:
        assert f"crc_{m}" in campos, f"{schema} no puede recibir crc_{m}"


@pytest.mark.parametrize("modulo,funcion", [
    ("opex_api", "bulk_replace_opex"),
    ("costs_api", "bulk_replace_costs"),
])
def test_una_linea_en_dolares_no_se_ensucia_con_colones(modulo, funcion):
    """El arreglo no puede volverse «todo es CRC»: lo que se subio en dolares
    tiene que seguir en dolares, con los colones en cero."""
    import asyncio
    import importlib
    mod = importlib.import_module(f"app.api.{modulo}")
    schema = mod.OpexBulkRow if modulo == "opex_api" else mod.CostBulkRow
    fila = schema(dept_code="0180", account_code="7380", account_name="Misc",
                  **{m: Decimal("0") for m in _MK})
    fila.jul = Decimal("1200")
    db = _DBQueEscribe(_escenario_editable())
    asyncio.run(getattr(mod, funcion)("esc-1", [fila], db))
    guardada = db.agregados[0]
    assert guardada.currency == "USD"
    assert guardada.get_crc(7) == Decimal("0")
    assert guardada.get_month(7) == Decimal("1200")


# ── El CORTE de meses cerrados no se pierde al subir el resumen ──────────────
#
# `actuals_through` es lo que hace que el FORECAST Working tome del ACTUAL
# enlazado sus meses ya cerrados y proyecte solo el resto. `import-pl-snapshot`
# lo ponia en CERO —en silencio, pegado al modo de reemplazo, que ademas era el
# default— y eso deshace el cierre del mes sin avisar. El P&L sigue cuadrando
# consigo mismo.
#
# Medido en produccion el 2026-08-16, antes del arreglo: `FORECAST Working 2026`
# tenia corte 6 y `FORECAST April 2026` corte 4. Una sola llamada por fuera de
# la pantalla —un curl, un script— los dejaba a los dos en 0.

def _forecast(corte: int, version: str = "Working", current: bool = True):
    import datetime as _dt
    from app.models.scenario import Scenario
    s = Scenario(id=f"fc-{version}", hotel_id="CWL", year=2026, type="FORECAST",
                 version=version, status="draft")
    s.actuals_through = corte
    s.is_current_forecast = current
    s.created_at = _dt.datetime(2026, 1, 1)
    return s


class _DBDelSnapshot:
    """Fake para correr `import_pl_snapshot` de punta a punta sin base."""

    def __init__(self, escenarios):
        self._escs = escenarios
        self.agregados: list = []
        self.commits = 0

    async def execute(self, *_a, **_k):
        return _FakeResult(self._escs)

    def add(self, obj):
        self.agregados.append(obj)

    async def flush(self):
        pass

    async def commit(self):
        self.commits += 1


def _bloque_forecast(meses=(7,)):
    return {"label": "Forecast Working 2026", "type": "FORECAST", "year": 2026,
            "stats": {m: {"rooms_occupied": Decimal("300")} for m in meses},
            "lines": {m: {"TOTAL_REVENUES": 111000.0} for m in meses},
            "unmapped": []}


def _subir_snapshot(monkeypatch, escenarios, bloque, **kw):
    import asyncio

    from app.api import scenarios_api
    from app.api.scenarios_api import _BytesFile
    from app.importers import pl_snapshot_importer

    monkeypatch.setattr(pl_snapshot_importer, "parse_pl_snapshot",
                        lambda _data: [bloque])
    db = _DBDelSnapshot(escenarios)
    # Llamada directa, sin FastAPI: los defaults son objetos `Query(...)` y
    # cualquiera de ellos es «truthy». Hay que pasar los valores a mano — es lo
    # mismo que hace `import-all` cuando encadena los dos importadores.
    args = {"merge": True, "dry_run": False, "mes_de_cierre": None,
            "apagar_corte": False, **kw}
    res = asyncio.run(scenarios_api.import_pl_snapshot(
        file=_BytesFile(b"x", "upload.xlsx"), db=db, **args))
    return res, db


def test_subir_el_resumen_no_apaga_el_corte(monkeypatch):
    """El defecto, en el caso real: el Forecast Working con julio cerrado."""
    fc = _forecast(corte=6)
    res, _ = _subir_snapshot(monkeypatch, [fc], _bloque_forecast())
    assert fc.actuals_through == 6, (
        "El corte se fue a 0: el Forecast Working deja de tomar del Actual sus "
        "meses cerrados y vuelve a proyectarlos, sin que nada lo diga.")
    assert res["blocks"][0]["actuals_through"] == 6


def test_el_reemplazo_total_tampoco_lo_apaga_pero_lo_AVISA(monkeypatch):
    """La semantica de «reemplazar todo» es legitima; apagar el cierre a la
    pasada no lo es. Se conserva, y la respuesta lo dice."""
    fc = _forecast(corte=6)
    res, _ = _subir_snapshot(monkeypatch, [fc], _bloque_forecast(), merge=False)
    assert fc.actuals_through == 6
    aviso = res["blocks"][0].get("aviso_corte")
    assert aviso and "apagar_corte" in aviso, (
        "Se conserva el corte pero no se avisa: quien sube un snapshot que ya "
        "trae su propio blend no tiene como enterarse de que puede apagarlo.")


def test_apagar_el_corte_se_pide_a_mano_y_queda_dicho(monkeypatch):
    """El caso legitimo sigue disponible — pero explicito y reportado."""
    fc = _forecast(corte=6)
    res, _ = _subir_snapshot(monkeypatch, [fc], _bloque_forecast(),
                             merge=False, apagar_corte=True)
    assert fc.actuals_through == 0
    assert res["blocks"][0]["corte_apagado"] == 6


def test_el_alcance_por_defecto_del_resumen_es_el_MES():
    """El gemelo de `import-gl-detail`, que ya arrancaba en el mes. Dos puertas
    al mismo dato con defaults opuestos es lo que nadie recuerda al escribir un
    curl."""
    import inspect as _insp
    from app.api.scenarios_api import import_pl_snapshot
    p = _insp.signature(import_pl_snapshot).parameters["merge"]
    assert p.default.default is True, (
        "El default del resumen volvio a borrar y refabricar el escenario entero.")


def test_el_resumen_tiene_el_camino_del_cierre_mensual(monkeypatch):
    """`mes_de_cierre` escribe SOLO ese mes: un mes ya cerrado no se toca por
    accidente. Vive en el backend porque un curl tiene que topar el mismo tope."""
    fc = _forecast(corte=6)
    bloque = _bloque_forecast(meses=(5, 6, 7, 8))
    res, db = _subir_snapshot(monkeypatch, [fc], bloque, mes_de_cierre=7)
    assert res["blocks"][0]["months_touched"] == [7]
    assert res["meses_descartados"] == [5, 6, 8]
    meses_escritos = {s.month for s in db.agregados}
    assert meses_escritos == {7}, (
        f"Se escribieron los meses {sorted(meses_escritos)}: el cierre mensual "
        "tiene que descartar el resto del archivo, no solo avisar.")


def test_la_carga_combinada_tampoco_arranca_en_reemplazo_total():
    """`import-all` pasaba su `merge` a los DOS importadores, asi que su default
    arrastraba a `import-gl-detail` de vuelta al reemplazo total aunque ese ya
    estuviera arreglado."""
    import inspect as _insp
    from app.api.scenarios_api import import_all
    p = _insp.signature(import_all).parameters["merge"]
    assert p.default.default is True
    assert "mes_de_cierre" in _insp.signature(import_all).parameters, (
        "el cierre mensual no llega a la puerta combinada: el archivo entero "
        "tiene que recortarse al mismo mes en el resumen y en el detalle")


def test_la_carga_combinada_pasa_TODOS_los_parametros_a_mano():
    """`import-all` llama a los importadores como FUNCIONES, no por HTTP.

    FastAPI no resuelve los defaults en ese camino: cada parametro omitido
    llega como el objeto `Query(...)`, y **`bool(Query(None))` es True**. Con
    `scenario_id` sin pasar, `if scenario_id:` daba verdadero, no encontraba
    ningun escenario con ese id y la carga combinada moria con un 404
    «Escenario (version) no encontrado». Lo mismo le pasaria a `mes_de_cierre`:
    `Query(None) is not None` es verdadero y recortaria el archivo a nada.
    """
    import ast
    import inspect as _insp
    from fastapi import Query
    from app.api import scenarios_api
    assert bool(Query(None)) is True, "cambio el comportamiento de FastAPI; revisar"

    arbol = ast.parse(_insp.getsource(scenarios_api.import_all))
    llamadas: dict[str, list[set]] = {"import_gl_detail": [], "import_pl_snapshot": []}
    for nodo in ast.walk(arbol):
        if isinstance(nodo, ast.Call) and isinstance(nodo.func, ast.Name):
            if nodo.func.id in llamadas:
                llamadas[nodo.func.id].append({k.arg for k in nodo.keywords})

    for nombre_fn, juegos in llamadas.items():
        assert juegos, f"{nombre_fn} ya no se llama desde import_all"
        firma = _insp.signature(getattr(scenarios_api, nombre_fn)).parameters
        obligatorios = {n for n, p in firma.items()
                        if isinstance(p.default, type(Query(None)))}
        for pasados in juegos:
            faltan = obligatorios - pasados
            assert not faltan, (
                f"import_all llama a {nombre_fn} sin pasar {sorted(faltan)}: "
                f"le llega el objeto Query, que es truthy.")
