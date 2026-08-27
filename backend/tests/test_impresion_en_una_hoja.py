# -*- coding: utf-8 -*-
"""EL REPORTE SALE EN UNA SOLA HOJA: EN PAPEL Y EN EXCEL.

Owner, 2026-08-27: «este reporte debe salir en una sola página al imprimir…
podemos hacerla horizontal, y el Excel debe ser en una sola página sin separar».

**En papel.** `@page { size: letter landscape }` fija el papel pero no achica
nada: si el reporte mide más que la hoja, el navegador lo parte. El
`.print-dashboard` que había usa un `zoom: 0.6` clavado, que apuesta dos veces y
pierde las dos — con poco contenido desperdicia media hoja, y con mucho parte
igual. Ahora la escala se MIDE contra el tamaño de la hoja.

**En Excel.** Un cuadro de 12 meses son 14 columnas: en vertical y sin ajuste,
Excel lo parte en tres o cuatro hojas y los meses quedan repartidos entre
papeles distintos.

La trampa de openpyxl que cuida `test_el_fitToPage_esta_encendido`: sin
`sheet_properties.pageSetUpPr.fitToPage`, los `fitToWidth`/`fitToHeight` quedan
escritos en el XML y **Excel los ignora**. Se ve bien en el archivo y sale
partido igual.
"""
from __future__ import annotations

import io
import pathlib

import openpyxl
import pytest

from app.export.cuadro_excel import build_cuadros_workbook

FRONT = pathlib.Path(__file__).resolve().parent.parent.parent / "frontend"


def _libro(n_filas: int = 3, n_cols: int = 9):
    cuadro = {
        "titulo": "P&L por Departamento",
        "hoja": "PL x Depto",
        "columnas": ([{"label": "Departamento", "formato": "texto"}]
                     + [{"label": f"c{i}", "formato": "usd"} for i in range(n_cols - 1)]),
        "filas": [{"label": f"f{i}", "valores": [1.0] * (n_cols - 1)}
                  for i in range(n_filas)],
    }
    return openpyxl.load_workbook(io.BytesIO(build_cuadros_workbook([cuadro])))


# ─── Excel ────────────────────────────────────────────────────────────────────
def test_la_hoja_sale_horizontal():
    ws = _libro()["PL x Depto"]
    assert ws.page_setup.orientation == "landscape"


def test_el_ajuste_es_a_UNA_hoja():
    ws = _libro()["PL x Depto"]
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 1, (
        "fitToHeight=0 son «las hojas que hagan falta»: vuelve a partirse")


def test_el_fitToPage_esta_encendido():
    """Sin esto, los dos ajustes de arriba son decorativos."""
    ws = _libro()["PL x Depto"]
    assert ws.sheet_properties.pageSetUpPr is not None
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True


def test_el_area_de_impresion_se_acota_a_lo_escrito():
    """Sin área, una celda tocada lejos de la tabla arrastra hojas en blanco."""
    ws = _libro(n_filas=3, n_cols=9)["PL x Depto"]
    assert ws.print_area, "no se fijó el área de impresión"
    assert ws.print_area.endswith("$I$7"), ws.print_area


def test_el_area_crece_con_las_filas():
    """Que no quede clavada: un cuadro más largo tiene que entrar entero."""
    ws = _libro(n_filas=50, n_cols=9)["PL x Depto"]
    assert ws.print_area.endswith("$I$54"), ws.print_area


def test_un_cuadro_vacio_no_rompe():
    """Sin filas no hay área que fijar; el libro tiene que salir igual."""
    wb = openpyxl.load_workbook(io.BytesIO(build_cuadros_workbook([
        {"titulo": "Vacío", "hoja": "Vacio",
         "columnas": [{"label": "Concepto"}], "filas": []}])))
    assert "Vacio" in wb.sheetnames


# ─── Pantalla ─────────────────────────────────────────────────────────────────
@pytest.fixture(scope="module")
def hook() -> str:
    f = FRONT / "lib" / "imprimirEnUnaHoja.ts"
    if not f.exists():
        pytest.skip("no está el front en este árbol")
    return io.open(f, encoding="utf-8").read()


@pytest.fixture(scope="module")
def pagina() -> str:
    f = FRONT / "app" / "reports" / "pl-by-dept" / "page.tsx"
    if not f.exists():
        pytest.skip("no está el front en este árbol")
    return io.open(f, encoding="utf-8").read()


def test_la_escala_se_mide_en_estado_de_impresion(hook):
    """LA trampa. En pantalla los contenedores con scroll RECORTAN la tabla con
    un `max-height`, así que el alto medido sale corto y la escala queda grande:
    el reporte se parte igual, que es justo lo que se venía a arreglar."""
    assert "midiendo-impresion" in hook
    i = hook.index("getBoundingClientRect")
    assert "classList.add(\"midiendo-impresion\")" in hook[:i], (
        "se mide ANTES de encender el estado de impresión")
    assert "classList.remove(\"midiendo-impresion\")" in hook[i:], (
        "el estado de medición queda encendido y se ve en pantalla")


def test_la_clase_de_medicion_suelta_los_dos_envoltorios_de_scroll():
    """`.fin-sticky` y `.fin-scroll-x`: los dos recortan. Con uno solo, la mitad
    de las pantallas sigue midiendo de menos."""
    css = io.open(FRONT / "app" / "globals.css", encoding="utf-8").read()
    i = css.index(".midiendo-impresion .no-print")
    trozo = css[i:i + 400]
    assert ".fin-sticky" in trozo
    assert ".fin-scroll-x" in trozo


def test_la_escala_nunca_agranda(hook):
    """`Math.min(1, ...)`: un reporte chico no se estira hasta llenar la hoja."""
    assert "Math.min(1," in hook


def test_hay_piso_de_escala(hook):
    """Una hoja ilegible no es mejor que dos legibles."""
    assert "ESCALA_MINIMA" in hook
    assert "Math.max(ESCALA_MINIMA" in hook


def test_la_escala_se_limpia_al_terminar(hook):
    """Si quedara puesta, la pantalla se vería encogida después de imprimir y no
    habría cómo devolverla sin recargar."""
    assert "afterprint" in hook
    assert 'style.zoom = ""' in hook


def test_el_reporte_usa_el_escalado_medido_y_no_el_zoom_clavado(pagina):
    assert "print-una-hoja" in pagina
    assert "print-dashboard" not in pagina, (
        "volvió al zoom 0.6 clavado, que parte los reportes grandes")
    assert "useImprimirEnUnaHoja(hoja)" in pagina


def test_los_controles_no_se_imprimen(pagina):
    """Botones y selectores ocupan alto y no dicen nada en papel. Si no se
    escondieran, además, la medición y la impresión no coincidirían."""
    assert 'className="no-print"' in pagina
