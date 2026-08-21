# -*- coding: utf-8 -*-
"""
EL REPORTE DEL DUEÑO: LOS GASTOS VAN CON SU SIGNO.

Apareció en la auditoría de exportaciones (2026-08-13). El código escribía el
valor ABSOLUTO y le ponía un formato con paréntesis literales para que
*pareciera* negativo:

    value=value if value >= 0 else -value
    number_format = '("$"#,##0)' if value < 0 else …

Visto en pantalla, el archivo estaba perfecto: los gastos con paréntesis. Pero
la celda de un gasto de $50,000 contenía **+50000**, así que un `SUM()` sobre la
columna sumaba los gastos en vez de restarlos — y el dueño se llevaba un total
inflado sin ninguna señal de que algo anduviera mal.

Es exactamente el defecto que peor se detecta: el que solo se ve si uno hace la
cuenta aparte.
"""
import io

from openpyxl import load_workbook

from app.export.owner_excel import export_owner_report_to_excel

FILAS = [
    {"label": "Total Revenue", "value": 4_000_000.0},
    {"label": "Payroll", "value": -1_200_000.0},
    {"label": "Operating Expenses", "value": -900_000.0},
    {"label": "GOP", "value": 1_900_000.0},
]


def _hoja():
    contenido = export_owner_report_to_excel(
        "BUDGET v1 2027",
        {"occ_pct": 62.5, "adr": 187.43, "revpar": 117.14, "pax": 12_410},
        FILAS, [],
    )
    return load_workbook(io.BytesIO(contenido)).worksheets[0]


def _por_etiqueta(ws) -> dict:
    return {ws.cell(f, 1).value: ws.cell(f, 2).value
            for f in range(1, ws.max_row + 1) if ws.cell(f, 1).value}


def test_los_gastos_quedan_negativos_en_la_celda():
    valores = _por_etiqueta(_hoja())
    assert valores["Payroll"] == -1_200_000.0
    assert valores["Operating Expenses"] == -900_000.0


def test_sumar_la_columna_da_el_gop():
    """La cuenta que haría cualquiera en Excel. Con el valor absoluto daba
    7,100,000 en vez de 1,900,000."""
    v = _por_etiqueta(_hoja())
    suma = v["Total Revenue"] + v["Payroll"] + v["Operating Expenses"]
    assert suma == v["GOP"] == 1_900_000.0


def test_el_negativo_igual_se_ve_entre_parentesis():
    """El formato de tres secciones ya lo hace sin tocar el número: no hacía
    falta romper el dato para que se viera bien."""
    ws = _hoja()
    fila = next(f for f in range(1, ws.max_row + 1)
                if ws.cell(f, 1).value == "Payroll")
    fmt = ws.cell(fila, 2).number_format
    assert fmt.count(";") == 2 and "(" in fmt.split(";")[1]
