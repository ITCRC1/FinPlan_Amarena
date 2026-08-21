# -*- coding: utf-8 -*-
"""
LAS PLANTILLAS DEVUELVEN EXACTAMENTE LO QUE SE SUBIÓ.

**La norma (owner, 2026-08-13):** «yo bajo, corrijo y subo lo que guardé».

Estas pantallas no exportan un reporte: exportan una plantilla que se baja, se
corrige y se vuelve a subir. Y la importación **borra las filas antes de
escribir las nuevas**, así que todo lo que la plantilla no lleve se pierde en el
viaje — sin dar error, porque lo que sobrevive alcanza para que los totales
cuadren.

Lo que se encontró y se corrigió:

* **no-operativos** — `account_code` no salía en la plantilla ni se restauraba
  al subir: bajar y subir dejaba en blanco la cuenta 8xxx de TODA línea
  below-GOP. No daba error porque la línea del P&L la decide
  `report_line_code`, no la cuenta.
* **no-operativos** — el correlativo se recalculaba en cada importación, así que
  el código de una fila cambiaba según en qué posición hubiera quedado.
* **no-operativos** — una fila con código y montos pero SIN descripción se
  descartaba. La pantalla permite dejarla en blanco.

Y una trampa que apareció al arreglarlo: la fila de encabezados se buscaba con
«contiene», y el texto de instrucciones nombra «Código» y «ENE-DIC» dentro de
una frase. Se la tragaba como cabecera y leía los encabezados de verdad como si
fueran datos.
"""
from decimal import Decimal

import pytest

from app.export.nonop_excel import export_nonop_to_excel, import_nonop_from_excel

MESES = ["jan", "feb", "mar", "apr", "may", "jun",
         "jul", "aug", "sep", "oct", "nov", "dec"]


def _fila(codigo: str, cuenta: str, desc: str, base: float = 0.0) -> dict:
    return {"detail_code": codigo, "account_code": cuenta, "detail_desc": desc,
            **{m: base + i for i, m in enumerate(MESES)}}


ENTRADAS = {
    "RENT": [_fila("007", "8010", "Patente municipal", 100.0),
             _fila("012", "8015", "", 0.0)],                # sin descripción, a propósito
    "DEPRECIATION": [_fila("001", "8040", "Depreciación equipo", 50.0)],
}


@pytest.fixture(scope="module")
def vuelta() -> list[dict]:
    return import_nonop_from_excel(export_nonop_to_excel(ENTRADAS, "BUDGET v1", 2027))


def test_no_se_pierde_ni_se_inventa_una_fila(vuelta):
    """Ni una de menos —el TOTAL colándose como dato sería una de más."""
    assert len(vuelta) == 3, [f["detail_desc"] for f in vuelta]
    assert not any(f["detail_code"].upper() in ("TOTAL", "SUBTOTAL") for f in vuelta)


def test_la_cuenta_sobrevive_el_viaje(vuelta):
    """El defecto más caro: sin esto, la cuenta 8xxx de toda línea below-GOP
    quedaba en blanco después de bajar y subir."""
    assert {f["account_code"] for f in vuelta} == {"8010", "8015", "8040"}


def test_el_codigo_no_se_renumera(vuelta):
    """Se respeta el que trae la fila. Antes se recalculaba por posición, así
    que el código de una fila cambiaba con solo reordenar el Excel."""
    assert {f["detail_code"] for f in vuelta} == {"007", "012", "001"}


def test_la_fila_sin_descripcion_no_desaparece(vuelta):
    """La pantalla deja dejarla en blanco; el viaje la borraba."""
    vacias = [f for f in vuelta if f["detail_desc"] == ""]
    assert len(vacias) == 1 and vacias[0]["account_code"] == "8015"


def test_los_montos_llegan_completos_y_exactos(vuelta):
    por_cuenta = {f["account_code"]: f for f in vuelta}
    assert Decimal(por_cuenta["8010"]["jan"]) == Decimal("100")
    assert Decimal(por_cuenta["8010"]["dec"]) == Decimal("111")   # los doce meses
    assert Decimal(por_cuenta["8040"]["dec"]) == Decimal("61")


def test_cada_hoja_conserva_su_linea(vuelta):
    assert {f["report_line_code"] for f in vuelta} == {"RENT", "DEPRECIATION"}


def test_un_archivo_sin_la_columna_cuenta_sigue_subiendo():
    """Las columnas se ubican por ENCABEZADO, no por posición.

    Si se leyeran por posición, agregar una columna rompería todos los archivos
    que la gente ya tenga bajados: los corrige durante días y al subirlos los
    datos entrarían corridos una celda, sin dar error. Acá se simula un archivo
    viejo borrándole la columna «Cuenta».
    """
    import io
    from openpyxl import load_workbook
    from app.export.excel_base import workbook_to_bytes

    wb = load_workbook(io.BytesIO(export_nonop_to_excel(ENTRADAS, "BUDGET v1", 2027)))
    for ws in wb.worksheets:
        ws.delete_cols(2)          # la columna «Cuenta»
    viejo = import_nonop_from_excel(workbook_to_bytes(wb))

    assert len(viejo) == 3, "el archivo viejo dejó de subir"
    # La cuenta no viene —no existe en ese archivo—, pero NADA se corrió.
    assert {f["detail_desc"] for f in viejo} == {
        "Patente municipal", "", "Depreciación equipo"}
    assert {f["detail_code"] for f in viejo} == {"007", "012", "001"}


# ── OPEX ─────────────────────────────────────────────────────────────────────
# La pantalla muestra una columna «#» (`detail_code`) que la plantilla no
# llevaba. Al reimportar se asignaba un correlativo nuevo POR POSICIÓN, así que
# reordenar el Excel le cambiaba el código a una fila que nadie había tocado.

from app.export.opex_excel import export_opex_to_excel, import_opex_from_excel


def _opex_entry(cuenta: str, nombre: str, desc: str, codigo: str, base: float) -> dict:
    return {"account_code": cuenta, "account_name": nombre, "detail_desc": desc,
            "detail_code": codigo, "currency": "USD",
            **{m: base + i for i, m in enumerate(MESES)},
            **{f"crc_{m}": 0.0 for m in MESES}}


OPEX = {"0110": [
    _opex_entry("7065", "Cleaning Supplies", "Detergente", "003", 200.0),
    _opex_entry("7065", "Cleaning Supplies", "Cloro", "007", 50.0),
    _opex_entry("7400", "Operating Supplies", "Varios", "001", 10.0),
]}


@pytest.fixture(scope="module")
def opex_vuelta() -> list[dict]:
    xlsx = export_opex_to_excel(OPEX, "BUDGET v1 2027", 2027,
                                dept_names={"0110": "Habitaciones"})
    return import_opex_from_excel(xlsx)


def test_opex_conserva_el_numero_de_cada_linea(opex_vuelta):
    """Dos filas de la MISMA cuenta con códigos distintos: si se renumerara por
    posición, la «007» volvería como «002»."""
    por_desc = {f["detail_desc"]: f for f in opex_vuelta}
    assert por_desc["Detergente"]["detail_code"] == "003"
    assert por_desc["Cloro"]["detail_code"] == "007"
    assert por_desc["Varios"]["detail_code"] == "001"


def test_opex_no_pierde_filas_ni_montos(opex_vuelta):
    assert len(opex_vuelta) == 3
    por_desc = {f["detail_desc"]: f for f in opex_vuelta}
    assert Decimal(por_desc["Detergente"]["dec"]) == Decimal("211.00")


def test_opex_un_archivo_sin_la_columna_del_numero_sigue_subiendo():
    """Un archivo bajado antes de que existiera la columna. Se lee con
    `len(row) >= COL_NUM`, igual que la moneda: si no viene, la fila entra sin
    código —y el correlativo se lo asigna la API— en vez de romper."""
    import io
    from openpyxl import load_workbook
    from app.export.excel_base import workbook_to_bytes
    from app.export.opex_excel import COL_NUM

    wb = load_workbook(io.BytesIO(export_opex_to_excel(
        OPEX, "BUDGET v1 2027", 2027, dept_names={"0110": "Habitaciones"})))
    for ws in wb.worksheets:
        ws.delete_cols(COL_NUM)
    viejo = import_opex_from_excel(workbook_to_bytes(wb))

    assert len(viejo) == 3, "el archivo viejo dejó de subir"
    assert {f["detail_desc"] for f in viejo} == {"Detergente", "Cloro", "Varios"}
    assert all(f["detail_code"] == "" for f in viejo)
