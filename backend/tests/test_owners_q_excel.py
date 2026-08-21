"""El export de Owners Q: que el archivo salga, y que salga donde SCP lo espera.

SCP consolida por POSICIÓN de fila. Si la fila 21 del archivo no es TOTAL
OPERATING REVENUE, su consolidación se desalinea entera y nadie se entera hasta
que los números de la cadena no cierran. Por eso esto se prueba.
"""
import io
import json
import pathlib
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.export.owners_q_excel import export_owners_q, nombre_archivo

SEED = pathlib.Path(__file__).parent.parent / "app" / "seed_data" / "owners_q.json"


@pytest.fixture(scope="module")
def datos():
    filas = json.loads(SEED.read_text(encoding="utf-8"))["report_lines"]
    from app.engine.owners_q import COLUMNAS
    return {
        "entidad": "CWL", "anio": 2026, "mes": 6, "convencion": "favorable",
        "filas": [{
            "row_no": f["row_no"], "report_code": f["report_code"],
            "label": f["label"], "indent": f["indent"],
            "line_type": f["line_type"], "nature": f["nature"],
            "estilo": f.get("estilo") or {},
            "celdas": {c: ("-1234.56" if c == "A" else "1000.5") for c in COLUMNAS},
        } for f in filas],
    }


def test_el_archivo_abre_y_tiene_la_hoja_de_scp(datos):
    wb = load_workbook(io.BytesIO(export_owners_q(datos)))
    assert wb.sheetnames[0] == "SCPCWL"
    assert wb["SCPCWL"]["A1"].value == "SCP CWL"
    assert wb["SCPCWL"]["A2"].value == "Statement of Income"


def test_cada_fila_cae_en_su_numero_de_fila_del_excel(datos):
    """`row_no` ES la fila del archivo — es lo que hace comparables los dos."""
    ws = load_workbook(io.BytesIO(export_owners_q(datos)))["SCPCWL"]
    for f in datos["filas"]:
        etiqueta = ws.cell(row=f["row_no"], column=17).value
        assert etiqueta.strip() == f["label"].strip(), \
            f"fila {f['row_no']} debería ser {f['label']!r} y es {etiqueta!r}"
    # Los anclajes que usa SCP para consolidar.
    assert ws.cell(row=21, column=17).value.strip() == "TOTAL OPERATING REVENUE"
    assert ws.cell(row=38, column=17).value.strip() == "GROSS OPERATING PROFIT"
    assert ws.cell(row=56, column=17).value.strip() == "Total NET INCOME BEFORE TAXES"


def test_la_etiqueta_va_en_la_columna_Q(datos):
    ws = load_workbook(io.BytesIO(export_owners_q(datos)))["SCPCWL"]
    assert ws.cell(row=21, column=17).column_letter == "Q"


def test_la_sangria_se_conserva(datos):
    ws = load_workbook(io.BytesIO(export_owners_q(datos)))["SCPCWL"]
    # f16 ROOMS REVENUE tiene sangría 2; f21 el subtotal, sangría 1.
    assert ws.cell(row=16, column=17).value.startswith("    ")
    assert not ws.cell(row=21, column=17).value.startswith("    ")


def test_los_negativos_van_en_rojo_y_entre_parentesis(datos):
    """Como en el archivo del owner: por FORMATO NUMÉRICO, no pintando la letra.

    Es la diferencia entre un rojo que se queda pegado y uno que sigue al valor:
    si el número deja de ser negativo, el formato lo devuelve a negro solo.
    """
    ws = load_workbook(io.BytesIO(export_owners_q(datos)))["SCPCWL"]
    c = ws.cell(row=21, column=1)          # columna A, sembrada en -1234.56
    assert c.value == pytest.approx(-1234.56)
    assert "[Red]" in c.number_format
    assert c.number_format.startswith('"$"#,##0.00')
    # La letra NO lleva color propio — el original tampoco.
    assert c.font.color is None or c.font.color.rgb in (None, "FF000000")


def test_el_nombre_no_dice_owners_q():
    """El nombre interno no viaja al entregable."""
    n = nombre_archivo("CWL", 2026, 6)
    assert n == "SCP_CWL_JUN26_Statement_of_Income.xlsx"
    assert "owners" not in n.lower()


@pytest.fixture
def anio_datos(datos):
    from app.api.owners_q_api import PERIODOS_ANIO

    def periodos(marca):
        return {c: {"etiqueta": e, "acumulado": a, "meses": m,
                    "valores": {f["report_code"]: marca for f in datos["filas"]}}
                for c, m, e, a in PERIODOS_ANIO}

    return {
        "entidad": "CWL", "anio": 2026,
        "orden_periodos": [c for c, _m, _e, _a in PERIODOS_ANIO],
        "filas": [{k: f[k] for k in ("row_no", "report_code", "label",
                                     "indent", "line_type", "nature")}
                  for f in datos["filas"]],
        "datasets": {
            "actual": {"escenario_id": "a", "periodos": periodos("100")},
            "budget": {"escenario_id": "b", "periodos": periodos("200")},
            "py": {"escenario_id": "p", "periodos": periodos("300")},
        },
    }


def test_hay_una_hoja_de_meses_y_otra_de_trimestres(datos, anio_datos):
    wb = load_workbook(io.BytesIO(export_owners_q(datos, anio_datos)))
    assert wb.sheetnames == ["SCPCWL", "Meses", "Trimestres y Año"]


def test_el_set_es_actual_budget_lastyear_y_sus_acumulados(datos, anio_datos):
    """Pedido del owner: seis columnas por período."""
    ws = load_workbook(io.BytesIO(export_owners_q(datos, anio_datos)))["Meses"]
    assert [ws.cell(row=3, column=j).value for j in range(2, 8)] == \
        ["ACTUAL", "BUDGET", "LAST YEAR", "ACUM ACT", "ACUM BUD", "ACUM LY"]
    assert ws.cell(row=2, column=2).value == "ENE"
    assert ws.cell(row=2, column=8).value == "FEB"          # el set siguiente
    # 12 meses × 6 columnas + la de etiquetas.
    assert ws.cell(row=2, column=2 + 11 * 6).value == "DIC"


def test_cada_columna_del_set_trae_su_dataset(datos, anio_datos):
    """Que ACTUAL traiga actual y BUDGET traiga budget — no todo lo mismo."""
    ws = load_workbook(io.BytesIO(export_owners_q(datos, anio_datos)))["Meses"]
    fila_gop = next(i for i, f in enumerate(anio_datos["filas"], start=5)
                    if f["report_code"] == "GOP")
    assert [ws.cell(row=fila_gop, column=j).value for j in range(2, 8)] == \
        [100, 200, 300, 100, 200, 300]


def test_los_trimestres_traen_el_ano_completo(datos, anio_datos):
    ws = load_workbook(io.BytesIO(export_owners_q(datos, anio_datos)))["Trimestres y Año"]
    bandas = [ws.cell(row=2, column=2 + k * 6).value for k in range(5)]
    assert bandas == ["Q1", "Q2", "Q3", "Q4", "FULL YEAR"]
    # El año completo no tiene acumulado aparte: sus tres columnas de acumulado
    # repiten el año, no salen vacías.
    fila_gop = next(i for i, f in enumerate(anio_datos["filas"], start=5)
                    if f["report_code"] == "GOP")
    c0 = 2 + 4 * 6
    assert [ws.cell(row=fila_gop, column=c0 + k).value for k in range(6)] == \
        [100, 200, 300, 100, 200, 300]


def test_los_periodos_del_anio_son_33(datos):
    from app.api.owners_q_api import PERIODOS_ANIO
    assert len(PERIODOS_ANIO) == 33           # 12×2 + 4×2 + 1


def test_los_periodos_acumulan_los_meses_correctos():
    from app.api.owners_q_api import PERIODOS_ANIO

    d = {c: m for c, m, _e, _a in PERIODOS_ANIO}
    assert d["M01"] == [1]
    assert d["M06"] == [6]
    assert d["M06_ACUM"] == [1, 2, 3, 4, 5, 6]
    assert d["Q2"] == [4, 5, 6]
    assert d["Q2_ACUM"] == [1, 2, 3, 4, 5, 6]   # el semestre
    assert d["Q4_ACUM"] == list(range(1, 13))
    assert d["FY"] == list(range(1, 13))
