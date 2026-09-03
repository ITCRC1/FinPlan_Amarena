# -*- coding: utf-8 -*-
"""Un solo Excel con TODOS los sub-tabs.

Owner, 2026-09-03: *«se podrá bajar en Excel todos los tabs; que bajen todos es
todos, en un solo archivo»*.
"""
import re
from pathlib import Path

FRONT = Path(__file__).resolve().parents[2] / "frontend"
PAGINA = FRONT / "app/month-end/pl/page.tsx"


def _src() -> str:
    return PAGINA.read_text(encoding="utf-8")


def _cuerpo() -> str:
    s = _src()
    return s[s.index("async function bajarExcel()"):s.index("async function bajarWord()")]


def test_sale_del_MISMO_registro_que_el_Word():
    """⚠️ Un segundo armado sería un segundo lugar donde olvidarse un sub-tab —
    que es exactamente el defecto que el Word acaba de tener."""
    assert "CAPITULOS[v.key]" in _cuerpo()


def test_NO_filtra_por_los_escondidos():
    """«Todos es todos». El Word es lo que ve el dueño y por eso respeta el
    panel de Vistas; esto es el respaldo para trabajar, y ahí esconder una hoja
    no ayuda a nadie."""
    cuerpo = _cuerpo()
    assert "for (const v of VISTAS)" in cuerpo
    assert "subOcultos" not in cuerpo, (
        "el Excel completo empezó a respetar el panel de Vistas: deja de ser "
        "«todos»")


def test_INCLUYE_los_cuadros_sin_datos():
    """En un Word una página en cero se lee como «el mes no tuvo movimiento»;
    en Excel una hoja vacía se ve vacía, y sacarla dejaría la duda de si el tab
    existe."""
    assert "tieneDatos" not in _cuerpo()


def test_cada_cuadro_va_a_su_HOJA():
    """El capítulo manda su nombre de pestaña si lo tiene; si no, el título."""
    assert "hoja: c.hoja || c.titulo" in _cuerpo()


def test_no_se_baja_con_la_pantalla_a_medio_cargar():
    """El mismo motivo que en el Word: la mitad de los capítulos lee el estado
    de la pantalla y la otra mitad pide lo suyo con `await`."""
    assert "if (!datos.length || !gastos.length)" in _cuerpo()


def test_avisa_cuales_no_se_pudieron_armar():
    """Un archivo con quince hojas se ve completo aunque falten dos."""
    cuerpo = _cuerpo()
    assert "fallaron.push" in cuerpo and "fallaron.join" in cuerpo


def test_es_EL_boton_de_Excel_y_no_uno_aparte():
    """Owner, 2026-09-03: «necesito que el botón que está a la par de download
    de Word, de Excel, ahí estén todos los tabs, en el mismo orden».

    ⚠️ Antes ese botón bajaba UNA sola hoja —el P&L— y el intento anterior fue
    agregar un segundo botón al lado. Dos descargas parecidas dejan al usuario
    eligiendo entre ellas, que es una decisión que nadie quiere tomar.
    """
    src = _src()
    assert "onClick={bajarExcel}" in src
    assert "bajarExcelTodo" not in src, "quedó el segundo botón"
    # Uno solo en la BARRA de arriba. Los otros «⬇ Excel» del archivo son los
    # de cada sub-tab —bajan su propio cuadro— y ésos sí van.
    assert src.count("onClick={bajarExcel}") == 1
    # Y el viejo de una sola hoja no puede volver.
    assert "[cuadroPL()]);" not in src


def test_el_ORDEN_es_el_de_la_pantalla():
    """«En el mismo orden»: se recorre `VISTAS`, la misma lista que dibuja la
    fila de sub-tabs."""
    assert "for (const v of VISTAS)" in _cuerpo()


def test_el_PL_Statement_aporta_sus_DOS_vistas_tambien_al_Excel():
    """Owner: «no veo que el detallado en Word salga».

    ⚠️ El capítulo del P&L Statement devuelve DOS cuadros —Totales y
    Departamental— y el Excel usa el mismo registro, así que las dos son dos
    hojas. Si alguien volviera a dejar uno solo, se perdería en los dos
    formatos a la vez.
    """
    src = _src()
    assert "[cuadroEstado(false), cuadroEstado(true)]" in src
    # Y el desglose por departamento está DENTRO del cuadro, no sólo en la
    # pantalla: es lo que hace distinta a la versión departamental.
    cuadro = src[src.index("function cuadroEstado("):src.index("function cuadroSummary")]
    assert "conDepto ? desglose(f.code) : []" in cuadro
    assert '" — Departamental" : " — Totales"' in cuadro


# ─── Que el libro se pueda LEER (owner: «bien profesional y claro») ─────────

def test_el_libro_trae_un_INDICE_adelante():
    """⚠️ El Word tiene su página de CONTENIDO; un libro de doce hojas sin
    índice obliga a recorrer las pestañas de abajo una por una — y los nombres
    van cortados a 31 caracteres, así que ni se leen enteros. El índice es
    donde el título completo cabe."""
    import io as _io

    import openpyxl

    from app.export.cuadro_excel import build_cuadros_workbook
    cuadro = lambda t: {
        "titulo": t, "subtitulo": "Julio 2026 · USD",
        "columnas": [{"label": "Line Item", "ancho": 30, "formato": "texto"},
                     {"label": "ACTUAL", "ancho": 16, "formato": "usd2"}],
        "filas": [{"label": "Rooms", "es_total": False, "valores": [1.0]}]}
    b = build_cuadros_workbook([cuadro("Uno"), cuadro("Dos")])
    wb = openpyxl.load_workbook(_io.BytesIO(b))
    assert wb.sheetnames[0] == "Índice"
    texto = " ".join(str(c.value) for r in wb["Índice"].iter_rows(values_only=True)
                     for c in [type("X", (), {"value": v})() for v in r] if c.value)
    assert "Uno" in texto and "Dos" in texto


def test_con_UNA_sola_hoja_no_hay_indice():
    """Una portada que dice «1. esa hoja» es un clic de más para llegar al
    único cuadro."""
    import io as _io

    import openpyxl

    from app.export.cuadro_excel import build_cuadros_workbook
    b = build_cuadros_workbook([{
        "titulo": "Solo", "columnas": [{"label": "A", "ancho": 20, "formato": "texto"}],
        "filas": [{"label": "x", "es_total": False, "valores": []}]}])
    wb = openpyxl.load_workbook(_io.BytesIO(b))
    assert "Índice" not in wb.sheetnames


def test_el_indice_muestra_el_nombre_REAL_de_la_pestana():
    """Si dos cuadros se llamaban parecido, el libro los desambiguó: el índice
    tiene que mostrar el nombre con el que quedó la pestaña o no sirve para
    encontrarla."""
    import inspect

    from app.export import cuadro_excel
    fuente = inspect.getsource(cuadro_excel.build_cuadros_workbook)
    assert "nombres.append(_hoja(wb, cuadro, usados).title)" in fuente


def test_las_dos_hojas_del_Statement_se_DISTINGUEN():
    """⚠️ Excel corta el nombre de pestaña en 31 caracteres. Con el título
    largo, las dos salían «Profit & Loss Statement YTD JUL» y «…YTD (2)»:
    imposible saber cuál es la departamental sin abrirlas."""
    src = _src()
    assert 'hoja: conDepto ? "P&L Departamental" : "P&L Totales"' in src
    assert "c.hoja || c.titulo" in _cuerpo(), (
        "el Excel dejó de respetar el nombre de pestaña que manda cada "
        "capítulo y volvió a cortar títulos largos")
