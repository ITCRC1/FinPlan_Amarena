# -*- coding: utf-8 -*-
"""El checkbook de gastos conserva el formato del archivo del owner.

**El encargo (owner, 2026-08-18).** Mandó el paquete completo —formato, motor
que genera, motor que lee— y después: «nunca pierdas la perspectiva del Excel
que subí para que no pierdas ese formato», con su
`CHECKBOOK MADRESAL 2026.xlsx` de referencia.

Por eso `app/checkbook/build.py` y `read.py` se instalaron **sin tocarles la
lógica**: están validados contra ese archivo (1.184 fórmulas, 0 errores) y
cualquier retoque nuestro se despegaría de la especificación. Lo único que se
adaptó fue el router.

**La geometría es un contrato**, no una preferencia. El SUMMARY resuelve sus
totales con referencias directas —`U10 ='BUDGET 2027 Detail'!S28`— así que mover
un bloque una fila rompe el archivo en silencio: las fórmulas siguen
calculando, pero de la celda equivocada.
"""
import os
import tempfile

import openpyxl
import pytest

from app.checkbook import build

CFG = {
    "departamento": "Gastos Operativos Club Madresal (600)",
    "codigo_departamento": "600",
    "anio_version": 2027,
    "detalles_por_cuenta": 11,
    "detalle_inicial": 800,
    "proteger": False,
    "estadisticas": {},
    "referencias": {},
    "cuentas": [{"cuenta": 7030, "descripcion": "Building"},
                {"cuenta": 7050, "descripcion": "Centralized Accounting Charges"}],
}


def _gen(cfg) -> openpyxl.Workbook:
    t = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    t.close()
    try:
        build(cfg, t.name, force=True)
        return openpyxl.load_workbook(t.name)
    finally:
        try:
            os.unlink(t.name)
        except OSError:
            pass


def test_las_dos_hojas_y_sus_nombres():
    wb = _gen(dict(CFG))
    assert wb.sheetnames == ["BUDGET 2027 Detail", "SUMMARY"]


def test_la_geometria_de_los_bloques_no_se_mueve():
    """headers 15, 33, 51 … y TOTAL 27, 45, 63 — igual que el archivo del owner.

    El SUMMARY apunta a esas filas por número. Moverlas rompe el archivo sin
    que ninguna fórmula falle.
    """
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert ws.cell(15, 2).value == "# Cuenta"
    assert ws.cell(27, 3).value == "TOTAL 2027"
    assert ws.cell(33, 7).value is not None      # header del 2º bloque
    assert ws.cell(45, 3).value == "TOTAL 2027"  # TOTAL del 2º bloque


def test_las_columnas_son_las_del_original():
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert [ws.cell(15, c).value for c in range(2, 7)] == [
        "# Cuenta", "Descripcion de Cuenta", "Departamento",
        "Detalle", "Detalle Descripcion"]


def test_los_anchos_del_original_se_respetan():
    """B 9 · C 27.57 · D 14 · E 7.43 · S 16.71 — medidos del archivo del owner.
    Son parte de que el archivo se vea como el suyo al abrirlo.

    ⚠️ **La F es la única que se desvía, y a propósito.** En el original mide
    64.29, pero ahí nada estaba congelado. Con el panel fijo en G16, las
    columnas A–F sumaban ~122 caracteres —más ancho que la ventana— y Excel
    deja de poder desplazarse: «no logro ver los 12 meses, se devuelve al
    principio el cursor» (owner, 18-ago-2026). A 42 el texto sigue legible y
    quedan ~60 caracteres de aire para los meses.
    """
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    esperados = {"B": 9, "C": 27.57, "D": 14, "E": 7.43, "S": 16.71}
    for col, ancho in esperados.items():
        assert round(ws.column_dimensions[col].width, 2) == pytest.approx(ancho, abs=0.02), col
    ancho_f = ws.column_dimensions["F"].width
    assert ancho_f == pytest.approx(42, abs=0.5), "la F se angostó a propósito"
    congelado = sum(ws.column_dimensions[c].width for c in "ABCDEF")
    assert congelado < 105, (
        f"las columnas congeladas suman {congelado:.0f} caracteres; por encima "
        f"de ~105 Excel no puede desplazarse a los últimos meses")


def test_la_fila_TOTAL_deja_la_columna_B_vacia():
    """⚠️ La regla crítica del formato. El SUMMARY suma con
    `SUMIF(B:B, cuenta, ...)`: si la fila TOTAL llevara el número de cuenta, el
    SUMIF la sumaría junto con sus once líneas y el total saldría al doble."""
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert ws.cell(27, 2).value is None
    assert ws.cell(28, 2).value is None      # TOTAL 2026 (referencia)


def test_el_total_de_cada_linea_es_una_formula():
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert ws.cell(16, 19).value == "=SUM(G16:R16)"
    assert ws.cell(27, 7).value == "=SUM(G16:G26)"


# ── La precarga: lo que convierte el archivo en una ida y vuelta ─────────────
#
# ⚠️ AGREGADO el 18-ago-2026. El motor escribía `None` en todas las celdas de
# captura, así que bajar el checkbook de un departamento que YA tiene
# presupuesto daba una hoja en blanco: había que volver a teclear lo que FinPlan
# ya sabe, y las descripciones de detalle que el owner tiene escritas en su
# archivo (columna F, «Mantenimiento…») se perdían en cada regeneración.
#
# Es OPCIONAL a propósito: sin `lineas`, el archivo sale idéntico al validado.

def test_sin_lineas_el_archivo_sale_como_estaba_validado():
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    for r in range(16, 27):
        assert ws.cell(r, 6).value is None                  # descripción
        assert all(ws.cell(r, c).value is None for c in range(7, 19))   # meses


def test_con_lineas_precarga_descripcion_y_montos():
    cfg = dict(CFG, lineas={"7030": {
        "800": {"descripcion": "Mantenimiento techos",
                "montos": [100, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 50]},
        "806": {"descripcion": "Pintura", "montos": [0] * 12}}})
    ws = _gen(cfg)["BUDGET 2027 Detail"]
    assert ws.cell(16, 6).value == "Mantenimiento techos"
    assert ws.cell(16, 7).value == 100          # enero
    assert ws.cell(16, 18).value == 50          # diciembre
    assert ws.cell(22, 6).value == "Pintura"    # detalle 806 → fila 16+6


def test_una_linea_sin_dato_queda_en_blanco_para_capturar():
    """No se rellena con ceros: una celda vacía invita a escribir, un cero
    parece un dato ya decidido."""
    cfg = dict(CFG, lineas={"7030": {"800": {"descripcion": "X", "montos": [1] * 12}}})
    ws = _gen(cfg)["BUDGET 2027 Detail"]
    assert ws.cell(18, 6).value is None          # detalle 802, sin precarga
    assert ws.cell(18, 7).value is None


def test_la_precarga_no_mueve_la_geometria():
    cfg = dict(CFG, lineas={"7030": {"800": {"descripcion": "X", "montos": [9] * 12}}})
    ws = _gen(cfg)["BUDGET 2027 Detail"]
    assert ws.cell(27, 3).value == "TOTAL 2027"
    assert ws.cell(16, 19).value == "=SUM(G16:R16)"


def test_el_router_precarga_desde_opex_entries():
    """`opex_entries` tiene la MISMA granularidad que el archivo. La
    especificación decía que esa dimensión no existía en FinPlan; se escribió
    para otro FinPlan."""
    import inspect

    from app.api import checkbook_api

    src = inspect.getsource(checkbook_api._lineas_actuales)
    assert "OpexEntry" in src
    assert "detail_desc" in src


def test_la_precarga_mapea_por_POSICION_y_no_por_codigo():
    """⚠️ El defecto que casi se entrega, y era del tipo silencioso.

    Medido en producción (Budget 2027 Final, 417 cuentas): FinPlan usa códigos
    de detalle `''`, `001`…`011`; el formato del owner usa `800`…`810`. Buscar
    por igualdad no habría coincidido con NINGUNA línea —y 73 ni siquiera
    tienen código—: el archivo habría salido en blanco y se habría visto
    perfectamente normal.

    La línea n-ésima de una cuenta va a la ranura n-ésima. Es lo que el campo
    significa: «Detalle» es un renglón dentro de la cuenta, no una llave con
    sentido propio.
    """
    import inspect

    from app.api import checkbook_api

    src = inspect.getsource(checkbook_api._lineas_actuales)
    assert "enumerate(lineas[:detalles])" in src, "tiene que ir por posición"
    assert "str(det_ini + j)" in src, "la ranura sale de la posición, no del código"
    assert "if not det:" not in src, "ya no se descartan las líneas sin código"


def test_una_cuenta_con_mas_lineas_que_ranuras_se_REPORTA():
    """Hoy ninguna las supera —el máximo medido son 11, justo las que trae el
    formato— pero un archivo que pierde renglones en silencio es exactamente lo
    que no puede pasar."""
    import inspect

    from app.api import checkbook_api

    src = inspect.getsource(checkbook_api._lineas_actuales)
    assert "desbordes.append" in src
    src_prev = inspect.getsource(checkbook_api.preview)
    assert '"desbordes"' in src_prev, "el preview tiene que avisarlo antes de bajar"


def test_el_bloque_dice_QUE_suma_y_donde_corta():
    """⚠️ «No sé si la cuenta suma arriba o abajo» (owner, 18-ago-2026).

    El TOTAL del año de versión cierra las once líneas de ARRIBA; los años de
    abajo son referencia y no entran en esa suma. Sin nada que lo separe, el
    bloque se lee ambiguo — y el que lo llena no sabe si su número va a sumar.

    Dos señales, porque hacen cosas distintas: la línea gruesa dice DÓNDE corta,
    el rótulo dice POR QUÉ.
    """
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert "suma las 11 lineas de arriba" in str(ws.cell(27, 6).value)
    assert "no suma" in str(ws.cell(28, 6).value)
    assert "2026" in str(ws.cell(28, 6).value)


def test_la_linea_gruesa_separa_el_total_de_las_referencias():
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert ws.cell(27, 3).border.bottom.style == "thick", (
        "el TOTAL tiene que cortar con lo de abajo")
    assert ws.cell(27, 3).border.top.style == "medium", (
        "y cerrar lo que suma, arriba")


def test_el_bloque_cierra_al_final_de_las_referencias():
    """Para ver dónde termina una cuenta y empieza la siguiente."""
    ws = _gen(dict(CFG))["BUDGET 2027 Detail"]
    assert ws.cell(29, 3).border.bottom.style == "medium"   # última referencia


# ── El formato no se deteriora con más o menos cuentas ───────────────────────
#
# «Cuando el departamento tiene más o menos cuentas, el formato no se
# deteriora» (owner, 18-ago-2026). Medido: 1, 3, 12 y 40 cuentas, y 5/11/20
# líneas por cuenta. El primer bloque siempre arranca en la 15 y el paso solo
# depende de las líneas por cuenta, nunca de cuántas cuentas haya.

@pytest.mark.parametrize("n", [1, 3, 12, 40])
def test_el_primer_bloque_siempre_arranca_en_la_15(n):
    cfg = dict(CFG, cuentas=[{"cuenta": 7000 + i, "descripcion": f"C{i}"} for i in range(n)])
    ws = _gen(cfg)["BUDGET 2027 Detail"]
    assert ws.cell(15, 2).value == "# Cuenta"
    assert ws.cell(27, 3).value == "TOTAL 2027"


@pytest.mark.parametrize("n", [1, 3, 12, 40])
def test_el_paso_no_depende_de_cuantas_cuentas_haya(n):
    cfg = dict(CFG, cuentas=[{"cuenta": 7000 + i, "descripcion": f"C{i}"} for i in range(n)])
    ws = _gen(cfg)["BUDGET 2027 Detail"]
    totales = [r for r in range(15, ws.max_row + 1)
               if str(ws.cell(r, 3).value or "") == "TOTAL 2027"]
    assert len(totales) == n
    assert all(b - a == 18 for a, b in zip(totales, totales[1:])), totales


@pytest.mark.parametrize("ndet,paso", [(5, 12), (11, 18), (20, 27)])
def test_el_paso_sale_de_las_lineas_por_cuenta(ndet, paso):
    """`paso = detalles + 2 + N_REFS + blancos`. Es lo único que lo mueve."""
    cfg = dict(CFG, detalles_por_cuenta=ndet)
    ws = _gen(cfg)["BUDGET 2027 Detail"]
    totales = [r for r in range(15, ws.max_row + 1)
               if str(ws.cell(r, 3).value or "") == "TOTAL 2027"]
    assert totales[1] - totales[0] == paso


@pytest.mark.parametrize("n", [1, 3, 12, 40])
def test_el_encabezado_del_SUMMARY_no_se_mueve(n):
    """⚠️ Estaba quemado en la fila 9. Al agregar el costo por habitación
    ocupada de los años de referencia pasó a la 12 — y una fila fija habría
    escrito la tabla de cuentas ENCIMA de las estadísticas, sin fallar."""
    cfg = dict(CFG, cuentas=[{"cuenta": 7000 + i, "descripcion": f"C{i}"} for i in range(n)])
    sm = _gen(cfg)["SUMMARY"]
    assert sm.cell(12, 4).value == "# Cuenta"
    assert sm.cell(13, 4).value == 7000          # la primera cuenta, justo debajo


def test_el_SUMMARY_trae_el_costo_por_habitacion_de_los_TRES_anios():
    """«Mete también para 2026-2025 el costo, tal como 2027» (owner).

    Cada uno divide el gran total de SU año por las noches de SU año: comparar
    el costo de 2027 contra el de 2026 solo significa algo si los dos están por
    habitación ocupada — un año con menos ocupación gasta menos en total sin ser
    más eficiente.
    """
    sm = _gen(dict(CFG))["SUMMARY"]
    etiquetas = [str(sm.cell(r, 4).value or "") for r in range(5, 12)]
    assert "Rooms Occupied 2027" in etiquetas
    assert "Rooms Occupied 2026" in etiquetas
    assert "Rooms Occupied 2025" in etiquetas
    assert "Cost per Room Occupied 2027" in etiquetas
    assert "Cost per Room Occupied 2026" in etiquetas
    assert "Cost per Room Occupied 2025" in etiquetas


def test_cada_costo_usa_las_noches_de_SU_anio():
    """Cruzar los años daría un número plausible y equivocado."""
    sm = _gen(dict(CFG))["SUMMARY"]
    fila = {str(sm.cell(r, 4).value or ""): r for r in range(5, 12)}
    for anio, f_gt in ((2027, 9), (2026, 10), (2025, 11)):
        f_cpo = fila[f"Cost per Room Occupied {anio}"]
        f_noc = fila[f"Rooms Occupied {anio}"]
        assert f"!G{f_gt}/G{f_noc}" in str(sm.cell(f_cpo, 7).value), (anio, sm.cell(f_cpo, 7).value)


def test_el_numero_de_cuenta_no_lleva_coma():
    """«7,030» no es una cuenta, es un número. El SUMIF busca por el valor, así
    que el formato es solo lo que se ve — pero lo que se ve tiene que parecerse
    a una cuenta."""
    sm = _gen(dict(CFG))["SUMMARY"]
    assert sm.cell(13, 4).number_format == "0"


def test_las_columnas_de_referencia_dicen_el_ESCENARIO():
    """«En la columna U debe poner Forecast 2026, en la V Actual 2025» (owner).

    Un mismo año tiene Working, Draft y Final: «TOTAL 2026» no dice cuál es.
    """
    cfg = dict(CFG, etiquetas_ref={"2026": "Forecast 2026", "2025": "Actual 2025"})
    sm = _gen(cfg)["SUMMARY"]
    assert sm.cell(12, 21).value == "Forecast 2026"     # U
    assert sm.cell(12, 22).value == "Actual 2025"       # V


def test_sin_etiquetas_se_cae_a_TOTAL_del_anio():
    sm = _gen(dict(CFG))["SUMMARY"]
    assert sm.cell(12, 21).value == "TOTAL 2026"


def test_el_SUMMARY_congela_en_enero_y_bajo_el_encabezado():
    sm = _gen(dict(CFG))["SUMMARY"]
    assert sm.freeze_panes == "G13"


@pytest.mark.parametrize("hoja", ["BUDGET 2027 Detail", "SUMMARY"])
def test_la_proteccion_impide_editar_pero_no_navegar(hoja):
    """⚠️ `selectLockedCells=True` significa lo CONTRARIO de lo que parece: que
    el usuario NO puede seleccionar celdas bloqueadas.

    Con eso puesto, las flechas y el Tab saltaban solo entre celdas de captura
    y al llegar a la última volvían a la primera: «llego al final y se vuelve al
    inicio» (owner, 18-ago-2026). No se podía recorrer la hoja ni leer los
    totales — y el archivo se veía perfectamente normal.
    """
    cfg = dict(CFG, proteger=True)
    ws = _gen(cfg)[hoja]
    assert ws.protection.sheet is True, "la hoja tiene que seguir protegida"
    assert ws.protection.selectLockedCells is False, (
        "con True no se puede navegar la hoja; la protección es para no editar")


def test_los_titulos_van_centrados():
    """La fila 1 del Detail y la 2 del SUMMARY son bandas que cruzan la hoja:
    centradas se leen como un título, alineadas a la izquierda parecen una
    celda más."""
    wb = _gen(dict(CFG))
    assert wb["BUDGET 2027 Detail"].cell(1, 2).alignment.horizontal == "center"
    assert wb["SUMMARY"].cell(2, 4).alignment.horizontal == "center"


def test_el_nombre_del_archivo_sale_de_la_fila_1():
    """«¿Será que el nombre del archivo puede adoptarse según la línea 1 del tab
    Budget?» (owner, 18-ago-2026).

    Esa fila ya identifica el archivo —departamento, año y para qué es— así que
    el nombre deja de ser una convención aparte que hay que recordar.
    """
    from app.api.checkbook_api import _nombre_archivo

    n = _nombre_archivo({"departamento": "Mantenimiento (0200)", "anio_version": 2027})
    assert n == "MANTENIMIENTO (0200) - PRESUPUESTO 2027 - CHECKBOOK DE GASTOS.xlsx"


def test_el_nombre_aguanta_un_departamento_con_barra():
    """`A&B / Restaurante` tiene una barra: Windows no la acepta en un nombre de
    archivo y la descarga fallaría sin decir por qué."""
    from app.api.checkbook_api import _nombre_archivo

    n = _nombre_archivo({"departamento": "A&B / Restaurante (0120)", "anio_version": 2028})
    assert "/" not in n and "\\" not in n
    assert n.endswith(".xlsx")


# ── Elegir a mano el escenario de cada año de referencia ─────────────────────

def test_los_pares_ano_escenario_se_leen_bien():
    """«Necesito que estos escenarios den la oportunidad de cambiarlas si yo
    quisiera» (owner, 18-ago-2026). La regla —Forecast el anterior, Actual el
    previo— es un default razonable, no una imposición."""
    from app.api.checkbook_api import _refs

    assert _refs(["2026:abc", "2025:def"]) == {"2026": "abc", "2025": "def"}
    assert _refs(["basura", "2026:", ":abc", ""]) == {}, "lo mal formado se ignora"
    assert _refs([]) == {}


def test_un_escenario_de_OTRO_ano_se_rechaza():
    """Elegir el Budget 2024 como referencia de 2026 daría un archivo que dice
    «TOTAL 2026» con números de 2024, y nada fallaría."""
    import inspect

    from app.api import checkbook_api

    src = inspect.getsource(checkbook_api._armar_config)
    assert "ref.year != anio" in src
    # El motivo ya no se escribe acá: vive en el catálogo bilingüe, y lo que
    # este archivo tiene que seguir haciendo es NOMBRARLO al cortar.
    assert "checkbook.referencia_de_otro_ano" in src


def test_el_preview_ofrece_las_opciones_de_cada_ano():
    """Sin la lista, cambiar el escenario obligaría a saberse los ids."""
    import inspect

    from app.api import checkbook_api

    src = inspect.getsource(checkbook_api.preview)
    assert '"opciones"' in src
    assert '"escenario_id"' in src


def test_la_etiqueta_distingue_Working_de_Final():
    """«Forecast 2026» no alcanza si hay Working y April: la columna del SUMMARY
    tiene que decir cuál."""
    import inspect

    from app.api import checkbook_api

    src = inspect.getsource(checkbook_api._armar_config)
    assert "ref.version" in src and "etiquetas[str(anio)]" in src
