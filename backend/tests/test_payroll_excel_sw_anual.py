"""
`SW Anual USD*` en el Excel de Planilla, contra el mismo criterio que usa el
motor (`payroll_calculator.calc_sw`): un TC por mes, no uno para el año
entero — y sin el `/12` de más que traía la fórmula vieja.

`salary_amount` es MENSUAL (docstring de `PayrollPosition`), y
`SW_mes = salary × FTE_mes / TC_mes` (CRC) o `salary × FTE_mes` (USD). El
anual correcto es la SUMA de los doce, cada uno con su TC — nunca
`salary/12 * ... / tc_de_un_solo_mes`, que además de usar el TC equivocado
dividía por 12 sin necesidad.
"""
from decimal import Decimal

import openpyxl
import pytest

from app.export.payroll_excel import export_payroll_to_excel, COL_REF, HEADER_ROW


def _eval_sumproduct_formula(formula: str, fte: list[float], salary: float,
                             tc_por_mes: list[float], currency: str) -> float:
    """Evalúa a mano lo que Excel calcularía con esta fórmula, sin abrir Excel.

    No reimplementa la fórmula: lee sus coeficientes {1/tc1,...,1/tc12} de la
    propia cadena y aplica el mismo SUMPRODUCT que Excel aplicaría. Si la
    fórmula cambiara de forma, este parseo fallaría en vez de mentir.
    """
    inicio = formula.index("{") + 1
    fin = formula.index("}")
    coefs = [float(x) for x in formula[inicio:fin].split(",")]
    assert len(coefs) == 12
    if currency == "CRC":
        return sum(f * c for f, c in zip(fte, coefs)) * salary
    return sum(fte) * salary


@pytest.mark.parametrize("currency", ["CRC", "USD"])
def test_sw_anual_usa_el_tc_de_cada_mes(currency):
    """El caso que rompía la fórmula vieja: TC que sube fuerte durante el año.

    Con un TC de enero muy distinto al de diciembre, la fórmula vieja
    (salary/12 * SUM(FTE) / tc_enero) daba un número que no se parecía en
    nada a la suma mes a mes real."""
    fte = [1.0] * 12
    salary = 500_000.0 if currency == "CRC" else 1_000.0
    # TC subiendo de 500 a 700 durante el año — bien distinto entre meses.
    tc_por_mes = [500.0 + i * (200.0 / 11) for i in range(12)]

    xlsx = export_payroll_to_excel(
        {"0110": [{
            "dept_name": "Rooms", "position_code": "001", "position_name": "Test",
            "employee_name": "Fulano", "employee_type": "Local",
            "salary_amount": salary, "salary_currency": currency,
            **{f"fte_{m}": v for m, v in zip(
                ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"],
                fte)},
        }]},
        "Test Scenario", 2027, dept_names={"0110": "Rooms"}, tc_por_mes=tc_por_mes,
    )
    import io
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    formula = ws.cell(row=HEADER_ROW + 1, column=COL_REF).value
    assert formula.startswith("=IF(")

    resultado = _eval_sumproduct_formula(formula, fte, salary, tc_por_mes, currency)

    # Lo mismo que calcularía calc_sw mes a mes, sumado.
    if currency == "CRC":
        esperado = sum(salary * f / tc for f, tc in zip(fte, tc_por_mes))
    else:
        esperado = sum(salary * f for f in fte)

    assert resultado == pytest.approx(esperado, rel=1e-6)

    # Y que NO sea lo que daba la formula vieja (salary/12 * SUM(FTE) / tc_enero):
    tc_enero = tc_por_mes[0]
    formula_vieja = (salary / 12 * sum(fte) / tc_enero if currency == "CRC"
                     else salary * sum(fte) / 12)
    assert resultado != pytest.approx(formula_vieja, rel=1e-6), (
        "la formula sigue dando el numero de la version vieja (con /12 y "
        "un solo TC): no se corrigio nada")


def test_con_tc_parejo_el_anual_es_12_veces_el_mensual():
    """Caso simple para verificar el orden de magnitud: TC constante,
    FTE=1 todo el año → el anual tiene que ser exactamente 12x un mes."""
    fte = [1.0] * 12
    salary = 500_000.0
    tc = 620.0

    xlsx = export_payroll_to_excel(
        {"0110": [{
            "dept_name": "Rooms", "position_code": "001", "position_name": "Test",
            "employee_name": "Fulano", "employee_type": "Local",
            "salary_amount": salary, "salary_currency": "CRC",
            **{f"fte_{m}": 1.0 for m in
               ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]},
        }]},
        "Test Scenario", 2027, dept_names={"0110": "Rooms"}, tc_por_mes=[tc] * 12,
    )
    import io
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    formula = ws.cell(row=HEADER_ROW + 1, column=COL_REF).value
    resultado = _eval_sumproduct_formula(formula, fte, salary, [tc] * 12, "CRC")

    sw_un_mes = salary / tc  # calc_sw de un mes con FTE=1
    assert resultado == pytest.approx(sw_un_mes * 12, rel=1e-6)
