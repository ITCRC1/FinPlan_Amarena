# -*- coding: utf-8 -*-
"""
EL EXPORTADOR GENÉRICO DE CUADROS.

Un solo camino para que las ~47 pantallas que hoy no bajan a Excel lo hagan con
el formato de la casa, en vez de escribir 47 exportadores que se desincronizan.

**Lo que estas pruebas protegen no es el estilo, es el DATO.** Los tres defectos
que la auditoría encontró en los exports que ya existían son de datos, no de
estética:

* `/reports/summary` manda `"$1,234.00"` como TEXTO — el Excel no se puede sumar.
* `/reports/expenses` simula la jerarquía con espacios dentro del texto — se
  pierde al ordenar o al copiar.
* `/reports/pl-ytd` escribe `""` donde va un número.

Por eso acá se verifica que los números lleguen como números, que la sangría sea
sangría de Excel, y que el cero no se imprima.
"""
import io

from openpyxl import load_workbook

from app.export.cuadro_excel import FORMATOS, build_cuadros_workbook

CUADRO = {
    "titulo": "Big Picture — Budget 2027",
    "subtitulo": "USD",
    "columnas": [
        {"label": "Concepto", "ancho": 42, "formato": "texto"},
        {"label": "2026", "ancho": 14, "formato": "usd"},
        {"label": "Var %", "ancho": 10, "formato": "pct"},
    ],
    "filas": [
        {"label": "Total Ingresos", "nivel": 0, "es_total": True, "valores": [4000000, 0.125]},
        {"label": "Habitaciones", "nivel": 1, "valores": [3560260.57, 0.0918]},
        {"label": "Private Bar", "nivel": 2, "valores": [0, None]},
    ],
}


def _abrir(cuadros):
    return load_workbook(io.BytesIO(build_cuadros_workbook(cuadros)))


def test_los_numeros_van_como_numero_no_como_texto():
    """La diferencia entre un Excel que se puede sumar y uno que no."""
    ws = _abrir([CUADRO]).worksheets[0]
    assert ws.cell(5, 2).value == 4000000
    assert isinstance(ws.cell(5, 2).value, (int, float))
    assert ws.cell(6, 2).value == 3560260.57
    # El porcentaje va como fracción con formato de %, no como la cadena "12.5%"
    assert ws.cell(5, 3).value == 0.125
    assert "%" in ws.cell(5, 3).number_format


def test_la_jerarquia_es_sangria_de_excel_no_espacios():
    ws = _abrir([CUADRO]).worksheets[0]
    assert ws.cell(5, 1).value == "Total Ingresos"      # sin espacios pegados
    assert ws.cell(5, 1).alignment.indent == 0
    assert ws.cell(6, 1).alignment.indent == 1
    assert ws.cell(7, 1).alignment.indent == 2


def test_los_totales_se_ven_como_totales():
    ws = _abrir([CUADRO]).worksheets[0]
    assert ws.cell(5, 1).font.bold is True
    assert ws.cell(5, 2).font.bold is True
    assert ws.cell(6, 1).font.bold is False


def test_el_cero_no_se_imprime_y_el_negativo_va_en_rojo():
    """Una grilla llena de ceros esconde las cifras que sí importan."""
    for fmt in ("usd", "usd2", "num", "num1", "pct"):
        assert FORMATOS[fmt].endswith('""'), fmt      # tercera sección vacía = cero no imprime
        assert "[Red]" in FORMATOS[fmt], fmt


def test_una_hoja_por_cuadro_y_los_nombres_no_chocan():
    wb = _abrir([CUADRO, {**CUADRO, "titulo": "Otro"}, CUADRO])
    # ⚠️ El libro de VARIAS hojas trae además un «Índice» adelante desde el
    # 2026-09-03 (owner: «que baje bien profesional y claro»): con doce hojas,
    # los nombres van cortados a 31 caracteres y no se leen enteros.
    assert wb.sheetnames[0] == "Índice"
    cuadros = [h for h in wb.sheetnames if h != "Índice"]
    assert len(cuadros) == 3
    assert len(set(cuadros)) == 3            # sin duplicados aunque el título repita


def test_la_celda_vacia_no_es_un_cero():
    """`None` deja la celda vacía; 0 escribe un cero (que el formato esconde)."""
    ws = _abrir([CUADRO]).worksheets[0]
    assert ws.cell(7, 2).value == 0          # Private Bar sin dato todavía
    assert ws.cell(7, 3).value is None       # var % que no aplica


def test_la_fila_puede_pisar_el_formato_de_la_columna():
    """El bloque de drivers del Big Picture pone noches, ocupación % y ADR en
    dólares una debajo de otra. Sin esto habría que partirlo en tres cuadros."""
    cuadro = {
        **CUADRO,
        "filas": [
            {"label": "Noches ocupadas", "formato": "num", "valores": [12410, None]},
            {"label": "Ocupación %", "formato": "pct", "valores": [0.625, None]},
            {"label": "ADR", "formato": "usd2", "valores": [287.5, None]},
            {"label": "Rooms Revenue", "valores": [3567875, None]},   # hereda la columna
        ],
    }
    ws = _abrir([cuadro]).worksheets[0]
    assert ws.cell(5, 2).number_format == FORMATOS["num"]
    assert ws.cell(6, 2).number_format == FORMATOS["pct"]
    assert ws.cell(7, 2).number_format == FORMATOS["usd2"]
    assert ws.cell(8, 2).number_format == FORMATOS["usd"]      # la de la columna


def test_una_columna_puede_llevar_texto_de_verdad():
    """Las pantallas de mapeo son casi todas de texto (cuenta · departamento ·
    línea del P&L · modo de ruteo). Sin esto había que amontonar cuatro datos
    dentro de la etiqueta de la fila. NO es permiso para mandar números
    formateados: para eso está la prueba de arriba."""
    cuadro = {
        "titulo": "Mapeo",
        "columnas": [
            {"label": "Cuenta", "formato": "texto"},
            {"label": "Departamento", "formato": "texto"},
            {"label": "Línea del P&L", "formato": "texto"},
            {"label": "Monto", "formato": "usd"},
        ],
        "filas": [{"label": "7065", "valores": ["0110", "OPEX_ROOMS", 41583.0]}],
    }
    ws = _abrir([cuadro]).worksheets[0]
    assert ws.cell(5, 2).value == "0110"
    assert ws.cell(5, 3).value == "OPEX_ROOMS"
    assert ws.cell(5, 4).value == 41583.0
    # El texto se lee a la izquierda; el número sigue a la derecha.
    assert ws.cell(5, 2).alignment.horizontal == "left"
    assert ws.cell(5, 4).alignment.horizontal == "right"


def test_se_congelan_la_cabecera_y_las_etiquetas():
    """Sin esto, un cuadro de 12 meses obliga a adivinar la fila en diciembre."""
    ws = _abrir([CUADRO]).worksheets[0]
    assert ws.freeze_panes == "B5"


def test_nunca_devuelve_un_libro_sin_hojas():
    wb = load_workbook(io.BytesIO(build_cuadros_workbook([])))
    assert wb.sheetnames
