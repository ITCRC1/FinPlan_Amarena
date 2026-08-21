# -*- coding: utf-8 -*-
"""La plantilla del Country Mix va y vuelve: bajo, corrijo, subo.

**El pedido (owner, 18-ago-2026).** «Debe haber la opción de editar por si
acaso un ajuste; se podría quizás bajar a Excel controlado y después subir con
el cambio.»

El «⬇ Excel» que ya existía es un REPORTE —trae variance en puntos porcentuales
contra el budget— y no se puede volver a subir. Esta plantilla es la grilla
cruda: una fila por (país, métrica), doce columnas de mes.

⚠️ La norma del proyecto: **las columnas se ubican por ENCABEZADO, nunca por
posición.** El owner trabaja sobre estos archivos y mover o insertar una
columna no puede romper la carga.
"""
import io

import openpyxl
import pytest

from app.export.country_mix_xlsx import HOJA_DE, MESES, construir_libro, leer_libro


def _bytes(wb) -> bytes:
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def _clave(filas):
    """El contenido, sin depender del orden.

    Con una hoja por métrica las filas salen agrupadas por pestaña, no
    intercaladas como se pasaron. El orden NO es parte del contrato —lo que
    importa es que no se pierda ni se altere ningún número—, así que compararlo
    haría fallar la prueba por algo que no es un defecto.
    """
    return sorted((f["pais"], f["metric"], tuple(float(x) for x in f["values"]))
                  for f in filas)


FILAS = [
    {"pais": "United States", "metric": "rooms", "values": [322, 300, 391, 304, 120, 145, 117, 0, 0, 0, 0, 0]},
    {"pais": "United States", "metric": "pax", "values": [610, 559, 742, 560, 214, 231, 240, 0, 0, 0, 0, 0]},
    {"pais": "Others", "metric": "rooms", "values": [37, 69, 63, 89, 30, 37, 63, 0, 0, 0, 0, 0]},
]


def test_lo_que_bajo_es_lo_que_vuelve():
    """La ida y vuelta sin tocar nada tiene que devolver exactamente lo mismo."""
    filas, problemas = leer_libro(_bytes(construir_libro("t", FILAS)))
    assert problemas == []
    assert _clave(filas) == _clave(FILAS)


def test_un_ajuste_a_mano_entra():
    """El caso real: bajo, corrijo una celda, subo."""
    wb = construir_libro("t", FILAS)
    ws = wb["Habitaciones"]
    ws.cell(row=5, column=2, value=999)      # United States / Ene
    filas, problemas = leer_libro(_bytes(wb))
    assert problemas == []
    us = next(f for f in filas if f["pais"] == "United States" and f["metric"] == "rooms")
    assert us["values"][0] == 999


def test_mover_una_columna_NO_rompe_la_carga():
    """⚠️ La norma. Si se leyera por posición, esto daría números corridos.

    Se inserta una columna al principio: todo se desplaza uno. Buscando por
    encabezado, el resultado tiene que ser idéntico.
    """
    wb = construir_libro("t", FILAS)
    wb["Habitaciones"].insert_cols(1)
    filas, problemas = leer_libro(_bytes(wb))
    assert problemas == []
    assert _clave(filas) == _clave(FILAS)


def test_un_pais_nuevo_se_crea():
    wb = construir_libro("t", FILAS)
    ws = wb["Habitaciones"]
    f = ws.max_row + 1
    ws.cell(row=f, column=1, value="Germany")
    ws.cell(row=f, column=2, value=64)
    filas, problemas = leer_libro(_bytes(wb))
    assert problemas == []
    assert any(x["pais"] == "Germany" and x["values"][0] == 64 for x in filas)


def test_borrar_la_fila_borra_el_pais():
    """Para que un país deje de existir se borra su fila — y eso tiene que
    reflejarse, no ignorarse."""
    wb = construir_libro("t", FILAS)
    wb["Habitaciones"].delete_rows(6)     # la fila de Others
    filas, _ = leer_libro(_bytes(wb))
    assert not any(x["pais"] == "Others" for x in filas)




def test_un_pais_repetido_se_avisa_en_vez_de_pisarse():
    """Dos filas del mismo país y métrica: una pisaría a la otra en silencio y
    el total quedaría mal sin que nada lo dijera."""
    wb = construir_libro("t", FILAS)
    ws = wb["Habitaciones"]
    f = ws.max_row + 1
    ws.cell(row=f, column=1, value="United States")
    ws.cell(row=f, column=2, value=1)
    _filas, problemas = leer_libro(_bytes(wb))
    assert any("repetido" in p for p in problemas)


def test_un_negativo_se_avisa():
    """No hay noches ni pax negativos: es un typo, no un dato."""
    wb = construir_libro("t", FILAS)
    wb["Habitaciones"].cell(row=5, column=2, value=-5)
    _filas, problemas = leer_libro(_bytes(wb))
    assert any("negativos" in p for p in problemas)


def test_un_archivo_que_no_es_la_plantilla_lo_dice():
    wb = openpyxl.Workbook()
    wb.active["A1"] = "cualquier cosa"
    filas, problemas = leer_libro(_bytes(wb))
    assert filas == []
    assert any("País" in p for p in problemas)


def test_los_doce_meses_estan_en_el_encabezado():
    for hoja in ("Habitaciones", "Pax"):
        ws = construir_libro("t", FILAS)[hoja]
        enc = [ws.cell(row=4, column=j).value for j in range(1, 14)]
        assert enc[0] == "País", f"{hoja}: {enc}"
        assert enc[1:] == MESES, f"{hoja}: {enc}"


def test_subir_sin_confirmar_no_guarda():
    """Reemplaza el mix entero del escenario: conviene ver qué cambia antes."""
    import inspect

    from app.api import revenue_api

    src = inspect.getsource(revenue_api.country_mix_subir_plantilla)
    assert "if not confirmar:" in src
    assert '"guardado": False' in src
    # y el borrado tiene que estar DESPUÉS de ese corte
    assert src.index("if not confirmar:") < src.index("delete(CountryMixEntry)")


def test_hay_UNA_HOJA_POR_METRICA():
    """⚠️ «Se me hace confuso subir si pax y rooms; ocupo que haya 2 tabs para
    esto» (owner, 18-ago-2026).

    La primera versión ponía las dos métricas en la misma hoja con una columna
    «Métrica». Para corregir las noches de un país había que encontrar SU fila
    entre el doble de filas, y una fila mal rotulada cambiaba la métrica
    equivocada sin que se notara.
    """
    wb = construir_libro("t", FILAS)
    assert wb.sheetnames == ["Habitaciones", "Pax"]
    assert set(HOJA_DE.values()) == set(wb.sheetnames)


def test_cada_hoja_trae_SOLO_su_metrica():
    wb = construir_libro("t", FILAS)
    hab = [wb["Habitaciones"].cell(row=r, column=1).value
           for r in range(5, wb["Habitaciones"].max_row + 1)]
    pax = [wb["Pax"].cell(row=r, column=1).value
           for r in range(5, wb["Pax"].max_row + 1)]
    assert hab == ["United States", "Others"]     # las dos filas `rooms`
    assert pax == ["United States"]               # la única fila `pax`


def test_la_hoja_ya_NO_lleva_columna_Metrica():
    """Si volviera, volvería la confusión que el owner reportó."""
    ws = construir_libro("t", FILAS)["Habitaciones"]
    enc = [ws.cell(row=4, column=j).value for j in range(1, 14)]
    assert "Métrica" not in enc


def test_el_formato_VIEJO_de_una_sola_hoja_sigue_subiendo():
    """Alguien puede tener bajado un archivo de antes del cambio. La métrica
    sale de la columna «Métrica» si está; si no, del nombre de la hoja."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CountryMix"
    for j, n in enumerate(["País", "Métrica"] + MESES, start=1):
        ws.cell(row=4, column=j, value=n)
    ws.cell(row=5, column=1, value="United States")
    ws.cell(row=5, column=2, value="Habitaciones")
    ws.cell(row=5, column=3, value=322)
    ws.cell(row=6, column=1, value="United States")
    ws.cell(row=6, column=2, value="Pax")
    ws.cell(row=6, column=3, value=612)
    filas, problemas = leer_libro(_bytes(wb))
    assert problemas == []
    assert _clave(filas) == _clave([
        {"pais": "United States", "metric": "rooms", "values": [322] + [0] * 11},
        {"pais": "United States", "metric": "pax", "values": [612] + [0] * 11},
    ])


def test_una_hoja_con_nombre_raro_lo_dice_en_vez_de_adivinar():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    for j, n in enumerate(["País"] + MESES, start=1):
        ws.cell(row=4, column=j, value=n)
    ws.cell(row=5, column=1, value="United States")
    ws.cell(row=5, column=2, value=10)
    filas, problemas = leer_libro(_bytes(wb))
    assert filas == []
    assert any("Habitaciones" in p and "Pax" in p for p in problemas)
