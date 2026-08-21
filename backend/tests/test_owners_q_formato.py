"""El export contra el ARCHIVO REAL del owner, celda por celda de formato.

Los valores ya los cubre `test_owners_q`. Esto cubre lo otro: que el archivo se
VEA como el que SCP recibe. No es cosmética — SCP consolida por posición de
fila y lee la jerarquía por sangría y resalte, así que el formato es parte del
contrato igual que los números.

La referencia es el `.xlsx` que el owner mandó, guardado tal cual en fixtures.
Si algún día SCP cambia el formato, se reemplaza ese archivo y esta prueba dice
exactamente qué quedó distinto.
"""
import io
import json
import pathlib

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.engine.owners_q import COLUMNAS
from app.export.owners_q_excel import export_owners_q

ORIGINAL = pathlib.Path(__file__).parent / "fixtures" / "SCP_CWL_JUN2026_original.xlsx"
SEED = pathlib.Path(__file__).parent.parent / "app" / "seed_data" / "owners_q.json"


@pytest.fixture(scope="module")
def original():
    return load_workbook(ORIGINAL)["SCPCWL"]


@pytest.fixture(scope="module")
def mio():
    filas = json.loads(SEED.read_text(encoding="utf-8"))["report_lines"]
    datos = {
        "entidad": "CWL", "anio": 2026, "mes": 6, "convencion": "raw",
        "filas": [{
            "row_no": f["row_no"], "report_code": f["report_code"],
            "label": f["label"], "indent": f["indent"],
            "line_type": f["line_type"], "nature": f["nature"],
            "estilo": f.get("estilo") or {},
            "celdas": {c: "-1234.56" for c in COLUMNAS},
        } for f in filas],
    }
    return load_workbook(io.BytesIO(export_owners_q(datos)))["SCPCWL"]


# ─── El bloque de encabezado ──────────────────────────────────────────────────
def test_la_hoja_se_llama_como_la_del_owner(mio):
    assert mio.title == "SCPCWL"


def test_identificacion_filas_1_a_4(original, mio):
    for ref in ("A1", "A2", "A3", "A4", "B4"):
        assert mio[ref].value == original[ref].value, ref
    # `As of Date` es el último día del mes.
    assert mio["B3"].value.date() == original["B3"].value.date()
    assert mio["B3"].number_format == original["B3"].number_format


def test_month_ending_y_year_to_date(original, mio):
    """Filas 6 y 7: el rótulo del bloque y su fecha, incluida la del año anterior."""
    for ref in ("A6", "E6", "K6", "R6", "V6", "AB6"):
        assert mio[ref].value == original[ref].value, ref
    for ref in ("A7", "E7", "K7", "R7", "V7", "AB7"):
        assert mio[ref].value.date() == original[ref].value.date(), ref
        assert mio[ref].number_format == original[ref].number_format, ref
    # La del bloque de año anterior es de 2025, no de 2026.
    assert mio["K7"].value.year == 2025
    assert mio["AB7"].value.year == 2025
    assert mio["A7"].value.year == 2026


def test_los_33_rotulos_de_la_fila_8(original, mio):
    for c in range(1, 34):
        letra = get_column_letter(c)
        assert mio[f"{letra}8"].value == original[f"{letra}8"].value, letra


# ─── El cuerpo ────────────────────────────────────────────────────────────────
def test_las_etiquetas_con_su_sangria(original, mio):
    """Los espacios de sangría son del archivo, no de un `indent` de Excel."""
    for r in range(9, 57):
        assert mio.cell(row=r, column=17).value == original.cell(row=r, column=17).value, \
            f"fila {r}"


def test_los_rellenos_de_subtotal_y_total(original, mio):
    """Los dos azules, en las mismas 11 filas."""
    for r in range(9, 57):
        a, b = mio.cell(row=r, column=17), original.cell(row=r, column=17)
        assert a.fill.fgColor.rgb == b.fill.fgColor.rgb, \
            f"fila {r}: {a.fill.fgColor.rgb} vs {b.fill.fgColor.rgb}"


def test_los_bordes(original, mio):
    """La 49 lleva línea arriba sin ser subtotal y la 52 la lleva doble: por eso
    el borde es dato de la fila y no una regla."""
    for r in range(9, 57):
        a, b = mio.cell(row=r, column=1), original.cell(row=r, column=1)
        assert (a.border.top.style, a.border.bottom.style) == \
               (b.border.top.style, b.border.bottom.style), f"fila {r}"


def test_los_formatos_numericos(original, mio):
    """Con signo de dólar; entero en las noches y porcentaje en la ocupación."""
    for r in range(9, 57):
        for c in range(1, 34):
            if c == 17:
                continue
            a = mio.cell(row=r, column=c).number_format
            b = original.cell(row=r, column=c).number_format
            assert a == b, f"{get_column_letter(c)}{r}: {a!r} vs {b!r}"


def test_la_fuente(original, mio):
    for ref in ("Q9", "Q21", "A16", "A38", "Q56"):
        a, b = mio[ref].font, original[ref].font
        assert (a.name, a.size, a.bold) == (b.name, b.size, b.bold), ref


def test_los_anchos_de_columna(original, mio):
    for c in range(1, 34):
        letra = get_column_letter(c)
        esperado = original.column_dimensions[letra].width
        if not esperado:
            continue
        assert mio.column_dimensions[letra].width == pytest.approx(esperado, abs=0.1), letra


def test_congela_en_B9_como_el_original(original, mio):
    assert mio.freeze_panes == original.freeze_panes == "B9"


def test_los_alto_de_fila(original, mio):
    for r in range(9, 57):
        esperado = original.row_dimensions[r].height
        if not esperado:
            continue
        assert mio.row_dimensions[r].height == pytest.approx(esperado, abs=0.01), f"fila {r}"


# ─── Y las hojas propias siguen ahí ───────────────────────────────────────────
def test_las_hojas_de_meses_y_trimestres_no_estorban(mio):
    """Lo que se manda a SCP es la PRIMERA hoja; las otras dos son de la casa."""
    wb = mio.parent
    assert wb.sheetnames[0] == "SCPCWL"


# ─── Flexibilidad de bloques (pedido del owner, 2026-08-18) ──────────────────
def _datos_con_bloques(bloques, es_estandar):
    filas = json.loads(SEED.read_text(encoding="utf-8"))["report_lines"]
    return {
        "entidad": "CWL", "anio": 2026, "mes": 6, "convencion": "raw",
        "bloques": bloques, "es_estandar": es_estandar,
        "filas": [{
            "row_no": f["row_no"], "report_code": f["report_code"],
            "label": f["label"], "indent": f["indent"],
            "line_type": f["line_type"], "nature": f["nature"],
            "estilo": f.get("estilo") or {},
            "celdas": {c: "1.0" for c in COLUMNAS},
        } for f in filas],
    }


ESTANDAR = {
    "actual": {"etiqueta": "Actual 2026", "anio": 2026, "mes": 6, "por_defecto": True},
    "budget": {"etiqueta": "Budget 2026", "anio": 2026, "mes": 6, "por_defecto": True},
    "py": {"etiqueta": "Actual 2025", "anio": 2025, "mes": 6, "por_defecto": True},
}


def test_el_estandar_no_cambia_nada(original):
    """Lo que se manda a SCP sigue idéntico: la fila 5 en blanco y las fechas
    de siempre. La flexibilidad no puede filtrarse al entregable."""
    ws = load_workbook(io.BytesIO(
        export_owners_q(_datos_con_bloques(ESTANDAR, True))))["SCPCWL"]
    assert ws["A5"].value == original["A5"].value == " "
    assert ws["A7"].value.date() == original["A7"].value.date()
    assert ws["K7"].value.date() == original["K7"].value.date()


def test_un_forecast_en_la_columna_de_budget_se_avisa_y_cambia_la_fecha():
    """Si alguien compara contra otra cosa, el archivo lo dice en su cara:
    quien lo recibe no puede deducirlo mirando los números."""
    bloques = {
        **ESTANDAR,
        "budget": {"etiqueta": "Forecast 2026 · working", "anio": 2026, "mes": 5,
                   "por_defecto": False},
    }
    ws = load_workbook(io.BytesIO(
        export_owners_q(_datos_con_bloques(bloques, False))))["SCPCWL"]
    assert "NO es el reporte estándar" in ws["A5"].value
    assert "Forecast 2026 · working" in ws["A5"].value
    # Y el `Month Ending` de esa columna sigue al escenario: mayo, no junio.
    assert ws["E7"].value.date().month == 5
    assert ws["A7"].value.date().month == 6      # la columna Actual no se movió


def test_el_ano_del_bloque_manda_sobre_la_aritmetica():
    """Un Forecast 2026 puesto en la columna de año anterior imprime 2026."""
    bloques = {**ESTANDAR,
               "py": {"etiqueta": "Forecast 2026", "anio": 2026, "mes": 6,
                      "por_defecto": False}}
    ws = load_workbook(io.BytesIO(
        export_owners_q(_datos_con_bloques(bloques, False))))["SCPCWL"]
    assert ws["K7"].value.date().year == 2026
    assert ws["AB7"].value.date().year == 2026


# ─── Períodos: trimestre y Full Year (pedido del owner, 2026-08-18) ──────────
def test_los_17_periodos():
    from app.engine.owners_q import periodos_disponibles, resolver_periodo
    p = periodos_disponibles()
    assert len(p) == 17
    assert [x["clave"] for x in p][-5:] == ["Q1", "Q2", "Q3", "Q4", "FY"]

    # Un mes: el bloque es ese mes y el acumulado va de enero hasta él.
    assert resolver_periodo("M06")[:2] == ([6], [1, 2, 3, 4, 5, 6])
    # Un trimestre: el bloque es el trimestre, el acumulado el año hasta su cierre.
    assert resolver_periodo("Q2")[:2] == ([4, 5, 6], [1, 2, 3, 4, 5, 6])
    assert resolver_periodo("Q4")[:2] == ([10, 11, 12], list(range(1, 13)))
    # El año completo: los dos son los doce meses.
    assert resolver_periodo("FY")[:2] == (list(range(1, 13)), list(range(1, 13)))
    # Solo el mes simple es el estándar de SCP.
    assert resolver_periodo("M06")[3] is True
    assert resolver_periodo("Q2")[3] is False
    assert resolver_periodo("FY")[3] is False


def test_un_periodo_desconocido_falla():
    from app.engine.owners_q import ReporteError, resolver_periodo
    for malo in ("Q5", "M13", "M00", "", "año"):
        with pytest.raises(ReporteError):
            resolver_periodo(malo)


def test_full_year_avisa_y_cambia_el_rotulo():
    """Con el año completo «Month Ending» sería mentira, y SCP pide UN mes."""
    datos = _datos_con_bloques(ESTANDAR, False)
    datos.update({"periodo": "FY", "periodo_etiqueta": "Full Year",
                  "es_un_mes": False, "mes": 12})
    ws = load_workbook(io.BytesIO(export_owners_q(datos)))["SCPCWL"]
    assert ws["A6"].value == "Period Ending"
    assert ws["R6"].value == "Year To Date"        # ese no cambia
    assert "Full Year" in ws["A5"].value and "SCP pide UN mes" in ws["A5"].value


def test_el_nombre_del_archivo_dice_el_periodo():
    from app.export.owners_q_excel import nombre_archivo
    assert nombre_archivo("CWL", 2026, {"periodo": "M06", "mes": 6}) == \
        "SCP_CWL_JUN26_Statement_of_Income.xlsx"
    assert nombre_archivo("CWL", 2026, {"periodo": "Q2", "mes": 6}) == \
        "SCP_CWL_Q2_26_Statement_of_Income.xlsx"
    assert nombre_archivo("CWL", 2026, {"periodo": "FY", "mes": 12}) == \
        "SCP_CWL_FY26_Statement_of_Income.xlsx"
    # La firma vieja (un mes suelto) sigue andando.
    assert nombre_archivo("CWL", 2026, 6) == "SCP_CWL_JUN26_Statement_of_Income.xlsx"


# ─── La etiqueta tiene que nombrar SU bloque ─────────────────────────────────

def test_la_posicion_viaja_junto_al_bloque():
    """⚠️ Encontrado el 2026-08-19, y es del tipo que no falla: miente.

    El pie del reporte arma «Posición = escenario» recorriendo las tres
    posiciones. Estaba escrito así:

        POSICIONES.map(p => rep.bloques?.[p.key])
          .filter(Boolean)
          .map((b, i) => POS_TITULO[POSICIONES[i].key] ...)

    `filter` COMPACTA el arreglo. Si el bloque de la primera posición viene
    nulo, el que queda en el índice 0 es el de la SEGUNDA — y se rotula con el
    título de la primera. La etiqueta nombra el escenario equivocado.

    No hay error, no hay celda vacía: hay un número real con el nombre de otro,
    en el reporte que se le manda a los dueños. Por eso la posición tiene que
    viajar junto a su bloque y no reencontrarse por índice.
    """
    from tests._rutas import FRONT

    import re

    crudo = (FRONT / "app" / "reports" / "owners-q" / "page.tsx").read_text(encoding="utf-8")
    # ⚠️ Sin los comentarios. La primera version de esta prueba fallaba por el
    # comentario que documenta el patron VIEJO: buscaba texto y no codigo, que
    # es exactamente el defecto que vino a cazar.
    src = re.sub(r"/\*.*?\*/", "", crudo, flags=re.S)
    src = re.sub(r"^\s*//.*$", "", src, flags=re.M)

    assert "POSICIONES[i]" not in src, (
        "volvió el reencuentro por índice: si un bloque viene nulo, la etiqueta "
        "nombra el escenario equivocado")
    assert "{ p, b: rep.bloques?.[p.key] }" in src, (
        "la posición tiene que viajar junto al bloque")
