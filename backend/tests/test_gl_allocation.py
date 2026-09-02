"""Los departamentos de reparto NO se descartan: netean solos, y el saldo se ve.

## La regla, y por qué cambió (owner, 2026-08-28)

*«cafetería y laundry tienen saldo — que salga ese saldo en overhead»* · *«si
tiene saldo que lo vea como normal y que aparezca esa diferencia en overhead;
hasta que se deje en 0, no pasa nada»*.

**Antes** `ALLOCATION_EXCLUDE = {"0220": {5,6,7}, "0161": {6,7}}` hacía que esas
filas **no llegaran a la base**. El razonamiento era que el crédito de
Distribución las dejaba en cero, así que tirarlas daba lo mismo.

**No daba lo mismo cuando no netean.** Lo que se tiraba entonces no era un
duplicado: era el SOBRANTE, y desaparecía sin que nada avisara — el P&L cerraba
consigo mismo sin él.

Se vio subiendo los actuales de 2026: marzo y abril entraron (no traían estos
departamentos) y mayo, junio y julio rebotaron con **409**, porque el bloque de
verificación del archivo incluía el gasto de cafetería y el detalle lo había
descartado.

## Lo que se blinda ahora

**El neteo lo hace la aritmética, no una lista.** `calculate_full_pl` suma
`planilla + costo + opex + reparto` por grupo, y `CAFETERIA` y `LAUNDRY_OPS` son
grupos de OVERHEAD. Entonces:

* un departamento que reparte todo su gasto da **cero** solo, y su línea ni se
  dibuja — el resultado visible es el mismo que antes;
* uno que deja saldo **lo muestra en overhead**, que es lo que antes se perdía.

Las dos situaciones dejan de verse iguales, que era el problema.
"""
import io

import openpyxl

from app.engine import pl_engine
from app.importers.gl_detail_importer import (ALLOC_EXCL_COST, ALLOC_EXCL_OPEX,
                                              ALLOC_EXCL_PAYROLL,
                                              ALLOCATION_EXCLUDE,
                                              DEPTOS_DE_REPARTO, parse_gl_detail)

ZERO = pl_engine._d(0)


def _gl_book(rows):
    """rows: (clase_lbl, dept_name, account_code, account_name, valor_enero)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Fila 15 = rótulo del bloque · Fila 14 = mes (el parser detecta por ambos).
    ws.cell(row=15, column=5, value="Actual 2099")
    ws.cell(row=14, column=5, value="January")
    r = 17
    for cls_lbl, dept, code, name, jan in rows:
        ws.cell(row=r, column=1, value=cls_lbl)
        ws.cell(row=r, column=2, value=dept)
        ws.cell(row=r, column=3, value=code)
        ws.cell(row=r, column=4, value=name)
        ws.cell(row=r, column=5, value=jan)   # mes 1
        r += 1
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def test_ya_no_se_descarta_nada_al_importar():
    """Lo que se descarta y quién reparte son DOS COSAS, y confundirlas fue el bug.

    `ALLOCATION_EXCLUDE` dice qué filas no llegan a la base: vacío, nada se
    descarta. `DEPTOS_DE_REPARTO` dice quién es origen de reparto, y de ahí
    salen los sets que el P&L por Departamento usa para restarle a cada uno lo
    que efectivamente repartió — ésos NO se vacían, o el reporte volvería a
    mostrar el gasto bruto y contaría dos veces.
    """
    assert ALLOCATION_EXCLUDE == {}
    assert DEPTOS_DE_REPARTO == {"0220": {"5", "6", "7"}, "0161": {"6", "7"}}
    assert ALLOC_EXCL_COST == {"0220"}
    assert ALLOC_EXCL_PAYROLL == {"0220", "0161"}
    assert ALLOC_EXCL_OPEX == {"0220", "0161"}
    assert pl_engine.ACTUAL_EXCLUDED_DEPTS == set(), (
        "la cafetería volvió a excluirse del P&L de actuales: su sobrante "
        "desaparecería otra vez sin dejar rastro")


def test_el_gasto_de_reparto_llega_a_la_base():
    """Lo que antes se tiraba en el parser ahora se importa entero."""
    data = _gl_book([
        ("Rev",  "Lavanderia",                   "4700", "Laundry Services rev", 100),
        ("Cost", "Lavanderia",                   "5603", "Laundry Services cost", 50),
        ("Pay",  "Lavanderia",                   "6000", "Laundry payroll",      200),
        ("Opex", "Lavanderia",                   "7320", "Laundry supplies",      80),
        ("Cost", "Cafeteria Empleados",          "5700", "Dining cost",          300),
        ("Pay",  "Cafeteria Empleados",          "6000", "Dining payroll",       150),
        ("Opex", "Cafeteria Empleados",          "7065", "Dining opex",           20),
        ("Pay",  "Departamento de Habitaciones", "6000", "Rooms payroll",        500),
    ])
    blk = parse_gl_detail(data)[0]

    def depts(key):
        return {r["dept_code"] for r in blk[key]}

    for clave in ("costs", "payroll", "opex"):
        assert "0220" in depts(clave), f"la cafetería se perdió en {clave}"
    for clave in ("payroll", "opex"):
        assert "0161" in depts(clave), f"la lavandería se perdió en {clave}"
    # Lo que ya entraba tiene que seguir entrando.
    assert "0161" in depts("revenue")
    assert "0161" in depts("costs")
    assert "0110" in depts("payroll")


def test_si_el_reparto_cubre_el_gasto_la_linea_da_cero():
    """El caso de siempre: el resultado visible no cambia."""
    lineas = pl_engine.calculate_full_pl(
        revenue_by_line={},
        payroll_by_dept={"0220": pl_engine._d(150)},
        cos_by_dept={"0220": pl_engine._d(300)},
        opex_by_dept={"0220": pl_engine._d(20)},
        # El crédito de Distribución: saca los 470 completos.
        alloc_by_dept={"0220": pl_engine._d(-470)},
    )
    caf = [l for l in lineas if l.line_code == "OVH_CAFETERIA"]
    # Sin saldo, la línea ni se dibuja — es ruido.
    assert not caf or caf[0].amount_usd == ZERO
    total = [l for l in lineas if l.line_code == "TOTAL_OVERHEAD"][0]
    assert total.amount_usd == ZERO


def test_si_sobra_saldo_aparece_en_overhead():
    """Lo que el owner pidió: la diferencia se ve, no se pierde."""
    lineas = pl_engine.calculate_full_pl(
        revenue_by_line={},
        payroll_by_dept={"0220": pl_engine._d(150)},
        cos_by_dept={"0220": pl_engine._d(300)},
        opex_by_dept={"0220": pl_engine._d(20)},
        # El reparto sólo cubre 400 de los 470: sobran 70.
        alloc_by_dept={"0220": pl_engine._d(-400)},
    )
    caf = [l for l in lineas if l.line_code == "OVH_CAFETERIA"]
    assert caf, "la cafetería con saldo no aparece en overhead"
    assert caf[0].amount_usd == pl_engine._d(70)
    assert caf[0].section == "OVERHEAD", (
        "el saldo tiene que salir en OVERHEAD, no arriba del GOP: la cafetería "
        "no es un departamento operativo")
    total = [l for l in lineas if l.line_code == "TOTAL_OVERHEAD"][0]
    assert total.amount_usd == pl_engine._d(70)


def test_la_lavanderia_tambien_deja_ver_su_saldo():
    lineas = pl_engine.calculate_full_pl(
        revenue_by_line={},
        payroll_by_dept={"0161": pl_engine._d(200)},
        cos_by_dept={},
        opex_by_dept={"0161": pl_engine._d(80)},
        alloc_by_dept={"0161": pl_engine._d(-250)},
    )
    lav = [l for l in lineas if l.line_code == "OVH_LAUNDRY_OPS"]
    assert lav and lav[0].amount_usd == pl_engine._d(30)
    assert lav[0].section == "OVERHEAD"
