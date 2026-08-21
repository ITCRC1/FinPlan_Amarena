# -*- coding: utf-8 -*-
"""La plantilla vuelve en el orden en que el owner la subio, con separadores.

Owner (2026-08-14): «debe quedar en el mismo orden, y que esten todas. mismo
orden.» Y despues: «que haya separadores por departamentos».

Antes se ordenaba por grupo del P&L y clase. Era determinista y estable entre
descargas, pero distinto del archivo historico del owner: cada vez que comparaba
la bajada contra el suyo tenia que cruzar dos listas.

Ahora manda `orden_archivo` (mig 111), que el importador guarda con la fila del
Excel de origen. Lo que no lo tenga cae al orden de siempre, asi que un escenario
que nunca se importo se ve igual que antes.
"""
import io

import openpyxl
import pytest

from app.export.detail_excel import build_detail_workbook

BLOQUE = "Actual actual 2025"
DEPTOS = {"0110": "Habitaciones", "0120": "A&B", "0180": "Administracion"}


def _cuenta(orden, dept, code, clase="Opex", grupo="ROOMS"):
    return {"clase": clase, "grupo": grupo, "dept_code": dept, "cuenta": code,
            "nombre": f"Cuenta {code}", "vals": {(BLOQUE, 1): 100.0}, "orden": orden}


def _leer(xls: bytes) -> list[tuple]:
    """(departamento, cuenta) de cada fila de datos, en el orden de la hoja."""
    ws = openpyxl.load_workbook(io.BytesIO(xls)).worksheets[0]
    fuera = []
    for r in range(16, ws.max_row + 1):
        clase = str(ws.cell(r, 1).value or "")
        dept, cta = ws.cell(r, 3).value, ws.cell(r, 4).value
        if dept is None and cta is None:
            continue
        # Las tres filas de KPI (9010/9020/9060) van siempre arriba y no son
        # cuentas: la hoja las escribe con clase «Stat».
        if clase == "Stat":
            continue
        fuera.append((str(dept or ""), str(cta or "")))
    return fuera


def test_respeta_el_orden_del_archivo():
    """El owner puso 7065 antes que 7000: asi tiene que volver.

    Aplica a las cuentas que NO estan en `orden_plantilla.json` — para las que si
    estan manda esa lista, que es la estructura oficial. Las cuentas de esta
    prueba usan un departamento inventado justamente para caer fuera.
    """
    accts = [_cuenta(50, "9990", "7065"), _cuenta(10, "9990", "7000")]
    filas = [f for f in _leer(build_detail_workbook([BLOQUE], accts, {}, DEPTOS))
             if f[1]]
    assert [c for _, c in filas] == ["7000", "7065"]

    # Y al reves: si el owner los subio en el otro orden, vuelven en ese.
    accts = [_cuenta(10, "9990", "7065"), _cuenta(50, "9990", "7000")]
    filas = [f for f in _leer(build_detail_workbook([BLOQUE], accts, {}, DEPTOS))
             if f[1]]
    assert [c for _, c in filas] == ["7065", "7000"]











def test_cada_seccion_se_anuncia_con_su_titulo():
    """Sin el titulo, la hoja se lee «todo junto» — que es la queja del owner."""
    import io as _io

    import openpyxl as _px

    accts = [_cuenta(10, "0110", "4000", clase="Revenue"),
             _cuenta(20, "0110", "6000", clase="Payroll"),
             _cuenta(30, "0110", "7000", clase="Opex")]
    ws = _px.load_workbook(
        _io.BytesIO(build_detail_workbook([BLOQUE], accts, {}, DEPTOS))).worksheets[0]
    titulos = [str(ws.cell(r, 1).value or "") for r in range(16, ws.max_row + 1)]
    for esperado in ("INGRESO", "PLANILLA", "GASTO OPERATIVO (OPEX)"):
        assert esperado in titulos, f"falta el titulo de seccion «{esperado}»"


def test_hay_un_separador_por_departamento():
    accts = [_cuenta(10, "0110", "7000"), _cuenta(20, "0110", "7065"),
             _cuenta(30, "0120", "7100"), _cuenta(40, "0180", "7200")]
    filas = _leer(build_detail_workbook([BLOQUE], accts, {}, DEPTOS))
    separadores = [d for d, c in filas if not c and "──" in d]
    assert len(separadores) == 3, f"esperaba 3 separadores, hubo {separadores}"
    for nombre in DEPTOS.values():
        assert any(nombre in s for s in separadores), f"falta el separador de {nombre}"


def test_el_separador_no_rompe_la_re_subida():
    """La fila separadora no trae cuenta NI monto.

    Si trajera monto, el importador la contaria como «plata sin cuenta» y se
    negaria a importar — el archivo que la app genera no puede ser rechazado por
    la app.
    """
    accts = [_cuenta(10, "0110", "7000"), _cuenta(30, "0120", "7100")]
    xls = build_detail_workbook([BLOQUE], accts, {}, DEPTOS)
    ws = openpyxl.load_workbook(io.BytesIO(xls)).worksheets[0]
    for r in range(16, ws.max_row + 1):
        if "──" in str(ws.cell(r, 3).value or ""):
            assert not ws.cell(r, 4).value, "el separador no puede traer cuenta"
            montos = [ws.cell(r, c).value for c in range(6, ws.max_column + 1)]
            assert not any(v for v in montos), "el separador no puede traer monto"


@pytest.mark.parametrize("campo", ["fila"])
def test_el_parser_guarda_la_fila_de_origen(campo):
    """Sin esto no hay orden que devolver."""
    import inspect

    from app.importers import gl_detail_importer as gl

    fuente = inspect.getsource(gl.parse_gl_detail)
    assert f'"{campo}": r0 + 1' in fuente, (
        "El parser dejo de guardar la fila del archivo: la plantilla vuelve a "
        "salir en un orden que no es el del owner.")


# ── El rotulo del departamento ──────────────────────────────────────────────

def test_el_departamento_muestra_codigo_y_nombre():
    """Owner: «por que Tienda no tiene numero de departamento».

    La columna mostraba el NOMBRE cuando el depto estaba en el catalogo y el
    CODIGO PELADO cuando no, en la misma columna: `0240` salia como numero y
    `Tienda` como texto. Leyendo la hoja no se podia saber si «0240» era un
    codigo sin nombre o un departamento que se llama asi.
    """
    from app.export.detail_excel import rotulo_depto

    assert rotulo_depto("0151", {"0151": "Tienda"}) == "0151 · Tienda"
    # Sin nombre, se ve que NO LO TIENE en vez de disfrazarse de nombre.
    r = rotulo_depto("0240", {"0151": "Tienda"})
    assert r.startswith("0240 · ") and "sin nombre" in r


def test_el_codigo_vuelve_exacto_al_re_subir():
    """La plantilla escribe «0165 · Gift Shop» y el parser lee 0165 SIN adivinar.

    Adivinar por palabra clave es fragil, y para un departamento sin nombre en el
    catalogo es imposible: no hay palabra. Antes esa fila se quedaba sin
    departamento al volver a subir, y se perdia.
    """
    from app.export.detail_excel import rotulo_depto
    from app.importers.gl_detail_importer import dept_code_from_name

    for code, nombres in (("0165", {"0165": "Gift Shop"}),
                          ("0151", {"0151": "Tienda"}),
                          ("260", {"260": "Club Madresal"}),
                          ("0240", {})):          # el que no tiene nombre
        assert dept_code_from_name(rotulo_depto(code, nombres)) == code


def test_el_archivo_escrito_a_mano_sigue_funcionando():
    """El fuzzy de siempre no se pierde: el owner arma archivos sin codigo."""
    from app.importers.gl_detail_importer import dept_code_from_name

    assert dept_code_from_name("Habitaciones") == "0110"
    assert dept_code_from_name("Restaurante A&B") == "0123"
    assert dept_code_from_name("Gift Shop") == "0165"
    assert dept_code_from_name("") is None


def test_el_departamento_manda_sobre_la_clase():
    """Owner: «para Rooms sale ingreso y sus cuentas, costo y sus cuentas,
    planilla y sus cuentas y opex y sus cuentas. Y despues sigues con F&B, Spa».

    Departamento AFUERA, clase adentro. Antes era al reves y las cuentas de un
    mismo departamento quedaban repartidas por toda la hoja.
    """
    accts = [_cuenta(None, "0110", "7000", clase="Opex"),
             _cuenta(None, "0120", "4100", clase="Revenue"),
             _cuenta(None, "0110", "4000", clase="Revenue")]
    filas = [f for f in _leer(build_detail_workbook([BLOQUE], accts, {}, DEPTOS))
             if f[1]]
    deptos = [d for d, _ in filas]
    # Las dos de 0110 juntas, y despues la de 0120.
    assert deptos == sorted(deptos, key=lambda x: 0 if "0110" in x else 1)


def test_el_orden_del_owner_es_una_LISTA_no_una_regla():
    """El orden de las clases CAMBIA segun el departamento.

    En el archivo del owner: Rooms no tiene costo; A&B va Ingreso-Costo-Planilla-
    Opex; Tours va Ingreso-Planilla-COSTO-Opex. Ninguna regla produce las tres a
    la vez — por eso el orden vive en `orden_plantilla.json` y no en el codigo.
    Tres intentos de inventar la regla fallaron.
    """
    from app.export.detail_excel import orden_canonico

    orden = orden_canonico()
    assert orden, "no se cargo orden_plantilla.json"

    def clases_de(dept):
        vistas = []
        for (d, cta), i in sorted(orden.items(), key=lambda kv: kv[1]):
            if d != dept:
                continue
            c = {"4": "Revenue", "5": "Cost", "6": "Payroll", "7": "Opex"}.get(cta[0])
            if c and (not vistas or vistas[-1] != c):
                vistas.append(c)
        return vistas

    ab = clases_de("0120")
    tours = clases_de("0150")
    assert ab and tours
    # La prueba del asunto: los dos departamentos NO tienen el mismo orden.
    assert ab != tours, (
        f"A&B {ab} y Tours {tours} coinciden: si siempre coincidieran, el orden "
        f"seria una regla y no haria falta la lista.")


def test_la_distribucion_tiene_su_propio_titulo():
    """Es clase 4 y NO es ingreso. Sin titulo propio salia rotulada «INGRESO»
    una segunda vez dentro del mismo departamento y parecia un error."""
    from app.export.detail_excel import CLASE_TITULO

    assert CLASE_TITULO["Distribucion"] != CLASE_TITULO["Revenue"]
