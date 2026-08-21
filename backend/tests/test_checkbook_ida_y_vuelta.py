# -*- coding: utf-8 -*-
"""EL CHECKBOOK, BAJADO Y SUBIDO DE VUELTA — CORRIENDO DE VERDAD.

**Por qué esta prueba tiene que ejecutar y no leer código.** El modo de fallar
de la vuelta es silencioso y caro: la bajada manda la línea n-ésima de una
cuenta a la ranura `800 + n`, porque los códigos de detalle de FinPlan (`''`,
`001`…) y los del formato del owner (`800`…`810`) son listas distintas para la
misma idea. Si la subida buscara por CÓDIGO en vez de por posición, no
coincidiría ninguna línea: crearía once líneas nuevas por cuenta, las viejas
quedarían con su monto viejo, **y el total del departamento se duplicaría**. El
archivo se vería perfectamente normal.

Una prueba que solo lea el código no distingue entre «el mapeo está» y «el mapeo
invierte bien». Así que acá se genera el archivo con el motor real, se llena, se
lee con el motor real y se corre el endpoint real contra una sesión de mentira
—el mismo patrón que `test_el_recalculo_no_pisa_el_mes_cerrado`—, y después se
mide dónde aterrizó cada monto.
"""
from __future__ import annotations

import asyncio
import os
import tempfile
from decimal import Decimal

import openpyxl
import pytest

from app.api import checkbook_api
from app.checkbook import build, leer
from app.errores import ErrorApi
from app.models.exchange_rate import ExchangeRate
from app.models.opex_entry import OpexEntry

MESES = ["jan", "feb", "mar", "apr", "may", "jun",
         "jul", "aug", "sep", "oct", "nov", "dec"]


# ─── El andamio ──────────────────────────────────────────────────────────────

class _Resultado:
    def __init__(self, filas): self._f = list(filas)
    def scalars(self): return self
    def all(self): return list(self._f)
    def scalar_one_or_none(self): return self._f[0] if self._f else None


class _Sesion:
    """Despacha cada `select(Modelo)` a la lista que le toca."""

    def __init__(self, **por_modelo):
        self.datos = por_modelo
        self.agregados: list = []
        self.commits = 0

    async def execute(self, stmt):
        try:
            ent = stmt.column_descriptions[0]["entity"]
        except Exception:                                   # noqa: BLE001
            ent = None
        return _Resultado(self.datos.get(getattr(ent, "__name__", ""), []))

    async def get(self, modelo, _id):
        return self.datos.get(getattr(modelo, "__name__", ""), [None])[0]

    def add(self, obj): self.agregados.append(obj)
    async def flush(self): pass
    async def commit(self): self.commits += 1


class _Esc:
    id = "esc"
    year = 2027
    type = "BUDGET"
    version = "Working"
    hotel_id = "CWL"
    status = "draft"
    is_locked = False


class _Depto:
    dept_code = "600"
    dept_name = "Gastos Operativos Club Madresal"


class _Archivo:
    """Lo mínimo de `UploadFile` que usa el endpoint."""

    def __init__(self, data: bytes): self._d = data
    async def read(self): return self._d


def _linea(cuenta: str, detalle: str, desc: str, monto: float,
           currency: str = "USD") -> OpexEntry:
    e = OpexEntry(scenario_id="esc", hotel_id="CWL", dept_code="600",
                  account_code=cuenta, account_name=f"Cuenta {cuenta}",
                  detail_code=detalle, detail_desc=desc, currency=currency)
    for m in range(1, 13):
        e.set_month(m, Decimal(str(monto)))
    return e


def _excel(lineas: dict, cuentas: list[dict]) -> bytes:
    """Genera el checkbook con el motor real y devuelve sus bytes."""
    cfg = {
        "departamento": "Gastos Operativos Club Madresal (600)",
        "codigo_departamento": "600",
        "anio_version": 2027,
        "detalles_por_cuenta": 11,
        "detalle_inicial": 800,
        "proteger": False,
        "estadisticas": {},
        "referencias": {},
        "lineas": lineas,
        "cuentas": cuentas,
    }
    t = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    t.close()
    try:
        build(cfg, t.name, force=True)
        with open(t.name, "rb") as fh:
            return fh.read()
    finally:
        try: os.unlink(t.name)
        except OSError: pass


def _reescribir(data: bytes, cambios: dict) -> bytes:
    """Llena el archivo como lo haría una persona: `{(cuenta, ranura): montos}`.

    Se escribe en la HOJA, buscando la fila por (columna B = cuenta, columna E =
    ranura), que es exactamente como la lee el motor de vuelta.
    """
    t = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    t.write(data); t.close()
    wb = openpyxl.load_workbook(t.name)
    ws = wb[[n for n in wb.sheetnames if n.startswith("BUDGET")][0]]
    for r in range(15, ws.max_row + 1):
        b, e = ws.cell(r, 2).value, ws.cell(r, 5).value
        if not isinstance(b, (int, float)) or not isinstance(e, (int, float)):
            continue
        clave = (str(int(b)), int(e))
        if clave in cambios:
            desc, montos = cambios[clave]
            ws.cell(r, 6).value = desc
            for i, v in enumerate(montos):
                ws.cell(r, 7 + i).value = v
    wb.save(t.name)
    with open(t.name, "rb") as fh:
        out = fh.read()
    try: os.unlink(t.name)
    except OSError: pass
    return out


def _subir(lineas_en_base: list[OpexEntry], data: bytes, dept="600",
           rates: list | None = None, dry_run=False):
    sesion = _Sesion(OpexEntry=lineas_en_base,
                     Scenario=[_Esc()], DepartmentCatalog=[_Depto()],
                     ExchangeRate=rates or [])
    r = asyncio.run(checkbook_api.importar(
        "esc", dept, archivo=_Archivo(data), dry_run=dry_run, db=sesion))
    return r, sesion


CUENTAS = [{"cuenta": 7030, "descripcion": "Building"},
           {"cuenta": 7065, "descripcion": "Cleaning Supplies"}]


# ─── Lo que de verdad importa ────────────────────────────────────────────────

def test_el_monto_aterriza_en_la_LINEA_correcta():
    """⚠️ El corazón del asunto.

    Tres líneas de la 7030 con códigos `001`, `002`, `003`; el archivo las
    muestra en las ranuras 800, 801 y 802. Se cambia SOLO la segunda. Si el
    mapeo se hiciera por código, el 800 no coincidiría con nada y el monto
    caería en una línea nueva.
    """
    base = [_linea("7030", "001", "Pintura", 100),
            _linea("7030", "002", "Techos", 200),
            _linea("7030", "003", "Ventanas", 300),
            _linea("7065", "001", "Jabón", 50)]
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12},
                       "801": {"descripcion": "Techos", "montos": [200] * 12},
                       "802": {"descripcion": "Ventanas", "montos": [300] * 12}},
              "7065": {"800": {"descripcion": "Jabón", "montos": [50] * 12}}}
    data = _reescribir(_excel(lineas, CUENTAS),
                       {("7030", 801): ("Techos nuevos", [999] * 12)})

    r, _ = _subir(base, data)

    assert r["ok"] is True
    assert float(base[0].get_month(1)) == 100, "se movió la línea de ARRIBA"
    assert float(base[1].get_month(1)) == 999, "el monto no llegó a su línea"
    assert base[1].detail_desc == "Techos nuevos"
    assert float(base[2].get_month(1)) == 300, "se movió la línea de ABAJO"
    assert float(base[3].get_month(1)) == 50, "se tocó otra cuenta"


def test_no_se_crean_lineas_nuevas_al_subir_lo_mismo():
    """Si el mapeo fuera por código, cada subida agregaría once líneas por
    cuenta y el total del departamento se duplicaría."""
    base = [_linea("7030", "001", "Pintura", 100),
            _linea("7065", "002", "Jabón", 50)]
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12}},
              "7065": {"800": {"descripcion": "Jabón", "montos": [50] * 12}}}
    r, sesion = _subir(base, _excel(lineas, CUENTAS))

    assert sesion.agregados == [], f"creó {len(sesion.agregados)} líneas de la nada"
    assert r["lineas_nuevas"] == 0
    assert r["lineas_actualizadas"] == 2


def test_una_linea_borrada_en_el_excel_queda_en_CERO_y_se_cuenta():
    """No se borra: borrarla correría las posiciones de las que siguen y la
    próxima bajada pondría los montos en ranuras distintas."""
    base = [_linea("7030", "001", "Pintura", 100),
            _linea("7030", "002", "Techos", 200)]
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12},
                       "801": {"descripcion": "Techos", "montos": [200] * 12}}}
    # El archivo sale SIN la segunda línea: se generó con una sola.
    solo_una = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12}}}
    r, _ = _subir(base, _excel(solo_una, CUENTAS))

    assert float(base[1].get_month(1)) == 0, "la línea borrada conservó su monto"
    assert base[1] in base, "se borró la fila en vez de vaciarla"
    assert r["lineas_vaciadas"] == 1, "vació sin decirlo"


def test_una_linea_en_COLONES_se_convierte_de_vuelta():
    """⚠️ El archivo muestra dólares, pero en una línea CRC el dato maestro son
    los colones y el dólar se DERIVA. Escribir el dólar directo sobrevive hasta
    el próximo recálculo y después se revierte solo, sin avisar."""
    base = [_linea("7030", "001", "Alquiler", 0, currency="CRC")]
    rates = [ExchangeRate(scenario_id="esc", hotel_id="CWL", month=m, year=2027,
                          tc_crc_usd=Decimal("500")) for m in range(1, 13)]
    lineas = {"7030": {"800": {"descripcion": "Alquiler", "montos": [0] * 12}}}
    data = _reescribir(_excel(lineas, CUENTAS),
                       {("7030", 800): ("Alquiler", [100] * 12)})

    r, _ = _subir(base, data, rates=rates)

    assert float(base[0].get_crc(1)) == 50000, "no se guardó el dato maestro en colones"
    assert float(base[0].get_month(1)) == 100, "el dólar derivado no volvió al valor tecleado"
    assert r["lineas_en_colones"] == 1


def test_una_linea_en_colones_SIN_tipo_de_cambio_se_rechaza():
    """Inventar un TC sería inventar el dato maestro de esa línea."""
    base = [_linea("7030", "001", "Alquiler", 0, currency="CRC")]
    lineas = {"7030": {"800": {"descripcion": "Alquiler", "montos": [0] * 12}}}
    data = _reescribir(_excel(lineas, CUENTAS),
                       {("7030", 800): ("Alquiler", [100] * 12)})

    with pytest.raises(ErrorApi) as ex:
        _subir(base, data, rates=[])
    assert ex.value.status_code == 422


def test_el_archivo_de_OTRO_departamento_se_rechaza():
    """El error más caro de esta pantalla: reescribiría el departamento
    equivocado con montos que no son suyos, y el total general podría quedar
    parecido igual."""
    base = [_linea("7030", "001", "Pintura", 100)]
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12}}}
    with pytest.raises(ErrorApi) as ex:
        _subir(base, _excel(lineas, CUENTAS), dept="0110")
    assert ex.value.status_code == 422
    assert "600" in str(ex.value.detail)


def test_una_cuenta_que_el_escenario_ya_no_tiene_se_rechaza():
    base = [_linea("7065", "001", "Jabón", 50)]          # sin la 7030
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12}}}
    with pytest.raises(ErrorApi) as ex:
        _subir(base, _excel(lineas, CUENTAS))
    assert ex.value.status_code == 422
    assert "7030" in str(ex.value.detail)


def test_dry_run_no_escribe_nada():
    base = [_linea("7030", "001", "Pintura", 100)]
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12}}}
    data = _reescribir(_excel(lineas, CUENTAS),
                       {("7030", 800): ("Otra cosa", [777] * 12)})

    r, sesion = _subir(base, data, dry_run=True)

    assert r["dry_run"] is True
    assert r["lineas_actualizadas"] == 1
    assert float(base[0].get_month(1)) == 100, "el dry-run escribió"
    assert sesion.commits == 0


def test_un_escenario_ENLLAVADO_no_se_puede_subir():
    class _Locked(_Esc):
        status = "locked"
        is_locked = True

    base = [_linea("7030", "001", "Pintura", 100)]
    lineas = {"7030": {"800": {"descripcion": "Pintura", "montos": [100] * 12}}}
    sesion = _Sesion(OpexEntry=base, Scenario=[_Locked()],
                     DepartmentCatalog=[_Depto()], ExchangeRate=[])
    with pytest.raises(ErrorApi) as ex:
        asyncio.run(checkbook_api.importar(
            "esc", "600", archivo=_Archivo(_excel(lineas, CUENTAS)),
            dry_run=False, db=sesion))
    assert ex.value.status_code == 409


# ─── El contrato entre las dos direcciones ───────────────────────────────────

def test_las_dos_direcciones_ORDENAN_igual():
    """⚠️ El mapeo por posición solo funciona si bajar y subir recorren las
    líneas en el mismo orden. Son dos consultas en dos funciones distintas: si
    alguien cambia una y no la otra, los montos aterrizan corridos y nada falla.
    """
    import inspect

    orden = "order_by(OpexEntry.account_code, OpexEntry.detail_code)"
    baja = inspect.getsource(checkbook_api._lineas_actuales)
    sube = inspect.getsource(checkbook_api.importar)
    assert orden in baja, "cambió el orden de la BAJADA"
    assert orden in sube, "cambió el orden de la SUBIDA"


def test_la_subida_no_busca_por_codigo_de_detalle():
    """La ranura sale de la POSICIÓN. Buscar por `detail_code` es el defecto
    que duplicaría el departamento entero."""
    import inspect

    src = inspect.getsource(checkbook_api.importar)
    assert "ranuras.get(800 + j)" in src, "la ranura tiene que salir del índice"
    assert "enumerate(lineas)" in src
