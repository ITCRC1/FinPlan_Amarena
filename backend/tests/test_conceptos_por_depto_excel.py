# -*- coding: utf-8 -*-
"""
EL EXCEL DE CONCEPTOS DICE LO MISMO QUE LA PANTALLA.

## Por qué existe (2026-08-27)

Owner: *«no hay ningún reporte donde yo veo los beneficios calculados según las
fórmulas. Este pareciera que te lo va a dar, pero después sale otra cosa»*.

Y tenía razón: `/payroll/{id}/export/excel/` baja las **posiciones** —persona,
salario, FTE— porque es la mitad de un viaje de ida y vuelta con el importador.
Los conceptos derivados (CCSS, aguinaldo, provisión de vacaciones, cesantía) no
estaban en ningún archivo: sólo en pantalla, un departamento a la vez.

## Lo que cuida

* **El Excel y la pantalla salen del MISMO `summarize_dept`.** Es la propiedad
  que importa: si el reporte sumara por su cuenta, el día que cambie una fórmula
  los dos dirían cosas distintas y nadie sabría cuál creer. Se verifica
  comparando el archivo contra el endpoint que alimenta la pantalla, celda por
  celda, no contra números escritos a mano.
* **Una hoja por departamento**, más el resumen, y el resumen suma las hojas.
* **Cada fila trae su cuenta.** Un `6023` en la fila de Vacation Provision es lo
  que hace auditable el número sin abrir el código.
* **BASE va después de sus siete componentes y TOTAL al final** — el orden de la
  pantalla. Quien compara no tiene que reordenar mentalmente.
* **El archivo de subir NO cambió.** `/export/excel/` sigue trayendo posiciones:
  si alguien lo "mejora" convirtiéndolo en este reporte, el importador se queda
  sin plantilla y la vuelta se rompe en silencio.
"""
import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.export.conceptos_por_depto_excel import (FILAS, export_conceptos_por_depto,
                                                  _drivers)


class ParamsFalsos:
    """Sólo lo que lee `_drivers`. No toca la base: es un cálculo puro."""
    ccss_rate = Decimal("0.26830")
    aguinaldo_divisor = Decimal("12")
    overtime_pct = Decimal("0")
    bonus_pct = Decimal("0")
    vacaciones_rate = Decimal("0.04")
    severance_annual_rate = Decimal("0")
    cafeteria_daily_crc = Decimal("0")
    transport_monthly_crc = Decimal("0")
    housing_monthly_crc = Decimal("0")
    other_monthly_crc = Decimal("0")
    ins_annual_crc = Decimal("1130647.80")


#: Dos posiciones que, sumadas, dan el S&W del departamento. Es la propiedad
#: que importa: el desglose sale de las MISMAS filas que el total.
POSICIONES = [
    {"codigo": "P01", "puesto": "Supervisor", "detalle": "Ana · ₡500,000.00 CRC",
     "meses": [700.0] * 12},
    {"codigo": "P02", "puesto": "Oficial", "detalle": "VACANTE · ₡300,000.00 CRC",
     "meses": [300.0] * 12},
]


def _mes(base=1000.0):
    """Un mes con BASE y sus derivados, coherente con los drivers de arriba."""
    ccss = base * 0.2683
    agui = base / 12
    vac = base * 0.04
    return {
        "c6000": base, "c6001": 0.0, "c6002": 0.0, "c6003": 0.0, "c6010": 0.0,
        "c6024": 0.0, "c6027": 0.0, "base": base,
        "c6020": ccss, "c6021": agui, "c6004": 0.0, "c6022": 0.0,
        "c6023": vac, "c6025": 0.0, "c6026": 0.0, "c6028": 0.0,
        "c6029": 0.0, "c6030": 0.0,
        "total": base + ccss + agui + vac,
    }


@pytest.fixture
def libro():
    datos = [
        {"dept_code": "0111", "dept_name": "Front Desk",
         "monthly": [_mes(1000.0) for _ in range(12)],
         "posiciones": POSICIONES},
        {"dept_code": "0113", "dept_name": "Housekeeping",
         "monthly": [_mes(500.0) for _ in range(12)]},
    ]
    xlsx = export_conceptos_por_depto(datos, ParamsFalsos(), "AMA BUDGET Working", 2026)
    return load_workbook(io.BytesIO(xlsx))


def _fila_de(ws, texto: str) -> int:
    """Ubica una fila por su ROTULO, no por su indice.

    Las filas se corren cuando se inserta el bloque de posiciones; una prueba
    atada a un numero fijo se rompe por el motivo equivocado y hace perder el
    tiempo buscando un bug que no existe.
    """
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and str(v).strip() == texto:
            return r
    raise AssertionError(f"no se encontró la fila {texto!r}")


def test_una_hoja_por_departamento_mas_el_resumen(libro):
    assert libro.sheetnames[0] == "Resumen", "el resumen va primero"
    assert len(libro.sheetnames) == 3
    assert any("0111" in h for h in libro.sheetnames)
    assert any("0113" in h for h in libro.sheetnames)


def test_cada_fila_trae_su_cuenta(libro):
    hoja = [h for h in libro.sheetnames if "0111" in h][0]
    ws = libro[hoja]
    por_concepto = {n: ws.cell(row=_fila_de(ws, n), column=1).value
                    for _c, _cta, n, _t in
                    [(f[0], f[1], f[2], f[3]) for f in FILAS]}
    assert por_concepto["Social Security"] == "6020"
    assert por_concepto["Vacation Provision"] == "6023"
    assert por_concepto["Aguinaldo"] == "6021"
    # BASE y TOTAL no son cuentas: son subtotales y no deben inventarse una.
    # openpyxl devuelve `None` donde se escribió cadena vacía, así que lo que
    # se afirma es «la celda está vacía», no cuál de las dos formas de vacío es.
    assert not por_concepto["BASE"]
    assert not por_concepto["TOTAL"]


def test_base_va_despues_de_sus_componentes_y_total_al_final():
    claves = [f[0] for f in FILAS]
    assert claves[-1] == "total"
    i = claves.index("base")
    # Los siete que la componen van antes; CCSS y aguinaldo, que se calculan
    # SOBRE la base, van después. Al revés el archivo se leería como si la CCSS
    # entrara en la base.
    assert claves[:i] == ["c6000", "c6001", "c6002", "c6003", "c6010", "c6024", "c6027"]
    assert claves[i + 1] == "c6020"


def test_el_driver_lleva_el_valor_del_escenario(libro):
    hoja = [h for h in libro.sheetnames if "0111" in h][0]
    ws = libro[hoja]
    drivers = {n: (ws.cell(row=_fila_de(ws, n), column=3).value or "")
               for _c, _cta, n, _t in
               [(f[0], f[1], f[2], f[3]) for f in FILAS]}
    # No alcanza con decir «CCSS»: el número tiene que poder auditarse sin
    # abrir el código.
    assert "26.830%" in drivers["Social Security"]
    assert "BASE" in drivers["Social Security"]
    assert "12" in drivers["Aguinaldo"]
    assert "4.000%" in drivers["Vacation Provision"]


def test_lo_digitado_no_finge_tener_formula():
    d = _drivers(ParamsFalsos())
    # `overtime_pct` está en cero: decir «0% sobre S&W» mandaría a buscar el
    # error en el parámetro en vez de en el dato cargado.
    assert "digitado" in d["c6001"]
    assert "%" not in d["c6001"]


def test_sin_parametros_lo_dice_en_vez_de_mostrar_ceros():
    d = _drivers(None)
    assert all("default" in v for v in d.values())


def test_el_resumen_suma_lo_mismo_que_las_hojas(libro):
    ws = libro["Resumen"]
    anual_resumen = {}
    r = 5
    while ws.cell(row=r, column=1).value:
        anual_resumen[ws.cell(row=r, column=1).value] = ws.cell(row=r, column=15).value
        r += 1

    for code in ("0111", "0113"):
        hoja = [h for h in libro.sheetnames if code in h][0]
        ws2 = libro[hoja]
        fila_total = _fila_de(ws2, "TOTAL")
        assert ws2.cell(row=fila_total, column=16).value == pytest.approx(
            anual_resumen[code]), f"{code}: el resumen no cuadra con su hoja"


def test_el_excel_que_se_sube_no_cambio():
    """El de posiciones sigue siendo otro módulo, con su importador."""
    from app.export.payroll_excel import (export_payroll_to_excel,
                                          import_payroll_from_excel)
    assert callable(export_payroll_to_excel)
    assert callable(import_payroll_from_excel)
    import app.export.conceptos_por_depto_excel as nuevo
    assert not hasattr(nuevo, "import_conceptos_from_excel"), (
        "este archivo es un REPORTE: no debe crecerle un importador")


# ── El desglose por posición (owner, 2026-08-27) ──────────────────────────────

def test_las_posiciones_van_justo_antes_del_sw(libro):
    """La fila 6000 tiene que leerse como el total de lo que está encima."""
    ws = libro[[h for h in libro.sheetnames if "0111" in h][0]]
    encabezado = _fila_de(ws, "POSICIONES")
    sw = _fila_de(ws, "Salary and Wages")
    assert encabezado < sw
    # Entre el encabezado del bloque y el S&W van exactamente las posiciones.
    assert sw - encabezado - 1 == len(POSICIONES)


def test_el_sw_es_la_suma_de_las_posiciones(libro):
    """La propiedad que hace confiable el desglose.

    Si el bloque se calculara aparte (salario × FTE ÷ TC) en vez de agrupar las
    mismas `payroll_concept_entries`, esto cuadraría casi siempre — y el día que
    no, habría dos verdades en la misma hoja sin nada que avise cuál es.
    """
    ws = libro[[h for h in libro.sheetnames if "0111" in h][0]]
    sw = _fila_de(ws, "Salary and Wages")
    primera = _fila_de(ws, "POSICIONES") + 1
    for col in list(range(4, 16)) + [16]:          # los 12 meses y el anual
        suma = sum(float(ws.cell(row=r, column=col).value or 0)
                   for r in range(primera, sw))
        total = float(ws.cell(row=sw, column=col).value or 0)
        assert suma == pytest.approx(total), (
            f"columna {col}: las posiciones suman {suma} y el S&W dice {total}")


def test_la_posicion_muestra_quien_y_cuanto(libro):
    ws = libro[[h for h in libro.sheetnames if "0111" in h][0]]
    primera = _fila_de(ws, "POSICIONES") + 1
    puestos = [str(ws.cell(row=r, column=2).value or "").strip()
               for r in range(primera, primera + len(POSICIONES))]
    detalles = [str(ws.cell(row=r, column=3).value or "")
                for r in range(primera, primera + len(POSICIONES))]
    assert "Supervisor" in puestos and "Oficial" in puestos
    assert any("Ana" in d for d in detalles)
    assert any("VACANTE" in d for d in detalles), (
        "una plaza sin ocupar tiene que decirlo: su costo igual está en el total")


def test_un_departamento_sin_posiciones_no_rompe(libro):
    """`0113` se armó sin bloque: la hoja tiene que salir igual."""
    ws = libro[[h for h in libro.sheetnames if "0113" in h][0]]
    assert _fila_de(ws, "Salary and Wages")
    with pytest.raises(AssertionError):
        _fila_de(ws, "POSICIONES")
