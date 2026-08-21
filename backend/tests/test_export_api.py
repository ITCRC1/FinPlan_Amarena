# -*- coding: utf-8 -*-
"""
EL ENDPOINT DE DESCARGA, DE PUNTA A PUNTA.

`test_cuadro_excel` prueba el armador del libro. Esto prueba el camino completo:
lo que la pantalla manda por HTTP → validación → `.xlsx` de vuelta, con el
nombre de archivo y el tipo de contenido que hacen que el navegador lo baje en
vez de mostrarlo.
"""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.export_api import XLSX, router

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")


@pytest.fixture(scope="module")
def cliente():
    from fastapi import FastAPI
    app = FastAPI()
    app.include_router(router, prefix="/api")
    return TestClient(app)


CUERPO = {
    "archivo": "Big_Picture_2027",
    "cuadros": [{
        "titulo": "Budget Big Picture 2027",
        "hoja": "Big Picture",
        "columnas": [
            {"label": "Concepto", "ancho": 48, "formato": "texto"},
            {"label": "Actual 2025", "formato": "usd"},
            {"label": "Crec. %", "formato": "pct"},
        ],
        "filas": [
            {"label": "INGRESOS", "es_total": True, "valores": [None, None]},
            {"label": "Rooms", "nivel": 1, "valores": [3560260.57, 0.092]},
            {"label": "Total Ingresos", "es_total": True, "valores": [4000000, None]},
        ],
    }],
}


def test_devuelve_un_xlsx_que_el_navegador_baja(cliente):
    r = cliente.post("/api/export/cuadros/", json=CUERPO)
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX
    # Sin `attachment`, el navegador intenta mostrar el archivo en vez de bajarlo.
    assert r.headers["content-disposition"].startswith("attachment;")
    assert "Big_Picture_2027" in r.headers["content-disposition"]
    assert ".xlsx" in r.headers["content-disposition"]


def test_el_archivo_llega_entero_y_con_los_numeros_intactos(cliente):
    r = cliente.post("/api/export/cuadros/", json=CUERPO)
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    assert ws.title == "Big Picture"
    assert ws.cell(6, 2).value == 3560260.57      # ni redondeado ni como texto
    assert ws.cell(6, 3).value == 0.092
    assert ws.cell(7, 1).font.bold is True


def test_el_nombre_lleva_la_propiedad(cliente):
    """Cuatro hoteles van a bajar el mismo reporte a la misma carpeta de
    Descargas. Sin el sufijo, el segundo pisa al primero."""
    from app.hotel_actual import hotel_slug
    r = cliente.post("/api/export/cuadros/", json=CUERPO)
    assert hotel_slug() in r.headers["content-disposition"]


def test_un_formato_inventado_se_rechaza_con_su_motivo(cliente):
    """Que falle acá y no en silencio con todas las celdas en formato de dólar."""
    malo = {"archivo": "x", "cuadros": [{
        "titulo": "t",
        "columnas": [{"label": "a", "formato": "euros"}],
        "filas": [{"label": "f", "valores": [1]}]}]}
    r = cliente.post("/api/export/cuadros/", json=malo)
    assert r.status_code == 422
    assert "euros" in r.text


def test_un_formato_inventado_en_la_fila_tambien_se_rechaza(cliente):
    malo = {"archivo": "x", "cuadros": [{
        "titulo": "t",
        "columnas": [{"label": "a", "formato": "usd"}],
        "filas": [{"label": "f", "formato": "quetzales", "valores": [1]}]}]}
    r = cliente.post("/api/export/cuadros/", json=malo)
    assert r.status_code == 422
    assert "quetzales" in r.text


def test_sin_cuadros_no_devuelve_un_libro_vacio(cliente):
    r = cliente.post("/api/export/cuadros/", json={"archivo": "x", "cuadros": []})
    assert r.status_code == 422


def test_un_cuadro_desmedido_se_rechaza_en_vez_de_colgar_el_servidor(cliente):
    from app.api.export_api import MAX_FILAS
    gigante = {"archivo": "x", "cuadros": [{
        "titulo": "t",
        "columnas": [{"label": "a", "formato": "usd"}],
        "filas": [{"label": str(i), "valores": [i]} for i in range(MAX_FILAS + 1)]}]}
    r = cliente.post("/api/export/cuadros/", json=gigante)
    assert r.status_code == 413


def test_un_numero_no_se_convierte_en_texto_al_pasar_por_pydantic(cliente):
    """`valores` acepta `float | str | None`. Si el orden del `|` se invirtiera,
    Pydantic pasaría los números a cadena y el Excel dejaría de sumar — sin dar
    error, que es lo peor que podría pasar acá."""
    cuerpo = {"archivo": "x", "cuadros": [{
        "titulo": "t",
        "columnas": [{"label": "Cuenta", "formato": "texto"},
                     {"label": "Línea del P&L", "formato": "texto"},
                     {"label": "Monto", "formato": "usd"}],
        "filas": [{"label": "7065", "valores": ["OPEX_ROOMS", 1234.5]}]}]}
    r = cliente.post("/api/export/cuadros/", json=cuerpo)
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    assert ws.cell(5, 2).value == "OPEX_ROOMS"
    assert ws.cell(5, 3).value == 1234.5
    assert isinstance(ws.cell(5, 3).value, float)


def test_una_fila_con_un_valor_de_mas_se_rechaza_en_vez_de_perderlo(cliente):
    """La primera columna es la etiqueta, así que a `valores` le tocan
    `len(columnas) - 1` celdas. Antes, el valor que sobraba se caía sin avisar —
    y una columna que está en pantalla pero no en el Excel es justo el defecto
    que todo esto viene a corregir."""
    cuerpo = {"archivo": "x", "cuadros": [{
        "titulo": "Mapeo",
        "columnas": [{"label": "Cuenta", "formato": "texto"},
                     {"label": "Monto", "formato": "usd"}],
        "filas": [{"label": "7065", "valores": ["OPEX_ROOMS", 1234.5]}]}]}
    r = cliente.post("/api/export/cuadros/", json=cuerpo)
    assert r.status_code == 422
    assert "7065" in r.text          # dice CUÁL fila, no solo que algo falló


def test_una_fila_corta_sigue_valiendo(cliente):
    """Las bandas de sección mandan `valores` vacío a propósito."""
    cuerpo = {"archivo": "x", "cuadros": [{
        "titulo": "t",
        "columnas": [{"label": "a", "formato": "texto"},
                     {"label": "b", "formato": "usd"},
                     {"label": "c", "formato": "usd"}],
        "filas": [{"label": "INGRESOS", "es_total": True, "valores": []},
                  {"label": "Rooms", "valores": [100]}]}]}
    assert cliente.post("/api/export/cuadros/", json=cuerpo).status_code == 200
