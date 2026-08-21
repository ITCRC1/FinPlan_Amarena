# -*- coding: utf-8 -*-
"""
EXCEL DE CONCEPTOS MANUALES, por departamento y posición.

No todo beneficio sale de una fórmula: horas extra, bono, cesantía, transporte,
vivienda y otros se negocian caso por caso. El owner los baja, los llena y los
sube. Lo que se fija aquí es el viaje completo: lo que se escribe es lo que se lee.
"""
from decimal import Decimal

import pytest

from app.export.beneficios_excel import (
    CONCEPTOS_MANUALES, export_beneficios, import_beneficios,
)


def _posiciones(n=3):
    return [{
        "dept_code": "0111" if i < 2 else "0200",
        "dept_name": "FRONT DESK" if i < 2 else "MANTENIMIENTO",
        "position_code": f"P{i:02d}",
        "position_name": f"PUESTO {i}",
        "employee_name": f"EMPLEADO {i}",
        "meses": [Decimal("0")] * 12,
    } for i in range(n)]


def _archivo(valores=None):
    datos = {c: _posiciones() for c, _, _ in CONCEPTOS_MANUALES}
    if valores:
        for col, i, mes, v in valores:
            datos[col][i]["meses"][mes] = Decimal(str(v))
    return export_beneficios(datos, "CWL Working", 2027)


def test_trae_una_hoja_por_concepto_manual():
    from openpyxl import load_workbook
    from io import BytesIO
    wb = load_workbook(BytesIO(_archivo()))
    for _, titulo, _ in CONCEPTOS_MANUALES:
        assert titulo[:31] in wb.sheetnames
    assert "Guia" in wb.sheetnames


def test_lo_que_se_escribe_es_lo_que_se_lee():
    xlsx = _archivo([("c6001_overtime", 0, 0, 150.50),
                     ("c6029_transport", 2, 11, 75.25)])
    datos, avisos = import_beneficios(xlsx)
    ot = datos["c6001_overtime"]
    assert len(ot) == 1                       # solo la fila con monto
    assert ot[0]["dept_code"] == "0111"
    assert ot[0]["position_code"] == "P00"
    assert ot[0]["meses"][0] == Decimal("150.50")
    tr = datos["c6029_transport"]
    assert tr[0]["position_code"] == "P02"
    assert tr[0]["meses"][11] == Decimal("75.25")
    assert not avisos


def test_las_filas_en_cero_no_viajan():
    """Un archivo sin llenar no debe borrar nada ni generar ruido."""
    datos, avisos = import_beneficios(_archivo())
    assert all(not filas for filas in datos.values())
    assert not avisos


def test_los_renglones_de_departamento_no_se_leen_como_puestos():
    datos, _ = import_beneficios(_archivo([("c6026_severance", 2, 3, 40)]))
    filas = datos["c6026_severance"]
    assert len(filas) == 1
    assert filas[0]["dept_code"] == "0200"    # no el renglón "0200 — MANTENIMIENTO"


@pytest.mark.parametrize("cuenta", ["c6000_sw", "c6020_ccss", "c6021_aguinaldo",
                                    "c6022_occ_hazard", "c6025_cafeteria"])
def test_lo_que_es_formula_o_reparto_no_esta_en_el_excel(cuenta):
    """6000/6020/6021 son fórmula; 6022 y 6025 son reparto. Digitarlos seria
    pisarlos en el siguiente recálculo, o contarlos dos veces."""
    assert cuenta not in [c for c, _, _ in CONCEPTOS_MANUALES]


def test_estan_los_que_el_owner_pidio():
    cols = [c for c, _, _ in CONCEPTOS_MANUALES]
    for c in ("c6001_overtime", "c6027_incentive_bonus", "c6026_severance",
              "c6029_transport", "c6028_housing", "c6030_other"):
        assert c in cols
